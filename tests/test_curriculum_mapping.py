"""Faza 2 kurikularnog mapiranja — deterministički testovi builda.

Sve nad LOKALNIM fajlovima (kanonski workbook, Faza 1, generisana Faza 2):
nula mrežnih poziva, nula modela. Testovi štite:
  • učitavanje oba ulazna workbooka i rekonsilijaciju 534 lekcije;
  • ćirilica→latinica normalizaciju i grubi stemer (parovi koje MORA spojiti);
  • stabilnost ID-jeva lekcija i netaknutost kanonskog workbooka;
  • enum-e mapiranja, postojanje ciljeva, nijednu tiho odbačenu stavku;
  • determinizam (bajt-identičan izlaz) i neovisnost od redoslijeda redova;
  • čuvanje dvosmislenosti kao needs_review (nikad tihi izbor);
  • primjere exact/neighbour/no_match ponašanja (živi gate slučaj djeljivosti);
  • pokrivenost pilot lista (25 lekcija dvije pilot oblasti).
"""
import hashlib
import json
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "scripts"))

import build_curriculum_mapping as bcm  # noqa: E402

PHASE2_PATH = ROOT / "reference" / "curriculum" / "semantics" / "MATBOT_Faza2_Mapiranje.xlsx"


def _sha(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


@pytest.fixture(scope="module")
def pipeline():
    """Jedan in-memory prolaz cijelog cjevovoda, dijeljen svim testovima."""
    lessons = bcm.load_lessons()
    items = bcm.load_items()
    bcm.prepare_items(items)
    mappings, issues = bcm.build_mappings(items, lessons)
    return lessons, items, mappings, issues


# ---------------------------------------------------------------------------
# Učitavanje i rekonsilijacija
# ---------------------------------------------------------------------------

def test_both_input_workbooks_load_with_expected_counts(pipeline):
    lessons, items, _mappings, _issues = pipeline
    assert len(lessons) == 534
    assert len({l.lesson_id for l in lessons}) == 534
    assert len(items) == 573
    assert len({i.item_id for i in items}) == 573


def test_canonical_lessons_reconcile_with_topics_json(pipeline):
    lessons, _items, _m, _i = pipeline
    data = json.loads((ROOT / "data" / "topics.json").read_text(encoding="utf-8"))
    json_ids = {
        (lesson["id"], str(grade), lesson["title"])
        for grade, grade_data in data["grades"].items()
        for lesson in grade_data["lessons"]
    }
    assert {(l.lesson_id, str(l.grade), l.title) for l in lessons} == json_ids


def test_item_totals_match_phase1_kontrola(pipeline):
    _lessons, items, _m, _i = pipeline
    from collections import Counter
    actual = Counter((i.source_id, i.grade) for i in items)
    assert dict(actual) == bcm.EXPECTED_ITEMS_BY_SOURCE_GRADE


def test_canonical_workbook_not_modified_by_build(tmp_path):
    before = _sha(bcm.CANONICAL_XLSX)
    bcm.main(["--out", str(tmp_path / "out.xlsx")])
    assert _sha(bcm.CANONICAL_XLSX) == before
    assert _sha(bcm.PHASE1_XLSX)  # ulaz Faze 1 i dalje postoji i čitljiv je


# ---------------------------------------------------------------------------
# Normalizacija pisma i stemer
# ---------------------------------------------------------------------------

def test_cyrillic_transliteration_covers_digraphs_and_diacritics():
    assert bcm.transliterate("угао") == "ugao"
    assert bcm.transliterate("дјељивост") == "djeljivost"
    assert bcm.transliterate("Џеп љуљашка њива") == "Džep ljuljaška njiva"
    assert bcm.transliterate("ћирилица чиста шђж") == "ćirilica čista šđž"
    # Latinica prolazi netaknuta.
    assert bcm.transliterate("razlomak π") == "razlomak π"


@pytest.mark.parametrize("a, b", [
    ("sabiranje", "sabirati"),
    ("upoređivanje", "uporediti"),
    ("proširivanje", "proširiti"),
    ("množenje", "množiti"),
    ("djelilac", "djelioca"),
    ("sadržilac", "višekratnik"),
    ("kut", "ugao"),
    ("nazivnik", "imenilac"),
    ("brojnik", "brojilac"),
    ("jednadžba", "jednačina"),
    ("kriterijumima", "pravila"),
    ("četvorougao", "četverougao"),
    ("zamjene", "supstitucije"),
])
def test_stemmer_and_aliases_join_required_pairs(a, b):
    fold = lambda word: bcm._fold(bcm.stem(word.lower()))
    assert fold(a) == fold(b), (fold(a), fold(b))


def test_generic_words_never_carry_a_mapping_alone():
    for word in ("zadatak", "primjena", "pojam", "računanje", "rješavanje",
                 "svojstva", "definisati", "učenik"):
        tokens = bcm.tokenize(word)
        assert tokens and tokens[0].kind == "generic", word


# ---------------------------------------------------------------------------
# Invarijante mapiranja
# ---------------------------------------------------------------------------

def test_every_phase1_item_has_at_least_one_row(pipeline):
    _l, items, mappings, _i = pipeline
    covered = {m.item.item_id for m in mappings}
    assert covered == {i.item_id for i in items}


def test_mapping_enums_and_target_existence(pipeline):
    lessons, _items, mappings, _i = pipeline
    lesson_ids = {l.lesson_id for l in lessons}
    ids = set()
    for m in mappings:
        assert m.mapping_id not in ids
        ids.add(m.mapping_id)
        assert m.review_status in bcm.REVIEW_STATUSES
        assert m.method in bcm.MAPPING_METHODS
        if m.lesson is None:
            assert m.review_status == "no_match"
            assert m.relation == ""
        else:
            assert m.relation in bcm.RELATIONS
            assert m.confidence in bcm.CONFIDENCES
            assert m.lesson.lesson_id in lesson_ids


def test_no_match_rows_only_for_items_without_target(pipeline):
    _l, _items, mappings, _i = pipeline
    from collections import defaultdict
    per_item = defaultdict(list)
    for m in mappings:
        per_item[m.item.item_id].append(m)
    for rows in per_item.values():
        no_match = [m for m in rows if m.review_status == "no_match"]
        if no_match:
            assert len(rows) == 1  # no_match je jedini red te stavke


def test_ambiguity_is_preserved_as_needs_review(pipeline):
    _l, _items, mappings, _i = pipeline
    ambiguous = [m for m in mappings
                 if "više uvjerljivih lekcija" in (m.ambiguity_note or "")]
    assert ambiguous, "očekivan bar jedan dokumentovan izjednačen slučaj"
    assert all(m.review_status in ("needs_review", "conflict") for m in ambiguous)


def test_non_content_sections_are_never_force_mapped(pipeline):
    _l, _items, mappings, _i = pipeline
    for m in mappings:
        if bcm.SECTION_POLICY.get(m.item.section) == "non_content":
            assert m.lesson is None
            assert m.review_status == "no_match"


def test_methodology_sections_cap_at_supporting(pipeline):
    _l, _items, mappings, _i = pipeline
    for m in mappings:
        if (bcm.SECTION_POLICY.get(m.item.section) == "methodology"
                and m.lesson is not None):
            assert m.relation in ("supporting", "prerequisite", "advanced",
                                  "excluded", "neighbour")


# ---------------------------------------------------------------------------
# Primjeri ponašanja (živi slučaj release gatea: pravila djeljivosti)
# ---------------------------------------------------------------------------

def _rows_for(mappings, item_id):
    return [m for m in mappings if m.item.item_id == item_id]


def test_divisibility_rules_item_maps_exact_to_the_rules_lesson(pipeline):
    _l, _items, mappings, _i = pipeline
    rows = _rows_for(mappings, "KS_2018-0040")   # „primjenjivati pravila za djeljivost…“
    exact = [m for m in rows if m.relation == "exact"]
    assert [m.lesson.lesson_id for m in exact] == ["6-03-004"]
    assert exact[0].confidence == "high"


def test_rs_kriterijumi_djeljivosti_reaches_rules_lesson_via_alias(pipeline):
    _l, _items, mappings, _i = pipeline
    rows = _rows_for(mappings, "RS_2014-0022")
    exact = [m for m in rows if m.relation == "exact"]
    assert any(m.lesson.lesson_id == "6-03-004" for m in exact)
    assert any(m.method == "terminology_alias" for m in exact)


def test_divisibility_confusion_pair_documents_neighbour(pipeline):
    _l, _items, mappings, _i = pipeline
    rows = _rows_for(mappings, "KS_2018-0040")
    neighbours = {m.lesson.lesson_id for m in rows if m.relation == "neighbour"}
    # Traženje djelilaca je susjedna, lako zamjenjiva vještina — dokumentovano.
    assert "6-03-001" in neighbours


def test_general_goal_item_is_no_match(pipeline):
    _l, _items, mappings, _i = pipeline
    rows = _rows_for(mappings, "RS_2014-0001")   # opšti cilj, nije gradivo
    assert len(rows) == 1
    assert rows[0].lesson is None
    assert rows[0].review_status == "no_match"


def test_substitution_item_does_not_map_to_graphical_method(pipeline):
    _l, _items, mappings, _i = pipeline
    rows = _rows_for(mappings, "KS_2018-0372")
    exact_targets = {m.lesson.lesson_id for m in rows if m.relation == "exact"}
    assert "9-05-007" in exact_targets
    assert "9-05-006" not in exact_targets


# ---------------------------------------------------------------------------
# Determinizam
# ---------------------------------------------------------------------------

def test_output_is_byte_reproducible(tmp_path):
    out1, out2 = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    bcm.main(["--out", str(out1)])
    bcm.main(["--out", str(out2)])
    assert _sha(out1) == _sha(out2)


def test_mapping_result_is_independent_of_item_row_order(pipeline):
    lessons, items, mappings, _i = pipeline
    reordered = list(reversed(items))
    remapped, _issues = bcm.build_mappings(reordered, lessons)

    def signature(ms):
        return sorted(
            (m.mapping_id, m.item.item_id,
             m.lesson.lesson_id if m.lesson else "", m.relation,
             m.confidence, m.method, m.review_status)
            for m in ms
        )

    assert signature(mappings) == signature(remapped)


# ---------------------------------------------------------------------------
# Generisani workbook Faze 2 (commitovani artefakt)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def phase2_workbook():
    assert PHASE2_PATH.exists(), "Faza 2 workbook nije izgrađen"
    wb = openpyxl.load_workbook(PHASE2_PATH, read_only=True, data_only=True)
    yield wb
    wb.close()

def test_phase2_workbook_has_all_required_sheets(phase2_workbook):
    assert phase2_workbook.sheetnames == [
        "README", "Izvori", "Lekcije_534", "Stavke_NPP", "Mapiranje",
        "Terminologija", "Praznine_i_sukobi", "Pokrivenost", "Pilot_25",
        "Kontrola",
    ]


def test_phase2_lessons_sheet_preserves_all_534(phase2_workbook):
    rows = list(phase2_workbook["Lekcije_534"].iter_rows(values_only=True))[1:]
    assert len(rows) == 534
    data = json.loads((ROOT / "data" / "topics.json").read_text(encoding="utf-8"))
    json_ids = {
        (lesson["id"], lesson["title"])
        for grade_data in data["grades"].values()
        for lesson in grade_data["lessons"]
    }
    assert {(r[0], r[3]) for r in rows} == json_ids


def test_phase2_items_sheet_preserves_all_573(phase2_workbook):
    rows = list(phase2_workbook["Stavke_NPP"].iter_rows(values_only=True))[1:]
    assert len(rows) == 573
    assert all(row[7] for row in rows)   # originalni tekst sačuvan


def test_phase2_kontrola_has_no_fail(phase2_workbook):
    rows = list(phase2_workbook["Kontrola"].iter_rows(values_only=True))[1:]
    statuses = {row[2] for row in rows}
    assert "FAIL" not in statuses
    assert "PASS" in statuses


def test_phase2_pilot_covers_exactly_the_25_pilot_lessons(phase2_workbook):
    rows = list(phase2_workbook["Pilot_25"].iter_rows(values_only=True))[1:]
    assert len(rows) == 25
    data = json.loads((ROOT / "data" / "topics.json").read_text(encoding="utf-8"))
    expected = {
        lesson["id"]
        for lesson in data["grades"]["6"]["lessons"]
        if lesson["oblast"] in ("Djeljivost brojeva", "Razlomci")
    }
    assert {r[0] for r in rows} == expected
    priorities = {r[12] for r in rows}
    assert priorities <= {1, 2, 3}


def test_phase2_terminology_preserves_cyrillic_and_explicit_aliases(phase2_workbook):
    rows = list(phase2_workbook["Terminologija"].iter_rows(values_only=True))[1:]
    scripts = {row[2] for row in rows}
    assert "ćirilica" in scripts and "latinica" in scripts
    pairs = {(row[0], row[1]) for row in rows}
    assert ("ugao", "kut") in pairs                # eksplicitni KS alias sačuvan
    assert any(row[2] == "ćirilica" and row[0] == "djeljivost" for row in rows)


def test_phase2_mapping_sheet_enums_are_valid(phase2_workbook):
    rows = list(phase2_workbook["Mapiranje"].iter_rows(values_only=True))[1:]
    assert rows
    for row in rows:
        relation, confidence, method, status = row[12], row[13], row[14], row[17]
        assert status in bcm.REVIEW_STATUSES
        assert method in bcm.MAPPING_METHODS
        if row[8]:  # ima cilj
            assert relation in bcm.RELATIONS
            assert confidence in bcm.CONFIDENCES
        else:
            assert status == "no_match"
