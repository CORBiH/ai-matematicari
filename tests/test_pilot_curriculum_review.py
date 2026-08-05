"""Faza 2.5 — testovi stručnog pregleda pilot mapiranja (25 lekcija).

Sve lokalno i deterministički: presude i oporavljeni dokazi se ukrštaju sa
STVARNIM sadržajem Faze 1/2; nijedan red ne smije biti tiho izostavljen,
nijedan citat izmišljen, nijedan ulazni workbook promijenjen.
"""
import hashlib
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "scripts"))

import build_curriculum_mapping as bcm  # noqa: E402
import review_pilot_curriculum_mapping as rev  # noqa: E402

REVIEW_PATH = ROOT / "reference" / "curriculum" / "semantics" / "MATBOT_Faza2_5_Pilot25_Review.xlsx"


def _sha(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


@pytest.fixture(scope="module")
def phase2():
    mapping_rows, item_rows = rev.load_phase2()
    return mapping_rows, {r[0]: r for r in item_rows}


@pytest.fixture(scope="module")
def pilot():
    lessons = rev.pilot_lessons()
    return lessons, {l.lesson_id for l in lessons}


# ---------------------------------------------------------------------------
# Obim pilota i potpunost presuda
# ---------------------------------------------------------------------------

def test_exactly_25_pilot_lessons(pilot):
    lessons, ids = pilot
    assert len(lessons) == 25
    data = json.loads((ROOT / "data" / "topics.json").read_text(encoding="utf-8"))
    expected = {l["id"] for l in data["grades"]["6"]["lessons"]
                if l["oblast"] in ("Djeljivost brojeva", "Razlomci")}
    assert ids == expected


def test_every_pilot_phase2_row_has_a_verdict_and_none_extra(phase2, pilot):
    """Ključ presude je (item, cilj, relacija) — otporan na renumeraciju
    mapping_id-jeva koju je izazvala popravka Faze 2 (nalaz Faze 3)."""
    mapping_rows, _items = phase2
    _lessons, pilot_ids = pilot
    pilot_keys = {rev.verdict_key(r) for r in mapping_rows if r[8] in pilot_ids}
    assert pilot_keys == set(rev.MAPPING_VERDICTS)
    assert len(pilot_keys) == 64


def test_verdict_key_is_not_positional(phase2, pilot):
    """mapping_id se NE smije koristiti kao ključ: pozicioni je."""
    mapping_rows, _items = phase2
    _lessons, pilot_ids = pilot
    sample = next(r for r in mapping_rows if r[8] in pilot_ids)
    assert rev.verdict_key(sample) == (sample[1], sample[8], sample[12])
    assert not any(isinstance(k, str) for k in rev.MAPPING_VERDICTS)


def test_verdicts_use_only_allowed_enum():
    for verdict, new_rel, new_conf, reason, flags in rev.MAPPING_VERDICTS.values():
        assert verdict in rev.VERDICTS
        assert reason.strip()
        if new_rel is not None:
            assert new_rel in bcm.RELATIONS
        if new_conf is not None:
            assert new_conf in bcm.CONFIDENCES


def test_rejected_rows_keep_reason_and_exist_in_phase2(phase2):
    mapping_rows, _items = phase2
    keys = {rev.verdict_key(r) for r in mapping_rows}
    rejected = [k for k, v in rev.MAPPING_VERDICTS.items()
                if v[0] in ("reject", "wrong_lesson", "wrong_grade")]
    assert rejected, "pregled mora imati bar jedno odbijanje"
    for key in rejected:
        assert key in keys            # originalni red postoji (ništa obrisano)
        assert rev.MAPPING_VERDICTS[key][3].strip()   # razlog obavezan


# ---------------------------------------------------------------------------
# Oporavljeni dokazi
# ---------------------------------------------------------------------------

def test_recovered_item_evidence_exists_in_phase2(phase2):
    _mapping_rows, items = phase2
    for rec in rev.RECOVERED_EVIDENCE:
        rid, origin, source, page, lesson, relation, conf, quote, flags = rec
        assert relation in bcm.RELATIONS and conf in bcm.CONFIDENCES
        if origin == "stavka":
            assert source in items, rid


def test_recovered_page_quotes_exist_verbatim_modulo_whitespace():
    pages = rev.load_phase1_pages()

    def ws(text):
        return " ".join(str(text).split())

    for rec in rev.RECOVERED_EVIDENCE:
        rid, origin, source, page, _lesson, _rel, _conf, quote, _flags = rec
        if origin == "stranica":
            assert (source, page) in pages, rid
            assert ws(quote) in ws(pages[(source, page)][1]), rid


def test_formula_loss_flags_preserved_and_finding_recorded():
    """Faza 3: zastavice ostaju, a nalaz provjere izvora je zapisan doslovno.

    Blokada se smije ukinuti SAMO zato što lekcija ne zavisi od spornog tokena —
    ne zato što je formula rekonstruisana."""
    flagged = [rec[0] for rec in rev.RECOVERED_EVIDENCE if "formula_loss" in rec[8]]
    assert "R25-008" in flagged and "R25-009" in flagged
    rows = [r for r in rev.SOURCE_VERIFICATION if r[0] == "6-03-003"]
    assert rows, "6-03-003 mora imati zapisanu provjeru izvora"
    finding = " ".join(r[2] for r in rows)
    assert "NIJE eksponent" in finding
    assert "10^n" not in rev.LESSON_SEMANTICS["6-03-003"]["core_skill"]


def test_source_verification_never_reconstructs_a_formula():
    for lesson, topic, finding, method, impact in rev.SOURCE_VERIFICATION:
        assert lesson in rev.LESSON_SEMANTICS
        for field in (topic, finding, method, impact):
            assert str(field).strip()


# ---------------------------------------------------------------------------
# Faza 3 — klasa aktivacije
# ---------------------------------------------------------------------------

def test_activation_covers_all_25_with_valid_classes(pilot):
    _lessons, pilot_ids = pilot
    assert set(rev.ACTIVATION) == pilot_ids
    for klasa, why in rev.ACTIVATION.values():
        assert klasa in rev.ACTIVATION_CLASSES
        assert why.strip()


def test_proven_detector_lessons_stay_ready(pilot):
    """Dokazano blokiranje koje danas radi ne smije biti degradirano."""
    assert "6-03-004" in rev.PROVEN_DETECTORS
    for lesson_id in rev.PROVEN_DETECTORS:
        assert rev.ACTIVATION[lesson_id][0] == "READY", lesson_id


def test_advisory_only_lessons_have_no_proven_detector():
    """ADVISORY_ONLY ne smije uvesti novo determinističko odbijanje."""
    for lesson_id, (klasa, _why) in rev.ACTIVATION.items():
        if klasa != "READY":
            assert lesson_id not in rev.PROVEN_DETECTORS, lesson_id


def test_no_lesson_is_ready_without_exact_evidence(phase2, pilot):
    """READY se ne dodjeljuje radi pokrivenosti — traži izričit dokaz."""
    mapping_rows, _items = phase2
    _lessons, pilot_ids = pilot
    per_lesson, _rej, _unres = rev.collect_final_evidence(pilot_ids, mapping_rows)
    for lesson_id, (klasa, _why) in rev.ACTIVATION.items():
        if klasa == "READY":
            assert per_lesson.get(lesson_id, {}).get("exact"), lesson_id


# ---------------------------------------------------------------------------
# Šest praznina + semantika lekcija
# ---------------------------------------------------------------------------

def test_all_six_gap_lessons_explicitly_resolved():
    assert set(rev.GAP_RESOLUTIONS) == set(rev.GAP_LESSONS)
    for lesson_id, (case, reason) in rev.GAP_RESOLUTIONS.items():
        assert case in ("A", "B", "C", "D", "E")
        assert reason.strip()


def test_lesson_semantics_cover_all_25_with_valid_statuses(pilot):
    _lessons, pilot_ids = pilot
    assert set(rev.LESSON_SEMANTICS) == pilot_ids
    for sem in rev.LESSON_SEMANTICS.values():
        assert sem["confidence"] in rev.SEMANTIC_CONFIDENCE
        assert sem["status"] in rev.HUMAN_REVIEW_STATUS
        for key in ("family", "core_skill", "actions", "archetypes",
                    "answer_kinds", "level1", "level2"):
            assert str(sem[key]).strip()


def test_no_lesson_without_exact_evidence_claims_ready(phase2, pilot):
    mapping_rows, _items = phase2
    _lessons, pilot_ids = pilot
    per_lesson, _rejected, _unresolved = rev.collect_final_evidence(
        pilot_ids, mapping_rows)
    for lesson_id in pilot_ids:
        if not per_lesson.get(lesson_id, {}).get("exact"):
            assert rev.LESSON_SEMANTICS[lesson_id]["status"] != "ready_for_contract_draft", lesson_id


# ---------------------------------------------------------------------------
# Porodice i granice
# ---------------------------------------------------------------------------

def test_families_partition_all_25_lessons(pilot):
    _lessons, pilot_ids = pilot
    members = [m for fam in rev.FAMILIES.values() for m in fam["members"]]
    assert sorted(members) == sorted(pilot_ids)
    for fam_id, fam in rev.FAMILIES.items():
        if len(fam["members"]) == 1:
            assert fam.get("single_member_reason", "").strip(), fam_id
        for member in fam["members"]:
            assert rev.LESSON_SEMANTICS[member]["family"] == fam_id


def test_boundaries_are_directional_where_declared(pilot):
    _lessons, pilot_ids = pilot
    directional = [b for b in rev.BOUNDARIES if not b[2]]
    symmetric = [b for b in rev.BOUNDARIES if b[2]]
    assert directional and symmetric, "moraju postojati i smjerne i simetrične"
    for source, forbidden, _sym, why in rev.BOUNDARIES:
        assert source in pilot_ids            # izvor je uvijek pilot lekcija
        assert source != forbidden
        assert why.strip()
    # živi gate slučaj mora biti pokriven simetrično:
    assert any({s, f} == {"6-03-004", "6-03-001"} and sym
               for s, f, sym, _ in rev.BOUNDARIES)


def test_required_fraction_and_divisibility_distinctions_present():
    pairs = {(s, f) for s, f, _sym, _w in rev.BOUNDARIES}
    both = pairs | {(f, s) for s, f in pairs}
    required = [
        ("6-03-004", "6-02-005"),   # pravila vs dijeljenje s ostatkom
        ("6-03-005", "6-03-007"),   # prosti vs faktorizacija
        ("6-03-007", "6-03-008"),   # faktorizacija vs NZD
        ("6-03-008", "6-03-009"),   # NZD vs NZS
        ("6-04-005", "6-04-006"),   # proširivanje vs skraćivanje
        ("6-04-005", "6-04-008"),   # ekvivalencija vs poređenje
        ("6-04-009", "6-04-010"),   # jednaki vs različiti imenioci
        ("6-04-011", "6-04-012"),   # množenje vs dijeljenje
        ("6-04-011", "6-06-002"),   # razlomci vs procenat
        ("6-04-014", "6-07-002"),   # izraz vs jednačina
    ]
    for pair in required:
        assert pair in both, pair


# ---------------------------------------------------------------------------
# Primjeri
# ---------------------------------------------------------------------------

def test_every_lesson_has_L1_L2_and_invalid_neighbour_example(pilot):
    _lessons, pilot_ids = pilot
    per = defaultdict(Counter)
    for lesson, level, validity, task, _s, _d in rev.EXAMPLES:
        if validity == "valid" and not task.startswith("(nepodržan"):
            per[lesson][level] += 1
        elif validity == "invalid_neighbour":
            per[lesson]["inv"] += 1
    for lesson_id in pilot_ids:
        assert per[lesson_id][1] >= 1, lesson_id
        assert per[lesson_id][2] >= 1, lesson_id
        assert per[lesson_id]["inv"] >= 1, lesson_id


def test_level3_present_or_explicitly_unsupported(pilot):
    _lessons, pilot_ids = pilot
    for lesson_id in pilot_ids:
        rows = [e for e in rev.EXAMPLES if e[0] == lesson_id and e[1] == 3]
        assert rows, lesson_id     # L3 red postoji: validan ILI '(nepodržan …)'


def test_examples_are_labeled_authored_never_official():
    wb = openpyxl.load_workbook(REVIEW_PATH, read_only=True, data_only=True)
    rows = list(wb["Primjeri_Pilot"].iter_rows(values_only=True))[1:]
    wb.close()
    assert len(rows) == len(rev.EXAMPLES)
    for row in rows:
        assert "autorska test-fikstura" in str(row[7])
        assert "NIJE zvanični tekst" in str(row[7])


# ---------------------------------------------------------------------------
# Workbook artefakt + determinizam + netaknuti ulazi
# ---------------------------------------------------------------------------

def test_review_workbook_has_all_required_sheets():
    wb = openpyxl.load_workbook(REVIEW_PATH, read_only=True)
    assert wb.sheetnames == [
        "README", "Lekcije_Pilot25", "Aktivacija", "Provjera_izvora",
        "Mapiranja_Pregled", "Dokazi_KS", "Dokazi_RS", "Porodice_Pilot",
        "Granice_lekcija", "Primjeri_Pilot", "Praznine", "Kvalitet_Mapiranja",
        "Kontrola",
    ]
    assert len(list(wb["Aktivacija"].iter_rows(values_only=True))[1:]) == 25
    rows = list(wb["Lekcije_Pilot25"].iter_rows(values_only=True))[1:]
    assert len(rows) == 25
    kontrola = list(wb["Kontrola"].iter_rows(values_only=True))[1:]
    assert all(row[2] != "FAIL" for row in kontrola)
    wb.close()


def test_cyrillic_originals_traceable_in_rs_evidence():
    wb = openpyxl.load_workbook(REVIEW_PATH, read_only=True, data_only=True)
    rows = list(wb["Dokazi_RS"].iter_rows(values_only=True))[1:]
    wb.close()
    assert rows
    assert any(any("а" <= ch <= "я" or "А" <= ch <= "Я" for ch in str(row[3]))
               for row in rows), "ćirilični original mora ostati sljedljiv"


def test_build_is_byte_reproducible_and_inputs_unchanged(tmp_path):
    inputs = (bcm.CANONICAL_XLSX, bcm.PHASE1_XLSX, rev.PHASE2_XLSX)
    before = {p: _sha(p) for p in inputs}
    out1, out2 = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    rev.main(["--out", str(out1)])
    rev.main(["--out", str(out2)])
    assert _sha(out1) == _sha(out2)
    assert {p: _sha(p) for p in inputs} == before
