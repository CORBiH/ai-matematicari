"""Faza 2.5: stručni pregled mapiranja za 25 pilot lekcija (6. razred:
Djeljivost brojeva + Razlomci).

    python scripts/review_pilot_curriculum_mapping.py            # izgradi workbook
    python scripts/review_pilot_curriculum_mapping.py --report   # + sažetak
    python scripts/review_pilot_curriculum_mapping.py --dry-run  # bez pisanja

ULAZI (samo čitanje): kanonski workbook, Faza 1, Faza 2, data/topics.json.
IZLAZ: reference/curriculum/semantics/MATBOT_Faza2_5_Pilot25_Review.xlsx

ŠTA JE OVAJ FAJL: ručno donesene stručne PRESUDE nad automatskim mapiranjem
Faze 2, zapisane kao deklarativne tabele. Skripta presude UKRŠTA sa stvarnim
sadržajem Faze 1/2 (svaki referencirani ID mora postojati, nijedan red Faze 2
za pilot ne smije ostati bez presude), računa metrike kvaliteta i emituje
workbook deterministički (bajt-identičan pri ponovnom pokretanju).

Granice: nikakva produkcijska logika, nikakav lesson_semantics.json, nikakvo
aktiviranje ugovora; autorski primjeri zadataka su TEST-FIKSTURE i nikad se ne
predstavljaju kao zvanični tekst kurikuluma; formule izgubljene u PDF
ekstrakciji se NE rekonstruišu — označene su za ručnu vizuelnu provjeru.
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import build_curriculum_mapping as bcm  # noqa: E402  (dijeli normalizaciju/ZIP)

PHASE2_XLSX = ROOT / "reference" / "curriculum" / "semantics" / "MATBOT_Faza2_Mapiranje.xlsx"
OUTPUT_XLSX = ROOT / "reference" / "curriculum" / "semantics" / "MATBOT_Faza2_5_Pilot25_Review.xlsx"

PILOT_AREAS = ("Djeljivost brojeva", "Razlomci")
PILOT_GRADE = 6

# Šest lekcija koje je Faza 2 ostavila bez exact dokaza — svaka MORA dobiti
# eksplicitno razrješenje (slovo slučaja iz specifikacije + obrazloženje).
GAP_LESSONS = ("6-03-002", "6-03-003", "6-03-006", "6-04-002", "6-04-007", "6-04-014")

VERDICTS = ("accept", "change_relation", "reject", "ambiguous", "broad_but_valid",
            "split_required", "missing_formula", "wrong_lesson", "wrong_grade",
            "duplicate_evidence", "recovered")

SEMANTIC_CONFIDENCE = ("high", "medium", "low")
HUMAN_REVIEW_STATUS = ("ready_for_contract_draft", "needs_manual_review",
                       "blocked_by_source_gap", "blocked_by_formula_loss",
                       "disputed_mapping")

# ---------------------------------------------------------------------------
# KORAK 1 — PRESUDE NAD SVAKOM PILOT VEZOM FAZE 2.
#
# KLJUČ JE (item_id, target_lesson_id, relation) — NIKAD mapping_id: mapping_id
# je pozicioni i pomjeri se čim se skup redova promijeni, pa bi popravka Faze 2
# tiho razvezala presude od redova na koje se odnose (nalaz Faze 3).
#
# Vrijednost: (verdict, nova_relacija|None, nova_pouzdanost|None, obrazloženje,
# zastavice). Svaka pilot veza MORA imati presudu — Kontrola to obara.
# ---------------------------------------------------------------------------

MAPPING_VERDICTS = {
    ('KS_2018-0035', '6-03-004', 'neighbour'): (
        'accept', None, None,
        "POPRAVLJEN DEFEKT 1 (Faza 3): susjed je sada KANONSKA lekcija pravila djeljivosti; ranije je pravilo biralo prvu lekciju po ID-ju čiji naslov sadrži 'djeljivost' (6-03-002), pa je granica pokazivala na pogrešnu vještinu",
        ()),
    ('KS_2018-0036', '6-03-001', 'exact'): (
        'accept', None, None,
        'upotreba pojmova djelilac/sadržilac/djeljivo je srž lekcije',
        ()),
    ('KS_2018-0036', '6-03-004', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica',
        ()),
    ('KS_2018-0038', '6-03-001', 'exact'): (
        'accept', None, 'high',
        'određivanje djelilaca datog broja JE osnovna radnja lekcije; potvrđeno',
        ()),
    ('KS_2018-0039', '6-03-001', 'exact'): (
        'accept', None, None,
        'odnos broja i njegovog sadržioca/djelioca = srž lekcije',
        ()),
    ('KS_2018-0039', '6-03-004', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica',
        ()),
    ('KS_2018-0040', '6-03-001', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica: stavka o pravilima pripada 6-03-004, traženje djelilaca je ne smije zamijeniti',
        ()),
    ('KS_2018-0040', '6-03-004', 'exact'): (
        'accept', None, None,
        "KS izričito: 'primjenjivati pravila za djeljivost sa 2, sa 3, sa 5, 6, 9, 4, 25 i sa 10n' — noseći dokaz; NAPOMENA: KS skup ne navodi 15, a navodi 10^n (MAT-BOT naslov: 2,3,4,5,6,9,10,15,25)",
        ()),
    ('KS_2018-0041', '6-03-005', 'exact'): (
        'accept', None, 'high',
        "'utvrđivati da li je broj prost ili složen' = srž lekcije; potvrđeno",
        ()),
    ('KS_2018-0042', '6-03-006', 'supporting'): (
        'change_relation', 'exact', 'high',
        "'utvrđivati jesu li dva data broja uzajamno (relativno) prosta' JE doslovno vještina lekcije — Faza 2 ju je potcijenila (supporting/low); lažno nisko, ne lažno visoko",
        ()),
    ('KS_2018-0043', '6-03-005', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica: prepoznavanje prostih nije faktorizacija',
        ()),
    ('KS_2018-0043', '6-03-007', 'exact'): (
        'accept', None, None,
        'noseći KS dokaz faktorizacije',
        ()),
    ('KS_2018-0043', '6-03-008', 'neighbour'): (
        'accept', None, None,
        'granica prema faktorizaciji; tačna',
        ()),
    ('KS_2018-0044', '6-03-001', 'exact'): (
        'wrong_lesson', None, None,
        'stavka traži NZD/NZS računanje — dokaz pripada 6-03-008/6-03-009; 6-03-001 je tek predznanje te stavke',
        ()),
    ('KS_2018-0044', '6-03-008', 'exact'): (
        'accept', None, 'high',
        "stavka izričito 'najveći zajednički djelilac'; izjednačenje razriješeno u korist 008+009 (vidi odbijeni red za 6-03-001)",
        ('split',)),
    ('KS_2018-0044', '6-03-009', 'exact'): (
        'accept', None, 'high',
        'ista stavka pokriva i NZS — legitiman dvostruki cilj (split)',
        ('split',)),
    ('KS_2018-0045', '6-03-001', 'exact'): (
        'accept', None, None,
        "KS tabela sadržaja nabraja 'Djelioci broja … Sadržioci broja'",
        ('broad',)),
    ('KS_2018-0045', '6-03-004', 'neighbour'): (
        'accept', None, None,
        'izvedena granica iz široke stavke; sadržajno tačna',
        ('broad',)),
    ('KS_2018-0045', '6-03-007', 'neighbour'): (
        'accept', None, None,
        'izvedena granica; tačna',
        ('broad',)),
    ('KS_2018-0045', '6-03-008', 'exact'): (
        'accept', None, None,
        "KS tabela: 'Zajednički djelioci … Najveći zajednički djelilac'",
        ('broad',)),
    ('KS_2018-0045', '6-03-009', 'exact'): (
        'accept', None, None,
        "KS tabela: 'Zajednički sadržioci … Najmanji zajednički sadržilac'",
        ('broad',)),
    ('KS_2018-0045', '6-03-010', 'exact'): (
        'accept', None, None,
        "vodeća klauzula 'rješavati tekstualne zadatke' STOJI u KS bloku djeljivosti — dokaz je tačan; frazni pogodak preko granice rečenice je artefakt ekstrakcije, ali cilja ispravno",
        ('broad',)),
    ('KS_2018-0045', '6-04-005', 'neighbour'): (
        'accept', None, None,
        'izvedena granica; tačna',
        ('broad',)),
    ('KS_2018-0045', '6-04-008', 'exact'): (
        'accept', None, None,
        "KS tabela: 'Upoređivanje razlomaka'",
        ('broad',)),
    ('KS_2018-0045', '6-04-009', 'exact'): (
        'accept', None, None,
        "KS tabela: 'Sabiranje i oduzimanje razlomaka jednakih imenilaca'",
        ('broad',)),
    ('KS_2018-0045', '6-04-010', 'exact'): (
        'accept', None, None,
        "KS tabela: '… razlomaka različitih imenilaca'",
        ('broad',)),
    ('KS_2018-0045', '6-04-011', 'exact'): (
        'accept', None, None,
        "KS tabela: 'Množenje razlomka prirodnim brojem. Množenje razlomka razlomkom.'",
        ('broad',)),
    ('KS_2018-0045', '6-04-012', 'exact'): (
        'accept', None, None,
        "KS tabela: 'Dijeljenje razlomka prirodnim brojem. Dijeljenje razlomka razlomkom.'",
        ('broad',)),
    ('KS_2018-0045', '6-04-014', 'neighbour'): (
        'accept', None, None,
        'granica jednačina↔izraz; tačna',
        ('broad',)),
    ('KS_2018-0046', '6-04-001', 'exact'): (
        'accept', None, None,
        'noseći KS dokaz: razlomak, brojilac, imenilac, razlomačka crta',
        ()),
    ('KS_2018-0046', '6-04-002', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica',
        ()),
    ('KS_2018-0049', '6-04-004', 'exact'): (
        'accept', None, None,
        "noseći dokaz; 'kao dio figure' dio stavke ide 6-04-002 (oporavljeno)",
        ()),
    ('KS_2018-0050', '6-04-004', 'supporting'): (
        'change_relation', 'supporting', 'medium',
        "'grafički prikaz' obuhvata i figure i polupravu — za polupravu je podrška; noseći dom stavke je 6-04-002 (oporavljeno tamo kao exact)",
        ()),
    ('KS_2018-0050', '6-04-013', 'supporting'): (
        'wrong_lesson', None, None,
        'stavka o čitanju grafičkog prikaza nema veze sa svojstvima operacija — mašinski šum niskog pouzdanja',
        ()),
    ('KS_2018-0051', '6-04-003', 'exact'): (
        'accept', None, 'high',
        "zapis nepravog razlomka mješovitim brojem i obratno = srž lekcije; izjednačenje s 6-05-001/6-06-001 je mašinski šum ('zapis'), potvrđeno",
        ()),
    ('KS_2018-0054', '6-04-005', 'exact'): (
        'accept', None, None,
        'invarijantnost vrijednosti pri proširivanju/skraćivanju — nosi OBJE lekcije (vidi promjenu za 6-04-006)',
        ('split',)),
    ('KS_2018-0054', '6-04-006', 'neighbour'): (
        'change_relation', 'exact', 'high',
        "stavka IZRIČITO navodi i skraćivanje ('proširivanjem i skraćivanjem'); mašina je dala samo susjedstvo — lažno nisko",
        ()),
    ('KS_2018-0054', '6-04-008', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica',
        ()),
    ('KS_2018-0055', '6-04-005', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica',
        ()),
    ('KS_2018-0055', '6-04-008', 'exact'): (
        'accept', None, None,
        "noseći dokaz: 'upoređivati razlomke'",
        ()),
    ('KS_2018-0057', '6-04-013', 'exact'): (
        'change_relation', 'supporting', 'medium',
        "'izvoditi osnovne računske operacije' dokazuje BLOK operacija (6-04-009…012), a lekciju o svojstvima samo podržava",
        ()),
    ('KS_2018-0060', '6-04-014', 'neighbour'): (
        'accept', None, None,
        'granica jednačina↔vrijednost izraza; nakon popravke pokazuje na izraze s RAZLOMCIMA (6-04-014) umjesto na 6-02-008 (izrazi s promjenljivim u N)',
        ()),
    ('KS_2018-0063', '6-04-001', 'exact'): (
        'accept', None, None,
        'KS tabela sadržaja nabraja pojmove lekcije',
        ('broad',)),
    ('KS_2018-0063', '6-04-002', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica',
        ('broad',)),
    ('KS_2018-0063', '6-04-005', 'exact'): (
        'accept', None, None,
        "KS tabela: 'Proširivanje razlomaka'",
        ('broad',)),
    ('KS_2018-0063', '6-04-006', 'exact'): (
        'accept', None, None,
        "KS tabela: 'Skraćivanje razlomaka'; NAPOMENA: 'nesvodivi razlomak' iz naslova nijedan izvor ne imenuje — MAT-BOT dekompozicija",
        ('broad',)),
    ('KS_2018-0063', '6-04-008', 'neighbour'): (
        'accept', None, None,
        'izvedena granica; tačna',
        ('broad',)),
    ('KS_2018-0063', '6-04-009', 'neighbour'): (
        'accept', None, None,
        'granica račun↔jednačine; tačna',
        ('broad',)),
    ('KS_2018-0063', '6-04-011', 'neighbour'): (
        'accept', None, None,
        'široka stavka stvarno nabraja i postotni zapis — granica razlomci↔procenti sadržajno stoji',
        ('broad',)),
    ('KS_2018-0063', '6-04-014', 'neighbour'): (
        'accept', None, None,
        'ista granica iz široke KS stavke; cilj je sada lekcija iste jedinice',
        ('broad',)),
    ('KS_2018-0073', '6-03-010', 'exact'): (
        'wrong_lesson', None, None,
        'KS-0073 pripada nizu ishoda za DECIMALNE brojeve (0064–0075): pravi dom je 6-05-011, ne 6-03-010',
        ()),
    ('KS_2018-0073', '6-04-015', 'exact'): (
        'wrong_lesson', None, None,
        'isti razlog kao za 6-03-010: KS-0073 pripada decimalnom nizu (dom 6-05-011)',
        ()),
    ('KS_2018-0075', '6-04-009', 'neighbour'): (
        'accept', None, None,
        'granica račun↔jednačine; tačna',
        ('broad',)),
    ('KS_2018-0075', '6-04-010', 'neighbour'): (
        'accept', None, None,
        'granica; tačna',
        ('broad',)),
    ('KS_2018-0075', '6-04-014', 'neighbour'): (
        'accept', None, None,
        'granica jednačina↔izraz; tačna',
        ('broad',)),
    ('RS_2014-0022', '6-03-001', 'neighbour'): (
        'accept', None, None,
        'ispravna smjerna granica prema RS stavci o kriterijumima',
        ()),
    ('RS_2014-0022', '6-03-004', 'exact'): (
        'accept', None, None,
        "RS 'kriterijumima djeljivosti' = pravila djeljivosti (terminološki alias)",
        ()),
    ('RS_2014-0023', '6-03-005', 'exact'): (
        'accept', None, 'high',
        'RS poseban cilj imenuje lekciju; potvrđeno',
        ()),
    ('RS_2014-0024', '6-03-007', 'neighbour'): (
        'accept', None, None,
        'granica faktorizacija↔NZD/NZS iz RS stavke; tačna',
        ()),
    ('RS_2014-0024', '6-03-008', 'exact'): (
        'accept', None, None,
        'RS cilj imenuje NZD',
        ()),
    ('RS_2014-0024', '6-03-009', 'exact'): (
        'accept', None, None,
        'RS cilj imenuje NZS',
        ()),
    ('RS_2014-0025', '6-04-013', 'exact'): (
        'change_relation', 'supporting', 'medium',
        'širok RS cilj (pojam + operacije); za svojstva je samo podrška; podjela stavke: exact za 6-04-001 (oporavljeno), podrška operacijama',
        ('split',)),
    ('RS_2014-0026', '6-04-009', 'neighbour'): (
        'accept', None, None,
        'RS stavka o jednačinama; granica prema direktnom računu tačna',
        ()),
    ('RS_2014-0026', '6-04-014', 'neighbour'): (
        'accept', None, None,
        'ista granica iz RS cilja o jednačinama s razlomcima',
        ()),
}


def verdict_key(row):
    """Stabilan ključ presude iz reda Faze 2: (item_id, cilj, relacija)."""
    return (row[1], row[8], row[12])


def verdict_for(row):
    return MAPPING_VERDICTS[verdict_key(row)]


# ---------------------------------------------------------------------------
# OPORAVLJENI DOKAZI — nalaz stručnog pregleda, NIKAD izlaz Faze 2.
# origin: 'stavka' (postojeći Stavke_NPP ID) ili 'stranica' (Stranice_KS/RS).
# Za 'stranica' se navodi izvor+stranica+doslovan citat (provjerava se da citat
# stvarno postoji u tekstu te stranice). relation/confidence su presuda pregleda.
# ---------------------------------------------------------------------------

RECOVERED_EVIDENCE = (
    # id, origin, item_or_source, page, lesson, relation, confidence, quote/razlog, flags
    ("R25-001", "stavka", "KS_2018-0037", None, "6-03-001", "exact", "medium",
     "'napamet odrediti nekoliko sadržilaca prostog broja' — sadržioci su "
     "polovina lekcije; Faza 2 ju je ostavila bez cilja", ()),
    ("R25-002", "stavka", "KS_2018-0036", None, "6-03-005", "supporting", "medium",
     "stavka traži i upotrebu pojma 'prost broj je'", ()),
    ("R25-003", "stranica", "KS_2018", 10, "6-03-002", "exact", "high",
     "Djeljivost  zbira, razlike i proizvoda.", ()),
    ("R25-004", "stranica", "KS_2018", 12, "6-03-002", "exact", "high",
     "Djeljivost zbira, razlike i  proizvoda prirodnih brojeva", ()),
    ("R25-005", "stranica", "KS_2018", 15, "6-03-002", "supporting", "high",
     "pokazati djeljivost zbira, odnosno, proizvoda brojem", ()),
    ("R25-006", "stranica", "KS_2018", 10, "6-03-003", "exact", "high",
     "djeljivost dekadskom jedinicom", ()),
    ("R25-007", "stranica", "KS_2018", 12, "6-03-003", "exact", "high",
     "Djeljivost dekadnim jedinicama  i brojevima: 2,3,4,6,9,25", ()),
    ("R25-008", "stavka", "KS_2018-0040", None, "6-03-003", "supporting", "medium",
     "'… i sa 10n' — eksponent 10^n izgubljen u PDF ekstrakciji; POTREBNA "
     "RUČNA VIZUELNA PROVJERA zapisa", ("formula_loss",)),
    ("R25-009", "stranica", "KS_2018", 8, "6-03-003", "supporting", "medium",
     "sa 10n , n ,...", ("formula_loss",)),
    ("R25-010", "stavka", "KS_2018-0047", None, "6-04-002", "exact", "high",
     "'dijeliti cijelo na jednake djelove, na modelu i na slici' — dio cjeline; "
     "Faza 2 stavku ostavila bez cilja", ()),
    ("R25-011", "stavka", "KS_2018-0050", None, "6-04-002", "exact", "medium",
     "prepoznavanje razlomka iz grafičkog prikaza = razlomak kao dio cjeline", ()),
    ("R25-012", "stavka", "KS_2018-0049", None, "6-04-002", "supporting", "medium",
     "'… i kao dio figure' — sekundarni dio stavke o brojevnoj polupravoj", ()),
    ("R25-013", "stranica", "RS_2014", 341, "6-04-002", "exact", "high",
     "Razlomak se  uvodi kao količnik dva prirodna broja", ()),
    ("R25-014", "stranica", "KS_2018", 16, "6-04-007", "exact", "high",
     "Uvježbati svođenje razlomaka na zajednički  nazivnik pa preći na sabiranje.",
     ()),
    ("R25-015", "stavka", "KS_2018-0045", None, "6-04-014", "exact", "medium",
     "KS tabela razlomaka nabraja 'Brojevni izrazi sa zagradama' (kontekst "
     "jedinice: razlomci)", ("broad",)),
    ("R25-016", "stavka", "KS_2018-0059", None, "6-04-014", "supporting", "medium",
     "'izračunavati vrijednost brojevnog izraza …' u nizu ishoda za razlomke", ()),
    ("R25-017", "stavka", "KS_2018-0045", None, "6-04-015", "exact", "medium",
     "KS tabela razlomaka nabraja 'Tekstualni zadaci' (kontekst jedinice: "
     "razlomci)", ("broad",)),
    ("R25-018", "stavka", "KS_2018-0058", None, "6-04-015", "supporting", "medium",
     "'povezivati … sa kontekstom problema' u nizu ishoda za razlomke", ()),
    ("R25-019", "stranica", "RS_2014", 341, "6-04-008", "supporting", "medium",
     "Upoređivanje razlomaka i  osnovne operacije s njima neophodno je", ()),
    ("R25-020", "stranica", "RS_2014", 340, "6-03-001", "exact", "medium",
     "pojmovima činioca i  sadržioca prirodno", ()),
    ("R25-021", "stavka", "KS_2018-0057", None, "6-04-009", "supporting", "medium",
     "'osnovne računske operacije sa razlomcima' pokriva blok operacija", ()),
    ("R25-022", "stavka", "KS_2018-0057", None, "6-04-010", "supporting", "medium",
     "isto — blok operacija", ()),
    ("R25-023", "stavka", "KS_2018-0057", None, "6-04-011", "supporting", "medium",
     "isto — blok operacija", ()),
    ("R25-024", "stavka", "KS_2018-0057", None, "6-04-012", "supporting", "medium",
     "isto — blok operacija", ()),
    ("R25-025", "stavka", "RS_2014-0025", None, "6-04-001", "exact", "medium",
     "'Da se upoznaju sa pojmom razlomka …' — jedini RS dokaz pojma razlomka "
     "(podjela široke stavke)", ("split",)),
    ("R25-026", "stranica", "RS_2014", 341, "6-04-009", "supporting", "medium",
     "osnovne operacije s njima neophodno je", ()),
    ("R25-027", "stranica", "RS_2014", 341, "6-04-010", "supporting", "medium",
     "osnovne operacije s njima neophodno je", ()),
    ("R25-028", "stranica", "RS_2014", 341, "6-04-011", "supporting", "medium",
     "osnovne operacije s njima neophodno je", ()),
    ("R25-029", "stranica", "RS_2014", 341, "6-04-012", "supporting", "medium",
     "osnovne operacije s njima neophodno je", ()),

    # --- FAZA 3: DOKAZI IZ KOLONE „SADRŽAJI PROGRAMA“ (KS str. 12) ----------
    # Faza 1 je stranicu čuvala u cijelosti, ali je njena EKSTRAKCIJA STAVKI
    # spljoštila tabelu u nečitljive blokove, pa mapiranje ove redove nikad
    # nije vidjelo. Ponovna ekstrakcija originalnog PDF-a s očuvanjem kolona
    # (`pdftotext -layout`) pokazuje da su ovo zasebni redovi SADRŽAJA jedinice
    # Razlomci/Djeljivost. Citati se i dalje provjeravaju doslovno nad tekstom
    # stranice iz Faze 1 — dokaz ostaje reproducibilan iz repozitorija.
    ("R25-030", "stranica", "KS_2018", 12, "6-04-013", "exact", "high",
     "Svojstva sabiranja razlomaka", ()),
    ("R25-031", "stranica", "KS_2018", 12, "6-04-013", "exact", "high",
     "Svojstva množenja", ()),
    ("R25-032", "stranica", "KS_2018", 12, "6-04-014", "exact", "high",
     "Brojevni izrazi sa zagradama", ()),
    ("R25-033", "stranica", "KS_2018", 12, "6-04-015", "exact", "high",
     "Tekstualni zadaci", ()),
)

# ---------------------------------------------------------------------------
# FAZA 3 — OBAVEZNA PROVJERA IZVORA NAD ORIGINALNIM PDF-om (bez mreže).
# Zapisuje se TAČNO ono što se na stranici vidi; ništa se ne rekonstruiše.
# ---------------------------------------------------------------------------

SOURCE_VERIFICATION = (
    # lesson, tema provjere, nalaz, metod, posljedica
    ("6-03-003", "zapis „10n“ (moguć izgubljen eksponent 10^n)",
     "NIJE eksponent. Na obje pojave (KS str. 8 i str. 12) slovo „n“ je "
     "iscrtano istom veličinom fonta i na ISTOJ osnovnoj liniji kao okolni "
     "tekst — str. 12: „…25 i sa 10“ F3 10.08pt y=623.210, zatim „n“ F5 "
     "10.08pt y=623.210; str. 8: „…sa 10“ F3 12.00pt y=502.940, zatim „n“ F5 "
     "12.00pt y=502.940. Dokument nigdje ne koristi operator pomaka teksta "
     "(Ts). Vidljiva notacija je dakle „10n“ (n u drugom fontu), NE „10^n“.",
     "analiza toka sadržaja PDF-a (font, veličina, osnovna linija) — "
     "renderovanje stranice u sliku nije bilo moguće bez instalacije alata, "
     "a mreža je zabranjena",
     "Lekcija 6-03-003 NE zavisi od ovog tokena: njen dokaz je zaseban red "
     "sadržaja „Djeljivost dekadnim jedinicama…“ (KS str. 12). Blokada zbog "
     "gubitka formule se ukida, a sam token ostaje zabilježen kao nerazriješena "
     "notacija u ishodu koji pripada lekciji 6-03-004."),
    ("6-03-003", "obim lekcije u izvoru",
     "KS red sadržaja glasi „Djeljivost dekadnim jedinicama i brojevima: "
     "2,3,4,6,9,25“ — dekadske jedinice i pojedinačna pravila su u JEDNOM redu.",
     "pdftotext -layout, KS str. 12, kolona „Sadržaji programa“",
     "Podjela na 6-03-003 i 6-03-004 je MAT-BOT dekompozicija, ne KS podjela; "
     "granica GR (003↔004) mora ostati eksplicitna."),
    ("6-04-013", "postoji li zaseban izvorni sadržaj",
     "DA — kolona sadržaja jedinice Razlomci nosi dva zasebna reda: „Svojstva "
     "sabiranja razlomaka“ i „Svojstva množenja razlomaka“.",
     "pdftotext -layout, KS str. 12; citati potvrđeni i u tekstu stranice Faze 1",
     "Dokaz je sada izričit (R25-030/031), a ne samo podržavajući. Ostaje "
     "ograničenje ENFORCEMENTA: koje je svojstvo upotrijebljeno nije "
     "deterministički provjerljivo, pa lekcija ide u ADVISORY_ONLY."),
    ("6-04-014", "postoji li zaseban izvorni sadržaj",
     "DA — red sadržaja „Brojevni izrazi sa zagradama“ u jedinici Razlomci; "
     "uz ishod „izračunavati vrijednost brojevnog izraza…“.",
     "pdftotext -layout, KS str. 12; citat potvrđen u tekstu stranice Faze 1",
     "Dokaz podignut na exact/high (R25-032)."),
    ("6-04-015", "postoji li zaseban izvorni sadržaj",
     "DA — red sadržaja „Tekstualni zadaci“ u jedinici Razlomci (odvojen od "
     "istoimenog ishoda u jedinici Djeljivost, koji pripada 6-03-010).",
     "pdftotext -layout, KS str. 12; citat potvrđen u tekstu stranice Faze 1",
     "Dokaz podignut na exact/high (R25-033); enforcement ostaje ograničen jer "
     "kvalitet priče nije deterministički provjerljiv."),
)

# ---------------------------------------------------------------------------
# KORAK 2 — RAZRJEŠENJE ŠEST PRAZNINA (slovo slučaja iz specifikacije).
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# FAZA 3 — KLASA AKTIVACIJE. Dvije NEZAVISNE ose, namjerno razdvojene:
#
#   activation_class  — smije li lekcija uopšte dobiti blokirajuću semantiku
#                       (READY / ADVISORY_ONLY / BLOCKED)
#   enforcement_now   — postoji li DANAS dokazan deterministički detektor
#
# READY ne znači „uključi odmah“: znači „dokaz je izričit i vještina je
# deterministički odlučiva, pa blokiranje SMIJE doći kad detektor bude
# implementiran i dokazan“. Lekcija se NIKAD ne diže u READY radi pokrivenosti.
# ---------------------------------------------------------------------------

ACTIVATION_CLASSES = ("READY", "ADVISORY_ONLY", "BLOCKED")

# Detektori koji danas STVARNO postoje i dokazano blokiraju u produkciji.
PROVEN_DETECTORS = {
    "6-03-004": "mcq_integrity divisibility oracle + lesson_fidelity "
                "semantic_task_requirement (aktivno blokira u produkciji)",
    "6-04-005": "contracts/verifiers.exact_rational + uključen ugovor lekcije",
    "6-04-006": "contracts/verifiers.exact_rational + uključen ugovor lekcije",
    "6-04-009": "contracts/verifiers.exact_rational + uključen ugovor lekcije",
    "6-04-010": "contracts/verifiers.exact_rational + uključen ugovor lekcije",
    "6-04-011": "contracts/verifiers.exact_rational + uključen ugovor lekcije",
    "6-04-012": "contracts/verifiers.exact_rational + uključen ugovor lekcije",
}

ACTIVATION = {
    "6-03-001": ("READY", "izričiti redovi sadržaja („Djelioci broja“, „Sadržioci "
                 "broja“) i ishodi; pripadnost skupu djelilaca/sadržilaca je "
                 "egzaktna cjelobrojna provjera"),
    "6-03-002": ("READY", "izričit red sadržaja „Djeljivost zbira, razlike i "
                 "proizvoda prirodnih brojeva“ (KS str. 12); djeljivost "
                 "zbira/razlike/proizvoda je egzaktno provjerljiva"),
    "6-03-003": ("READY", "izričit red sadržaja „Djeljivost dekadnim jedinicama…“; "
                 "djeljivost sa 10^k je egzaktna (završne nule). NAPOMENA: KS taj "
                 "red dijeli s pravilima za 2,3,4,6,9,25 — granica 003↔004 je "
                 "MAT-BOT dekompozicija i mora ostati eksplicitna"),
    "6-03-004": ("READY", "dokazano i VEĆ AKTIVNO blokiranje u produkciji; ovaj "
                 "pregled ga ne mijenja — ponašanje mora ostati ekvivalentno"),
    "6-03-005": ("READY", "izričit ishod „utvrđivati da li je broj prost ili "
                 "složen“; prostost je egzaktno provjerljiva"),
    "6-03-006": ("READY", "izričit ishod „utvrđivati jesu li dva data broja "
                 "uzajamno (relativno) prosta“; NZD(a,b)==1 je egzaktan"),
    "6-03-007": ("READY", "izričit ishod i red sadržaja; potpuna faktorizacija je "
                 "egzaktno provjerljiva"),
    "6-03-008": ("READY", "izričit red sadržaja i ishod; NZD je egzaktan"),
    "6-03-009": ("READY", "izričit red sadržaja i ishod; NZS je egzaktan"),
    "6-03-010": ("ADVISORY_ONLY", "izvorni ishod postoji („rješavati tekstualne "
                 "zadatke“ u jedinici Djeljivost), ali KVALITET PRIČE nije "
                 "deterministički provjerljiv — smije samo voditi prompt"),
    "6-04-001": ("READY", "izričit ishod (razlomak, brojilac, imenilac, razlomačka "
                 "crta); imenovanje dijelova zapisa je zatvoren skup"),
    "6-04-002": ("READY", "obje polovine imaju izvor: „dijeliti cijelo na jednake "
                 "djelove“ (KS) i „Razlomak se uvodi kao količnik dva prirodna "
                 "broja“ (RS str. 341); konverzija količnik↔razlomak je egzaktna"),
    "6-04-003": ("READY", "izričit ishod (nepravi ↔ mješoviti) i red sadržaja "
                 "„Vrste razlomaka“; konverzija i klasifikacija su egzaktne. "
                 "NAPOMENA: „prividni“ razlomci nisu izričito imenovani u KS"),
    "6-04-004": ("READY", "izričit red sadržaja „Pridruživanje tačaka brojevne "
                 "poluprave razlomcima“ i ishod; pozicija je egzaktna"),
    "6-04-005": ("READY", "izričit red sadržaja; već postoji dokazan "
                 "deterministički ugovor lekcije"),
    "6-04-006": ("READY", "isti red sadržaja („Proširivanje i skraćivanje“); već "
                 "postoji dokazan deterministički ugovor lekcije"),
    "6-04-007": ("ADVISORY_ONLY", "jedini dokaz je METODIČKI tekst (KS str. 16), a "
                 "KS ga vodi kao KORAK sabiranja različitih imenilaca, ne kao "
                 "zasebnu lekciju; blokiranje granice svođenje↔sabiranje bi "
                 "prejudiciralo odluku koju izvor ne donosi"),
    "6-04-008": ("READY", "izričit red sadržaja i ishod „upoređivati razlomke“; "
                 "poređenje racionalnih je egzaktno"),
    "6-04-009": ("READY", "izričit red sadržaja; postoji dokazan ugovor lekcije"),
    "6-04-010": ("READY", "izričit red sadržaja; postoji dokazan ugovor lekcije"),
    "6-04-011": ("READY", "izričit red sadržaja; postoji dokazan ugovor lekcije"),
    "6-04-012": ("READY", "izričit red sadržaja; postoji dokazan ugovor lekcije"),
    "6-04-013": ("ADVISORY_ONLY", "dokaz je nakon provjere izvora IZRIČIT (dva reda "
                 "sadržaja), ali KOJE je svojstvo upotrijebljeno nije "
                 "deterministički provjerljivo — granica „svojstvo vs goli račun“ "
                 "nije dokazana, pa nema blokiranja"),
    "6-04-014": ("READY", "izričit red sadržaja „Brojevni izrazi sa zagradama“ u "
                 "jedinici Razlomci; vrijednost izraza je egzaktno provjerljiva"),
    "6-04-015": ("ADVISORY_ONLY", "izričit red sadržaja „Tekstualni zadaci“, ali "
                 "kvalitet priče nije deterministički provjerljiv — isto kao "
                 "6-03-010"),
}

GAP_RESOLUTIONS = {
    "6-03-002": ("B", "sadržaj postoji u tekstu KS stranica 10/12 (tabele sadržaja) "
                      "i 15 (metodika), ali ekstrakcija Faze 1 te tabele nije "
                      "pretvorila u stavke; oporavljeno: R25-003..005"),
    "6-03-003": ("B", "KS stranice 10 i 12 izričito imenuju djeljivost dekadskim "
                      "jedinicama; uz to KS-0040 sadrži '… i sa 10n' gdje je "
                      "eksponent izgubljen (zastavica formula_loss, ručna "
                      "vizuelna provjera); oporavljeno: R25-006..009"),
    "6-03-006": ("A", "sadržaj je POSTOJAO u stavci KS-0042, ali ju je Faza 2 "
                      "potcijenila (supporting/low); presuda: exact/high"),
    "6-04-002": ("A", "dio cjeline: KS-0047 (no_match u Fazi 2) i KS-0050; "
                      "količnik: RS stranica 341 'Razlomak se uvodi kao količnik "
                      "dva prirodna broja'; oporavljeno: R25-010..013"),
    "6-04-007": ("B", "KS stranica 16 (metodika): 'Uvježbati svođenje razlomaka "
                      "na zajednički nazivnik pa preći na sabiranje' — KS ga vodi "
                      "kao korak sabiranja različitih imenilaca, MAT-BOT kao "
                      "zasebnu lekciju (opravdana dekompozicija); R25-014"),
    "6-04-014": ("A", "KS tabela razlomaka (u širokoj stavci KS-0045) nabraja "
                      "'Brojevni izrazi sa zagradama', a KS-0059 traži "
                      "izračunavanje vrijednosti brojevnog izraza u nizu ishoda "
                      "za razlomke; oporavljeno: R25-015..016"),
}

# ---------------------------------------------------------------------------
# KORAK 4 — SEMANTIČKI RED PO LEKCIJI (25 redova).
# Autorski sadržaj stručnog pregleda; dokazne liste se RAČUNAJU iz presuda.
# ---------------------------------------------------------------------------

LESSON_SEMANTICS = {
    "6-03-001": {
        "family": "common_divisors_multiples",
        "core_skill": "odrediti djelioce i sadržioce datog prirodnog broja i "
                      "provjeriti odnos 'je djelilac / je sadržilac'",
        "actions": "nabrojati djelioce; nabrojati prvih nekoliko sadržilaca; "
                   "provjeriti da li je a djelilac (sadržilac) broja b",
        "concepts": "djelilac/faktor; sadržilac/višekratnik; djeljivo je",
        "archetypes": "enumerate_set; verify_membership; select_by_property",
        "answer_kinds": "integer; integer_list; yes_no",
        "level1": "jedan broj ≤ 50: nabrojati/prepoznati djelioce ili prvih "
                  "nekoliko sadržilaca",
        "level2": "provjera odnosa za dva broja ili izbor broja koji jeste/nije "
                  "djelilac datog broja",
        "level3": "kombinovanje uslova (npr. broj koji je sadržilac od a, a "
                  "djelilac od b) — bez NZD/NZS terminologije",
        "prerequisites": "množenje i dijeljenje u N (6-02-004)",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "RS dokaz koristi termin 'činilac' (=faktor) — dodati u "
                "terminološke aliase pri izradi ugovora.",
    },
    "6-03-002": {
        "family": "divisibility_predicate_application",
        "core_skill": "primijeniti stavove o djeljivosti zbira, razlike i "
                      "proizvoda BEZ računanja vrijednosti",
        "actions": "zaključiti djeljivost zbira/razlike/proizvoda iz djeljivosti "
                   "sabiraka/faktora; odabrati/obrazložiti tačan zaključak",
        "concepts": "djeljivost zbira; djeljivost razlike; djeljivost proizvoda",
        "archetypes": "yes_no_rule_check; select_by_rule; justify_rule",
        "answer_kinds": "yes_no; option_statement",
        "level1": "jedan stav, direktna da/ne provjera (npr. 12+18 djeljivo sa 3?)",
        "level2": "dva stava ili zaključak bez izračunavanja velikog zbira",
        "level3": "kombinacija stavova s obrazloženjem (npr. zašto je 6·k+12 "
                  "djeljivo sa 6)",
        "prerequisites": "pojam djeljivosti (6-03-001)",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "FAZA 3: potvrđen izričit red sadržaja „Dranica (R25-003..005) — "
                "prije ugovora potvrditi obim (KS metodika pominje i teoreme).",
    },
    "6-03-003": {
        "family": "divisibility_predicate_application",
        "core_skill": "primijeniti djeljivost dekadskim jedinicama (10, 100, "
                      "1000, …)",
        "actions": "provjeriti/odabrati broj djeljiv sa 10^k po broju nula",
        "concepts": "dekadska jedinica; završne nule broja",
        "archetypes": "yes_no_rule_check; select_by_rule",
        "answer_kinds": "yes_no; integer",
        "level1": "da/ne za djeljivost sa 10 ili 100",
        "level2": "izbor među kandidatima za djeljivost sa 100/1000",
        "level3": "dopuna broja nulama da bude djeljiv sa 10^k uz još jedan uslov",
        "prerequisites": "mjesna vrijednost; 6-03-001",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "FAZA 3: provjereno u originalnom PDF-u —' (R25-008/009) — "
                "eksponent tražiti vizuelno u PDF-u prije finalnog ugovora.",
    },
    "6-03-004": {
        "family": "divisibility_predicate_application",
        "core_skill": "primijeniti pravila djeljivosti sa 2, 3, 4, 5, 6, 9, 10, "
                      "15 i 25 na dat broj",
        "actions": "primijeniti imenovano pravilo (zadnja cifra / zbir cifara / "
                   "zadnje dvije cifre / kombinacija); odabrati broj koji "
                   "zadovoljava pravilo; obrazložiti pravilom",
        "concepts": "pravilo djeljivosti; zadnja cifra; zbir cifara",
        "archetypes": "yes_no_rule_check; select_by_rule; justify_rule; "
                      "digit_completion",
        "answer_kinds": "yes_no; integer; digit",
        "level1": "jedno izričito pravilo, direktna da/ne provjera ili izbor",
        "level2": "dva istovremena uslova ili pravilo + obrazloženje",
        "level3": "dopuna cifre / najmanji-najveći broj uz više uslova",
        "prerequisites": "6-03-001",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Razlika skupova djelilaca: KS navodi 10^n a ne 15; RS govori o "
                "'kriterijumima'. MAT-BOT skup {2,3,4,5,6,9,10,15,25} je unija — "
                "za 15 osloniti se na kombinaciju 3 i 5 (bez posebnog izvora).",
    },
    "6-03-005": {
        "family": "prime_structure",
        "core_skill": "klasifikovati broj kao prost ili složen",
        "actions": "odlučiti prost/složen; odabrati prost (složen) broj iz skupa",
        "concepts": "prost broj; složen broj; broj 1 nije ni prost ni složen",
        "archetypes": "classify; select_by_property; justify_rule",
        "answer_kinds": "yes_no; integer; option_statement",
        "level1": "klasifikacija jednog broja ≤ 30",
        "level2": "izbor prostog broja u skupu; brojevi do 100",
        "level3": "tvrdnje o prostim brojevima (npr. jedini paran prost) s "
                  "obrazloženjem",
        "prerequisites": "6-03-001",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "",
    },
    "6-03-006": {
        "family": "prime_structure",
        "core_skill": "utvrditi da li su dva broja uzajamno (relativno) prosta",
        "actions": "provjeriti zajedničke djelioce para; odabrati par uzajamno "
                   "prostih brojeva",
        "concepts": "uzajamno/relativno prosti; zajednički djelilac 1",
        "archetypes": "verify_property; select_by_property",
        "answer_kinds": "yes_no; pair",
        "level1": "da/ne za mali par (npr. 8 i 15)",
        "level2": "izbor uzajamno prostog para među ponuđenim",
        "level3": "dopuna broja tako da par bude uzajamno prost uz dodatni uslov",
        "prerequisites": "6-03-001; 6-03-005",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Faza 2 lažno nisko (supporting/low) — ispravljeno presudom.",
    },
    "6-03-007": {
        "family": "prime_structure",
        "core_skill": "rastaviti složen broj na proste faktore",
        "actions": "izvesti/prepoznati potpunu faktorizaciju; uočiti grešku u "
                   "tuđoj faktorizaciji",
        "concepts": "prost faktor; potpuna faktorizacija; stepen faktora",
        "archetypes": "decompose; identify_error; verify_property",
        "answer_kinds": "factorization; option_statement",
        "level1": "faktorizacija broja ≤ 50 (izbor tačnog zapisa)",
        "level2": "broj ≤ 200 ili prepoznavanje NEpotpune faktorizacije",
        "level3": "greška u tuđem postupku ili faktorizacija s obrazloženjem",
        "prerequisites": "6-03-005",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "",
    },
    "6-03-008": {
        "family": "common_divisors_multiples",
        "core_skill": "odrediti zajedničke djelioce i najveći zajednički "
                      "djelilac (NZD) dva ili više brojeva",
        "actions": "nabrojati zajedničke djelioce; izračunati NZD (nabrajanjem "
                   "ili preko faktorizacije)",
        "concepts": "zajednički djelilac; NZD",
        "archetypes": "enumerate_set; compute_value; word_context_simple",
        "answer_kinds": "integer; integer_list",
        "level1": "NZD dva mala broja (≤ 30) nabrajanjem",
        "level2": "NZD preko faktorizacije; tri broja",
        "level3": "primjena (najveća jednaka podjela) bez pune tekstualne priče",
        "prerequisites": "6-03-001; 6-03-007",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "",
    },
    "6-03-009": {
        "family": "common_divisors_multiples",
        "core_skill": "odrediti zajedničke sadržioce i najmanji zajednički "
                      "sadržilac (NZS)",
        "actions": "nabrojati zajedničke sadržioce; izračunati NZS",
        "concepts": "zajednički sadržilac; NZS",
        "archetypes": "enumerate_set; compute_value; word_context_simple",
        "answer_kinds": "integer; integer_list",
        "level1": "NZS dva mala broja nabrajanjem",
        "level2": "NZS preko faktorizacije; tri broja",
        "level3": "primjena (istovremeni događaji) bez pune tekstualne priče",
        "prerequisites": "6-03-001; 6-03-007",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "",
    },
    "6-03-010": {
        "family": "word_problems_pilot",
        "core_skill": "riješiti tekstualni (životni) zadatak modelovan "
                      "djeljivošću, NZD-om ili NZS-om",
        "actions": "prepoznati model (podjela bez ostatka / NZD / NZS) iz priče "
                   "i izračunati traženu vrijednost",
        "concepts": "životni kontekst; izbor modela; djeljivost/NZD/NZS",
        "archetypes": "word_problem",
        "answer_kinds": "integer; short_text",
        "level1": "jednokoračna priča s malim brojevima (podjela bez ostatka)",
        "level2": "priča koja traži NZD ili NZS",
        "level3": "priča s dva uslova (npr. NZS pa uslov opsega)",
        "prerequisites": "6-03-004; 6-03-008; 6-03-009",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "FAZA 3: potvrđeno — ishod „-rješavati teroke KS stavke — "
                "potvrditi da KS zaista traži tekstualne zadatke bas u ovoj "
                "jedinici (stranica 12).",
    },
    "6-04-001": {
        "family": "fraction_concept_representation",
        "core_skill": "prepoznati razlomak i imenovati brojilac, imenilac i "
                      "razlomačku crtu",
        "actions": "imenovati dijelove zapisa; pročitati/zapisati razlomak",
        "concepts": "razlomak; brojilac/brojnik; imenilac/nazivnik; razlomačka crta",
        "archetypes": "recognize; name_parts",
        "answer_kinds": "short_text; option_statement; fraction",
        "level1": "imenovanje dijela zapisa ili čitanje razlomka",
        "level2": "zapis razlomka iz riječi ('tri sedmine') i obratno",
        "level3": "(nepodržan) — dublji zadaci prelaze u 6-04-002/6-04-003",
        "prerequisites": "dijeljenje u N (6-02-004)",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Level 3 namjerno nepodržan: sve iznad terminologije i zapisa "
                "napušta lekciju.",
    },
    "6-04-002": {
        "family": "fraction_concept_representation",
        "core_skill": "interpretirati razlomak kao dio cjeline i kao količnik "
                      "dva prirodna broja",
        "actions": "odrediti razlomak prikazan figurom/modelom; podijeliti "
                   "cjelinu na jednake dijelove; zapisati količnik a:b kao a/b",
        "concepts": "dio cjeline; jednaki dijelovi; količnik kao razlomak",
        "archetypes": "recognize; convert_representation",
        "answer_kinds": "fraction; option_statement",
        "level1": "razlomak sa osjenčene figure (imenilac ≤ 8)",
        "level2": "količnik a:b zapisan razlomkom; dio skupa predmeta",
        "level3": "poređenje dviju interpretacija (koja figura prikazuje 3/4) "
                  "uz obrazloženje",
        "prerequisites": "6-04-001",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Količnik-polovina dokazana RS stranicom 341; RS dodaje i 'kao "
                "razmjera' — to NE ulazi u ovu lekciju (6-06-003).",
    },
    "6-04-003": {
        "family": "fraction_concept_representation",
        "core_skill": "klasifikovati razlomke (pravi/nepravi/prividni) i "
                      "pretvarati neprave razlomke u mješovite brojeve i obratno",
        "actions": "klasifikovati razlomak; zapisati nepravi razlomak mješovitim "
                   "brojem i obratno",
        "concepts": "pravi/nepravi/prividni razlomak; mješoviti broj",
        "archetypes": "classify; convert_representation",
        "answer_kinds": "option_statement; mixed_number; fraction",
        "level1": "klasifikacija jednog razlomka",
        "level2": "pretvaranje nepravi ↔ mješoviti (imenilac ≤ 12)",
        "level3": "prividni razlomci i kombinovana klasifikacija više zapisa",
        "prerequisites": "6-04-001",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "",
    },
    "6-04-004": {
        "family": "fraction_concept_representation",
        "core_skill": "pridružiti razlomak tački brojevne poluprave i pročitati "
                      "razlomak s poluprave",
        "actions": "smjestiti razlomak na polupravu s datom podjelom; odrediti "
                   "razlomak pridružen označenoj tački",
        "concepts": "brojevna poluprava; jedinična duž; podjela na jednake dijelove",
        "archetypes": "convert_representation; recognize",
        "answer_kinds": "fraction; option_statement",
        "level1": "tačka na polupravoj s imeniocem podjele ≤ 8",
        "level2": "smještanje razlomka između cijelih brojeva; mješoviti broj",
        "level3": "više razlomaka na istoj polupravoj (redoslijed tačaka)",
        "prerequisites": "6-04-001; 6-04-002",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "MCQ oblik: opcije su pozicije/vrijednosti — bez crtanja.",
    },
    "6-04-005": {
        "family": "fraction_equivalence",
        "core_skill": "proširiti razlomak datim brojem ili na dati imenilac",
        "actions": "proširiti zadanim faktorom; odrediti faktor proširenja; "
                   "dopuniti brojilac/imenilac u jednakosti",
        "concepts": "proširivanje; nepromijenjena vrijednost; faktor proširenja",
        "archetypes": "compute_value; find_missing_value; identify_equivalent",
        "answer_kinds": "fraction; integer",
        "level1": "proširivanje malim faktorom (2–5)",
        "level2": "proširivanje na zadani imenilac; nedostajući brojilac",
        "level3": "izbor svih ekvivalentnih zapisa / obrazloženje invarijantnosti",
        "prerequisites": "6-04-001; množenje u N",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Postojeći deterministički pilot ugovor 6-04-005 (K1/K3) ostaje "
                "referenca — ovaj red ga ne mijenja.",
    },
    "6-04-006": {
        "family": "fraction_equivalence",
        "core_skill": "skratiti razlomak i prepoznati nesvodivi oblik",
        "actions": "skratiti zajedničkim djeliocem; skratiti do kraja; odlučiti "
                   "da li je razlomak nesvodiv",
        "concepts": "skraćivanje; zajednički djelilac; nesvodivi razlomak",
        "archetypes": "compute_value; verify_property; identify_equivalent",
        "answer_kinds": "fraction; yes_no",
        "level1": "skraćivanje očiglednim djeliocem (2, 3, 5)",
        "level2": "potpuno skraćivanje; provjera nesvodivosti",
        "level3": "skraćivanje preko NZD-a s obrazloženjem",
        "prerequisites": "6-04-005; 6-03-008",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "'Nesvodivi razlomak' nijedan izvor ne imenuje izričito — "
                "MAT-BOT dekompozicija (slučaj D), sadržajno neupitna.",
    },
    "6-04-007": {
        "family": "fraction_equivalence",
        "core_skill": "svesti dva ili više razlomaka na zajednički imenilac",
        "actions": "odrediti zajednički imenilac (najmanji ili bilo koji); "
                   "svesti oba razlomka; dopuniti brojioce",
        "concepts": "zajednički imenilac; NZS imenilaca",
        "archetypes": "compute_value; find_missing_value",
        "answer_kinds": "fraction_pair; integer",
        "level1": "imenioci gdje jedan dijeli drugi (2 i 4)",
        "level2": "uzajamno prosti imenioci; NZS kao imenilac",
        "level3": "tri razlomka ili najmanji zajednički imenilac s obrazloženjem",
        "prerequisites": "6-04-005; 6-03-009",
        "confidence": "medium", "status": "needs_manual_review",
        "note": "KS ga vodi kao korak sabiranja (stranica 16), RS implicitno; "
                "kao samostalna lekcija oslonjena na metodički tekst — "
                "potvrditi obim prije ugovora (slučaj B/C).",
    },
    "6-04-008": {
        "family": "fraction_compare_order",
        "core_skill": "uporediti dva razlomka i poredati više razlomaka po "
                      "veličini",
        "actions": "uporediti (isti imenilac / isti brojilac / svođenjem); "
                   "poredati rastuće ili opadajuće; odabrati veći/manji",
        "concepts": "poređenje; uređivanje; odnos prema 1",
        "archetypes": "compare_pair; order_set; select_by_property",
        "answer_kinds": "relation_symbol; ordering; fraction",
        "level1": "isti imenioci ili isti brojioci",
        "level2": "svođenje na zajednički imenilac; poređenje s 1",
        "level3": "uređivanje 3–4 razlomka miješanih oblika",
        "prerequisites": "6-04-005; 6-04-007",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Jednočlana porodica opravdana: radnja (poređenje/uređivanje) i "
                "vrsta odgovora (relacija/redoslijed) različite od ekvivalencije.",
    },
    "6-04-009": {
        "family": "fraction_arithmetic_direct",
        "core_skill": "sabrati i oduzeti razlomke jednakih imenilaca",
        "actions": "sabrati/oduzeti brojioce uz isti imenilac; skratiti rezultat",
        "concepts": "isti imenilac; sabiranje/oduzimanje brojilaca",
        "archetypes": "direct_computation; find_missing_value; identify_error",
        "answer_kinds": "fraction",
        "level1": "jedan zbir/razlika, mali imenioci, bez skraćivanja",
        "level2": "rezultat traži skraćivanje ili mješoviti zapis",
        "level3": "tri člana ili nedostajući član jednakosti",
        "prerequisites": "6-04-001",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Poklapa se s postojećim pilot ugovorom 6-04-009 (K1).",
    },
    "6-04-010": {
        "family": "fraction_arithmetic_direct",
        "core_skill": "sabrati i oduzeti razlomke različitih imenilaca",
        "actions": "svesti na zajednički imenilac pa sabrati/oduzeti; skratiti",
        "concepts": "različiti imenioci; svođenje; NZS imenilaca",
        "archetypes": "direct_computation; find_missing_value; identify_error",
        "answer_kinds": "fraction",
        "level1": "jedan imenilac dijeli drugi",
        "level2": "uzajamno prosti imenioci",
        "level3": "tri člana ili kombinacija sa oduzimanjem i skraćivanjem",
        "prerequisites": "6-04-007; 6-04-009",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Invarijanta ugovora: imenioci RAZLIČITI (denominator_relation).",
    },
    "6-04-011": {
        "family": "fraction_arithmetic_direct",
        "core_skill": "pomnožiti razlomak prirodnim brojem i razlomkom",
        "actions": "pomnožiti brojioce i imenioce; skratiti prije/poslije",
        "concepts": "množenje razlomaka; skraćivanje u toku množenja",
        "archetypes": "direct_computation; find_missing_value; identify_error",
        "answer_kinds": "fraction",
        "level1": "razlomak × prirodan broj, mali brojevi",
        "level2": "razlomak × razlomak uz skraćivanje",
        "level3": "tri faktora ili nedostajući faktor",
        "prerequisites": "6-04-006",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Poklapa se s postojećim pilot ugovorom 6-04-011.",
    },
    "6-04-012": {
        "family": "fraction_arithmetic_direct",
        "core_skill": "podijeliti razlomak prirodnim brojem i razlomkom",
        "actions": "pomnožiti recipročnim; skratiti rezultat",
        "concepts": "recipročan razlomak; dijeljenje kao množenje recipročnim",
        "archetypes": "direct_computation; find_missing_value; identify_error",
        "answer_kinds": "fraction",
        "level1": "razlomak : prirodan broj, mali brojevi",
        "level2": "razlomak : razlomak",
        "level3": "kombinacija množenja i dijeljenja ili nedostajući djelilac",
        "prerequisites": "6-04-011",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "Poklapa se s postojećim pilot ugovorom 6-04-012.",
    },
    "6-04-013": {
        "family": "fraction_operation_properties",
        "core_skill": "primijeniti komutativnost, asocijativnost i "
                      "distributivnost pri računanju s razlomcima",
        "actions": "prepoznati upotrijebljeno svojstvo; odabrati zapis jednak "
                   "datom po svojstvu; iskoristiti svojstvo za lakši račun",
        "concepts": "komutativnost; asocijativnost; distributivnost",
        "archetypes": "identify_equivalent; select_by_property; justify_rule",
        "answer_kinds": "option_statement; fraction",
        "level1": "prepoznavanje svojstva u jednom koraku",
        "level2": "izbor pogodnog pregrupisavanja za lakši račun",
        "level3": "primjena distributivnosti s obrazloženjem",
        "prerequisites": "6-04-009; 6-04-011",
        "confidence": "low", "status": "ready_for_contract_draft",
        "note": 'FAZA 3: izvor JE imenuje — kolona sadržaja nosi „Svojstva sabiranja razlomaka“ i „Svojstva množenja razlomaka“ (KS str. 12; R25-030/031). Obim: komutativnost i asocijativnost sabiranja i množenja; distributivnost KS ne imenuje za razlomke. Enforcement ostaje ADVISORY_ONLY.',
    },
    "6-04-014": {
        "family": "numeric_expressions_multi_step",
        "core_skill": "izračunati vrijednost brojevnog izraza s razlomcima uz "
                      "poštovanje zagrada i redoslijeda operacija",
        "actions": "izračunati izraz sa 2–3 operacije i zagradama; odabrati "
                   "ispravan sljedeći korak",
        "concepts": "redoslijed operacija; zagrade; višekoračni račun",
        "archetypes": "direct_computation; identify_next_step; identify_error",
        "answer_kinds": "fraction",
        "level1": "dvije operacije bez zagrada",
        "level2": "zagrade + dvije-tri operacije",
        "level3": "ugniježđene zagrade ili greška u tuđem postupku",
        "prerequisites": "6-04-009; 6-04-010; 6-04-011; 6-04-012",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "FAZA 3: potvrđeno na KS str. 12 (ne 13):zi sa zagradama' u "
                "jedinici razlomaka) — potvrditi na stranici 13 vizuelno.",
    },
    "6-04-015": {
        "family": "word_problems_pilot",
        "core_skill": "riješiti tekstualni (životni) zadatak s razlomcima",
        "actions": "prepoznati operaciju iz priče; izračunati dio cjeline; "
                   "izračunati cjelinu iz dijela",
        "concepts": "životni kontekst; dio od broja; ostatak",
        "archetypes": "word_problem",
        "answer_kinds": "fraction; integer; short_text",
        "level1": "jednokoračna priča (dio od broja)",
        "level2": "dvokoračna priča (ostatak; poređenje dijelova)",
        "level3": "cjelina iz dijela ili dva uslova",
        "prerequisites": "6-04-009…012",
        "confidence": "high", "status": "ready_for_contract_draft",
        "note": "FAZA 3: potvrđeno — „Tekstualni zadaci“ nom nizu) exact dokaz "
                "je enumeracija 'Tekstualni zadaci' u KS tabeli razlomaka — "
                "slabiji oblik dokaza, potvrditi vizuelno.",
    },
}

# ---------------------------------------------------------------------------
# KORAK 5 — PORODICE (10 porodica; 3 jednočlane s obrazloženjem).
# ---------------------------------------------------------------------------

FAMILIES = {
    "divisibility_predicate_application": {
        "name": "Primjena pravila/stavova djeljivosti",
        "members": ("6-03-002", "6-03-003", "6-03-004"),
        "core": "primijeniti imenovano pravilo ili stav djeljivosti bez "
                "izračunavanja količnika",
        "archetypes": "yes_no_rule_check; select_by_rule; justify_rule; "
                      "digit_completion (samo 6-03-004 L3)",
        "forbidden": "traženje djelilaca/faktora; dijeljenje s ostatkom; "
                     "faktorizacija; NZD/NZS",
        "difficulty": "broj istovremenih pravila (1→2→3); veličina broja; "
                      "da-ne → izbor → dopuna cifre/obrazloženje",
        "answer_kinds": "yes_no; integer; digit; option_statement",
        "required_params": "rule_set (podskup {2,3,4,5,6,9,10,15,25} ∪ {10^k}; "
                           "sum|difference|product za 6-03-002)",
        "optional_params": "raspon broja; broj kandidata",
        "overrides": "6-03-002 (stavovi umjesto cifarskih pravila); 6-03-003 "
                     "(dekadske jedinice, formula_loss)",
        "detectors": "divisibility_predicate (mcq_integrity — POSTOJI); "
                     "integer_options oracle (POSTOJI); digit_completion "
                     "placeholder-guard (POSTOJI kao skip)",
        "model_only": "kvalitet obrazloženja; priča oko brojeva",
        "evidence": "high (004) / high-page (002) / medium+formula_loss (003)",
        "status": "ready_for_contract_draft",
    },
    "common_divisors_multiples": {
        "name": "Djelioci, sadržioci, NZD i NZS",
        "members": ("6-03-001", "6-03-008", "6-03-009"),
        "core": "nabrojati djelioce/sadržioce i odrediti zajedničke te "
                "najveći/najmanji",
        "archetypes": "enumerate_set; verify_membership; compute_value; "
                      "select_by_property; word_context_simple (L3)",
        "forbidden": "pravila djeljivosti kao glavni cilj; faktorizacija kao "
                     "glavni cilj; tekstualni zadatak s pričom (6-03-010)",
        "difficulty": "veličina brojeva; broj brojeva (2→3); nabrajanje → "
                      "NZD/NZS → primjena",
        "answer_kinds": "integer; integer_list; yes_no",
        "required_params": "target (divisors|multiples|gcd|lcm); operand_range",
        "optional_params": "broj operanada; metoda (nabrajanje|faktorizacija)",
        "overrides": "6-03-001 bez NZD/NZS terminologije (to su 008/009)",
        "detectors": "exact integer arithmetic (NZD/NZS provjerljivi); "
                     "set-membership provjera",
        "model_only": "formulacija; izbor konteksta u L3",
        "evidence": "high (001, 008, 009; RS str. 340 za 001)",
        "status": "ready_for_contract_draft",
    },
    "prime_structure": {
        "name": "Prosti brojevi i faktorizacija",
        "members": ("6-03-005", "6-03-006", "6-03-007"),
        "core": "klasifikovati po prostosti, provjeriti uzajamnu prostost, "
                "rastaviti na proste faktore",
        "archetypes": "classify; verify_property; select_by_property; "
                      "decompose; identify_error",
        "forbidden": "NZD/NZS računanje kao glavni cilj; pravila djeljivosti "
                     "kao glavni cilj",
        "difficulty": "veličina broja; klasifikacija → par → potpuna "
                      "faktorizacija/greška",
        "answer_kinds": "yes_no; integer; pair; factorization; option_statement",
        "required_params": "task (prime_check|coprime_check|factorize); range",
        "optional_params": "oblik zapisa faktorizacije",
        "overrides": "—",
        "detectors": "prostost/faktorizacija egzaktno provjerljive; "
                     "coprime = NZD(a,b)==1",
        "model_only": "obrazloženja; distraktori kategorija grešaka",
        "evidence": "high (005, 006, 007)",
        "status": "ready_for_contract_draft",
    },
    "word_problems_pilot": {
        "name": "Tekstualni zadaci (djeljivost; razlomci)",
        "members": ("6-03-010", "6-04-015"),
        "core": "iz životne priče prepoznati model i izračunati odgovor",
        "archetypes": "word_problem",
        "forbidden": "gola računska operacija bez priče; zadatak druge lekcije "
                     "samo umotan u rečenicu",
        "difficulty": "broj koraka (1→2); vrsta modela; dodatni uslov",
        "answer_kinds": "integer; fraction; short_text",
        "required_params": "domain (djeljivost|razlomci); model",
        "optional_params": "kontekst priče",
        "overrides": "—",
        "detectors": "story_context (postoji obrazac u task_family_validation); "
                     "numerička provjera odgovora kad je model poznat",
        "model_only": "kvalitet i prirodnost priče (glavni rizik)",
        "evidence": "medium (obje: enumeracija/vodeća klauzula u širokoj stavci)",
        "status": "needs_manual_review",
    },
    "fraction_concept_representation": {
        "name": "Pojam, vrste i prikazi razlomka",
        "members": ("6-04-001", "6-04-002", "6-04-003", "6-04-004"),
        "core": "prepoznati, imenovati, klasifikovati i predstaviti razlomak "
                "(figura, količnik, brojevna poluprava, mješoviti broj)",
        "archetypes": "recognize; name_parts; classify; convert_representation",
        "forbidden": "računske operacije s razlomcima; ekvivalencija kao "
                     "glavni cilj; poređenje kao glavni cilj",
        "difficulty": "veličina imenioca; jedna → dvije reprezezentacije; "
                      "klasifikacija → konverzija → kombinacija",
        "answer_kinds": "fraction; mixed_number; option_statement; short_text",
        "required_params": "representation (figure|quotient|number_line|words); "
                           "denominator_range",
        "optional_params": "vrsta figure",
        "overrides": "6-04-001 bez L3 (sve dublje napušta lekciju)",
        "detectors": "konverzije egzaktno provjerljive (nepravi↔mješoviti; "
                     "količnik↔razlomak); klasifikacija provjerljiva",
        "model_only": "opis figure riječima (nema slika u Practice)",
        "evidence": "high (001, 002, 003, 004)",
        "status": "ready_for_contract_draft",
    },
    "fraction_equivalence": {
        "name": "Ekvivalencija razlomaka (proširivanje, skraćivanje, zajednički "
                "imenilac)",
        "members": ("6-04-005", "6-04-006", "6-04-007"),
        "core": "proizvesti/prepoznati ekvivalentan zapis razlomka",
        "archetypes": "compute_value; find_missing_value; identify_equivalent; "
                      "verify_property",
        "forbidden": "poređenje veličina kao glavni cilj; sabiranje/oduzimanje "
                     "kao glavni cilj",
        "difficulty": "veličina faktora; smjer (zadat faktor → zadat imenilac → "
                      "nesvodivost/NZS); broj razlomaka",
        "answer_kinds": "fraction; integer; fraction_pair; yes_no",
        "required_params": "direction (expand|reduce|common_denominator); range",
        "optional_params": "denominator_relation (za 007)",
        "overrides": "6-04-007 (dva razlomka istovremeno; L1 ograničen na "
                     "imenioce u odnosu djeljivosti)",
        "detectors": "exact_rational (POSTOJI u contracts/verifiers); "
                     "nesvodivost = NZD(brojilac, imenilac)==1",
        "model_only": "formulacija 'dopuni' zadataka",
        "evidence": "high (005, 006) / medium-metodika (007)",
        "status": "ready_for_contract_draft",
    },
    "fraction_compare_order": {
        "name": "Upoređivanje i uređivanje razlomaka",
        "members": ("6-04-008",),
        "core": "uporediti dva razlomka ili urediti više njih po veličini",
        "archetypes": "compare_pair; order_set; select_by_property",
        "forbidden": "sam ekvivalentan zapis kao odgovor; računanje zbira/"
                     "razlike kao glavni cilj",
        "difficulty": "isti imenioci → isti brojioci → svođenje; 2 → 4 razlomka",
        "answer_kinds": "relation_symbol; ordering; fraction",
        "required_params": "comparison_basis; count",
        "optional_params": "odnos prema 1",
        "overrides": "—",
        "detectors": "egzaktno poređenje racionalnih (POSTOJI)",
        "model_only": "raznolikost formulacija",
        "evidence": "high",
        "status": "ready_for_contract_draft",
        "single_member_reason": "radnja (poređenje/uređivanje) i vrste odgovora "
            "(relacijski simbol, redoslijed) bitno različite i od ekvivalencije "
            "i od računa — spajanje bi zamaglilo upravo granicu koju su živi "
            "nalazi tražili (ekvivalencija vs poređenje)",
    },
    "fraction_arithmetic_direct": {
        "name": "Direktan račun s razlomcima",
        "members": ("6-04-009", "6-04-010", "6-04-011", "6-04-012"),
        "core": "izvesti jednu imenovanu računsku operaciju s razlomcima",
        "archetypes": "direct_computation; find_missing_value; identify_error",
        "forbidden": "jednačine s razlomcima (6-07); brojevni izrazi s više "
                     "operacija (6-04-014); pretvaranje u decimalni/postotni "
                     "zapis (6-05/6-06)",
        "difficulty": "odnos imenilaca; potreba skraćivanja; broj članova (2→3); "
                      "nedostajući član",
        "answer_kinds": "fraction",
        "required_params": "operation (add_sub|mul|div); denominator_relation "
                           "(equal|unlike — invarijanta 009/010)",
        "optional_params": "sign_policy (uvijek non_negative u 6. razredu); "
                           "integer_range",
        "overrides": "6-04-010 zahtijeva RAZLIČITE imenioce (invarijanta)",
        "detectors": "exact_rational (POSTOJI; već aktivan u pilot ugovorima "
                     "009/010/011/012)",
        "model_only": "proza oko zadatka",
        "evidence": "high (sva četiri; KS tabela + RS str. 341)",
        "status": "ready_for_contract_draft",
    },
    "fraction_operation_properties": {
        "name": "Svojstva računskih operacija s razlomcima",
        "members": ("6-04-013",),
        "core": "prepoznati i iskoristiti komutativnost/asocijativnost/"
                "distributivnost",
        "archetypes": "identify_equivalent; select_by_property; justify_rule",
        "forbidden": "goli račun bez svojstva (6-04-009…012)",
        "difficulty": "jedno svojstvo → izbor pregrupisavanja → distributivnost",
        "answer_kinds": "option_statement; fraction",
        "required_params": "property (comm|assoc|distr)",
        "optional_params": "—",
        "overrides": "—",
        "detectors": "ekvivalencija izraza provjerljiva egzaktnim računom "
                     "(ograničeno); izbor svojstva NIJE deterministički dokaziv",
        "model_only": "imenovanje svojstva; obrazloženja",
        "evidence": "low — samo podržavajući dokazi",
        "status": "needs_manual_review",
        "single_member_reason": "jedina lekcija pilota čiji je predmet SVOJSTVO "
            "operacije, a ne rezultat; slaba dokazna podloga traži poseban "
            "ljudski pregled prije bilo kakvog ugovora",
    },
    "numeric_expressions_multi_step": {
        "name": "Brojevni izrazi s razlomcima (zagrade, redoslijed)",
        "members": ("6-04-014",),
        "core": "izračunati višekoračni izraz uz redoslijed operacija",
        "archetypes": "direct_computation; identify_next_step; identify_error",
        "forbidden": "jednačine (6-07); jedna operacija (6-04-009…012)",
        "difficulty": "broj operacija (2→3); zagrade (bez→jedne→ugniježđene)",
        "answer_kinds": "fraction",
        "required_params": "operation_count; bracket_depth",
        "optional_params": "dozvoljene operacije",
        "overrides": "—",
        "detectors": "egzaktna evaluacija izraza (AST, POSTOJI princip u "
                     "mathcheck); identify_next_step NIJE deterministički",
        "model_only": "izbor 'sljedećeg koraka' distraktora",
        "evidence": "medium (enumeracija u KS tabeli + KS-0059)",
        "status": "needs_manual_review",
        "single_member_reason": "u pilotu jedina višekoračna računska lekcija; "
            "porodica se prirodno širi na 6-05-011 i 7-03-014 IZVAN pilota — "
            "jednočlanost je artefakt granice pilota, ne dizajna",
    },
}

# ---------------------------------------------------------------------------
# KORAK 3 — SMJERNE GRANICE (izvorna lekcija → zabranjena zamjena).
# symmetric=True znači da zabrana važi u OBA smjera.
# ---------------------------------------------------------------------------

BOUNDARIES = (
    ("6-03-004", "6-02-005", False,
     "zadatak o pravilima djeljivosti ne smije postati računanje količnika i "
     "ostatka; obrnuto NIJE granica pilota (6-02-005 nije u pilotu) — "
     "ispravlja mašinski red M-KS_2018-0035-02 koji je za susjeda uzeo 6-03-002"),
    ("6-03-004", "6-03-001", True,
     "primjena pravila ↔ traženje djelilaca/faktora: živi gate nalaz "
     "('Koji od ponuđenih brojeva je djelilac broja 84?' nije primjena pravila)"),
    ("6-03-004", "6-03-002", True,
     "cifarska pravila ↔ stavovi o zbiru/razlici/proizvodu: srodno, ali "
     "različit predmet provjere"),
    ("6-03-003", "6-03-004", True,
     "dekadske jedinice (10^k) ↔ pojedinačna pravila (2,5,25…): dekadska "
     "lekcija ne smije degradirati u opšte pravilo i obratno"),
    ("6-03-001", "6-03-008", False,
     "nabrajanje djelilaca ne smije postati NZD zadatak (NZD terminologija "
     "pripada 6-03-008)"),
    ("6-03-001", "6-03-009", False,
     "nabrajanje sadržilaca ne smije postati NZS zadatak"),
    ("6-03-005", "6-03-007", True,
     "klasifikacija prost/složen ↔ potpuna faktorizacija: prepoznavanje nije "
     "rastavljanje"),
    ("6-03-007", "6-03-008", True,
     "faktorizacija je ALAT za NZD, ne sam NZD; NZD zadatak ne smije tražiti "
     "samo faktorizaciju"),
    ("6-03-007", "6-03-009", True, "isto za NZS"),
    ("6-03-006", "6-03-005", False,
     "uzajamna prostost je svojstvo PARA — ne smije postati klasifikacija "
     "jednog broja"),
    ("6-03-008", "6-03-009", True,
     "NZD ↔ NZS: najčešća učenička zamjena; zadatak mora jasno tražiti jedno"),
    ("6-03-010", "6-03-004", False,
     "tekstualni zadatak mora imati stvarnu priču — golo pravilo djeljivosti "
     "umotano u rečenicu nije tekstualni zadatak; obrnuto: 6-03-004 ne smije "
     "dobiti punu priču (to je 6-03-010)"),
    ("6-04-001", "6-04-002", False,
     "imenovanje dijelova zapisa ne smije postati interpretacija dijela "
     "cjeline/količnika"),
    ("6-04-002", "6-04-004", True,
     "dio cjeline (figura) ↔ tačka na brojevnoj polupravoj: dvije različite "
     "reprezentacije"),
    ("6-04-002", "6-02-005", False,
     "razlomak kao količnik a/b ≠ dijeljenje s ostatkom a=bq+r (KS str. 15 "
     "izričito razdvaja)"),
    ("6-04-003", "6-04-001", False,
     "klasifikacija vrsta ne smije degradirati u čisto imenovanje dijelova"),
    ("6-04-005", "6-04-006", True,
     "proširivanje ↔ skraćivanje: suprotni smjerovi ekvivalencije (postojeći "
     "pilot ugovori ovo već razdvajaju kao scaling_direction)"),
    ("6-04-005", "6-04-008", True,
     "ekvivalentan zapis ↔ poređenje veličina: proširivanje u zadatku "
     "poređenja je alat, ne cilj"),
    ("6-04-006", "6-03-008", False,
     "potpuno skraćivanje koristi NZD kao alat — zadatak ne smije postati "
     "'izračunaj NZD'"),
    ("6-04-007", "6-04-010", False,
     "svođenje na zajednički imenilac je PRIPREMA — ne smije postati puno "
     "sabiranje različitih imenilaca (KS ih izričito niže: 'pa preći na "
     "sabiranje')"),
    ("6-04-007", "6-03-009", False,
     "NZS je alat za zajednički imenilac — zadatak ne smije postati goli NZS"),
    ("6-04-008", "6-05-006", True,
     "poređenje razlomaka ↔ poređenje decimalnih zapisa: druga lekcija (i "
     "drugi validator)"),
    ("6-04-009", "6-04-010", True,
     "jednaki ↔ različiti imenioci: invarijanta postojećih ugovora "
     "(denominator_relation)"),
    ("6-04-009", "6-07-002", False,
     "direktan zbir/razlika ne smije postati jednačina x±a=b; obrnuto važi u "
     "6-07 (izvan pilota)"),
    ("6-04-010", "6-04-007", False,
     "puno sabiranje ne smije stati na svođenju — svođenje bez završnog "
     "računa je 6-04-007"),
    ("6-04-011", "6-04-012", True,
     "množenje ↔ dijeljenje (recipročni): postojeća invarijanta "
     "allowed_operations u pilot ugovorima"),
    ("6-04-011", "6-06-002", False,
     "množenje razlomkom ne smije postati 'procenat broja' (6-06-002), iako "
     "je matematika srodna"),
    ("6-04-012", "6-04-011", False,
     "dijeljenje se ne smije zadati tako da se stvarno provjerava samo "
     "množenje (bez recipročnog koraka)"),
    ("6-04-013", "6-04-009", False,
     "zadatak o svojstvu mora ispitivati SVOJSTVO — goli račun je "
     "6-04-009…012"),
    ("6-04-014", "6-07-002", True,
     "izraz ↔ jednačina: izraz nema nepoznatu koju treba odrediti"),
    ("6-04-014", "6-04-009", False,
     "višekoračni izraz ne smije degradirati u jednu operaciju"),
    ("6-04-015", "6-04-009", False,
     "tekstualni zadatak mora imati stvarnu priču — gola operacija u "
     "rečenici nije priča"),
    ("6-04-015", "6-03-010", True,
     "tekstualni s razlomcima ↔ tekstualni iz djeljivosti: domen modela mora "
     "odgovarati lekciji"),
    ("6-04-004", "6-05-005", True,
     "poluprava s razlomcima ↔ poluprava s decimalnim brojevima: zapis "
     "vrijednosti određuje lekciju"),
)

# ---------------------------------------------------------------------------
# KORAK 6 — AUTORSKI PRIMJERI (test-fiksture; NISU zvanični tekst kurikuluma).
# (lesson, level, validity, task, expected_skill|reject_reason, det_check)
# validity: 'valid' | 'invalid_neighbour'; level: 1|2|3|None(za invalid)
# det_check: 'da' | 'djelimično' | 'ne' + kratko čime.
# ---------------------------------------------------------------------------

EXAMPLES = (
    ("6-03-001", 1, "valid", "Koji od ponuđenih brojeva je djelilac broja 84?",
     "prepoznavanje djelioca", "da: cjelobrojna provjera 84 % d == 0"),
    ("6-03-001", 2, "valid", "Koji broj je sadržilac broja 12, a manji od 50?",
     "sadržilac uz dodatni uslov", "da: k·12 i opseg"),
    ("6-03-001", 3, "valid", "Koji broj je istovremeno sadržilac broja 6 i "
     "djelilac broja 36?", "kombinovani odnos djelilac/sadržilac",
     "da: obje relacije egzaktne"),
    ("6-03-001", None, "invalid_neighbour", "Da li je broj 84 djeljiv sa 4? "
     "(primjena pravila)", "pripada 6-03-004: primjena pravila djeljivosti, ne "
     "traženje djelilaca", "da: divisibility_predicate"),

    ("6-03-002", 1, "valid", "Brojevi 12 i 18 su djeljivi sa 3. Da li je njihov "
     "zbir djeljiv sa 3?", "stav o djeljivosti zbira", "da: 3 | 12+18"),
    ("6-03-002", 2, "valid", "Ne računajući vrijednost, odredi da li je "
     "proizvod 15 · 7 djeljiv sa 5.", "stav o djeljivosti proizvoda",
     "da: 5 | 15·7"),
    ("6-03-002", 3, "valid", "Broj n je djeljiv sa 6. Obrazloži zašto je i "
     "n + 12 djeljiv sa 6 i odaberi tačno obrazloženje.",
     "stav + obrazloženje", "djelimično: tačna opcija provjerljiva, "
     "obrazloženje modelsko"),
    ("6-03-002", None, "invalid_neighbour", "Izračunaj 12 + 18 i podijeli "
     "rezultat sa 3.", "goli račun količnika — nije stav o zbiru (6-02-004)",
     "da: story/goal check"),

    ("6-03-003", 1, "valid", "Da li je broj 4300 djeljiv sa 100?",
     "djeljivost dekadskom jedinicom", "da: završne nule"),
    ("6-03-003", 2, "valid", "Koji od ponuđenih brojeva je djeljiv sa 1000?",
     "izbor po broju završnih nula", "da"),
    ("6-03-003", 3, "valid", "Kojom najmanjom dekadskom jedinicom treba "
     "pomnožiti 37 da proizvod bude djeljiv sa 100?",
     "dekadska jedinica + uslov", "da: egzaktno"),
    ("6-03-003", None, "invalid_neighbour", "Da li je 4300 djeljivo sa 4?",
     "pojedinačno pravilo (zadnje dvije cifre) pripada 6-03-004",
     "da: divisibility_predicate"),

    ("6-03-004", 1, "valid", "Da li je broj 234 djeljiv sa 3?",
     "jedno pravilo, da/ne", "da: mcq_integrity oracle POSTOJI"),
    ("6-03-004", 2, "valid", "Koji od ponuđenih brojeva je djeljiv i sa 6 i "
     "sa 25?", "dva istovremena pravila", "da: oracle POSTOJI"),
    ("6-03-004", 3, "valid", "Odredi cifru x tako da broj 3x5 bude djeljiv "
     "sa 9.", "dopuna cifre po pravilu", "djelimično: oracle namjerno "
     "preskače mjestodržač (živi nalaz baef3fd) — provjera uvrštavanjem"),
    ("6-03-004", None, "invalid_neighbour", "Nabroji sve djelioce broja 84.",
     "traženje djelilaca pripada 6-03-001 (živi gate nalaz)", "da"),

    ("6-03-005", 1, "valid", "Da li je broj 17 prost?",
     "klasifikacija prost/složen", "da: egzaktno"),
    ("6-03-005", 2, "valid", "Koji od ponuđenih brojeva je složen: 31, 37, 51, "
     "41?", "izbor složenog broja", "da"),
    ("6-03-005", 3, "valid", "Koja tvrdnja o prostim brojevima je tačna? (npr. "
     "'2 je jedini paran prost broj')", "svojstva prostih brojeva",
     "djelimično: istinitost opcija provjerljiva za konkretne tvrdnje"),
    ("6-03-005", None, "invalid_neighbour", "Rastavi broj 51 na proste "
     "faktore.", "faktorizacija pripada 6-03-007", "da"),

    ("6-03-006", 1, "valid", "Da li su brojevi 8 i 15 uzajamno prosti?",
     "provjera uzajamne prostosti", "da: NZD(8,15)==1"),
    ("6-03-006", 2, "valid", "Koji par brojeva je uzajamno prost: (12,18), "
     "(9,16), (10,15), (14,21)?", "izbor uzajamno prostog para", "da"),
    ("6-03-006", 3, "valid", "Odredi najmanji broj veći od 10 koji je uzajamno "
     "prost sa 12.", "uslov + uzajamna prostost", "da: egzaktno"),
    ("6-03-006", None, "invalid_neighbour", "Da li je broj 15 prost?",
     "klasifikacija JEDNOG broja pripada 6-03-005 — uzajamna prostost je "
     "svojstvo para", "da"),

    ("6-03-007", 1, "valid", "Koji zapis je potpuna faktorizacija broja 36?",
     "prepoznavanje potpune faktorizacije", "da: egzaktno"),
    ("6-03-007", 2, "valid", "Rastavi broj 126 na proste faktore.",
     "faktorizacija većeg broja", "da"),
    ("6-03-007", 3, "valid", "Učenik je zapisao 60 = 2 · 3 · 10. Šta je "
     "pogriješio?", "greška u faktorizaciji (10 nije prost)",
     "djelimično: kategorija greške strukturno izvodiva (princip POSTOJI u "
     "contracts/verifiers)"),
    ("6-03-007", None, "invalid_neighbour", "Odredi NZD brojeva 36 i 60 preko "
     "faktorizacije.", "NZD pripada 6-03-008 — faktorizacija je samo alat",
     "da"),

    ("6-03-008", 1, "valid", "Odredi najveći zajednički djelilac brojeva 12 "
     "i 18.", "NZD nabrajanjem", "da: egzaktno"),
    ("6-03-008", 2, "valid", "Odredi NZD brojeva 36, 60 i 84.",
     "NZD tri broja", "da"),
    ("6-03-008", 3, "valid", "Od 36 crvenih i 60 plavih perlica prave se "
     "jednake narukvice bez ostatka. Najviše koliko narukvica?",
     "NZD u kratkoj primjeni", "da: NZD(36,60)"),
    ("6-03-008", None, "invalid_neighbour", "Odredi najmanji zajednički "
     "sadržilac brojeva 12 i 18.", "NZS pripada 6-03-009 — najčešća zamjena",
     "da"),

    ("6-03-009", 1, "valid", "Odredi najmanji zajednički sadržilac brojeva 4 "
     "i 6.", "NZS nabrajanjem", "da"),
    ("6-03-009", 2, "valid", "Odredi NZS brojeva 8, 12 i 20.",
     "NZS tri broja", "da"),
    ("6-03-009", 3, "valid", "Dva autobusa kreću zajedno; jedan polazi svakih "
     "12, drugi svakih 18 minuta. Nakon koliko minuta ponovo kreću zajedno?",
     "NZS u kratkoj primjeni", "da: NZS(12,18)"),
    ("6-03-009", None, "invalid_neighbour", "Odredi NZD brojeva 8 i 12.",
     "NZD pripada 6-03-008", "da"),

    ("6-03-010", 1, "valid", "U kutiji je 24 bombona. Može li se pravedno "
     "podijeliti na 6 djece bez ostatka?", "priča → djeljivost",
     "da: 6 | 24 + story check"),
    ("6-03-010", 2, "valid", "Pločice dimenzija 12 cm slažu se u red dužine "
     "180 cm bez rezanja. Koliko pločica stane?",
     "priča → model dijeljenja bez ostatka", "da"),
    ("6-03-010", 3, "valid", "Tri zvona zvone svakih 6, 9 i 12 minuta i sad su "
     "zazvonila zajedno. Za koliko minuta će prvi put ponovo zajedno, ako je "
     "to prije podneva?", "priča → NZS + dodatni uslov", "da: NZS + uslov"),
    ("6-03-010", None, "invalid_neighbour", "Da li je 180 djeljivo sa 12?",
     "golo pravilo bez priče pripada 6-03-004", "da: story_context"),

    ("6-04-001", 1, "valid", "U zapisu 3/7, kako se zove broj 7?",
     "imenovanje imenioca", "da: zatvoren skup naziva"),
    ("6-04-001", 2, "valid", "Zapiši razlomak 'pet devetina'.",
     "zapis iz riječi", "da: egzaktno"),
    ("6-04-001", 3, "valid", "(nepodržan nivo — sve dublje napušta lekciju)",
     "—", "—"),
    ("6-04-001", None, "invalid_neighbour", "Koji dio figure je osjenčen?",
     "interpretacija dijela cjeline pripada 6-04-002", "djelimično"),

    ("6-04-002", 1, "valid", "Pizza je podijeljena na 8 jednakih dijelova i "
     "pojedena su 3. Koji razlomak pizze je pojeden?",
     "dio cjeline iz opisa", "da: 3/8"),
    ("6-04-002", 2, "valid", "Zapiši količnik 3 : 4 u obliku razlomka.",
     "količnik kao razlomak", "da: egzaktno"),
    ("6-04-002", 3, "valid", "Marko je podijelio 2 čokolade na 5 drugova "
     "jednako. Koji razlomak čokolade dobija svaki?",
     "količnik u kontekstu", "da: 2/5"),
    ("6-04-002", None, "invalid_neighbour", "Izračunaj 3 : 4 s ostatkom.",
     "dijeljenje s ostatkom pripada 6-02-005 (KS str. 15 izričito razdvaja)",
     "da"),

    ("6-04-003", 1, "valid", "Da li je razlomak 7/4 pravi ili nepravi?",
     "klasifikacija", "da: 7>4"),
    ("6-04-003", 2, "valid", "Zapiši 11/4 u obliku mješovitog broja.",
     "nepravi → mješoviti", "da: egzaktno"),
    ("6-04-003", 3, "valid", "Koji od zapisa je prividan razlomak i čemu je "
     "jednak: 12/4, 7/5, 5/7, 9/2?", "prividni razlomci", "da"),
    ("6-04-003", None, "invalid_neighbour", "Skrati razlomak 12/4.",
     "skraćivanje pripada 6-04-006", "da"),

    ("6-04-004", 1, "valid", "Poluprava je podijeljena na četvrtine. Koji "
     "razlomak je pridružen trećoj tački iza nule?",
     "čitanje s poluprave", "da: 3/4"),
    ("6-04-004", 2, "valid", "Između kojih uzastopnih prirodnih brojeva na "
     "polupravoj leži 7/3?", "smještanje mješovitog", "da: 2 i 3"),
    ("6-04-004", 3, "valid", "Poredaj tačke pridružene razlomcima 1/2, 5/4 i "
     "3/4 po udaljenosti od nule.", "više tačaka", "da: poredak"),
    ("6-04-004", None, "invalid_neighbour", "Na polupravoj označi 0,75.",
     "decimalni zapis pripada 6-05-005", "da: oblik zapisa"),

    ("6-04-005", 1, "valid", "Proširi razlomak 2/3 brojem 4.",
     "proširivanje zadatim faktorom", "da: exact_rational (aktivni ugovor)"),
    ("6-04-005", 2, "valid", "Proširi 3/5 tako da imenilac bude 30.",
     "proširivanje na zadani imenilac", "da"),
    ("6-04-005", 3, "valid", "Koji su svi zapisi jednaki razlomku 2/3: 4/6, "
     "6/9, 8/15, 10/15?", "prepoznavanje svih ekvivalentnih", "da"),
    ("6-04-005", None, "invalid_neighbour", "Koji je razlomak veći: 2/3 ili "
     "4/5?", "poređenje pripada 6-04-008 (živa granica ekvivalencija↔poređenje)",
     "da"),

    ("6-04-006", 1, "valid", "Skrati razlomak 6/8.",
     "skraćivanje očiglednim djeliocem", "da: exact_rational"),
    ("6-04-006", 2, "valid", "Skrati 36/48 do nesvodivog oblika.",
     "potpuno skraćivanje", "da"),
    ("6-04-006", 3, "valid", "Da li je razlomak 35/64 nesvodiv? Obrazloži "
     "izborom tačnog razloga.", "nesvodivost preko zajedničkih djelilaca",
     "djelimično: nesvodivost egzaktna, obrazloženje modelsko"),
    ("6-04-006", None, "invalid_neighbour", "Proširi razlomak 3/4 brojem 5.",
     "suprotan smjer — pripada 6-04-005 (scaling_direction invarijanta)",
     "da"),

    ("6-04-007", 1, "valid", "Svedi razlomke 1/2 i 3/4 na zajednički imenilac.",
     "imenioci u odnosu djeljivosti", "da: exact_rational"),
    ("6-04-007", 2, "valid", "Svedi 2/3 i 3/5 na najmanji zajednički imenilac.",
     "uzajamno prosti imenioci", "da"),
    ("6-04-007", 3, "valid", "Svedi 1/4, 5/6 i 3/8 na najmanji zajednički "
     "imenilac.", "tri razlomka preko NZS", "da"),
    ("6-04-007", None, "invalid_neighbour", "Izračunaj 2/3 + 3/5.",
     "puno sabiranje pripada 6-04-010 — svođenje je samo priprema (KS str. "
     "16: 'pa preći na sabiranje')", "da"),

    ("6-04-008", 1, "valid", "Koji je razlomak veći: 3/7 ili 5/7?",
     "isti imenioci", "da: egzaktno poređenje"),
    ("6-04-008", 2, "valid", "Uporedi 5/6 i 7/9.",
     "svođenje pa poređenje", "da"),
    ("6-04-008", 3, "valid", "Poredaj po veličini: 2/3, 3/4, 5/8, 7/12.",
     "uređivanje četiri razlomka", "da: poredak egzaktan"),
    ("6-04-008", None, "invalid_neighbour", "Proširi 2/3 na imenilac 12.",
     "ekvivalentan zapis pripada 6-04-005", "da"),

    ("6-04-009", 1, "valid", "Izračunaj 2/7 + 3/7.",
     "zbir jednakih imenilaca", "da: aktivni ugovor 6-04-009"),
    ("6-04-009", 2, "valid", "Izračunaj 7/10 − 3/10 i skrati rezultat.",
     "razlika + skraćivanje", "da"),
    ("6-04-009", 3, "valid", "Odredi nedostajući razlomak: 3/8 + □ = 7/8.",
     "nedostajući član", "da"),
    ("6-04-009", None, "invalid_neighbour", "Riješi jednačinu x + 3/8 = 7/8.",
     "jednačina pripada 6-07-002 (granica račun↔jednačina)", "da"),

    ("6-04-010", 1, "valid", "Izračunaj 1/2 + 1/4.",
     "imenioci u odnosu djeljivosti", "da: aktivni ugovor 6-04-010"),
    ("6-04-010", 2, "valid", "Izračunaj 2/3 + 3/5.",
     "uzajamno prosti imenioci", "da"),
    ("6-04-010", 3, "valid", "Izračunaj 5/6 − 3/4 + 1/3.",
     "tri člana", "da"),
    ("6-04-010", None, "invalid_neighbour", "Izračunaj 2/9 + 5/9.",
     "jednaki imenioci pripadaju 6-04-009 (invarijanta denominator_relation)",
     "da: denominator_relation provjerljiv"),

    ("6-04-011", 1, "valid", "Izračunaj 3 · 2/5.",
     "razlomak puta prirodan broj", "da: aktivni ugovor 6-04-011"),
    ("6-04-011", 2, "valid", "Izračunaj 3/4 · 8/9 i skrati.",
     "razlomak puta razlomak", "da"),
    ("6-04-011", 3, "valid", "Odredi nedostajući faktor: 2/3 · □ = 1/2.",
     "nedostajući faktor", "da"),
    ("6-04-011", None, "invalid_neighbour", "Koliko je 40% od 60?",
     "procenat broja pripada 6-06-002, iako je račun srodan množenju "
     "razlomkom", "da"),

    ("6-04-012", 1, "valid", "Izračunaj 4/5 : 2.",
     "razlomak podijeljen prirodnim brojem", "da: aktivni ugovor 6-04-012"),
    ("6-04-012", 2, "valid", "Izračunaj 3/8 : 9/4.",
     "dijeljenje razlomkom (recipročni)", "da"),
    ("6-04-012", 3, "valid", "Odredi djelilac: 5/6 : □ = 5/12.",
     "nedostajući djelilac", "da"),
    ("6-04-012", None, "invalid_neighbour", "Izračunaj 3/8 · 4/9.",
     "množenje pripada 6-04-011 (allowed_operations invarijanta)", "da"),

    ("6-04-013", 1, "valid", "Koje svojstvo je upotrijebljeno u zapisu "
     "2/5 + 3/7 = 3/7 + 2/5?", "prepoznavanje komutativnosti",
     "djelimično: jednakost egzaktna, IME svojstva modelsko"),
    ("6-04-013", 2, "valid", "Koji redoslijed računanja najlakše daje "
     "vrijednost izraza 2/7 + 5/9 + 5/7?", "asocijativno pregrupisavanje",
     "djelimično"),
    ("6-04-013", 3, "valid", "Primjenom distributivnosti izračunaj "
     "3/4 · (8/9 + 4/9).", "distributivnost", "da: rezultat egzaktan"),
    ("6-04-013", None, "invalid_neighbour", "Izračunaj 2/7 + 5/7.",
     "goli račun bez svojstva pripada 6-04-009", "ne: razlika je u CILJU "
     "pitanja — modelska granica"),

    ("6-04-014", 1, "valid", "Izračunaj 1/2 + 1/3 · 3/4.",
     "redoslijed operacija bez zagrada", "da: egzaktna evaluacija"),
    ("6-04-014", 2, "valid", "Izračunaj (2/3 − 1/4) · 6/5.",
     "zagrade pa množenje", "da"),
    ("6-04-014", 3, "valid", "Učenik je u izrazu 1/2 + 1/3 · 3/4 prvo sabrao. "
     "Koji je ispravan prvi korak?", "greška redoslijeda",
     "djelimično: tačan rezultat egzaktan, 'sljedeći korak' modelski"),
    ("6-04-014", None, "invalid_neighbour", "Riješi jednačinu (x − 1/4) · 2 "
     "= 1/2.", "jednačina pripada 6-07 (izraz nema nepoznatu)", "da"),

    ("6-04-015", 1, "valid", "Ana je pojela 2/8 torte. Koji dio torte je "
     "ostao?", "priča → dopuna do cjeline", "da: 1 − 2/8 + story check"),
    ("6-04-015", 2, "valid", "Od 24 učenika njih 3/8 trenira košarku. Koliko "
     "učenika trenira?", "dio od broja u priči", "da: 24·3/8"),
    ("6-04-015", 3, "valid", "Biciklista je prešao 2/5 puta, pa još 12 km, i "
     "stigao na pola puta. Koliko je put dug?", "cjelina iz dijela",
     "da: jednačinski model provjerljiv uvrštavanjem"),
    ("6-04-015", None, "invalid_neighbour", "Izračunaj 3/8 · 24.",
     "gola operacija bez priče pripada 6-04-011", "da: story_context"),
)

# ---------------------------------------------------------------------------
# IZGRADNJA WORKBOOKA
# ---------------------------------------------------------------------------


def load_phase2():
    wb = openpyxl.load_workbook(PHASE2_XLSX, read_only=True, data_only=True)
    mapping_rows = list(wb["Mapiranje"].iter_rows(values_only=True))[1:]
    item_rows = list(wb["Stavke_NPP"].iter_rows(values_only=True))[1:]
    wb.close()
    return mapping_rows, item_rows


def load_phase1_pages():
    wb = openpyxl.load_workbook(bcm.PHASE1_XLSX, read_only=True, data_only=True)
    pages = {}
    for sheet in ("Stranice_KS", "Stranice_RS"):
        for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
            src, _grade, page, _printed, _n, orig, latin = r
            pages[(str(src), int(page))] = (str(orig or ""), str(latin or orig or ""))
    wb.close()
    return pages


def pilot_lessons():
    lessons = [l for l in bcm.load_lessons()
               if l.grade == PILOT_GRADE and l.oblast in PILOT_AREAS]
    assert len(lessons) == 25
    return lessons


def collect_final_evidence(pilot_ids, mapping_rows):
    """(prihvaćeni Faza-2 dokazi + oporavljeni) po lekciji, po vrsti."""
    per_lesson = defaultdict(lambda: defaultdict(list))
    rejected = defaultdict(list)
    unresolved = defaultdict(list)
    for r in mapping_rows:
        if r[8] not in pilot_ids:
            continue
        item_id, relation = r[1], r[12]
        verdict, new_rel, _newconf, _reason, _flags = verdict_for(r)
        final_rel = new_rel or relation
        if verdict in ("reject", "wrong_lesson", "wrong_grade", "duplicate_evidence"):
            rejected[r[8]].append(r[0])
        elif verdict == "ambiguous":
            unresolved[r[8]].append(r[0])
        else:
            per_lesson[r[8]][final_rel].append((item_id, r[2]))
    for rec in RECOVERED_EVIDENCE:
        rid, origin, source, page, lesson, relation, _conf, _quote, _flags = rec
        label = source if origin == "stavka" else f"{source.split('_')[0]}:str.{page}"
        per_lesson[lesson][relation].append((label, source.split("_")[0] if origin == "stranica" else source.split("-")[0]))
    return per_lesson, rejected, unresolved


def build(dry_run=False, out_path=OUTPUT_XLSX):
    mapping_rows, item_rows = load_phase2()
    pages = load_phase1_pages()
    lessons = pilot_lessons()
    pilot_ids = {l.lesson_id for l in lessons}
    lessons_by_id = {l.lesson_id: l for l in lessons}
    items_by_id = {r[0]: r for r in item_rows}
    mapping_by_id = {r[0]: r for r in mapping_rows}

    pilot_mapping_rows = [r for r in mapping_rows if r[8] in pilot_ids]
    per_lesson, rejected, unresolved = collect_final_evidence(pilot_ids, mapping_rows)

    input_hashes = [
        ("kanonski", bcm._sha256(bcm.CANONICAL_XLSX)),
        ("faza1", bcm._sha256(bcm.PHASE1_XLSX)),
        ("faza2", bcm._sha256(PHASE2_XLSX)),
    ]

    sheets = {}

    # --- README ---
    sheets["README"] = (("MAT-BOT Faza 2.5",), [
        ("Svrha", "Stručni pregled automatskog mapiranja Faze 2 za 25 pilot "
                  "lekcija (6. razred: Djeljivost brojeva, Razlomci) + nacrt "
                  "semantike po lekciji i porodici."),
        ("Ulazi", "kanonski workbook + Faza 1 + Faza 2 (svi samo čitani); "
                  "SHA-256: " + "; ".join(f"{n}={h[:16]}…" for n, h in input_hashes)),
        ("Presude", "accept / change_relation / reject / ambiguous / "
                    "broad_but_valid / split_required / missing_formula / "
                    "wrong_lesson / wrong_grade / duplicate_evidence — svaka od "
                    "62 pilot veze Faze 2 ima presudu; odbijeni redovi su "
                    "ZADRŽANI s razlogom, ništa nije tiho obrisano."),
        ("Oporavljeni dokazi", "R25-### redovi su nalaz OVOG pregleda (stavka "
                    "koju je Faza 2 promašila ili doslovan citat sa stranice "
                    "Faze 1). Nikad se ne prikazuju kao izlaz Faze 2."),
        ("Stranični dokaz", "origin='stranica' nosi izvor+stranicu+doslovan "
                    "citat; Kontrola provjerava da citat stvarno postoji u "
                    "tekstu te stranice u Fazi 1."),
        ("Formula_loss", "Zapis izgubljen u PDF ekstrakciji (npr. '10n' za "
                    "10^n) se NE rekonstruiše — označen je za ručnu vizuelnu "
                    "provjeru u PDF-u; lekcija 6-03-003 je zbog toga "
                    "blocked_by_formula_loss."),
        ("Primjeri", "List Primjeri_Pilot sadrži AUTORSKE TEST-FIKSTURE "
                    "(pisane za ovaj pregled, konzistentne s potvrđenim "
                    "dokazima). NISU zvanični tekst kurikuluma i nikad se ne "
                    "smiju citirati kao KS/RS sadržaj."),
        ("auto_high_confidence", "oznaka Faze 2 znači samo 'deterministički "
                    "visok skor' — NIJE ljudska potvrda; ljudsku potvrdu nosi "
                    "tek human_review_status u ovom workbooku."),
        ("Statusi lekcije", "ready_for_contract_draft / needs_manual_review / "
                    "blocked_by_source_gap / blocked_by_formula_loss / "
                    "disputed_mapping"),
        ("Granice", "List Granice_lekcija je SMJERAN: red kaže 'zadatak "
                    "lekcije A ne smije skliznuti u vještinu B'; symmetric=da "
                    "znači da važi i obratno."),
        ("Ništa se ne aktivira", "Ovo su prijedlozi za izradu ugovora — bez "
                    "produkcijskog JSON-a, bez izmjena runtime koda."),
    ])

    # --- Lekcije_Pilot25 ---
    lesson_rows = []
    for lesson in lessons:
        sem = LESSON_SEMANTICS[lesson.lesson_id]
        ev = per_lesson.get(lesson.lesson_id, {})

        def _ids(relation, source_prefix=None):
            values = [label for label, src in ev.get(relation, ())
                      if source_prefix is None or str(src).startswith(source_prefix)]
            return "; ".join(sorted(set(values)))

        gap = lesson.lesson_id in GAP_LESSONS
        gap_case, gap_reason = GAP_RESOLUTIONS.get(lesson.lesson_id, ("", ""))
        lesson_rows.append([
            lesson.lesson_id, lesson.grade, lesson.oblast, lesson.title,
            sem["family"], sem["core_skill"], sem["actions"], sem["concepts"],
            sem["archetypes"],
            _boundary_summary(lesson.lesson_id),
            sem["prerequisites"], sem["answer_kinds"],
            sem["level1"], sem["level2"], sem["level3"],
            _ids("exact", "KS"), _ids("exact", "RS"),
            "; ".join(sorted(set(
                [label for label, _ in ev.get("supporting", ())]
                + [label for label, _ in ev.get("prerequisite", ())]))),
            "; ".join(sorted(rejected.get(lesson.lesson_id, []))),
            "; ".join(sorted(unresolved.get(lesson.lesson_id, []))),
            "da" if gap else "ne",
            (f"slučaj {gap_case}: {gap_reason}" if gap else ""),
            sem["confidence"], sem["status"], sem["note"],
        ])
    sheets["Lekcije_Pilot25"] = (
        ("lesson_id", "grade", "area", "lesson_title", "candidate_family_id",
         "canonical_core_skill", "required_student_actions",
         "required_visible_concepts", "allowed_task_archetypes",
         "forbidden_neighbour_skills", "prerequisite_skills",
         "expected_answer_kinds", "Level_1_scope", "Level_2_scope",
         "Level_3_scope", "KS_exact_evidence_ids", "RS_exact_evidence_ids",
         "supporting_evidence_ids", "rejected_mapping_ids",
         "unresolved_evidence_ids", "source_gap", "source_gap_reason",
         "semantic_confidence", "human_review_status", "reviewer_note"),
        lesson_rows,
    )

    # --- Mapiranja_Pregled: svih 62 originalna reda + presuda + oporavljeni ---
    review_rows = []
    for r in sorted(pilot_mapping_rows, key=lambda r: (r[8], r[0])):
        verdict, new_rel, new_conf, reason, flags = verdict_for(r)
        review_rows.append([
            r[0], r[1], r[2], r[3], r[8], r[11], r[12], r[13], r[14], r[17],
            verdict, new_rel or ("" if verdict in ("reject", "wrong_lesson") else r[12]),
            new_conf or ("" if verdict in ("reject", "wrong_lesson") else r[13]),
            ", ".join(flags), reason,
        ])
    for rec in RECOVERED_EVIDENCE:
        rid, origin, source, page, lesson, relation, conf, quote, flags = rec
        lesson_title = lessons_by_id[lesson].title
        review_rows.append([
            rid, source if origin == "stavka" else "",
            source if origin == "stranica" else source.split("-")[0],
            PILOT_GRADE, lesson, lesson_title,
            "", "", "review_recovery", "recovered",
            "recovered", relation, conf,
            ", ".join(("page_evidence",) + flags if origin == "stranica" else flags),
            (f"str. {page}: „{quote}“" if origin == "stranica" else quote),
        ])
    sheets["Mapiranja_Pregled"] = (
        ("mapping_id", "item_id", "source_id", "source_grade", "target_lesson_id",
         "target_lesson_title", "phase2_relation", "phase2_confidence",
         "phase2_method", "phase2_status", "review_verdict", "final_relation",
         "final_confidence", "flags", "review_reason"),
        review_rows,
    )

    # --- Dokazi_KS / Dokazi_RS (puni izvorni tekst, oba pisma za RS) ---
    def _evidence_sheet(source_prefix):
        rows = []
        seen = set()
        for r in sorted(pilot_mapping_rows, key=lambda r: (r[1], r[0])):
            verdict = verdict_for(r)[0]
            if not str(r[2]).startswith(source_prefix):
                continue
            item = items_by_id.get(r[1])
            if item is None or r[1] in seen:
                continue
            seen.add(r[1])
            rows.append([r[1], item[3], item[4], item[7], item[8],
                        "korišten" if verdict not in ("reject", "wrong_lesson")
                         else "svi redovi odbijeni ili preusmjereni"])
        for rec in RECOVERED_EVIDENCE:
            rid, origin, source, page, lesson, relation, conf, quote, flags = rec
            if origin == "stavka" and source.startswith(source_prefix):
                item = items_by_id[source]
                if source not in seen:
                    seen.add(source)
                    rows.append([source, item[3], item[4], item[7], item[8],
                                 f"oporavljen ({rid})"])
            elif origin == "stranica" and source.startswith(source_prefix):
                orig, latin = pages[(source, page)]
                idx = latin.find(quote)
                snippet_latin = latin[max(0, idx - 60): idx + len(quote) + 60] if idx >= 0 else quote
                snippet_orig = orig[max(0, idx - 60): idx + len(quote) + 60] if idx >= 0 else ""
                rows.append([f"{rid} (stranica {page})", page, "stranični tekst",
                             snippet_orig.strip(), snippet_latin.strip(),
                             f"oporavljen stranični dokaz za {lesson}"])
        return rows

    header_ev = ("evidence_id", "pdf_page", "sekcija", "originalni_tekst",
                 "latinica_normalizovano", "status_u_pregledu")
    sheets["Dokazi_KS"] = (header_ev, _evidence_sheet("KS"))
    sheets["Dokazi_RS"] = (header_ev, _evidence_sheet("RS"))

    # --- Porodice_Pilot ---
    family_rows = []
    for family_id, fam in FAMILIES.items():
        family_rows.append([
            family_id, fam["name"], "; ".join(fam["members"]), fam["core"],
            fam["archetypes"], fam["forbidden"], fam["difficulty"],
            fam["answer_kinds"], fam["required_params"], fam["optional_params"],
            fam["overrides"], fam["detectors"], fam["model_only"],
            fam["evidence"], fam["status"],
            fam.get("single_member_reason", ""),
        ])
    family_rows.sort(key=lambda r: r[0])
    sheets["Porodice_Pilot"] = (
        ("family_id", "family_name", "member_lesson_ids", "shared_core_skill",
         "shared_allowed_archetypes", "shared_forbidden_neighbours",
         "difficulty_dimensions", "expected_answer_kinds", "required_parameters",
         "optional_parameters", "lessons_requiring_override",
         "deterministic_detector_candidates", "model_guidance_only",
         "evidence_quality", "review_status", "single_member_justification"),
        family_rows,
    )

    # --- Granice_lekcija ---
    boundary_rows = []
    for index, (source, forbidden, symmetric, why) in enumerate(BOUNDARIES, 1):
        boundary_rows.append([
            f"GR-{index:03d}", source,
            lessons_by_id[source].title if source in lessons_by_id else source,
            forbidden,
            lessons_by_id[forbidden].title if forbidden in lessons_by_id
            else bcm_title(forbidden),
            "da" if symmetric else "ne (smjerna)", why,
        ])
    sheets["Granice_lekcija"] = (
        ("boundary_id", "source_lesson_id", "source_lesson_title",
         "forbidden_replacement_id", "forbidden_replacement_title",
         "symmetric", "why_confusion_is_wrong"),
        boundary_rows,
    )

    # --- Primjeri_Pilot ---
    example_rows = []
    counters = Counter()
    for lesson, level, validity, task, skill_or_reason, det in EXAMPLES:
        counters[lesson] += 1
        example_id = f"EX-{lesson}-{counters[lesson]:02d}"
        unsupported = level == 3 and task.startswith("(nepodržan")
        example_rows.append([
            lesson, example_id,
            level if level is not None else "",
            ("level3_unsupported" if unsupported else validity),
            task,
            skill_or_reason if validity == "valid" else "",
            skill_or_reason if validity != "valid" else "",
            "autorska test-fikstura (Faza 2.5); uporište: potvrđeni dokazi "
            "lekcije — NIJE zvanični tekst kurikuluma",
            det, "authored_needs_review",
        ])
    sheets["Primjeri_Pilot"] = (
        ("lesson_id", "example_id", "level", "validity", "task_text",
         "expected_skill", "rejection_reason", "source_basis",
         "deterministic_check_possible", "review_status"),
        example_rows,
    )

    # --- Praznine ---
    gap_rows = []
    for lesson_id in GAP_LESSONS:
        case, reason = GAP_RESOLUTIONS[lesson_id]
        sem = LESSON_SEMANTICS[lesson_id]
        gap_rows.append([
            lesson_id, lessons_by_id[lesson_id].title, f"slučaj {case}", reason,
            sem["status"],
            "riješeno pregledom" if sem["status"] == "ready_for_contract_draft"
            else "ostaje uslovno — vidi status",
        ])
    gap_rows.append([
        "", "OSTALA ZAPAŽANJA", "", "", "", "",
    ])
    for note in (
        "KS_2018-0073 ('rješavati tekstualne zadatke', decimalni niz ishoda) "
        "pravi dom je 6-05-011 — IZVAN pilota; oba pilot reda odbijena kao "
        "wrong_lesson, preporuka za Fazu 3: preusmjeriti.",
        "KS_2018-0034 (količnik i a=bq+r) ostaje bez pilot cilja: pripada "
        "kontekstu dijeljenja s ostatkom (6-02-005), ne razlomcima.",
        "RS koristi termin 'činilac' za faktor (str. 340) — dodati u "
        "terminološke aliase pri izradi ugovora.",
        "Skupovi djelilaca u pravilima: KS {2,3,4,5,6,9,25,10^n} bez 15; "
        "MAT-BOT naslov ima 15 — pokriti kombinacijom pravila za 3 i 5.",
        "Zapis '10n' (10^n) i uslov 'n ∈ …' izgubljeni u ekstrakciji (KS str. "
        "8): NE rekonstruisati iz teksta — ručna vizuelna provjera PDF-a.",
    ):
        gap_rows.append(["", "zapažanje", "", note, "", ""])
    sheets["Praznine"] = (
        ("lesson_id", "lesson_title", "gap_case", "resolution",
         "human_review_status", "outcome"),
        gap_rows,
    )

    # --- Provjera_izvora (Faza 3, obavezna provjera originalnog PDF-a) ---
    sheets["Provjera_izvora"] = (
        ("lesson_id", "tema_provjere", "nalaz_tacno_kako_se_vidi", "metod",
         "posljedica_za_lekciju"),
        [list(row) for row in SOURCE_VERIFICATION],
    )

    # --- Aktivacija (Faza 3) ---
    activation_rows = []
    for lesson in lessons:
        klasa, why = ACTIVATION[lesson.lesson_id]
        detector = PROVEN_DETECTORS.get(lesson.lesson_id, "")
        ev = per_lesson.get(lesson.lesson_id, {})
        unresolved_note = "; ".join(sorted(unresolved.get(lesson.lesson_id, []))) or "-"
        activation_rows.append([
            lesson.lesson_id, lesson.title,
            LESSON_SEMANTICS[lesson.lesson_id]["status"], klasa,
            "da" if (klasa == "READY" and detector) else "ne",
            detector or "nema dokazanog detektora danas",
            ("blokirajuća semantika dozvoljena tek kad detektor bude "
             "implementiran i dokazan" if klasa == "READY" and not detector
             else "samo kompaktno vođenje prompta; bez novog determinističkog "
                  "odbijanja" if klasa == "ADVISORY_ONLY"
             else "zadržava postojeće dokazano ponašanje"),
            why,
            len(ev.get("exact", ())), len(ev.get("supporting", ())),
            unresolved_note,
            "; ".join(sorted(rejected.get(lesson.lesson_id, []))) or "-",
        ])
    sheets["Aktivacija"] = (
        ("lesson_id", "lesson_title", "human_review_status", "activation_class",
         "enforcement_allowed_now", "proven_detector_today", "enforcement_policy",
         "justification", "exact_evidence_count", "supporting_evidence_count",
         "unresolved_notes", "rejected_mapping_ids"),
        activation_rows,
    )

    # --- Kvalitet_Mapiranja ---
    quality_rows = _quality_metrics(pilot_mapping_rows)
    sheets["Kvalitet_Mapiranja"] = (
        ("kategorija", "vrijednost", "napomena"), quality_rows)

    # --- Kontrola ---
    checks = _run_checks(pilot_mapping_rows, mapping_by_id, items_by_id,
                         pages, lessons, per_lesson)
    content_digest = bcm._content_hash(sheets)
    kontrola_rows = [(name, value, status) for name, value, status in checks]
    kontrola_rows.append(("hash sadržaja (svi listovi, prije upisa)",
                          content_digest, "INFO"))
    kontrola_rows.append(("ulazni SHA-256",
                          "; ".join(f"{n}={h[:16]}…" for n, h in input_hashes),
                          "INFO"))
    sheets["Kontrola"] = (("provjera", "vrijednost", "status"), kontrola_rows)

    failed = [row for row in kontrola_rows if row[2] == "FAIL"]
    if failed:
        raise SystemExit(f"Kontrolne provjere pale: {failed}")

    if dry_run:
        return sheets, content_digest, None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "MAT-BOT Faza 2.5 review"
    wb.properties.created = bcm._FIXED_DOC_DATE
    wb.properties.modified = bcm._FIXED_DOC_DATE
    wb.properties.lastModifiedBy = "review_pilot_curriculum_mapping.py"
    order = ("README", "Lekcije_Pilot25", "Aktivacija", "Provjera_izvora",
             "Mapiranja_Pregled", "Dokazi_KS", "Dokazi_RS", "Porodice_Pilot",
             "Granice_lekcija", "Primjeri_Pilot", "Praznine",
             "Kvalitet_Mapiranja", "Kontrola")
    widths = {
        "README": {"A": 24, "B": 120},
        "Lekcije_Pilot25": {"D": 45, "F": 55, "G": 55, "J": 50, "M": 40,
                            "N": 40, "O": 40, "V": 60, "Y": 60},
        "Mapiranja_Pregled": {"F": 45, "O": 70},
        "Dokazi_KS": {"D": 70, "E": 70},
        "Dokazi_RS": {"D": 70, "E": 70},
        "Porodice_Pilot": {"B": 40, "D": 50, "F": 50, "L": 50, "P": 60},
        "Granice_lekcija": {"C": 45, "E": 45, "G": 80},
        "Primjeri_Pilot": {"E": 70, "F": 40, "G": 55, "H": 45, "I": 45},
        "Praznine": {"B": 30, "D": 90},
        "Kvalitet_Mapiranja": {"A": 55, "C": 70},
        "Kontrola": {"A": 55, "B": 70},
        "Aktivacija": {"B": 45, "F": 55, "G": 55, "H": 70},
        "Provjera_izvora": {"B": 40, "C": 90, "D": 50, "E": 70},
    }
    for name in order:
        header, rows = sheets[name]
        bcm._sheet(wb, name, header, rows, widths.get(name))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    bcm._normalize_zip(out_path)
    return sheets, content_digest, bcm._sha256(out_path)


def bcm_title(lesson_id):
    """Naslov i za lekcije izvan pilota (granice smiju pokazati susjeda vani)."""
    for lesson in bcm.load_lessons():
        if lesson.lesson_id == lesson_id:
            return lesson.title
    return lesson_id


def _boundary_summary(lesson_id):
    parts = []
    for source, forbidden, symmetric, _why in BOUNDARIES:
        if source == lesson_id:
            parts.append(forbidden)
        elif symmetric and forbidden == lesson_id:
            parts.append(source)
    return "; ".join(dict.fromkeys(parts))


def _quality_metrics(pilot_mapping_rows):
    """Metrike kvaliteta Faze 2 nad pilotom — PILOT DOKAZ, ne statistika za 534."""
    rows = [("NAPOMENA", "", "pilot dokaz sa 25 lekcija — NIJE statistička "
             "preciznost za svih 534 lekcije")]
    verdict_counter = Counter(verdict_for(r)[0] for r in pilot_mapping_rows)
    rows.append(("ukupno pilot redova Faze 2", len(pilot_mapping_rows), ""))
    for verdict in ("accept", "change_relation", "reject", "wrong_lesson",
                    "ambiguous", "duplicate_evidence"):
        rows.append((f"presuda: {verdict}", verdict_counter.get(verdict, 0), ""))
    exact_rows = [r for r in pilot_mapping_rows if r[12] == "exact"]
    exact_ok = sum(1 for r in exact_rows
                   if verdict_for(r)[0] in ("accept",))
    exact_changed = sum(1 for r in exact_rows
                        if verdict_for(r)[0] == "change_relation")
    exact_rejected = len(exact_rows) - exact_ok - exact_changed
    rows.append(("exact prihvaćeno / oslabljeno / odbijeno",
                 f"{exact_ok} / {exact_changed} / {exact_rejected}", ""))
    neighbour_rows = [r for r in pilot_mapping_rows if r[12] == "neighbour"]
    n_ok = sum(1 for r in neighbour_rows if verdict_for(r)[0] == "accept")
    rows.append(("neighbour prihvaćeno / odbijeno",
                 f"{n_ok} / {len(neighbour_rows) - n_ok}", ""))
    false_high = [r[0] for r in pilot_mapping_rows
                  if r[17] == "auto_high_confidence"
                  and verdict_for(r)[0] in ("reject", "wrong_lesson")]
    rows.append(("lažno visoko pouzdanje (auto_high → odbijeno)",
                 len(false_high), "; ".join(false_high)))
    false_low = [r[0] for r in pilot_mapping_rows
                 if verdict_for(r)[0] == "change_relation"
                 and verdict_for(r)[1] == "exact"]
    rows.append(("lažno nisko (podignuto na exact pregledom)",
                 len(false_low), "; ".join(false_low)))
    rows.append(("oporavljeni dokazi (novi, iz pregleda)",
                 len(RECOVERED_EVIDENCE),
                 f"{sum(1 for r in RECOVERED_EVIDENCE if r[1] == 'stavka')} iz "
                 f"stavki + {sum(1 for r in RECOVERED_EVIDENCE if r[1] == 'stranica')} "
                 "iz teksta stranica"))
    rows.append(("— PO METODI FAZE 2 (prihvaćeno+promijenjeno / ukupno) —", "", ""))
    by_method = defaultdict(lambda: [0, 0])
    for r in pilot_mapping_rows:
        verdict = verdict_for(r)[0]
        by_method[r[14]][1] += 1
        if verdict in ("accept", "change_relation"):
            by_method[r[14]][0] += 1
    for method in sorted(by_method):
        ok, total = by_method[method]
        rows.append((f"metoda {method}", f"{ok}/{total}",
                     f"{100.0 * ok / total:.0f}% pilot tačnost"))
    return rows


def _run_checks(pilot_mapping_rows, mapping_by_id, items_by_id, pages,
                lessons, per_lesson):
    checks = []

    def check(name, ok, value):
        checks.append((name, str(value), "PASS" if ok else "FAIL"))

    check("tačno 25 pilot lekcija", len(lessons) == 25, len(lessons))
    pilot_ids = {verdict_key(r) for r in pilot_mapping_rows}
    verdict_ids = set(MAPPING_VERDICTS)
    check("svaki pilot red Faze 2 ima presudu (nijedan tiho izostavljen)",
          pilot_ids == verdict_ids,
          f"{len(pilot_ids & verdict_ids)}/{len(pilot_ids)}; višak presuda: "
          f"{sorted(verdict_ids - pilot_ids) or '—'}")
    check("sve presude su iz dozvoljenog skupa",
          all(v[0] in VERDICTS for v in MAPPING_VERDICTS.values()), "enum ok")
    bad_items = [rec[0] for rec in RECOVERED_EVIDENCE
                 if rec[1] == "stavka" and rec[2] not in items_by_id]
    check("oporavljene stavke postoje u Fazi 2", not bad_items, bad_items or "sve")
    bad_pages = [rec[0] for rec in RECOVERED_EVIDENCE
                 if rec[1] == "stranica" and (rec[2], rec[3]) not in pages]
    check("oporavljene stranice postoje u Fazi 1", not bad_pages, bad_pages or "sve")
    def _ws(text):
        # PDF ekstrakcija nosi nepravilne razmake/prelome — citat se provjerava
        # uz sažimanje bjelina; sadržaj (riječi i redoslijed) mora biti doslovan.
        return " ".join(str(text).split())

    bad_quotes = [rec[0] for rec in RECOVERED_EVIDENCE
                  if rec[1] == "stranica"
                  and _ws(rec[7]) not in _ws(pages[(rec[2], rec[3])][1])]
    check("svaki stranični citat postoji u tekstu stranice (uz sažete bjeline)",
          not bad_quotes, bad_quotes or "svi")
    check("svih šest praznina ima razrješenje sa slovom slučaja",
          set(GAP_RESOLUTIONS) == set(GAP_LESSONS), sorted(GAP_RESOLUTIONS))
    check("semantika postoji za svih 25 lekcija",
          set(LESSON_SEMANTICS) == {l.lesson_id for l in lessons},
          len(LESSON_SEMANTICS))
    family_members = [m for fam in FAMILIES.values() for m in fam["members"]]
    check("porodice pokrivaju svih 25 lekcija tačno jednom",
          sorted(family_members) == sorted(l.lesson_id for l in lessons),
          f"{len(family_members)} članstava u {len(FAMILIES)} porodica")
    check("svaka lekcija u semantici pokazuje na svoju porodicu",
          all(LESSON_SEMANTICS[m]["family"] == fid
              for fid, fam in FAMILIES.items() for m in fam["members"]), "ok")
    per_lesson_examples = defaultdict(lambda: {"1": 0, "2": 0, "3": 0, "inv": 0})
    for lesson, level, validity, task, _s, _d in EXAMPLES:
        if validity == "valid" and not task.startswith("(nepodržan"):
            per_lesson_examples[lesson][str(level)] += 1
        elif validity == "invalid_neighbour":
            per_lesson_examples[lesson]["inv"] += 1
    missing = [lid for l in lessons
               for lid in [l.lesson_id]
               if per_lesson_examples[lid]["1"] < 1
               or per_lesson_examples[lid]["2"] < 1
               or per_lesson_examples[lid]["inv"] < 1]
    check("svaka lekcija ima L1, L2 i bar jedan invalid_neighbour primjer",
          not missing, missing or "sve")
    l3 = sum(1 for l in lessons if per_lesson_examples[l.lesson_id]["3"] >= 1)
    check("L3 pokriven ili izričito nepodržan",
          all(per_lesson_examples[l.lesson_id]["3"] >= 1
              or any(e[0] == l.lesson_id and e[1] == 3 for e in EXAMPLES)
              for l in lessons), f"L3 podržan za {l3}/25")
    check("granice: bar jedna smjerna (nesimetrična) relacija postoji",
          any(not b[2] for b in BOUNDARIES),
          f"{sum(1 for b in BOUNDARIES if not b[2])} smjernih / {len(BOUNDARIES)}")
    formula_rows = [rec[0] for rec in RECOVERED_EVIDENCE if "formula_loss" in rec[8]]
    verified = {row[0] for row in SOURCE_VERIFICATION}
    check("formula_loss zastavice sačuvane i nalaz zabilježen (10n slučaj)",
          len(formula_rows) >= 2 and "6-03-003" in verified
          and any("NIJE eksponent" in row[2] for row in SOURCE_VERIFICATION),
          formula_rows)
    check("svaka pilot lekcija ima klasu aktivacije iz dozvoljenog skupa",
          set(ACTIVATION) == {l.lesson_id for l in lessons}
          and all(v[0] in ACTIVATION_CLASSES for v in ACTIVATION.values()),
          len(ACTIVATION))
    unproven = [lid for lid, (klasa, _why) in sorted(ACTIVATION.items())
                if klasa != "READY" and lid in PROVEN_DETECTORS]
    check("nijedna ne-READY lekcija ne nosi dokazan detektor kao dozvolu",
          not unproven, unproven or "dosljedno")
    still_enforced = [lid for lid in PROVEN_DETECTORS
                      if ACTIVATION[lid][0] != "READY"]
    check("postojeće dokazano blokiranje ostaje READY (nema regresije)",
          not still_enforced, still_enforced or "6-03-004 + 6 ugovorenih lekcija")
    # Lekcija bez exact dokaza SMIJE postojati (to je upravo nalaz pregleda,
    # npr. 6-04-013 sa samo podržavajućim dokazima) — ali tada NE SMIJE nositi
    # status 'ready_for_contract_draft'.
    dishonest = [l.lesson_id for l in lessons
                 if not per_lesson.get(l.lesson_id, {}).get("exact")
                 and LESSON_SEMANTICS[l.lesson_id]["status"] == "ready_for_contract_draft"]
    check("nijedna lekcija bez exact dokaza nije proglašena spremnom za ugovor",
          not dishonest, dishonest or "dosljedno")
    return checks


def main(argv=None):
    parser = argparse.ArgumentParser(description="MAT-BOT Faza 2.5 pilot pregled")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report", action="store_true")
    parser.add_argument("--out", type=Path, default=OUTPUT_XLSX)
    args = parser.parse_args(argv)

    for path in (bcm.CANONICAL_XLSX, bcm.PHASE1_XLSX, PHASE2_XLSX, bcm.TOPICS_JSON):
        if not path.exists():
            raise SystemExit(f"Ulaz ne postoji: {path}")
    before = {p: bcm._sha256(p) for p in (bcm.CANONICAL_XLSX, bcm.PHASE1_XLSX, PHASE2_XLSX)}
    sheets, digest, file_digest = build(dry_run=args.dry_run, out_path=args.out)
    after = {p: bcm._sha256(p) for p in before}
    if before != after:
        raise SystemExit("ULAZNI WORKBOOK PROMIJENJEN — ovo ne smije da se desi.")
    if args.report or args.dry_run:
        print("=== FAZA 2.5 — sažetak ===")
        _header, quality = sheets["Kvalitet_Mapiranja"]
        for row in quality:
            print("  " + " | ".join(str(c) for c in row if c != ""))
        _h, lessons_sheet = sheets["Lekcije_Pilot25"]
        status_counts = Counter(r[23] for r in lessons_sheet)
        print("statusi lekcija:", dict(sorted(status_counts.items())))
        print("hash sadržaja:", digest[:32], "…")
        if file_digest:
            print("sha256 fajla:", file_digest)
    if not args.dry_run:
        print(f"OK: {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
