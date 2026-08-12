"""Konvertuje reference/curriculum/MATBOT_Sve_Lekcije_6_7_8_9.xlsx u data/topics.json.

Razvojni izvor podataka je Excel; produkcijski backend čita samo generisani
JSON i nikad ne otvara Excel u runtime-u. Pokreni ručno kad se kurikulum
promijeni:

    python scripts/build_topics_json.py
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "reference" / "curriculum" / "MATBOT_Sve_Lekcije_6_7_8_9.xlsx"
JSON_PATH = ROOT / "data" / "topics.json"
SCOPE_OVERRIDES_PATH = ROOT / "data" / "lesson_scope_overrides.json"
SHEET_NAME = "Sve lekcije"
EXPECTED_HEADER = ("Razred", "Oblast", "Lekcija", "ID lekcije")

# Polja koja se smiju spojiti iz overrides fajla u generisani zapis lekcije.
# Zatvorena lista: runtime čita tačno ova dva (matbot/topics.py::lesson_info),
# a `evidence_ids`/`evidence_note` ostaju SAMO u izvoru kao provenijencija.
SCOPE_OVERRIDE_FIELDS = ("lesson_scope", "objectives")


def load_scope_overrides(path=None):
    """Kurikularni opseg po lekciji, ili prazno kad fajl ne postoji.

    Kanonski Excel nosi samo četiri kolone i binaran je, pa se opseg lekcije
    drži u verzionisanom JSON-u pored njega (isti obrazac kao
    data/routing_overrides.json). Odsustvo fajla NIJE greška — tada se
    generiše tačno ono što se generisalo i ranije."""
    path = path or SCOPE_OVERRIDES_PATH
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    overrides = payload.get("overrides", {})
    resolved = {}
    for topic_id, entry in overrides.items():
        if not entry.get("evidence_ids"):
            raise ValueError(
                f"opseg lekcije bez kurikularnog dokaza: {topic_id!r}")
        fields = {name: entry[name] for name in SCOPE_OVERRIDE_FIELDS
                  if name in entry}
        if not fields:
            raise ValueError(f"prazan opseg lekcije: {topic_id!r}")
        resolved[str(topic_id)] = fields
    return resolved


def build_grades(rows, scope_overrides=None):
    """Redovi → strukture razreda. ČISTA funkcija nad datim redovima.

    Namjerno NE provjerava da je svaki override iskorišten: pozivalac smije
    proslijediti proizvoljan podskup redova (postojeći testovi to rade), pa bi
    takva provjera tu bila netačna. Pokrivenost overrides-a čuva kanonski build
    (`main`), gdje su redovi cio kurikulum."""
    if scope_overrides is None:
        scope_overrides = load_scope_overrides()
    grades = {}
    seen_ids = set()
    seen_combos = set()
    for razred, oblast, lekcija, topic_id in rows:
        if razred is None and oblast is None and lekcija is None and topic_id is None:
            continue
        grade_key = str(int(razred)).strip()
        oblast = (oblast or "").strip()
        lekcija = (lekcija or "").strip()
        topic_id = (topic_id or "").strip()
        if not grade_key or not oblast or not lekcija or not topic_id:
            raise ValueError(f"Nepotpun red u Excelu: {(razred, oblast, lekcija, topic_id)!r}")
        if topic_id in seen_ids:
            raise ValueError(f"Dupli ID lekcije: {topic_id!r}")
        seen_ids.add(topic_id)
        combo = (grade_key, oblast, lekcija)
        if combo in seen_combos:
            raise ValueError(f"Dupla kombinacija Razred+Oblast+Lekcija: {combo!r}")
        seen_combos.add(combo)

        grade = grades.setdefault(grade_key, {"oblast_order": [], "lessons": []})
        if oblast not in grade["oblast_order"]:
            grade["oblast_order"].append(oblast)
        lesson = {"id": topic_id, "oblast": oblast, "title": lekcija}
        # Spajanje je determinističko: uvijek isti ključevi, uvijek istim
        # redom, i SAMO za lekciju koja u izvoru stvarno ima red. Lekcija bez
        # opsega ostaje bajt za bajt kao prije (bez ijednog dodatnog ključa).
        override = scope_overrides.get(topic_id, {})
        for name in SCOPE_OVERRIDE_FIELDS:
            if name in override:
                lesson[name] = override[name]
        grade["lessons"].append(lesson)
    return grades


def unapplied_scope_overrides(grades, scope_overrides):
    """ID-jevi opsega koje kanonski kurikulum uopšte ne poznaje.

    Opseg za nepostojeću lekciju je tiha smrt podatka: nikad se ne ćuti, ali
    se sudi SAMO nad kompletnim kurikulumom (vidi `build_grades`)."""
    known = {lesson["id"] for grade in grades.values()
             for lesson in grade["lessons"]}
    return sorted(set(scope_overrides) - known)


def load_rows(xlsx_path):
    """Otvori workbook, provjeri sheet i header, vrati redove podataka (bez headera)."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel fajl ne postoji: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' ne postoji u {xlsx_path}")
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet je prazan.")

    header = rows[0]
    header_normalized = tuple((c or "").strip() if isinstance(c, str) else c for c in header[:4])
    if header_normalized != EXPECTED_HEADER:
        raise ValueError(
            f"Neočekivan header: {header_normalized!r}, očekivano {EXPECTED_HEADER!r}"
        )

    return rows[1:]


def main():
    try:
        data_rows = load_rows(XLSX_PATH)
        scope_overrides = load_scope_overrides()
        grades = build_grades(data_rows, scope_overrides)
        unapplied = unapplied_scope_overrides(grades, scope_overrides)
        if unapplied:
            raise ValueError(
                "opseg lekcije za nepoznat ID: " + ", ".join(unapplied))
    except (FileNotFoundError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)

    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump({"grades": grades}, f, ensure_ascii=False, indent=2)
        f.write("\n")

    total_lessons = sum(len(g["lessons"]) for g in grades.values())
    print(f"OK: {JSON_PATH} — {len(grades)} razreda, {total_lessons} lekcija.")


if __name__ == "__main__":
    main()
