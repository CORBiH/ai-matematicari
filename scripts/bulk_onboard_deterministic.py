"""Masovno uključivanje determinističkih lekcija (kapacitetna ekspanzija).

    python scripts/bulk_onboard_deterministic.py            # generiši + kompajliraj
    python scripts/bulk_onboard_deterministic.py --check    # provjeri da su artefakti u koraku
    python scripts/bulk_onboard_deterministic.py --report   # + sažetak pokrivenosti

ULAZI:
    data/topics.json                                   (kanonskih 534 lekcija)
    reference/curriculum/semantics/MATBOT_Faza2_Mapiranje.xlsx   (dokazi, opciono)
    OVA DATOTEKA — tabela ACTIVATIONS je pregledani izvor istine o tome koja
    lekcija pripada kojoj porodici s kojim parametrima.

IZLAZI:
    data/semantic_families.json            (dodane kapacitetne porodice)
    data/lesson_semantic_assignments.json  (dodane dodjele lekcija)
    data/lesson_semantics.compiled.json    (preko postojećeg kompajlera)
    reference/curriculum/semantics/deterministic_coverage_report.json

PRINCIPI (isti kao serverski validatori):
  • fail closed: nepoznata lekcija, naslov koji se ne poklapa s očekivanim,
    porodica bez registrovanog generatora ili parametri koje generator NE
    podržava u potpunosti OBARAJU build — lekcija se nikad ne aktivira tiho;
  • lekcija je PODATAK: nijedan ID lekcije ne postoji u matbot/ kodu;
  • dokazi se čitaju iz Faze 2 (exact mapiranja); lekcija bez izvora dobija
    izričitu oznaku CANON-TITLE (aktivacija na osnovu jednoznačnog naslova),
    nikad izmišljeni izvor;
  • dvostruko pokretanje daje bajt-identične artefakte.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

FAMILIES_PATH = ROOT / "data" / "semantic_families.json"
ASSIGNMENTS_PATH = ROOT / "data" / "lesson_semantic_assignments.json"
TOPICS_PATH = ROOT / "data" / "topics.json"
FAZA2_XLSX = ROOT / "reference" / "curriculum" / "semantics" / "MATBOT_Faza2_Mapiranje.xlsx"
REPORT_PATH = (ROOT / "reference" / "curriculum" / "semantics"
               / "deterministic_coverage_report.json")

CONTRACT_VERSION = "5F.1"

_HEADER = "SEMANTIČKI UGOVOR LEKCIJE (server ga provjerava determinističkim putem):"
_CLOSING = "- ako prekršiš ovo, server odbija paket prije objave"


def _template(main_parameter=None, main_text=None, operations=None,
              fixed=(), parameter_lines=None):
    template = {"header": _HEADER}
    if operations:
        template["operations"] = operations
    if main_parameter:
        template["main_line"] = {"parameter": main_parameter, "text": main_text}
    if fixed:
        template["fixed_lines"] = list(fixed)
    if parameter_lines:
        template["parameter_lines"] = parameter_lines
    template["closing"] = _CLOSING
    return template


_ARITHMETIC_OPERATION_VALUES = ["add", "subtract", "multiply", "divide"]


def _arithmetic_family(domain, domain_label, operation_labels, fixed,
                       extra_schema=None, parameter_lines=None):
    domains = domain if isinstance(domain, list) else [domain]
    schema = {
        "allowed_operations": {"kind": "enum_set",
                               "values": _ARITHMETIC_OPERATION_VALUES,
                               "required": True},
        "number_domain": {"kind": "enum", "values": domains, "required": True},
        "expression_shape": {"kind": "enum",
                             "values": ["single_operation", "multi_factor",
                                        "order_of_operations",
                                        "division_with_remainder",
                                        "complex_fraction"],
                             "required": False},
    }
    schema.update(extra_schema or {})
    return {
        "family_version": 1,
        "family_name": f"Direktan račun — {domain_label}",
        "detector": "numeric_arithmetic",
        "core_skill": f"izvesti imenovanu računsku operaciju ({domain_label})",
        "parameter_schema": schema,
        "enforced_parameters": ["allowed_operations", "number_domain"],
        "advisory_parameters": ["expression_shape"] + list((extra_schema or {})),
        "prompt_template": _template(
            operations="- glavna vidljiva operacija mora biti: {operations}",
            fixed=fixed, parameter_lines=parameter_lines),
        "operation_labels": operation_labels,
    }


def _concept_family(name, detector, core_skill, concept_values, value_labels,
                    fixed=(), main_text=None, extra_schema=None,
                    concept_parameter="concepts", parameter_lines=None):
    schema = {
        concept_parameter: {"kind": "enum_set", "values": concept_values,
                            "required": True},
    }
    schema.update(extra_schema or {})
    return {
        "family_version": 1,
        "family_name": name,
        "detector": detector,
        "core_skill": core_skill,
        "parameter_schema": schema,
        "enforced_parameters": [concept_parameter] + sorted(extra_schema or {}),
        "advisory_parameters": [],
        "prompt_template": _template(
            main_parameter=concept_parameter,
            main_text=main_text or "- glavna vidljiva radnja mora biti: {values}",
            fixed=fixed, parameter_lines=parameter_lines),
        "value_labels": value_labels,
    }


NEW_FAMILIES = {
    "natural_arithmetic_direct": _arithmetic_family(
        "natural", "prirodni brojevi",
        {"add": "sabiranje prirodnih brojeva",
         "subtract": "oduzimanje prirodnih brojeva",
         "multiply": "množenje prirodnih brojeva",
         "divide": "dijeljenje prirodnih brojeva"},
        ["- svi operandi i svi međurezultati moraju ostati prirodni brojevi "
         "(bez razlomaka i bez negativnih vrijednosti)"],
        parameter_lines={"expression_shape": {
            "order_of_operations": "- izraz mora tražiti poštovanje redoslijeda "
                                   "operacija (i zagrade gdje su zadate)",
            "multi_factor": "- izraz je proizvod više faktora"}}),
    "integer_arithmetic_direct": _arithmetic_family(
        "integer", "cijeli brojevi",
        {"add": "sabiranje cijelih brojeva",
         "subtract": "oduzimanje cijelih brojeva",
         "multiply": "množenje cijelih brojeva",
         "divide": "dijeljenje cijelih brojeva"},
        ["- operandi su cijeli brojevi (s negativnim vrijednostima gdje nivo "
         "traži) i rezultat je cio broj"],
        extra_schema={"sign_scope": {"kind": "enum",
                                     "values": ["any", "same_signs",
                                                "different_signs"],
                                     "required": False}},
        parameter_lines={"sign_scope": {
            "same_signs": "- svi sabirci moraju imati ISTE predznake",
            "different_signs": "- sabirci moraju imati RAZLIČITE predznake"},
            "expression_shape": {
                "multi_factor": "- izraz je proizvod više cijelih faktora"}}),
    "decimal_arithmetic_direct": _arithmetic_family(
        "decimal", "decimalni brojevi",
        {"add": "sabiranje decimalnih brojeva",
         "subtract": "oduzimanje decimalnih brojeva",
         "multiply": "množenje decimalnih brojeva",
         "divide": "dijeljenje decimalnih brojeva"},
        ["- operandi su decimalni brojevi i svaki vidljivi rezultat mora imati "
         "KONAČAN decimalan zapis"]),
    "rational_arithmetic_direct": _arithmetic_family(
        ["rational_signed", "rational_nonneg"], "racionalni brojevi",
        {"add": "sabiranje racionalnih brojeva",
         "subtract": "oduzimanje racionalnih brojeva",
         "multiply": "množenje racionalnih brojeva",
         "divide": "dijeljenje racionalnih brojeva"},
        ["- operandi su racionalni brojevi (razlomci s predznakom) i rezultat "
         "je racionalan broj"]),
    "number_comparison_order": _concept_family(
        "Poređenje i uređenje brojeva", "number_ordering",
        "uporediti ili urediti brojeve po veličini",
        ["natural", "fraction", "decimal", "integer", "rational"],
        {"natural": "prirodnih brojeva", "fraction": "razlomaka",
         "decimal": "decimalnih brojeva", "integer": "cijelih brojeva",
         "rational": "racionalnih brojeva"},
        fixed=["- zadatak traži izbor najvećeg/najmanjeg, poredak ili položaj "
               "između datih brojeva — ne računanje nove vrijednosti"],
        main_text="- glavna vidljiva radnja mora biti poređenje ili uređenje: "
                  "{values}",
        concept_parameter="number_domain"),
    "absolute_value_opposite": _concept_family(
        "Apsolutna vrijednost i suprotan broj", "absolute_opposite",
        "odrediti apsolutnu vrijednost ili suprotan broj",
        ["absolute_value", "opposite"],
        {"absolute_value": "apsolutna vrijednost broja",
         "opposite": "suprotan broj"},
        extra_schema={"number_domain": {"kind": "enum",
                                        "values": ["integer", "rational"],
                                        "required": True}}),
    "divisibility_predicate_application": _concept_family(
        "Primjena pravila djeljivosti", "divisibility_predicate",
        "primijeniti pravila djeljivosti na dat broj",
        [2, 3, 4, 5, 6, 9, 10, 15, 25],
        {},
        fixed=["- pitanje glasi »djeljiv sa N« (bez negacije i bez ‘ili’); "
               "računanje količnika smije biti samo pomoćni korak"],
        main_text="- zadatak mora tražiti primjenu pravila djeljivosti sa: "
                  "{values}",
        concept_parameter="divisors"),
    "common_divisors_multiples": _concept_family(
        "Djelioci, sadržioci, NZD i NZS", "common_divisors_multiples",
        "odrediti djelioce/sadržioce, NZD ili NZS",
        ["divisor_membership", "multiple_membership", "gcd", "lcm"],
        {"divisor_membership": "prepoznavanje djelioca datog broja",
         "multiple_membership": "prepoznavanje sadržioca datog broja",
         "gcd": "najveći zajednički djelilac",
         "lcm": "najmanji zajednički sadržilac"}),
    "prime_structure": _concept_family(
        "Prosti brojevi i prosta struktura", "prime_structure",
        "prepoznati proste brojeve i rastaviti broj na proste faktore",
        ["prime_classification", "coprime_pairs", "prime_factorization"],
        {"prime_classification": "prepoznavanje prostog i složenog broja",
         "coprime_pairs": "prepoznavanje relativno prostih brojeva",
         "prime_factorization": "rastavljanje na proste faktore"}),
    "percent_basic": _concept_family(
        "Osnovni procentni račun", "percent_basic",
        "izračunati procenat broja ili postotni zapis razlomka",
        ["percent_of_number", "fraction_to_percent"],
        {"percent_of_number": "procenat datog broja",
         "fraction_to_percent": "postotni zapis razlomka"},
        fixed=["- znak % piše se u prozi, nikad unutar $...$"]),
    "arithmetic_mean_direct": _concept_family(
        "Aritmetička sredina", "arithmetic_mean",
        "izračunati aritmetičku sredinu datih brojeva",
        ["mean"],
        {"mean": "aritmetička sredina datih brojeva"}),
    "power_arithmetic_direct": _concept_family(
        "Stepeni i zakoni stepena", "power_arithmetic",
        "izračunati vrijednost stepena ili primijeniti zakon stepena",
        ["square_value", "power_value", "zero_negative_exponent",
         "same_base_product_quotient", "power_of_power_product"],
        {"square_value": "kvadrat racionalnog broja",
         "power_value": "vrijednost stepena s prirodnim izložiocem",
         "zero_negative_exponent": "nulti i negativni izložilac",
         "same_base_product_quotient": "množenje i dijeljenje stepena "
                                       "jednakih osnova",
         "power_of_power_product": "stepen stepena, proizvoda i količnika"}),
    "square_root_direct": _concept_family(
        "Kvadratni korijen", "square_root",
        "izračunati kvadratni korijen ili prepoznati savršeni kvadrat",
        ["square_root_value", "perfect_square_recognition"],
        {"square_root_value": "kvadratni korijen savršenog kvadrata",
         "perfect_square_recognition": "prepoznavanje savršenog kvadrata"},
        fixed=["- potkorjena vrijednost mora biti savršen kvadrat (prirodan "
               "broj, razlomak ili decimalni zapis) — bez približnih "
               "vrijednosti"]),
    "linear_equation_direct": _concept_family(
        "Jednostavne linearne jednačine", "linear_equation",
        "riješiti ili provjeriti jednostavnu linearnu jednačinu",
        ["one_step_additive", "one_step_multiplicative", "parentheses",
         "check_solution", "check_inequality"],
        {"one_step_additive": "jednačina s jednim sabiranjem/oduzimanjem",
         "one_step_multiplicative": "jednačina s jednim množenjem/dijeljenjem",
         "parentheses": "jednačina sa zagradom",
         "check_solution": "provjera ponuđenog rješenja jednačine",
         "check_inequality": "provjera ponuđenog rješenja nejednačine"},
        fixed=["- jedna nepoznata i tačno jedno rješenje"],
        extra_schema={"number_domain": {"kind": "enum",
                                        "values": ["integer", "rational"],
                                        "required": True}},
        concept_parameter="shapes",
        parameter_lines={"number_domain": {
            "integer": "- koeficijenti i rješenje su cijeli brojevi",
            "rational": "- koeficijenti i rješenje smiju biti razlomci"}}),
    "classical_probability_basic": _concept_family(
        "Klasična vjerovatnoća", "classical_probability",
        "izračunati klasičnu vjerovatnoću (povoljni kroz svi ishodi)",
        ["classical_probability"],
        {"classical_probability": "klasična vjerovatnoća: broj povoljnih kroz "
                                  "broj svih ishoda"},
        fixed=["- svi ishodi su konačni i jednako vjerovatni; vjerovatnoća se "
               "piše riječju, nikad simbolom P"]),
}



def _kinds_family(name, detector, core_skill, kind_values, fixed=()):
    """Batch #3: porodica čiji ugovor nosi listu VRSTA zadataka (`kinds`)."""
    return {
        "family_version": 1,
        "family_name": name,
        "detector": detector,
        "core_skill": core_skill,
        "parameter_schema": {
            "kinds": {"kind": "enum_set", "values": list(kind_values),
                      "required": True},
            "fraction_coefficients": {"kind": "enum",
                                      "values": ["yes"], "required": False},
        },
        "enforced_parameters": ["kinds"],
        "advisory_parameters": ["fraction_coefficients"],
        "prompt_template": _template(
            main_parameter="kinds",
            main_text="- vidljivi zadatak mora pripadati vrstama: {values}",
            fixed=fixed),
        "value_labels": {},
    }


_GEO2D_KINDS = ["square_perimeter", "square_area", "rectangle_perimeter",
                "rectangle_area", "triangle_perimeter", "triangle_area",
                "parallelogram_area", "parallelogram_perimeter",
                "trapezoid_area", "trapezoid_midline", "triangle_midline",
                "rhombus_area_diagonals", "rhombus_perimeter", "deltoid_area",
                "orthodiagonal_area", "quad_perimeter", "circle_circumference",
                "circle_area", "arc_length", "sector_area", "annulus_area",
                "polygon_interior_sum", "polygon_diagonals",
                "regular_polygon_angle", "regular_polygon_perimeter"]
_PYTH_KINDS = ["hypotenuse", "leg", "verify_triple", "square_diagonal",
               "rectangle_diagonal", "isosceles_height", "equilateral_height",
               "equilateral_area", "rhombus_side",
               "isosceles_trapezoid_height", "right_trapezoid_leg",
               "chord_distance"]
_SOLID_KINDS = ["cube_surface", "cube_volume", "cube_space_diagonal",
                "cuboid_surface", "cuboid_volume", "prism4_lateral",
                "prism4_surface", "prism4_volume", "prism4_space_diagonal",
                "prism3_volume", "prism3_lateral", "pyramid4_apothem",
                "pyramid4_lateral", "pyramid4_surface", "pyramid4_volume",
                "cylinder_surface", "cylinder_volume",
                "cylinder_axial_section", "cone_slant", "cone_surface",
                "cone_volume", "cone_axial_section", "sphere_surface",
                "ball_volume", "prism_pyramid_ratio", "cylinder_cone_ratio",
                "density_mass", "polyhedron_elements"]
_ANGLE_KINDS = ["classify_angle", "central_peripheral", "central_fraction",
                "compare_angles", "dms_add_sub", "angle_times_n",
                "adjacent_vertical", "comp_supp", "transversal",
                "parallel_normal_arms", "triangle_third_angle",
                "exterior_angle", "exterior_from_interior",
                "classify_triangle_sides", "classify_triangle_angles",
                "side_angle_order", "triangle_inequality", "isosceles_angles",
                "right_triangle_acute", "quad_fourth_angle", "quad_exterior"]
_SYSTEM_KINDS = ["solve", "verify_pair", "single_equation", "classify",
                 "equivalent_system"]

NEW_FAMILIES["geometry_formula_2d"] = _kinds_family(
    "Formula-geometrija u ravni", "geometry_formula_2d",
    "izračunati obim/površinu/element ravne figure po kanonskoj formuli",
    _GEO2D_KINDS,
    fixed=["- sve tražene veličine su ZADATE brojevima — crtež nije potreban",
           "- notacija: P površina, O obim, d dijagonala, r poluprečnik, "
           "R prečnik (R = 2r)"])
NEW_FAMILIES["pythagoras_direct"] = _kinds_family(
    "Pitagorina teorema — direktne primjene", "pythagoras_direct",
    "primijeniti Pitagorinu teoremu na zadate dužine",
    _PYTH_KINDS,
    fixed=["- sve dužine su zadate brojevima; rezultat je egzaktan "
           "(cio broj ili korijen), nikad približna decimala"])
NEW_FAMILIES["solid_geometry_direct"] = _kinds_family(
    "Geometrijska tijela — formule", "solid_geometry_direct",
    "izračunati M/P/V/element tijela po kanonskoj formuli",
    _SOLID_KINDS,
    fixed=["- notacija: B baza, M omotač, P ukupna površina, V zapremina, "
           "H visina tijela, h_a apotema, D prostorna dijagonala",
           "- π ostaje simbolički u rezultatu; korijeni ostaju egzaktni"])
NEW_FAMILIES["angle_relationships_direct"] = _kinds_family(
    "Brojevni odnosi među uglovima", "angle_relationships_direct",
    "izračunati ili klasifikovati ugao iz zadatih brojevnih odnosa",
    _ANGLE_KINDS,
    fixed=["- svi uglovi su zadati mjerama — crtež nije potreban"])
NEW_FAMILIES["linear_system_direct"] = _kinds_family(
    "Sistem dvije linearne jednačine", "linear_system_direct",
    "riješiti, provjeriti ili klasifikovati sistem dvije linearne jednačine",
    _SYSTEM_KINDS,
    fixed=["- sistem ima cjelobrojne (ili razlomljene, gdje ugovor kaže) "
           "koeficijente; rješenje je uređeni par ili klasifikacija"])


# Porodica poređenja ima JEDAN domen po lekciji — parametar je enum, ne skup.
NEW_FAMILIES["number_comparison_order"]["parameter_schema"]["number_domain"]["kind"] = "enum"

# ---------------------------------------------------------------------------
# BATCH #2: prošireni koncepti postojećih porodica
# ---------------------------------------------------------------------------
NEW_FAMILIES["number_comparison_order"]["parameter_schema"]["forms"] = {
    "kind": "enum_set", "values": ["ordering", "place_value"],
    "required": False}
NEW_FAMILIES["number_comparison_order"]["advisory_parameters"] = ["forms"]

_pow = NEW_FAMILIES["power_arithmetic_direct"]
_pow["parameter_schema"]["concepts"]["values"] += [
    "scientific_notation", "unit_prefix_powers"]
_pow["value_labels"].update({
    "scientific_notation": "naučni zapis broja (a·10^n)",
    "unit_prefix_powers": "prefiksi mjernih jedinica kao stepeni broja 10"})

_root = NEW_FAMILIES["square_root_direct"]
_root["parameter_schema"]["concepts"]["values"] += [
    "root_product_quotient", "root_between_integers"]
_root["value_labels"].update({
    "root_product_quotient": "korijen proizvoda i količnika",
    "root_between_integers": "između koja dva uzastopna prirodna broja je korijen"})

_pct = NEW_FAMILIES["percent_basic"]
_pct["parameter_schema"]["concepts"]["values"] += [
    "percent_amount", "percent_rate"]
_pct["value_labels"].update({
    "percent_amount": "procentni iznos iz osnovice i stope",
    "percent_rate": "procentna stopa ili osnovica iz preostala dva podatka"})

_prob = NEW_FAMILIES["classical_probability_basic"]
_prob["parameter_schema"]["concepts"]["values"] += [
    "complement_probability", "outcome_counting"]
_prob["value_labels"].update({
    "complement_probability": "vjerovatnoća komplementarnog događaja",
    "outcome_counting": "broj elementarnih ishoda ogleda"})

_eq = NEW_FAMILIES["linear_equation_direct"]
_eq["parameter_schema"]["shapes"]["values"] += [
    "subtract_from", "fraction_form", "parentheses_combine",
    "solve_inequality_additive", "solve_inequality_multiplicative",
    "solve_inequality_sign_flip", "solve_inequality_parentheses",
    "absolute_value_equation", "absolute_value_inequality",
    "classification", "solution_count", "equivalence_choice"]
_eq["parameter_schema"]["number_domain"]["values"] += [
    "rational_nonneg", "decimal"]
_eq["value_labels"].update({
    "subtract_from": "jednačina oblika a - x = b",
    "fraction_form": "jednačina s razlomkom uz nepoznatu",
    "parentheses_combine": "jednačina sa zagradom i svođenjem sličnih članova",
    "solve_inequality_additive": "rješavanje nejednačine sa sabiranjem/oduzimanjem",
    "solve_inequality_multiplicative": "rješavanje nejednačine s množenjem/dijeljenjem",
    "solve_inequality_sign_flip": "nejednačina s promjenom smjera uz negativan koeficijent",
    "solve_inequality_parentheses": "rješavanje nejednačine sa zagradom",
    "absolute_value_equation": "jednačina s apsolutnom vrijednošću",
    "absolute_value_inequality": "nejednačina s apsolutnom vrijednošću",
    "classification": "prepoznavanje vrste zapisa",
    "solution_count": "broj rješenja jednačine",
    "equivalence_choice": "izbor ekvivalentne jednačine"})
_eq["prompt_template"]["parameter_lines"]["number_domain"].update({
    "rational_nonneg": "- koeficijenti i rješenja su POZITIVNI racionalni brojevi (Q+)",
    "decimal": "- koeficijenti i rješenja su decimalni brojevi s konačnim zapisom"})

# ---------------------------------------------------------------------------
# BATCH #2: nove porodice
# ---------------------------------------------------------------------------
NEW_FAMILIES.update({
    "divisibility_value_properties": _concept_family(
        "Djeljivost vrijednosti izraza i dekadske jedinice",
        "divisibility_value", "primijeniti djeljivost na vrijednost izraza",
        ["expression_divisibility", "decade_unit_divisibility"],
        {"expression_divisibility": "djeljivost vrijednosti zbira, razlike i proizvoda",
         "decade_unit_divisibility": "najveća dekadska jedinica koja dijeli broj"}),
    "frequency_basic": _concept_family(
        "Frekvencije malog skupa podataka", "frequency_basic",
        "očitati frekvenciju i relativnu frekvenciju",
        ["frequency", "relative_frequency", "frequency_table"],
        {"frequency": "frekvencija vrijednosti u nizu podataka",
         "relative_frequency": "relativna frekvencija kao razlomak",
         "frequency_table": "ukupan broj podataka iz tabele frekvencija"}),
    "decimal_rounding": _concept_family(
        "Zaokruživanje decimalnih brojeva", "decimal_rounding",
        "zaokružiti decimalni broj na dato mjesto",
        ["round_decimal", "round_then_estimate"],
        {"round_decimal": "zaokruživanje na dato decimalno mjesto",
         "round_then_estimate": "procjena rezultata zaokruživanjem"},
        fixed=["- školsko pravilo: cifra 5 i veće zaokružuju naviše",
               "- približenja se NIKAD ne pišu znakom jednakosti"]),
    "fraction_decimal_conversion": _concept_family(
        "Pretvaranje razlomak ↔ decimalni zapis", "fraction_decimal",
        "pretvoriti razlomak u decimalni zapis i obratno",
        ["fraction_to_decimal", "decimal_to_fraction", "decimal_place_value"],
        {"fraction_to_decimal": "razlomak u KONAČAN decimalni zapis",
         "decimal_to_fraction": "decimalni broj u SKRAĆEN razlomak",
         "decimal_place_value": "cifra na datom decimalnom mjestu"},
        extra_schema={"number_scope": {"kind": "enum",
                                       "values": ["nonneg", "signed"],
                                       "required": False}}),
    "linear_function_direct": _concept_family(
        "Linearna funkcija", "linear_function",
        "izračunati i protumačiti linearnu funkciju",
        ["evaluate", "table", "find_coefficient", "zero", "membership",
         "monotonicity", "sign_analysis", "from_two_points",
         "implicit_to_explicit"],
        {"evaluate": "vrijednost funkcije u tački",
         "table": "tabela vrijednosti funkcije",
         "find_coefficient": "koeficijent funkcije iz poznate tačke",
         "zero": "nula funkcije",
         "membership": "pripadnost tačke grafiku",
         "monotonicity": "rastuća/opadajuća funkcija (znak koeficijenta)",
         "sign_analysis": "znak funkcije",
         "from_two_points": "jednačina prave kroz dvije tačke",
         "implicit_to_explicit": "prelazak u eksplicitni oblik"},
        extra_schema={"function_kind": {"kind": "enum",
                                        "values": ["affine", "direct",
                                                   "inverse"],
                                        "required": False}}),
    "ratio_proportion_direct": _concept_family(
        "Razmjera i proporcija", "ratio_proportion",
        "raditi s razmjerama i proporcijama egzaktno",
        ["ratio_simplification", "proportion_property", "missing_term",
         "proportionality_recognition", "proportional_division"],
        {"ratio_simplification": "razmjera u najjednostavnijem obliku",
         "proportion_property": "osnovno svojstvo proporcije",
         "missing_term": "nepoznati član proporcije",
         "proportionality_recognition": "prepoznavanje direktne/obrnute proporcionalnosti",
         "proportional_division": "podjela veličine u datoj razmjeri"}),
    "polynomial_basic": _concept_family(
        "Polinomi i algebarski izrazi", "polynomial_basic",
        "sređivati, sabirati i množiti polinome; brojna vrijednost izraza",
        ["expression_evaluation", "structure_count", "monomial_structure",
         "combine_like_terms", "add_subtract", "multiply"],
        {"expression_evaluation": "brojna vrijednost izraza",
         "structure_count": "broj članova izraza",
         "monomial_structure": "koeficijent i stepen monoma",
         "combine_like_terms": "sređivanje polinoma",
         "add_subtract": "sabiranje i oduzimanje polinoma",
         "multiply": "množenje polinoma"},
        extra_schema={"number_domain": {"kind": "enum",
                                        "values": ["natural", "integer"],
                                        "required": False}}),
    "unit_conversion_direct": _concept_family(
        "Pretvaranje mjernih jedinica", "unit_conversion",
        "pretvoriti mjerne jedinice unutar iste dimenzije",
        ["length", "mass", "time", "area", "volume", "speed", "angle"],
        {"length": "jedinice dužine", "mass": "jedinice mase",
         "time": "jedinice vremena",
         "area": "jedinice površine (kvadrirani faktori)",
         "volume": "jedinice zapremine (kubirani faktori)",
         "speed": "brzina m/s ↔ km/h", "angle": "ugaone jedinice (po 60)"},
        fixed=["- pretvaranje NIKAD ne prelazi iz jedne fizičke dimenzije u drugu",
               "- kvadratne jedinice nose kvadriran, a kubne kubiran faktor"],
        concept_parameter="dimensions"),
    "simple_quadratic_equation": _concept_family(
        "Jednostavne kvadratne jednačine", "simple_quadratic",
        "riješiti kvadratnu jednačinu bez formule",
        ["x_squared_equals_a", "factor_out_x", "perfect_square_trinomial"],
        {"x_squared_equals_a": "jednačina oblika x² = a",
         "factor_out_x": "jednačina rješiva izlučivanjem x",
         "perfect_square_trinomial": "potpun kvadrat x² ± 2ax + a² = 0"},
        fixed=["- rješenja se navode kao POTPUN skup (oba rješenja, odnosno "
               "dvostruko rješenje)"],
        concept_parameter="shapes"),
})

# Batch #3: koncepti faktorizacije i algebarskih identiteta u polynomial_basic
# (isti mehanizam kao proširenja _pow/_root iznad — porodica ostaje jedna).
_poly = NEW_FAMILIES["polynomial_basic"]
_poly["parameter_schema"]["concepts"]["values"] += [
    "like_terms_select", "monomial_mul_div", "square_of_binomial",
    "cube_of_binomial", "factor_difference_squares", "factor_common",
    "factor_grouping", "factor_identity", "sum_diff_cubes", "zero_product",
    "fraction_domain"]
_poly["value_labels"].update({
    "like_terms_select": "prepoznavanje sličnih monoma",
    "monomial_mul_div": "množenje i dijeljenje monoma",
    "square_of_binomial": "kvadrat binoma (a ± b)²",
    "cube_of_binomial": "kub binoma (a ± b)³",
    "factor_difference_squares": "faktorizacija razlike kvadrata",
    "factor_common": "izlučivanje zajedničkog faktora",
    "factor_grouping": "faktorizacija grupisanjem",
    "factor_identity": "faktorizacija pomoću kvadrata binoma",
    "sum_diff_cubes": "zbir i razlika kubova",
    "zero_product": "jednačina data proizvodom faktora",
    "fraction_domain": "uslov definisanosti algebarskog razlomka"})


_LEVEL_BOUNDS = {
    "geometry_formula_2d": {
        "1": "direktna primjena jedne formule malim brojevima",
        "2": "inverzni zadatak (traži se stranica/visina) ili veći brojevi",
        "3": "dvokoračne veličine, koeficijenti i složeniji oblici"},
    "pythagoras_direct": {
        "1": "osnovne Pitagorine trojke",
        "2": "veće trojke i primjene na figure",
        "3": "egzaktni korijeni i složenije primjene"},
    "solid_geometry_direct": {
        "1": "jedna formula, male dimenzije",
        "2": "kombinovane veličine (P = 2B + M) ili veće dimenzije",
        "3": "dijagonale/apoteme preko Pitagorine teoreme, korijeni"},
    "angle_relationships_direct": {
        "1": "jedan imenovani odnos, male mjere",
        "2": "dvokoračni odnosi i minute",
        "3": "trokoračni odnosi i odnosi zadati razlikom"},
    "linear_system_direct": {
        "1": "izbor/provjera uređenog para",
        "2": "direktna supstitucija (koeficijent 1)",
        "3": "puno rješavanje i klasifikacija"},
    "natural_arithmetic_direct": {
        "1": "jedna operacija malim brojevima",
        "2": "dvije povezane operacije ili veći brojevi",
        "3": "tri povezane operacije (odnosno prioritet i zagrade)"},
    "integer_arithmetic_direct": {
        "1": "jedna operacija malim brojevima",
        "2": "dvije povezane operacije ili strože pravilo znakova",
        "3": "tri povezane operacije s miješanim predznacima"},
    "decimal_arithmetic_direct": {
        "1": "jedna operacija, jedno decimalno mjesto",
        "2": "dvije povezane operacije ili dva decimalna mjesta",
        "3": "tri povezane operacije"},
    "rational_arithmetic_direct": {
        "1": "jedna operacija jednostavnim razlomcima",
        "2": "dvije operacije ili nejednaki imenioci s predznacima",
        "3": "tri povezane operacije"},
    "number_comparison_order": {
        "1": "male, jasno razdvojene vrijednosti",
        "2": "bliže vrijednosti ili miješani zapisi",
        "3": "vrlo bliske vrijednosti, položaj između"},
    "absolute_value_opposite": {
        "1": "jedno direktno očitavanje",
        "2": "dva člana ili dvostruka promjena predznaka",
        "3": "tri člana izraza"},
    "divisibility_predicate_application": {
        "1": "jedno pravilo djeljivosti",
        "2": "dva pravila istovremeno",
        "3": "dva-tri pravila, veći brojevi"},
    "common_divisors_multiples": {
        "1": "jedan dati broj, mali djelioci/sadržioci",
        "2": "dva data broja (zajednički djelilac/sadržilac)",
        "3": "veći brojevi ili tri data broja"},
    "prime_structure": {
        "1": "mali brojevi (< 50)",
        "2": "brojevi do 100, „varljivi“ složeni brojevi",
        "3": "faktorizacija s tri prosta faktora"},
    "percent_basic": {
        "1": "osnovni procenti (10 %, 20 %, 25 %, 50 %)",
        "2": "prošireni procenti (5 %, 15 %, 30 %, ...)",
        "3": "necjelobrojni postoci ili traženje cjeline"},
    "arithmetic_mean_direct": {
        "1": "tri broja, cjelobrojna sredina",
        "2": "četiri broja",
        "3": "pet brojeva, decimalna sredina"},
    "power_arithmetic_direct": {
        "1": "jedan stepen/zakon malim brojevima",
        "2": "negativna osnova, negativan izložilac ili količnik",
        "3": "kombinovana dva zakona ili zbir stepena"},
    "square_root_direct": {
        "1": "korijen malog savršenog kvadrata",
        "2": "korijen razlomka ili decimalnog broja",
        "3": "zbir dva korijena"},
    "linear_equation_direct": {
        "1": "jedan računski korak do rješenja",
        "2": "dva koraka ili negativni koeficijenti",
        "3": "tri koraka (sređivanje pa rješavanje)"},
    "classical_probability_basic": {
        "1": "dvije boje, direktan događaj",
        "2": "tri boje",
        "3": "komplementaran događaj"},
}

_LEVEL_BOUNDS.update({
    "divisibility_value_properties": {
        "1": "zbir dva mala broja / jedna dekadska jedinica",
        "2": "razlika i proizvod / veće dekadske jedinice",
        "3": "veći brojevi, tri-četiri cifre"},
    "frequency_basic": {
        "1": "sedam podataka", "2": "deset podataka",
        "3": "dvanaest podataka"},
    "decimal_rounding": {
        "1": "cio broj ili desetinke", "2": "desetinke ili stotinke",
        "3": "stotinke/hiljaditke, procjena zbira"},
    "fraction_decimal_conversion": {
        "1": "polovine, četvrtine, desetinke",
        "2": "osmine, dvadesetine, obje smjera",
        "3": "tri decimale i mješovite vrijednosti"},
    "linear_function_direct": {
        "1": "mali cjelobrojni koeficijenti",
        "2": "negativni koeficijenti", "3": "razlomljene vrijednosti"},
    "ratio_proportion_direct": {
        "1": "mali članovi razmjere", "2": "veći faktori",
        "3": "veće vrijednosti, tri para u tabeli"},
    "polynomial_basic": {
        "1": "dva-tri člana, mali koeficijenti",
        "2": "četiri člana, negativni koeficijenti",
        "3": "viši stepeni i pet članova"},
    "unit_conversion_direct": {
        "1": "susjedne jedinice, cijeli brojevi",
        "2": "decimalni iznosi, oba smjera",
        "3": "preskok jedinice / složene jedinice"},
    "simple_quadratic_equation": {
        "1": "mali savršeni kvadrati",
        "2": "negativna rješenja i izlučivanje",
        "3": "veći koeficijenti"},
})

_REVIEWER_NOTES = {
    "geometry_formula_2d": "Invarijanta: kanonske formule i notacija (P/O/d/r/R=2r); egzaktan račun.",
    "pythagoras_direct": "Invarijanta: c² = a² + b² egzaktno; rezultat cio broj ili korijen.",
    "solid_geometry_direct": "Invarijanta: kanonske formule tijela (B/M/P/V/H/h_a/D); π i korijeni egzaktni.",
    "angle_relationships_direct": "Invarijanta: imenovani brojevni odnos uglova; zbirovi 90/180/360 egzaktni.",
    "linear_system_direct": "Invarijanta: rješenje uvršteno u OBJE jednačine; klasifikacija po determinanti.",  
    "natural_arithmetic_direct": "Invarijanta: svi operandi i međurezultati prirodni.",
    "integer_arithmetic_direct": "Invarijanta: cjelobrojni operandi i rezultat; pravilo znakova.",
    "decimal_arithmetic_direct": "Invarijanta: konačan decimalan zapis svake vidljive vrijednosti.",
    "rational_arithmetic_direct": "Invarijanta: egzaktan racionalan račun s predznacima.",
    "number_comparison_order": "Invarijanta: tačno jedna ekstremna/tražena vrijednost među opcijama.",
    "absolute_value_opposite": "Invarijanta: apsolutna vrijednost nenegativna; suprotan broj mijenja samo predznak.",
    "divisibility_predicate_application": "Invarijanta: tačno jedna opcija zadovoljava SVA navedena pravila (uski orakl to nezavisno dokazuje).",
    "common_divisors_multiples": "Invarijanta: NZD/NZS/djelilac/sadržilac dokazani dijeljenjem bez ostatka.",
    "prime_structure": "Invarijanta: tačno jedna opcija ima traženo svojstvo proste strukture.",
    "percent_basic": "Invarijanta: procenat je stoti dio; znak % nikad u $...$.",
    "arithmetic_mean_direct": "Invarijanta: sredina = zbir kroz broj podataka, egzaktno.",
    "power_arithmetic_direct": "Invarijanta: zakoni stepena nad jednakim osnovama; opcije različitih vrijednosti.",
    "square_root_direct": "Invarijanta: korijen savršenog kvadrata, provjerljiv kvadriranjem.",
    "linear_equation_direct": "Invarijanta: jedinstveno rješenje, provjera uvrštavanjem egzaktna.",
    "classical_probability_basic": "Invarijanta: vjerovatnoća = povoljni/svi, u [0,1].",
    "divisibility_value_properties": "Invarijanta: djeljivost se dokazuje nad IZRAČUNATOM vrijednošću izraza.",
    "frequency_basic": "Invarijanta: frekvencije se broje egzaktno iz navedenog niza.",
    "decimal_rounding": "Invarijanta: školsko half-up zaokruživanje nad egzaktnim razlomcima; bez a≈b jednakosti.",
    "fraction_decimal_conversion": "Invarijanta: samo konačne decimale; razlomak uvijek skraćen; opcije ne miješaju zapise.",
    "linear_function_direct": "Invarijanta: svaka tvrdnja o funkciji dokazana uvrštavanjem; tačke uvijek s imenom.",
    "ratio_proportion_direct": "Invarijanta: unakrsni proizvodi egzaktni; količnici opcija međusobno različiti.",
    "polynomial_basic": "Invarijanta: polinom je rječnik koeficijenata; rezultat provjeren uvrštavanjem broja.",
    "unit_conversion_direct": "Invarijanta: faktori po dimenziji; kvadratni 100, kubni 1000 po koraku; bez miješanja dimenzija.",
    "simple_quadratic_equation": "Invarijanta: potpun skup rješenja, provjerljiv uvrštavanjem.",
}


# ---------------------------------------------------------------------------
# PREGLEDANA TABELA AKTIVACIJA — jedina lista lekcija; kôd nikad ne grana po ID-ju
# ---------------------------------------------------------------------------
# (lesson_id, očekivani naslov, family_id, parametri, izvor_povjerenja)
# izvor_povjerenja: "pilot"  = Faza 2.5 pregled (READY),
#                   "exact"  = Faza 2 exact mapiranje + jednoznačan naslov,
#                   "title"  = jednoznačan kanonski naslov (bez NPP izvora).

ACTIVATIONS = [
    # --- prirodni brojevi (6. razred) -------------------------------------
    ("6-02-003", "Sabiranje i oduzimanje u skupu N0",
     "natural_arithmetic_direct",
     {"allowed_operations": ["add", "subtract"], "number_domain": "natural"},
     "exact"),
    ("6-02-004", "Množenje i dijeljenje u skupu N0",
     "natural_arithmetic_direct",
     {"allowed_operations": ["multiply", "divide"], "number_domain": "natural"},
     "exact"),
    ("6-02-007", "Redoslijed računskih operacija i zagrade",
     "natural_arithmetic_direct",
     {"allowed_operations": ["add", "subtract", "multiply", "divide"],
      "number_domain": "natural", "expression_shape": "order_of_operations"},
     "exact"),
    # --- djeljivost (6. razred, pilot Faze 2.5) ---------------------------
    ("6-03-001", "Djelilac/faktor i sadržilac/višekratnik prirodnog broja",
     "common_divisors_multiples",
     {"concepts": ["divisor_membership", "multiple_membership"]}, "pilot"),
    ("6-03-004", "Pravila djeljivosti sa 2, 3, 4, 5, 6, 9, 10, 15 i 25",
     "divisibility_predicate_application",
     {"divisors": [2, 3, 4, 5, 6, 9, 10, 15, 25]}, "pilot"),
    ("6-03-005", "Prosti i složeni brojevi",
     "prime_structure", {"concepts": ["prime_classification"]}, "pilot"),
    ("6-03-006", "Relativno prosti brojevi",
     "prime_structure", {"concepts": ["coprime_pairs"]}, "pilot"),
    ("6-03-007", "Rastavljanje složenih brojeva na proste faktore",
     "prime_structure", {"concepts": ["prime_factorization"]}, "pilot"),
    ("6-03-008", "Zajednički djelioci i najveći zajednički djelilac / NZD",
     "common_divisors_multiples", {"concepts": ["gcd"]}, "pilot"),
    ("6-03-009", "Zajednički sadržioci i najmanji zajednički sadržilac / NZS",
     "common_divisors_multiples", {"concepts": ["lcm"]}, "pilot"),
    # --- razlomci: poređenje (6. razred, pilot) ---------------------------
    ("6-04-008", "Upoređivanje razlomaka",
     "number_comparison_order", {"number_domain": "fraction"}, "pilot"),
    # --- decimalni brojevi (6. razred) ------------------------------------
    ("6-05-006", "Upoređivanje decimalnih brojeva",
     "number_comparison_order", {"number_domain": "decimal"}, "exact"),
    ("6-05-008", "Sabiranje i oduzimanje decimalnih brojeva",
     "decimal_arithmetic_direct",
     {"allowed_operations": ["add", "subtract"], "number_domain": "decimal"},
     "exact"),
    ("6-05-009", "Množenje decimalnih brojeva",
     "decimal_arithmetic_direct",
     {"allowed_operations": ["multiply"], "number_domain": "decimal"}, "exact"),
    ("6-05-010", "Dijeljenje decimalnih brojeva",
     "decimal_arithmetic_direct",
     {"allowed_operations": ["divide"], "number_domain": "decimal"}, "exact"),
    # --- postotak i sredina (6. razred) -----------------------------------
    ("6-06-001", "Postotni zapis razlomka",
     "percent_basic", {"concepts": ["fraction_to_percent"]}, "exact"),
    ("6-06-002", "Postotak/procenat i procenat broja",
     "percent_basic", {"concepts": ["percent_of_number", "fraction_to_percent"]},
     "exact"),
    ("6-06-004", "Aritmetička sredina danih brojeva / dva broja",
     "arithmetic_mean_direct", {"concepts": ["mean"]}, "exact"),
    # --- cijeli brojevi (7. razred) ---------------------------------------
    ("7-02-004", "Suprotni cijeli brojevi",
     "absolute_value_opposite",
     {"concepts": ["opposite"], "number_domain": "integer"}, "exact"),
    ("7-02-005", "Apsolutna vrijednost cijelog broja",
     "absolute_value_opposite",
     {"concepts": ["absolute_value"], "number_domain": "integer"}, "exact"),
    ("7-02-006", "Upoređivanje i uređenje cijelih brojeva",
     "number_comparison_order", {"number_domain": "integer"}, "exact"),
    ("7-02-007", "Sabiranje cijelih brojeva istih znakova",
     "integer_arithmetic_direct",
     {"allowed_operations": ["add"], "number_domain": "integer",
      "sign_scope": "same_signs"}, "exact"),
    ("7-02-008", "Sabiranje cijelih brojeva različitih znakova",
     "integer_arithmetic_direct",
     {"allowed_operations": ["add"], "number_domain": "integer",
      "sign_scope": "different_signs"}, "exact"),
    ("7-02-009", "Oduzimanje cijelih brojeva",
     "integer_arithmetic_direct",
     {"allowed_operations": ["subtract"], "number_domain": "integer"}, "exact"),
    ("7-02-011", "Množenje cijelih brojeva i pravilo znakova",
     "integer_arithmetic_direct",
     {"allowed_operations": ["multiply"], "number_domain": "integer"}, "exact"),
    ("7-02-012", "Dijeljenje cijelih brojeva i pravilo znakova",
     "integer_arithmetic_direct",
     {"allowed_operations": ["divide"], "number_domain": "integer"}, "exact"),
    ("7-02-013", "Proizvod više cijelih faktora",
     "integer_arithmetic_direct",
     {"allowed_operations": ["multiply"], "number_domain": "integer",
      "expression_shape": "multi_factor"}, "title"),
    ("7-02-015", "Brojevni izrazi sa cijelim brojevima",
     "integer_arithmetic_direct",
     {"allowed_operations": ["add", "subtract", "multiply"],
      "number_domain": "integer"}, "exact"),
    ("7-02-016", "Jednačine sa sabiranjem i oduzimanjem u Z",
     "linear_equation_direct",
     {"shapes": ["one_step_additive"], "number_domain": "integer"}, "exact"),
    ("7-02-017", "Jednačine sa množenjem i dijeljenjem u Z",
     "linear_equation_direct",
     {"shapes": ["one_step_multiplicative"], "number_domain": "integer"},
     "exact"),
    # --- racionalni brojevi (7. razred) -----------------------------------
    ("7-03-004", "Suprotan racionalni broj",
     "absolute_value_opposite",
     {"concepts": ["opposite"], "number_domain": "rational"}, "exact"),
    ("7-03-005", "Apsolutna vrijednost racionalnog broja",
     "absolute_value_opposite",
     {"concepts": ["absolute_value"], "number_domain": "rational"}, "exact"),
    ("7-03-006", "Upoređivanje racionalnih brojeva",
     "number_comparison_order", {"number_domain": "rational"}, "exact"),
    ("7-03-009", "Sabiranje racionalnih brojeva",
     "rational_arithmetic_direct",
     {"allowed_operations": ["add"], "number_domain": "rational_signed"},
     "exact"),
    ("7-03-010", "Oduzimanje racionalnih brojeva",
     "rational_arithmetic_direct",
     {"allowed_operations": ["subtract"], "number_domain": "rational_signed"},
     "exact"),
    ("7-03-011", "Množenje racionalnih brojeva",
     "rational_arithmetic_direct",
     {"allowed_operations": ["multiply"], "number_domain": "rational_signed"},
     "exact"),
    ("7-03-012", "Dijeljenje racionalnih brojeva",
     "rational_arithmetic_direct",
     {"allowed_operations": ["divide"], "number_domain": "rational_signed"},
     "exact"),
    ("7-03-016", "Jednačine u Q sa sabiranjem i oduzimanjem",
     "linear_equation_direct",
     {"shapes": ["one_step_additive"], "number_domain": "rational"}, "exact"),
    ("7-03-017", "Jednačine u Q sa množenjem i dijeljenjem",
     "linear_equation_direct",
     {"shapes": ["one_step_multiplicative"], "number_domain": "rational"},
     "exact"),
    # --- realni brojevi, stepeni, korijeni (8. razred) --------------------
    ("8-01-005", "Apsolutna vrijednost realnog broja",
     "absolute_value_opposite",
     {"concepts": ["absolute_value"], "number_domain": "rational"}, "exact"),
    ("8-01-006", "Kvadrat racionalnog broja",
     "power_arithmetic_direct", {"concepts": ["square_value"]}, "exact"),
    ("8-01-007", "Savršeni kvadrati i procjena",
     "square_root_direct", {"concepts": ["perfect_square_recognition"]},
     "title"),
    ("8-01-008", "Kvadratni korijen nenegativnog racionalnog broja",
     "square_root_direct", {"concepts": ["square_root_value"]}, "exact"),
    ("8-01-013", "Stepen sa cijelim izložiocem",
     "power_arithmetic_direct", {"concepts": ["power_value"]}, "exact"),
    ("8-01-014", "Nulti i negativni eksponent",
     "power_arithmetic_direct", {"concepts": ["zero_negative_exponent"]},
     "exact"),
    ("8-01-015", "Množenje i dijeljenje stepena jednakih osnova",
     "power_arithmetic_direct", {"concepts": ["same_base_product_quotient"]},
     "exact"),
    ("8-01-016", "Stepen proizvoda, količnika i stepena",
     "power_arithmetic_direct", {"concepts": ["power_of_power_product"]},
     "exact"),
    # --- podaci i vjerovatnoća (8. razred) --------------------------------
    ("8-06-009", "Aritmetička sredina",
     "arithmetic_mean_direct", {"concepts": ["mean"]}, "exact"),
    ("8-06-012", "Povoljni ishodi i klasična vjerovatnoća",
     "classical_probability_basic", {"concepts": ["classical_probability"]},
     "exact"),
    # --- 9. razred --------------------------------------------------------
    ("9-04-003", "Jednačina sa zagradama",
     "linear_equation_direct",
     {"shapes": ["parentheses"], "number_domain": "integer"}, "exact"),
    ("9-04-019", "Provjera rješenja jednačine i nejednačine",
     "linear_equation_direct",
     {"shapes": ["check_solution", "check_inequality"],
      "number_domain": "integer"}, "title"),
    ("9-08-002", "Vjerovatnoća slučajnog događaja",
     "classical_probability_basic", {"concepts": ["classical_probability"]},
     "exact"),
    ("9-08-010", "Aritmetička sredina i interpretacija podataka",
     "arithmetic_mean_direct", {"concepts": ["mean"]}, "exact"),
    # =====================================================================
    # BATCH #2 — proširenja postojećih motora
    # =====================================================================
    ("6-02-002", "Čitanje, zapisivanje i upoređivanje prirodnih brojeva",
     "number_comparison_order",
     {"number_domain": "natural", "forms": ["ordering", "place_value"]},
     "exact"),
    ("6-02-005", "Dijeljenje s ostatkom: a = bq + r",
     "natural_arithmetic_direct",
     {"allowed_operations": ["divide"], "number_domain": "natural",
      "expression_shape": "division_with_remainder"}, "exact"),
    ("6-03-002", "Djeljivost zbira, razlike i proizvoda",
     "divisibility_value_properties",
     {"concepts": ["expression_divisibility"]}, "pilot"),
    ("6-03-003", "Djeljivost dekadskim jedinicama",
     "divisibility_value_properties",
     {"concepts": ["decade_unit_divisibility"]}, "pilot"),
    ("6-04-014", "Brojevni izrazi s razlomcima",
     "rational_arithmetic_direct",
     {"allowed_operations": ["add", "subtract", "multiply", "divide"],
      "number_domain": "rational_nonneg",
      "expression_shape": "order_of_operations"}, "exact"),
    ("7-03-014", "Brojevni izrazi sa zagradama u Q",
     "rational_arithmetic_direct",
     {"allowed_operations": ["add", "subtract", "multiply", "divide"],
      "number_domain": "rational_signed",
      "expression_shape": "order_of_operations"}, "exact"),
    ("7-03-015", "Dvojni razlomci", "rational_arithmetic_direct",
     {"allowed_operations": ["divide"], "number_domain": "rational_signed",
      "expression_shape": "complex_fraction"}, "title"),
    ("8-01-004", "Uređenost i poređenje realnih brojeva",
     "number_comparison_order", {"number_domain": "rational"}, "exact"),
    ("8-01-010", "Korijen proizvoda i količnika", "square_root_direct",
     {"concepts": ["root_product_quotient"]}, "exact"),
    ("8-01-011", "Približne vrijednosti kvadratnog korijena",
     "square_root_direct", {"concepts": ["root_between_integers"]}, "exact"),
    ("8-01-012", "Računske operacije u skupu R", "rational_arithmetic_direct",
     {"allowed_operations": ["add", "subtract", "multiply", "divide"],
      "number_domain": "rational_signed",
      "expression_shape": "order_of_operations"}, "exact"),
    ("8-01-017", "Naučni zapis broja", "power_arithmetic_direct",
     {"concepts": ["scientific_notation"]}, "exact"),
    ("9-08-012", "Prefiksi mjernih jedinica kao stepeni broja 10",
     "power_arithmetic_direct", {"concepts": ["unit_prefix_powers"]}, "title"),
    ("8-06-013", "Komplementarni događaj", "classical_probability_basic",
     {"concepts": ["complement_probability"]}, "exact"),
    ("9-08-001", "Uvod u vjerovatnoću i skup elementarnih događaja",
     "classical_probability_basic", {"concepts": ["outcome_counting"]},
     "exact"),
    ("8-03-017", "Procentni iznos, osnovica i stopa", "percent_basic",
     {"concepts": ["percent_amount", "percent_rate"]}, "exact"),
    ("8-03-018", "Određivanje osnovice ili procentne stope", "percent_basic",
     {"concepts": ["percent_rate"]}, "exact"),
    ("8-06-002", "Frekvencija", "frequency_basic",
     {"concepts": ["frequency"]}, "exact"),
    ("8-06-003", "Relativna frekvencija", "frequency_basic",
     {"concepts": ["relative_frequency"]}, "exact"),
    ("8-06-004", "Tabela frekvencija", "frequency_basic",
     {"concepts": ["frequency_table"]}, "exact"),
    ("6-05-007", "Zaokruživanje decimalnih brojeva", "decimal_rounding",
     {"concepts": ["round_decimal"]}, "exact"),
    ("7-03-021", "Procjena, zaokruživanje i približan račun",
     "decimal_rounding", {"concepts": ["round_decimal",
                                       "round_then_estimate"]}, "exact"),
    # =====================================================================
    # BATCH #2 — pretvaranje razlomak ↔ decimala
    # =====================================================================
    ("6-05-001", "Decimalni zapis razlomka", "fraction_decimal_conversion",
     {"concepts": ["fraction_to_decimal"]}, "exact"),
    ("6-05-002", "Decimalni broj: cijeli dio, decimalni dio i decimalna mjesta",
     "fraction_decimal_conversion", {"concepts": ["decimal_place_value"]},
     "exact"),
    ("6-05-003", "Pretvaranje razlomka u decimalni broj i obratno",
     "fraction_decimal_conversion",
     {"concepts": ["fraction_to_decimal", "decimal_to_fraction"]}, "exact"),
    ("7-03-007", "Decimalni zapis racionalnog broja",
     "fraction_decimal_conversion",
     {"concepts": ["fraction_to_decimal"], "number_scope": "signed"},
     "exact"),
    ("7-03-008", "Pretvaranje decimalnog broja u razlomak",
     "fraction_decimal_conversion",
     {"concepts": ["decimal_to_fraction"], "number_scope": "signed"},
     "exact"),
    # =====================================================================
    # BATCH #2 — jednačine i nejednačine
    # =====================================================================
    ("6-07-001", "Jednakost, jednačina, nejednakost i nejednačina",
     "linear_equation_direct",
     {"shapes": ["classification"], "number_domain": "integer"}, "exact"),
    ("6-07-002", "Jednačine s razlomcima oblika x ± a = b i a ± x = b",
     "linear_equation_direct",
     {"shapes": ["one_step_additive", "subtract_from"],
      "number_domain": "rational_nonneg"}, "exact"),
    ("6-07-003", "Nejednačine s razlomcima oblika x ± a < b / > b i a ± x < b / > b",
     "linear_equation_direct",
     {"shapes": ["solve_inequality_additive"],
      "number_domain": "rational_nonneg"}, "exact"),
    ("6-07-004", "Jednačine s množenjem i dijeljenjem razlomaka",
     "linear_equation_direct",
     {"shapes": ["one_step_multiplicative"],
      "number_domain": "rational_nonneg"}, "exact"),
    ("6-07-005", "Nejednačine s množenjem i dijeljenjem razlomaka",
     "linear_equation_direct",
     {"shapes": ["solve_inequality_multiplicative"],
      "number_domain": "rational_nonneg"}, "exact"),
    ("6-07-006", "Jednačine i nejednačine s decimalnim brojevima",
     "linear_equation_direct",
     {"shapes": ["one_step_additive", "one_step_multiplicative"],
      "number_domain": "decimal"}, "exact"),
    ("7-02-018", "Jednačine sa apsolutnom vrijednošću",
     "linear_equation_direct",
     {"shapes": ["absolute_value_equation"], "number_domain": "integer"},
     "exact"),
    ("7-02-019", "Nejednačine sa sabiranjem i oduzimanjem u Z",
     "linear_equation_direct",
     {"shapes": ["solve_inequality_additive"], "number_domain": "integer"},
     "exact"),
    ("7-02-020", "Nejednačine sa množenjem i dijeljenjem u Z",
     "linear_equation_direct",
     {"shapes": ["solve_inequality_multiplicative",
                 "solve_inequality_sign_flip"],
      "number_domain": "integer"}, "exact"),
    ("7-03-018", "Jednačine sa zagradama i svođenjem sličnih članova",
     "linear_equation_direct",
     {"shapes": ["parentheses_combine"], "number_domain": "rational"},
     "exact"),
    ("7-03-019", "Nejednačine u Q", "linear_equation_direct",
     {"shapes": ["solve_inequality_additive",
                 "solve_inequality_multiplicative"],
      "number_domain": "rational"}, "exact"),
    ("9-04-001", "Osnovni pojmovi linearne jednačine",
     "linear_equation_direct",
     {"shapes": ["classification"], "number_domain": "integer"}, "exact"),
    ("9-04-002", "Ekvivalentne jednačine", "linear_equation_direct",
     {"shapes": ["equivalence_choice"], "number_domain": "integer"},
     "exact"),
    ("9-04-004", "Jednačina sa razlomcima", "linear_equation_direct",
     {"shapes": ["fraction_form"], "number_domain": "rational"}, "exact"),
    ("9-04-008", "Jednačina bez rješenja", "linear_equation_direct",
     {"shapes": ["solution_count"], "number_domain": "integer"}, "exact"),
    ("9-04-009", "Identitet sa beskonačno mnogo rješenja",
     "linear_equation_direct",
     {"shapes": ["solution_count"], "number_domain": "integer"}, "exact"),
    ("9-04-012", "Osnovni pojmovi linearne nejednačine",
     "linear_equation_direct",
     {"shapes": ["classification"], "number_domain": "integer"}, "exact"),
    ("9-04-014", "Nejednačina sa zagradama", "linear_equation_direct",
     {"shapes": ["solve_inequality_parentheses"],
      "number_domain": "integer"}, "exact"),
    ("9-04-016", "Promjena znaka nejednakosti pri množenju ili dijeljenju negativnim brojem",
     "linear_equation_direct",
     {"shapes": ["solve_inequality_sign_flip"], "number_domain": "integer"},
     "exact"),
    ("9-04-020", "Linearna jednačina sa apsolutnom vrijednošću",
     "linear_equation_direct",
     {"shapes": ["absolute_value_equation"], "number_domain": "integer"},
     "exact"),
    ("9-04-021", "Linearna nejednačina sa apsolutnom vrijednošću",
     "linear_equation_direct",
     {"shapes": ["absolute_value_inequality"], "number_domain": "integer"},
     "exact"),
    # =====================================================================
    # BATCH #2 — jednostavne kvadratne jednačine
    # =====================================================================
    ("8-01-009", "Jednačina x²=a, a≥0", "simple_quadratic_equation",
     {"shapes": ["x_squared_equals_a"]}, "exact"),
    ("8-07-013", "Jednačine rješive faktorizacijom",
     "simple_quadratic_equation", {"shapes": ["factor_out_x"]}, "exact"),
    ("9-06-013", "Jednostavna kvadratna jednačina ax² + bx = 0",
     "simple_quadratic_equation", {"shapes": ["factor_out_x"]}, "exact"),
    ("9-06-014", "Jednostavna kvadratna jednačina x² - a = 0",
     "simple_quadratic_equation", {"shapes": ["x_squared_equals_a"]},
     "exact"),
    ("9-06-015", "Kvadratna jednačina oblika x² ± 2ax + a² = 0",
     "simple_quadratic_equation", {"shapes": ["perfect_square_trinomial"]},
     "exact"),
    # =====================================================================
    # BATCH #2 — linearna funkcija
    # =====================================================================
    ("8-02-005", "Pojam linearne funkcije y=kx+n", "linear_function_direct",
     {"concepts": ["evaluate", "find_coefficient"]}, "exact"),
    ("8-02-006", "Tabela vrijednosti linearne funkcije",
     "linear_function_direct", {"concepts": ["table"]}, "exact"),
    ("8-02-010", "Nula linearne funkcije", "linear_function_direct",
     {"concepts": ["zero"]}, "exact"),
    ("8-02-011", "Znak i tok linearne funkcije", "linear_function_direct",
     {"concepts": ["monotonicity", "sign_analysis"]}, "exact"),
    ("8-03-006", "Funkcija direktne proporcionalnosti y=kx",
     "linear_function_direct",
     {"concepts": ["evaluate", "find_coefficient"],
      "function_kind": "direct"}, "exact"),
    ("8-03-007", "Funkcija obrnute proporcionalnosti y=k/x",
     "linear_function_direct",
     {"concepts": ["evaluate", "find_coefficient"],
      "function_kind": "inverse"}, "exact"),
    ("9-03-001", "Funkcija i vrijednost funkcije", "linear_function_direct",
     {"concepts": ["evaluate"]}, "exact"),
    ("9-03-002", "Linearna funkcija y = kx + n", "linear_function_direct",
     {"concepts": ["evaluate", "find_coefficient"]}, "exact"),
    ("9-03-003", "Tabela vrijednosti linearne funkcije",
     "linear_function_direct", {"concepts": ["table"]}, "exact"),
    ("9-03-007", "Nula linearne funkcije", "linear_function_direct",
     {"concepts": ["zero"]}, "exact"),
    ("9-03-008", "Rastuća i opadajuća linearna funkcija",
     "linear_function_direct", {"concepts": ["monotonicity"]}, "exact"),
    ("9-03-009", "Znak linearne funkcije", "linear_function_direct",
     {"concepts": ["sign_analysis"]}, "exact"),
    ("9-03-011", "Da li tačka pripada grafiku funkcije",
     "linear_function_direct", {"concepts": ["membership"]}, "exact"),
    ("9-03-012", "Eksplicitni i implicitni oblik jednačine prave",
     "linear_function_direct", {"concepts": ["implicit_to_explicit"]},
     "exact"),
    ("9-03-013", "Prelazak iz implicitnog u eksplicitni oblik",
     "linear_function_direct", {"concepts": ["implicit_to_explicit"]},
     "exact"),
    ("9-03-014", "Jednačina prave kroz dvije tačke",
     "linear_function_direct", {"concepts": ["from_two_points"]}, "exact"),
    # =====================================================================
    # BATCH #2 — razmjera i proporcija
    # =====================================================================
    ("6-06-003", "Razmjera/omjer", "ratio_proportion_direct",
     {"concepts": ["ratio_simplification"]}, "exact"),
    ("8-03-001", "Razmjera i omjer veličina", "ratio_proportion_direct",
     {"concepts": ["ratio_simplification"]}, "exact"),
    ("8-03-002", "Proporcija i osnovno svojstvo", "ratio_proportion_direct",
     {"concepts": ["proportion_property"]}, "exact"),
    ("8-03-003", "Nepoznati član proporcije", "ratio_proportion_direct",
     {"concepts": ["missing_term"]}, "exact"),
    ("8-03-004", "Prepoznavanje direktne proporcionalnosti",
     "ratio_proportion_direct",
     {"concepts": ["proportionality_recognition"]}, "exact"),
    ("8-03-005", "Prepoznavanje obrnute proporcionalnosti",
     "ratio_proportion_direct",
     {"concepts": ["proportionality_recognition"]}, "exact"),
    ("8-03-010", "Četvrta geometrijska proporcionala",
     "ratio_proportion_direct", {"concepts": ["missing_term"]}, "exact"),
    ("8-03-012", "Dijeljenje duži u datoj razmjeri",
     "ratio_proportion_direct", {"concepts": ["proportional_division"]},
     "exact"),
    ("8-03-020", "Proporcionalna podjela", "ratio_proportion_direct",
     {"concepts": ["proportional_division"]}, "exact"),
    # =====================================================================
    # BATCH #2 — polinomi i algebarski izrazi
    # =====================================================================
    ("6-02-008", "Izrazi s promjenljivim i brojna vrijednost izraza",
     "polynomial_basic",
     {"concepts": ["expression_evaluation"], "number_domain": "natural"},
     "exact"),
    ("8-07-001", "Algebarski racionalni izraz i brojna vrijednost",
     "polynomial_basic", {"concepts": ["expression_evaluation"]}, "exact"),
    ("8-07-002", "Monom, koeficijent i stepen monoma", "polynomial_basic",
     {"concepts": ["monomial_structure"]}, "exact"),
    ("8-07-006", "Polinom i sređivanje polinoma", "polynomial_basic",
     {"concepts": ["combine_like_terms"]}, "exact"),
    ("8-07-007", "Sabiranje i oduzimanje polinoma", "polynomial_basic",
     {"concepts": ["add_subtract"]}, "exact"),
    ("8-07-008", "Množenje polinoma", "polynomial_basic",
     {"concepts": ["multiply"]}, "exact"),
    ("9-01-001", "Algebarski izraz: konstante, promjenljive i članovi",
     "polynomial_basic",
     {"concepts": ["structure_count", "expression_evaluation"]}, "exact"),
    ("9-06-001", "Monom i polinom u skupu realnih brojeva",
     "polynomial_basic", {"concepts": ["monomial_structure"]}, "exact"),
    ("9-06-002", "Sređivanje polinoma", "polynomial_basic",
     {"concepts": ["combine_like_terms"]}, "exact"),
    ("9-06-003", "Sabiranje i oduzimanje polinoma", "polynomial_basic",
     {"concepts": ["add_subtract"]}, "exact"),
    ("9-06-004", "Množenje polinoma", "polynomial_basic",
     {"concepts": ["multiply"]}, "exact"),
    # =====================================================================
    # BATCH #2 — mjerne jedinice
    # =====================================================================
    ("6-13-001", "Mjerne jedinice za dužinu", "unit_conversion_direct",
     {"dimensions": ["length"]}, "exact"),
    ("6-13-002", "Mjerne jedinice za masu", "unit_conversion_direct",
     {"dimensions": ["mass"]}, "exact"),
    ("6-13-003", "Mjerne jedinice za vrijeme", "unit_conversion_direct",
     {"dimensions": ["time"]}, "exact"),
    ("6-13-004", "Mjerne jedinice za površinu", "unit_conversion_direct",
     {"dimensions": ["area"]}, "exact"),
    ("6-13-005", "Preračunavanje mjernih jedinica", "unit_conversion_direct",
     {"dimensions": ["length", "mass", "time", "area"]}, "exact"),
    ("6-09-010", "Ugaone jedinice: stepen, minuta, sekunda",
     "unit_conversion_direct", {"dimensions": ["angle"]}, "exact"),
    ("9-08-011", "Pretvaranje mjernih jedinica za dužinu, površinu i zapreminu",
     "unit_conversion_direct",
     {"dimensions": ["length", "area", "volume"]}, "exact"),
    ("9-08-013", "Pretvaranje brzine m/s i km/h", "unit_conversion_direct",
     {"dimensions": ["speed"]}, "exact"),
]



_BATCH3_ACTIVATIONS = [
    # --- uglovi (6. razred) ------------------------------------------------
    ("6-09-003", "Vrste uglova: nula, oštri, pravi, tupi, opruženi i puni ugao",
     "angle_relationships_direct", {"kinds": ["classify_angle"]}, "exact"),
    ("6-09-004", "Centralni/središnji ugao",
     "angle_relationships_direct", {"kinds": ["central_fraction"]}, "exact"),
    ("6-09-005", "Periferijski/obodni ugao i odnos sa centralnim uglom",
     "angle_relationships_direct", {"kinds": ["central_peripheral"]}, "exact"),
    ("6-09-007", "Upoređivanje uglova",
     "angle_relationships_direct", {"kinds": ["compare_angles"]}, "exact"),
    ("6-09-011", "Sabiranje i oduzimanje mjernih brojeva za uglove",
     "angle_relationships_direct", {"kinds": ["dms_add_sub"]}, "exact"),
    ("6-09-012", "Množenje i dijeljenje uglova prirodnim brojem",
     "angle_relationships_direct", {"kinds": ["angle_times_n"]}, "exact"),
    ("6-09-013", "Susjedni, uporedni i unakrsni uglovi",
     "angle_relationships_direct", {"kinds": ["adjacent_vertical"]}, "exact"),
    ("6-09-014", "Komplementni i suplementni uglovi",
     "angle_relationships_direct", {"kinds": ["comp_supp"]}, "exact"),
    ("6-09-015", "Uglovi uz transverzalu paralelnih pravih",
     "angle_relationships_direct", {"kinds": ["transversal"]}, "exact"),
    ("6-09-016", "Uglovi s paralelnim i normalnim kracima",
     "angle_relationships_direct", {"kinds": ["parallel_normal_arms"]},
     "exact"),
    # --- ugao i trougao (7. razred) ----------------------------------------
    ("7-04-001", "Uglovi sa paralelnim kracima",
     "angle_relationships_direct", {"kinds": ["parallel_normal_arms"]},
     "exact"),
    ("7-04-002", "Uglovi sa normalnim kracima",
     "angle_relationships_direct", {"kinds": ["parallel_normal_arms"]},
     "exact"),
    ("7-04-006", "Vrste trouglova prema stranicama",
     "angle_relationships_direct", {"kinds": ["classify_triangle_sides"]},
     "exact"),
    ("7-04-007", "Vrste trouglova prema uglovima",
     "angle_relationships_direct", {"kinds": ["classify_triangle_angles"]},
     "exact"),
    ("7-04-008", "Zbir unutrašnjih uglova trougla",
     "angle_relationships_direct", {"kinds": ["triangle_third_angle"]},
     "exact"),
    ("7-04-009", "Vanjski ugao trougla",
     "angle_relationships_direct", {"kinds": ["exterior_angle"]}, "exact"),
    ("7-04-010", "Zbir vanjskih uglova trougla",
     "angle_relationships_direct", {"kinds": ["exterior_from_interior"]},
     "exact"),
    ("7-04-011", "Odnos stranica i naspramnih uglova",
     "angle_relationships_direct", {"kinds": ["side_angle_order"]}, "exact"),
    ("7-04-012", "Nejednakost trougla",
     "angle_relationships_direct", {"kinds": ["triangle_inequality"]},
     "exact"),
    ("7-04-018", "Jednakokraki trougao - svojstva",
     "angle_relationships_direct", {"kinds": ["isosceles_angles"]}, "exact"),
    ("7-04-019", "Pravougli trougao i posebni uglovi",
     "angle_relationships_direct", {"kinds": ["right_triangle_acute"]},
     "exact"),
    ("7-04-025", "Srednja linija trougla",
     "geometry_formula_2d", {"kinds": ["triangle_midline"]}, "exact"),
    # --- četverougao, obim i površina (7. razred) --------------------------
    ("7-05-002", "Zbir unutrašnjih uglova četverougla",
     "angle_relationships_direct", {"kinds": ["quad_fourth_angle"]}, "exact"),
    ("7-05-003", "Zbir vanjskih uglova četverougla",
     "angle_relationships_direct", {"kinds": ["quad_exterior"]}, "exact"),
    ("7-05-012", "Srednja linija trapeza",
     "geometry_formula_2d", {"kinds": ["trapezoid_midline"]}, "exact"),
    ("7-05-016", "Obim trougla",
     "geometry_formula_2d", {"kinds": ["triangle_perimeter"]}, "exact"),
    ("7-05-017", "Obim četverougla",
     "geometry_formula_2d", {"kinds": ["quad_perimeter"]}, "exact"),
    ("7-05-019", "Površina pravougaonika i kvadrata - obnova",
     "geometry_formula_2d",
     {"kinds": ["square_area", "rectangle_area", "square_perimeter",
                "rectangle_perimeter"]}, "exact"),
    ("7-05-020", "Površina paralelograma",
     "geometry_formula_2d",
     {"kinds": ["parallelogram_area", "parallelogram_perimeter"]}, "exact"),
    ("7-05-021", "Površina trougla",
     "geometry_formula_2d", {"kinds": ["triangle_area"]}, "exact"),
    ("7-05-022", "Površina trapeza",
     "geometry_formula_2d", {"kinds": ["trapezoid_area"]}, "exact"),
    ("7-05-023", "Površina romba i deltoida preko dijagonala",
     "geometry_formula_2d",
     {"kinds": ["rhombus_area_diagonals", "deltoid_area",
                "rhombus_perimeter"]}, "exact"),
    ("7-05-024", "Površina četverougla sa normalnim dijagonalama",
     "geometry_formula_2d", {"kinds": ["orthodiagonal_area"]}, "exact"),
    # --- mnogougao, kružnica i krug (8. razred) ----------------------------
    ("8-08-002", "Zbir unutrašnjih uglova mnogougla",
     "geometry_formula_2d", {"kinds": ["polygon_interior_sum"]}, "exact"),
    ("8-08-004", "Broj dijagonala mnogougla",
     "geometry_formula_2d", {"kinds": ["polygon_diagonals"]}, "exact"),
    ("8-08-005", "Pravilni mnogougao",
     "geometry_formula_2d",
     {"kinds": ["regular_polygon_angle", "regular_polygon_perimeter"]},
     "exact"),
    ("8-08-007", "Obim i površina mnogougla",
     "geometry_formula_2d",
     {"kinds": ["regular_polygon_perimeter", "quad_perimeter"]}, "exact"),
    ("8-08-009", "Broj π i obim kruga",
     "geometry_formula_2d", {"kinds": ["circle_circumference"]}, "exact"),
    ("8-08-010", "Dužina kružnog luka",
     "geometry_formula_2d", {"kinds": ["arc_length"]}, "exact"),
    ("8-08-011", "Površina kruga",
     "geometry_formula_2d", {"kinds": ["circle_area"]}, "exact"),
    ("8-08-012", "Kružni prsten i kružni isječak",
     "geometry_formula_2d", {"kinds": ["annulus_area", "sector_area"]},
     "exact"),
    # --- Pitagorina teorema (8. razred) ------------------------------------
    ("8-04-001", "Pitagorina teorema - formulacija",
     "pythagoras_direct", {"kinds": ["hypotenuse", "verify_triple"]},
     "exact"),
    ("8-04-002", "Obrat Pitagorine teoreme",
     "pythagoras_direct", {"kinds": ["verify_triple"]}, "exact"),
    ("8-04-003", "Određivanje nepoznate katete",
     "pythagoras_direct", {"kinds": ["leg"]}, "exact"),
    ("8-04-004", "Dijagonala kvadrata",
     "pythagoras_direct", {"kinds": ["square_diagonal"]}, "exact"),
    ("8-04-005", "Dijagonala pravougaonika",
     "pythagoras_direct", {"kinds": ["rectangle_diagonal"]}, "exact"),
    ("8-04-006", "Visina jednakokrakog trougla",
     "pythagoras_direct", {"kinds": ["isosceles_height"]}, "exact"),
    ("8-04-007", "Visina jednakostraničnog trougla",
     "pythagoras_direct", {"kinds": ["equilateral_height"]}, "exact"),
    ("8-04-008", "Površina jednakostraničnog trougla",
     "pythagoras_direct", {"kinds": ["equilateral_area"]}, "exact"),
    ("8-04-009", "Primjena na romb",
     "pythagoras_direct", {"kinds": ["rhombus_side"]}, "exact"),
    ("8-04-010", "Primjena na jednakokraki trapez",
     "pythagoras_direct", {"kinds": ["isosceles_trapezoid_height"]}, "exact"),
    ("8-04-011", "Primjena na pravougli trapez",
     "pythagoras_direct", {"kinds": ["right_trapezoid_leg"]}, "exact"),
    ("8-04-012", "Tetiva i udaljenost od centra",
     "pythagoras_direct", {"kinds": ["chord_distance"]}, "exact"),
    # --- prizme i piramide (8. razred) -------------------------------------
    ("8-05-002", "Poliedar i osnovni elementi",
     "solid_geometry_direct", {"kinds": ["polyhedron_elements"]}, "exact"),
    ("8-05-003", "Prizma - baze, omotač, ivice i visina",
     "solid_geometry_direct", {"kinds": ["polyhedron_elements"]}, "exact"),
    ("8-05-004", "Piramida - baza, bočne strane, vrh i visina",
     "solid_geometry_direct", {"kinds": ["polyhedron_elements"]}, "exact"),
    ("8-05-005", "Pravilna uspravna trostrana prizma",
     "solid_geometry_direct", {"kinds": ["prism3_lateral", "prism3_volume"]},
     "exact"),
    ("8-05-006", "Pravilna uspravna četverostrana prizma",
     "solid_geometry_direct",
     {"kinds": ["prism4_lateral", "prism4_surface", "prism4_volume"]},
     "exact"),
    ("8-05-008", "Pravilna uspravna četverostrana piramida",
     "solid_geometry_direct",
     {"kinds": ["pyramid4_apothem", "pyramid4_lateral", "pyramid4_surface",
                "pyramid4_volume"]}, "exact"),
    ("8-05-013", "Površina omotača pravilne prizme",
     "solid_geometry_direct", {"kinds": ["prism4_lateral", "prism3_lateral"]},
     "exact"),
    ("8-05-014", "Ukupna površina pravilne prizme",
     "solid_geometry_direct", {"kinds": ["prism4_surface"]}, "exact"),
    ("8-05-015", "Zapremina pravilne prizme",
     "solid_geometry_direct", {"kinds": ["prism4_volume", "prism3_volume"]},
     "exact"),
    ("8-05-016", "Prostorna dijagonala pravilne četverostrane prizme",
     "solid_geometry_direct", {"kinds": ["prism4_space_diagonal"]}, "exact"),
    ("8-05-017", "Apotema pravilne piramide",
     "solid_geometry_direct", {"kinds": ["pyramid4_apothem"]}, "exact"),
    ("8-05-018", "Površina omotača pravilne piramide",
     "solid_geometry_direct", {"kinds": ["pyramid4_lateral"]}, "exact"),
    ("8-05-019", "Ukupna površina pravilne piramide",
     "solid_geometry_direct", {"kinds": ["pyramid4_surface"]}, "exact"),
    ("8-05-020", "Zapremina pravilne piramide",
     "solid_geometry_direct", {"kinds": ["pyramid4_volume"]}, "exact"),
    # --- geometrijska tijela (9. razred) -----------------------------------
    ("9-07-001", "Pojam poliedra i njegovi elementi",
     "solid_geometry_direct", {"kinds": ["polyhedron_elements"]}, "exact"),
    ("9-07-002", "Prizma i njeni elementi",
     "solid_geometry_direct", {"kinds": ["polyhedron_elements"]}, "exact"),
    ("9-07-004", "Površina uspravne prizme",
     "solid_geometry_direct", {"kinds": ["prism4_surface", "prism3_lateral"]},
     "exact"),
    ("9-07-005", "Zapremina prizme",
     "solid_geometry_direct",
     {"kinds": ["prism4_volume", "prism3_volume", "cuboid_volume",
                "cube_volume"]}, "exact"),
    ("9-07-007", "Primjena Pitagorine teoreme na pravilnu prizmu",
     "solid_geometry_direct",
     {"kinds": ["prism4_space_diagonal", "cube_space_diagonal"]}, "exact"),
    ("9-07-008", "Piramida i njeni elementi",
     "solid_geometry_direct", {"kinds": ["polyhedron_elements"]}, "exact"),
    ("9-07-010", "Površina pravilne piramide",
     "solid_geometry_direct",
     {"kinds": ["pyramid4_surface", "pyramid4_lateral"]}, "exact"),
    ("9-07-011", "Zapremina pravilne piramide",
     "solid_geometry_direct", {"kinds": ["pyramid4_volume"]}, "exact"),
    ("9-07-012", "Apotema pravilne piramide",
     "solid_geometry_direct", {"kinds": ["pyramid4_apothem"]}, "exact"),
    ("9-07-014", "Primjena Pitagorine teoreme na piramidu",
     "solid_geometry_direct", {"kinds": ["pyramid4_apothem"]}, "exact"),
    ("9-07-017", "Površina valjka",
     "solid_geometry_direct", {"kinds": ["cylinder_surface"]}, "exact"),
    ("9-07-018", "Zapremina valjka",
     "solid_geometry_direct", {"kinds": ["cylinder_volume"]}, "exact"),
    ("9-07-019", "Osni presjek valjka",
     "solid_geometry_direct", {"kinds": ["cylinder_axial_section"]}, "exact"),
    ("9-07-022", "Površina kupe",
     "solid_geometry_direct", {"kinds": ["cone_surface"]}, "exact"),
    ("9-07-023", "Zapremina kupe",
     "solid_geometry_direct", {"kinds": ["cone_volume"]}, "exact"),
    ("9-07-024", "Izvodnica kupe",
     "solid_geometry_direct", {"kinds": ["cone_slant"]}, "exact"),
    ("9-07-025", "Osni presjek kupe",
     "solid_geometry_direct", {"kinds": ["cone_axial_section"]}, "exact"),
    ("9-07-027", "Površina sfere",
     "solid_geometry_direct", {"kinds": ["sphere_surface"]}, "exact"),
    ("9-07-028", "Zapremina lopte",
     "solid_geometry_direct", {"kinds": ["ball_volume"]}, "exact"),
    ("9-07-030", "Odnos zapremina prizme i piramide",
     "solid_geometry_direct", {"kinds": ["prism_pyramid_ratio"]}, "exact"),
    ("9-07-031", "Odnos zapremina valjka i kupe",
     "solid_geometry_direct", {"kinds": ["cylinder_cone_ratio"]}, "exact"),
    ("9-07-038", "Masa tijela iz gustine i zapremine",
     "solid_geometry_direct", {"kinds": ["density_mass"]}, "exact"),
    # --- sistemi (9. razred) -----------------------------------------------
    ("9-05-001", "Linearna jednačina sa dvije nepoznate i njena rješenja",
     "linear_system_direct", {"kinds": ["single_equation"]}, "exact"),
    ("9-05-003", "Pojam sistema dvije linearne jednačine",
     "linear_system_direct", {"kinds": ["verify_pair"]}, "exact"),
    ("9-05-004", "Provjera uređenog para u sistemu",
     "linear_system_direct", {"kinds": ["verify_pair"]}, "exact"),
    ("9-05-005", "Ekvivalentni sistemi",
     "linear_system_direct", {"kinds": ["equivalent_system"]}, "exact"),
    ("9-05-007", "Metoda supstitucije",
     "linear_system_direct", {"kinds": ["solve"]}, "exact"),
    ("9-05-008", "Metoda suprotnih koeficijenata (Gausova metoda)",
     "linear_system_direct", {"kinds": ["solve"]}, "exact"),
    ("9-05-009", "Sistem sa jednim rješenjem",
     "linear_system_direct", {"kinds": ["solve", "classify"]}, "exact"),
    ("9-05-010", "Sistem bez rješenja",
     "linear_system_direct", {"kinds": ["classify"]}, "exact"),
    ("9-05-011", "Sistem sa beskonačno mnogo rješenja",
     "linear_system_direct", {"kinds": ["classify"]}, "exact"),
    ("9-05-012", "Odnos koeficijenata i broj rješenja sistema",
     "linear_system_direct", {"kinds": ["classify"]}, "exact"),
    ("9-05-014", "Geometrijsko tumačenje rješenja sistema",
     "linear_system_direct", {"kinds": ["classify"]}, "exact"),
    ("9-05-015", "Provjera rješenja sistema",
     "linear_system_direct", {"kinds": ["verify_pair"]}, "exact"),
    ("9-05-016", "Sistem sa razlomljenim koeficijentima",
     "linear_system_direct",
     {"kinds": ["solve"], "fraction_coefficients": "yes"}, "exact"),
    # --- polinomi (8. razred) ----------------------------------------------
    ("8-07-003", "Slični monomi",
     "polynomial_basic",
     {"concepts": ["like_terms_select", "combine_like_terms"]}, "exact"),
    ("8-07-004", "Sabiranje i oduzimanje monoma",
     "polynomial_basic", {"concepts": ["combine_like_terms"]}, "exact"),
    ("8-07-005", "Množenje i dijeljenje monoma",
     "polynomial_basic", {"concepts": ["monomial_mul_div"]}, "exact"),
    ("8-07-009", "Kvadrat zbira i razlike",
     "polynomial_basic", {"concepts": ["square_of_binomial"]}, "exact"),
    ("8-07-010", "Razlika kvadrata",
     "polynomial_basic", {"concepts": ["factor_difference_squares"]},
     "exact"),
    ("8-07-011", "Kub binoma, zbir i razlika kubova",
     "polynomial_basic",
     {"concepts": ["cube_of_binomial", "sum_diff_cubes"]}, "exact"),
    ("8-07-012", "Rastavljanje izdvajanjem zajedničkog faktora",
     "polynomial_basic", {"concepts": ["factor_common"]}, "exact"),
    # --- polinomi i kvadratne (9. razred) ----------------------------------
    ("9-06-005", "Kvadrat binoma",
     "polynomial_basic", {"concepts": ["square_of_binomial"]}, "exact"),
    ("9-06-006", "Razlika kvadrata",
     "polynomial_basic", {"concepts": ["factor_difference_squares"]},
     "exact"),
    ("9-06-007", "Kub binoma",
     "polynomial_basic", {"concepts": ["cube_of_binomial"]}, "exact"),
    ("9-06-008", "Zbir i razlika kubova",
     "polynomial_basic", {"concepts": ["sum_diff_cubes"]}, "exact"),
    ("9-06-009", "Izlučivanje zajedničkog faktora",
     "polynomial_basic", {"concepts": ["factor_common"]}, "exact"),
    ("9-06-010", "Rastavljanje grupisanjem",
     "polynomial_basic", {"concepts": ["factor_grouping"]}, "exact"),
    ("9-06-011", "Faktorizacija primjenom identiteta",
     "polynomial_basic", {"concepts": ["factor_identity"]}, "exact"),
    ("9-06-012", "Domena algebarskog razlomka pomoću faktorizacije",
     "polynomial_basic", {"concepts": ["fraction_domain"]}, "exact"),
    ("9-06-016", "Nula proizvoda i provjera rješenja",
     "polynomial_basic", {"concepts": ["zero_product"]}, "exact"),
]
ACTIVATIONS = ACTIVATIONS + _BATCH3_ACTIVATIONS

# ---------------------------------------------------------------------------
# BATCH #4 — racionalni izrazi, strukturisani tekstualni zadaci, skupovi,
# događaji/uzorak, finansije, parametarska diskusija, nejednačine, svojstva
# operacija, pojmovi razlomaka, sličnost, uglovi mnogougla i prava (tekstualno)
# ---------------------------------------------------------------------------

_BATCH4_FAMILIES = {
    "rational_expression_direct": _concept_family(
        "Razlomljeni racionalni izrazi", "rational_expression",
        "izvesti imenovanu radnju nad algebarskim razlomkom",
        ["domain_condition", "numeric_value", "expand", "reduce",
         "equal_fractions", "common_denominator", "add", "subtract",
         "multiply", "divide", "compound_fraction", "simplify_combined"],
        {"domain_condition": "uslov definisanosti (nule imenioca)",
         "numeric_value": "brojna vrijednost algebarskog razlomka",
         "expand": "proširivanje algebarskog razlomka",
         "reduce": "skraćivanje algebarskog razlomka faktorizacijom",
         "equal_fractions": "prepoznavanje jednakih algebarskih razlomaka",
         "common_denominator": "najmanji zajednički imenilac",
         "add": "sabiranje algebarskih razlomaka",
         "subtract": "oduzimanje algebarskih razlomaka",
         "multiply": "množenje algebarskih razlomaka",
         "divide": "dijeljenje algebarskih razlomaka",
         "compound_fraction": "pojednostavljivanje dvojnog razlomka",
         "simplify_combined": "sređivanje razlomljenog racionalnog izraza"},
        fixed=["- domen je dio identiteta: faktor koji se krati zadržava svoj "
               "uslov x ≠ vrijednost i u konačnom odgovoru"]),
    "rational_equation_direct": _concept_family(
        "Jednačine s algebarskim razlomcima", "rational_equation",
        "riješiti linearno rješivu razlomljenu jednačinu",
        ["fraction_equation", "double_fraction_equation"],
        {"fraction_equation": "jednačina s algebarskim razlomcima",
         "double_fraction_equation": "jednačina s dvojnim razlomkom"},
        fixed=["- prije rješavanja se zapisuju uslovi definisanosti; kandidat "
               "koji upada u zabranjene vrijednosti NIJE rješenje"]),
    "structured_word_problem": _concept_family(
        "Strukturisani tekstualni zadaci", "structured_word_problem",
        "riješiti tekstualni zadatak zadatog semantičkog tipa",
        ["equal_sharing", "sharing_remainder", "fraction_of_quantity",
         "fraction_remainder", "money_total", "money_change", "signed_change",
         "number_equation", "sum_difference_system", "sum_multiple_system",
         "box_volume", "cube_surface", "pythagoras_distance",
         "pythagoras_leg"],
        {"equal_sharing": "podjela na jednake dijelove",
         "sharing_remainder": "dijeljenje s ostatkom u priči",
         "fraction_of_quantity": "razlomak od date veličine",
         "fraction_remainder": "ostatak poslije uzetog dijela",
         "money_total": "ukupan iznos kupovine",
         "money_change": "kusur pri plaćanju",
         "signed_change": "promjene s predznakom (temperatura)",
         "number_equation": "priča koja se svodi na linearnu jednačinu",
         "sum_difference_system": "zbir i razlika dva broja (sistem)",
         "sum_multiple_system": "zbir i višekratnik (sistem)",
         "box_volume": "zapremina kvadra iz priče",
         "cube_surface": "površina kocke iz priče",
         "pythagoras_distance": "dijagonala/udaljenost Pitagorinom teoremom",
         "pythagoras_leg": "kateta Pitagorinom teoremom (merdevine)"},
        concept_parameter="problem_types",
        fixed=["- sve poznate veličine moraju biti izričito navedene u "
               "tekstu zadatka; odgovor slijedi isključivo iz njih"]),
    "finite_set_direct": _concept_family(
        "Konačni skupovi i skupovne operacije", "finite_set",
        "izvesti imenovanu skupovnu radnju nad konačnim skupovima",
        ["element_membership", "set_builder_match", "cardinality",
         "subset_equality", "union", "intersection", "difference",
         "complement", "ordered_pair", "cartesian_product"],
        {"element_membership": "pripadnost elementa skupu",
         "set_builder_match": "zadavanje skupa nabrajanjem prema opisu",
         "cardinality": "brojnost skupa (i prazan skup)",
         "subset_equality": "podskup i jednakost skupova",
         "union": "unija skupova", "intersection": "presjek skupova",
         "difference": "razlika skupova",
         "complement": "komplement uz izričit univerzalni skup",
         "ordered_pair": "uređeni par (poredak je bitan)",
         "cartesian_product": "Dekartov proizvod skupova"},
        fixed=["- skupovi se porede kao skupovi: poredak zapisivanja i "
               "ponovljeni elementi ne mijenjaju skup"]),
    "number_set_membership": _concept_family(
        "Brojevni skupovi i pripadnost", "number_set_membership",
        "klasifikovati broj prema brojevnim skupovima",
        ["natural_sets", "set_classification", "irrational_recognition",
         "sqrt_between"],
        {"natural_sets": "skupovi N i N0",
         "set_classification": "pripadnost skupovima N, Z, Q i R",
         "irrational_recognition": "prepoznavanje iracionalnog broja",
         "sqrt_between": "položaj korijena između uzastopnih cijelih brojeva"},
        fixed=["- korijen prirodnog broja je racionalan SAMO za potpun "
               "kvadrat; svaka klasifikacija se obrazlaže"]),
    "event_probability_facts": _concept_family(
        "Događaji i vjerovatnosne odluke", "event_probability",
        "klasifikovati događaj ili uporediti egzaktne vjerovatnoće",
        ["elementary_outcomes", "event_classification",
         "probability_decision", "population_sample"],
        {"elementary_outcomes": "broj elementarnih ishoda ogleda",
         "event_classification": "siguran, nemoguć ili slučajan događaj",
         "probability_decision": "odluka poređenjem egzaktnih vjerovatnoća",
         "population_sample": "populacija naspram uzorka"},
        fixed=["- sastav ogleda (kuglice, strane kockice) mora biti u "
               "potpunosti naveden u zadatku"]),
    "financial_arithmetic_direct": _concept_family(
        "Finansijska aritmetika", "financial_arithmetic",
        "izvesti egzaktan novčani račun sa stopama navedenim u zadatku",
        ["currency_conversion", "simple_interest", "credit_repayment",
         "budget_balance", "percent_change_price"],
        {"currency_conversion": "preračunavanje valuta po navedenom kursu",
         "simple_interest": "prosta kamata na štednju",
         "credit_repayment": "ukupan povrat kredita (glavnica + kamata)",
         "budget_balance": "stanje ličnog budžeta",
         "percent_change_price": "sniženje ili poskupljenje za procenat"},
        fixed=["- svaki kurs i svaka stopa moraju biti navedeni u zadatku; "
               "stvarni bankarski uslovi i složena kamata nisu u obimu"]),
    "parametric_linear_discussion": _concept_family(
        "Diskusija jednačina s parametrom", "parametric_linear",
        "diskutovati linearnu jednačinu ili sistem s parametrom",
        ["parameter_case", "parameter_value_solve",
         "parametric_system_classification"],
        {"parameter_case": "vrijednost parametra za traženi broj rješenja",
         "parameter_value_solve": "rješavanje uz zadanu vrijednost parametra",
         "parametric_system_classification": "broj rješenja sistema s "
                                             "parametrom"},
        fixed=["- podjela slučajeva mora biti potpuna i međusobno "
               "isključiva (koeficijent nula naspram različit od nule)"]),
    "linear_inequality_direct": _concept_family(
        "Linearne nejednačine", "linear_inequality",
        "riješiti ili transformisati linearnu nejednačinu",
        ["equivalent_inequality", "fraction_inequality", "interval_solution"],
        {"equivalent_inequality": "ekvivalentne nejednačine",
         "fraction_inequality": "nejednačina s razlomljenim koeficijentima",
         "interval_solution": "skup rješenja zapisan intervalom"},
        fixed=["- množenje ili dijeljenje negativnim brojem OBRĆE znak "
               "nejednakosti; transformacija se provjerava probnom "
               "vrijednošću"]),
    "operation_property_recognition": _concept_family(
        "Svojstva računskih operacija", "operation_property",
        "prepoznati komutativnost, asocijativnost ili distributivnost",
        ["commutativity_add", "commutativity_mul", "associativity_add",
         "associativity_mul", "distributivity"],
        {"commutativity_add": "komutativnost sabiranja",
         "commutativity_mul": "komutativnost množenja",
         "associativity_add": "asocijativnost sabiranja",
         "associativity_mul": "asocijativnost množenja",
         "distributivity": "distributivnost množenja prema sabiranju"},
        concept_parameter="properties",
        extra_schema={"number_domain": {"kind": "enum",
                                        "values": ["natural", "integer",
                                                   "fraction"],
                                        "required": True}},
        fixed=["- svaka prikazana jednakost mora biti numerički tačna; "
               "opcije se razlikuju po svojstvu, ne po tačnosti"]),
    "fraction_concept_direct": _concept_family(
        "Pojmovi razlomaka i decimalnih zapisa", "fraction_concept",
        "primijeniti imenovani pojam razlomka",
        ["part_of_whole", "fraction_types", "common_denominator_numeric",
         "decimal_type"],
        {"part_of_whole": "razlomak kao dio cjeline i kao količnik",
         "fraction_types": "pravi, nepravi i prividni razlomci; mješoviti "
                           "brojevi",
         "common_denominator_numeric": "svođenje na najmanji zajednički "
                                       "nazivnik",
         "decimal_type": "konačan naspram periodičnog decimalnog zapisa"},
        fixed=["- vrsta decimalnog zapisa dokazuje se faktorizacijom "
               "imenioca skraćenog razlomka (samo 2 i 5 daju konačan zapis)"]),
    "similarity_direct": _concept_family(
        "Sličnost i proporcionalne duži", "similarity",
        "primijeniti koeficijent sličnosti na duži, obim ili površinu",
        ["proportional_segments", "similarity_coefficient",
         "similar_perimeter", "similar_area"],
        {"proportional_segments": "nepoznata duž iz proporcije",
         "similarity_coefficient": "koeficijent sličnosti iz stranica",
         "similar_perimeter": "obim slične figure (skalira se sa k)",
         "similar_area": "površina slične figure (skalira se sa k²)"},
        fixed=["- obim se skalira koeficijentom k, a površina njegovim "
               "kvadratom; svaka tvrdnja se dokazuje množenjem"]),
    "polygon_angle_direct": _concept_family(
        "Uglovi mnogougla", "polygon_angle",
        "primijeniti formule za zbirove uglova mnogougla",
        ["exterior_sum", "interior_sum"],
        {"exterior_sum": "zbir vanjskih uglova (uvijek 360°)",
         "interior_sum": "zbir unutrašnjih uglova (n-2)·180°"},
        fixed=["- zbir vanjskih uglova konveksnog mnogougla ne zavisi od "
               "broja stranica"]),
    "coordinate_line_direct": _concept_family(
        "Prava i linearna funkcija (tekstualno)", "coordinate_line",
        "izvesti imenovanu tekstualnu radnju o pravoj ili funkciji",
        ["point_distance", "slope_meaning", "intercept_meaning",
         "implicit_explicit", "parallel_lines", "line_intersection",
         "dependency_type"],
        {"point_distance": "udaljenost tačaka zadatih koordinatama",
         "slope_meaning": "značenje koeficijenta k",
         "intercept_meaning": "značenje slobodnog člana n",
         "implicit_explicit": "prevođenje implicitnog u eksplicitni oblik",
         "parallel_lines": "paralelne i podudarne prave preko k i n",
         "line_intersection": "presjek dvije prave računom",
         "dependency_type": "direktna, obrnuta ili linearna zavisnost"},
        fixed=["- svi podaci su zadati u tekstu (koordinate, jednačine); "
               "crtanje i čitanje grafika nisu dio zadatka"]),
}
NEW_FAMILIES.update(_BATCH4_FAMILIES)

_LEVEL_BOUNDS.update({
    "rational_expression_direct": {
        "1": "monomski imenioci, jedna direktna radnja",
        "2": "binomni (linearni) imenioci, jedno kraćenje",
        "3": "razlika kvadrata / dva povezana koraka"},
    "rational_equation_direct": {
        "1": "a/x = b, cjelobrojno rješenje",
        "2": "(x+a)/(x+b) = c",
        "3": "razlomak = razlomak, moguć isključen kandidat (nema rješenja)"},
    "structured_word_problem": {
        "1": "jedna direktna relacija, mali brojevi",
        "2": "jedan dodatni korak (ostatak, kusur, druga promjena)",
        "3": "veće vrijednosti ili više povezanih relacija"},
    "finite_set_direct": {
        "1": "skupovi do 4 elementa, jedna radnja",
        "2": "skupovi do 5 elemenata, komplement/podskup",
        "3": "veći skupovi, Dekartov proizvod"},
    "number_set_membership": {
        "1": "mali brojevi, direktna pripadnost",
        "2": "klasifikacija s obrazloženjem",
        "3": "veći korijeni i granični slučajevi"},
    "event_probability_facts": {
        "1": "poznati ogledi (novčić, kockica)",
        "2": "kutije zadanog sastava",
        "3": "poređenje vjerovatnoća i uzorkovanje"},
    "financial_arithmetic_direct": {
        "1": "mali iznosi, jedna godina, okrugle stope",
        "2": "veći iznosi, dvije-tri godine",
        "3": "necjelobrojni iznosi i više stavki"},
    "parametric_linear_discussion": {
        "1": "nema rješenja / beskonačno mnogo (direktan slučaj)",
        "2": "rješavanje uz zadan parametar",
        "3": "potpuna klasifikacija sva tri slučaja"},
    "linear_inequality_direct": {
        "1": "jedan korak, pozitivni koeficijenti",
        "2": "negativan koeficijent (okretanje znaka)",
        "3": "nestroge nejednakosti i intervali"},
    "operation_property_recognition": {
        "1": "mali brojevi, jedno svojstvo",
        "2": "negativni brojevi ili razlomci",
        "3": "kombinovane jednakosti"},
    "fraction_concept_direct": {
        "1": "mali imenioci, direktan pojam",
        "2": "mješoviti brojevi i svođenje",
        "3": "veći imenioci, periodični zapisi"},
    "similarity_direct": {
        "1": "cjelobrojan koeficijent, male duži",
        "2": "veći koeficijenti i obimi",
        "3": "razlomljen koeficijent, površine (k²)"},
    "polygon_angle_direct": {
        "1": "trougao/četverougao/petougao",
        "2": "pravilan mnogougao (vanjski ugao)",
        "3": "mnogouglovi do dvanaest stranica"},
    "coordinate_line_direct": {
        "1": "male Pitagorine trojke / direktno očitavanje",
        "2": "negativni koeficijenti, podudarne prave",
        "3": "razlomljeni koeficijenti i presjeci"},
})

_REVIEWER_NOTES.update({
    "rational_expression_direct": "Invarijanta: domen je dio identiteta — skraćeni oblik zadržava uslove x ≠ nula imenioca; kanonsko poređenje simboličko.",
    "rational_equation_direct": "Invarijanta: uslovi definisanosti prije rješavanja; kandidat u zabranjenoj vrijednosti nije rješenje.",
    "structured_word_problem": "Invarijanta: činjenice prije proze; svaki broj u tekstu odgovara IR veličini; odgovor izveden isključivo iz činjenica.",
    "finite_set_direct": "Invarijanta: jednakost skupova je skupovna (poredak/duplikati nebitni); komplement traži izričit univerzum.",
    "number_set_membership": "Invarijanta: korijen je racionalan samo za potpun kvadrat; klasifikacija dokazana definicijom skupa.",
    "event_probability_facts": "Invarijanta: sastav ogleda naveden u zadatku; vjerovatnoća = povoljni/svi, egzaktno.",
    "financial_arithmetic_direct": "Invarijanta: sve stope u zadatku; egzaktan decimalan račun (nikad float); prosta kamata K = G·p·t/100.",
    "parametric_linear_discussion": "Invarijanta: potpuna i međusobno isključiva podjela slučajeva po koeficijentu uz x.",
    "linear_inequality_direct": "Invarijanta: negativan množilac obrće znak; rješenje potvrđeno probnom vrijednošću.",
    "operation_property_recognition": "Invarijanta: sve jednakosti numerički tačne; opcije se razlikuju po svojstvu.",
    "fraction_concept_direct": "Invarijanta: vrsta decimalnog zapisa dokazana faktorizacijom imenioca (2^a·5^b).",
    "similarity_direct": "Invarijanta: obim ∝ k, površina ∝ k²; proporcije preko unakrsnih proizvoda.",
    "polygon_angle_direct": "Invarijanta: vanjski zbir 360°, unutrašnji (n-2)·180°; pravilan n-ugao samo za n | 360.",
    "coordinate_line_direct": "Invarijanta: sve tvrdnje dokazane uvrštavanjem; udaljenost preko Pitagorine teoreme egzaktno.",
})

_BATCH4_ACTIVATIONS = [
    # --- racionalni izrazi (9. razred) ------------------------------------
    ("9-01-002", "Algebarski razlomak i uslov definiranosti",
     "rational_expression_direct", {"concepts": ["domain_condition"]}, "exact"),
    ("9-01-003", "Brojna vrijednost algebarskog razlomka",
     "rational_expression_direct", {"concepts": ["numeric_value"]}, "exact"),
    ("9-01-004", "Proširivanje algebarskog razlomka",
     "rational_expression_direct", {"concepts": ["expand"]}, "exact"),
    ("9-01-005", "Skraćivanje algebarskog razlomka faktorizacijom",
     "rational_expression_direct", {"concepts": ["reduce"]}, "exact"),
    ("9-01-006", "Jednaki algebarski razlomci",
     "rational_expression_direct", {"concepts": ["equal_fractions"]}, "exact"),
    ("9-01-007", "Zajednički imenilac algebarskih razlomaka",
     "rational_expression_direct", {"concepts": ["common_denominator"]},
     "exact"),
    ("9-01-008", "Sabiranje algebarskih razlomaka",
     "rational_expression_direct", {"concepts": ["add"]}, "exact"),
    ("9-01-009", "Oduzimanje algebarskih razlomaka",
     "rational_expression_direct", {"concepts": ["subtract"]}, "exact"),
    ("9-01-010", "Množenje algebarskih razlomaka",
     "rational_expression_direct", {"concepts": ["multiply"]}, "exact"),
    ("9-01-011", "Dijeljenje algebarskih razlomaka",
     "rational_expression_direct", {"concepts": ["divide"]}, "exact"),
    ("9-01-012", "Složeni razlomljeni racionalni izraz",
     "rational_expression_direct", {"concepts": ["compound_fraction"]},
     "exact"),
    ("9-01-013", "Sređivanje razlomljenog racionalnog izraza",
     "rational_expression_direct", {"concepts": ["simplify_combined"]},
     "exact"),
    ("9-01-014", "Jednačine sa algebarskim razlomcima",
     "rational_equation_direct", {"concepts": ["fraction_equation"]}, "exact"),
    ("9-04-005", "Jednačina sa algebarskim razlomcima",
     "rational_equation_direct", {"concepts": ["fraction_equation"]}, "exact"),
    ("9-04-006", "Jednačina sa dvojnim razlomkom",
     "rational_equation_direct", {"concepts": ["double_fraction_equation"]},
     "exact"),
    # --- strukturisani tekstualni zadaci ----------------------------------
    ("6-03-010", "Tekstualni zadaci iz djeljivosti",
     "structured_word_problem",
     {"problem_types": ["equal_sharing", "sharing_remainder"]}, "exact"),
    ("6-04-015", "Tekstualni zadaci s razlomcima",
     "structured_word_problem",
     {"problem_types": ["fraction_of_quantity", "fraction_remainder"]},
     "exact"),
    ("6-05-011", "Brojevni izrazi i tekstualni zadaci s decimalnim brojevima",
     "structured_word_problem",
     {"problem_types": ["money_total", "money_change"]}, "exact"),
    ("7-02-021", "Tekstualni zadaci s cijelim brojevima",
     "structured_word_problem", {"problem_types": ["signed_change"]}, "exact"),
    ("7-03-020", "Primjena linearne jednačine u tekstualnom zadatku",
     "structured_word_problem", {"problem_types": ["number_equation"]},
     "exact"),
    ("9-04-011", "Tekstualni zadatak koji se svodi na linearnu jednačinu",
     "structured_word_problem", {"problem_types": ["number_equation"]},
     "exact"),
    ("9-05-013", "Tekstualni zadatak sa sistemom",
     "structured_word_problem",
     {"problem_types": ["sum_difference_system", "sum_multiple_system"]},
     "exact"),
    ("9-07-033", "Tekstualni zadaci sa površinom i zapreminom",
     "structured_word_problem",
     {"problem_types": ["box_volume", "cube_surface"]}, "exact"),
    ("8-04-016", "Praktični problemski zadaci",
     "structured_word_problem",
     {"problem_types": ["pythagoras_distance", "pythagoras_leg"]}, "exact"),
    # --- skupovi (6. razred) + brojevni skupovi ---------------------------
    ("6-01-001", "Pojam skupa, elementi skupa i označavanje",
     "finite_set_direct", {"concepts": ["element_membership"]}, "exact"),
    ("6-01-002", "Načini zadavanja skupa",
     "finite_set_direct", {"concepts": ["set_builder_match"]}, "exact"),
    ("6-01-003", "Brojnost skupa i prazan skup",
     "finite_set_direct", {"concepts": ["cardinality"]}, "exact"),
    ("6-01-004", "Podskup i jednakost skupova",
     "finite_set_direct", {"concepts": ["subset_equality"]}, "exact"),
    ("6-01-006", "Unija skupova",
     "finite_set_direct", {"concepts": ["union"]}, "exact"),
    ("6-01-007", "Presjek skupova",
     "finite_set_direct", {"concepts": ["intersection"]}, "exact"),
    ("6-01-008", "Razlika skupova",
     "finite_set_direct", {"concepts": ["difference"]}, "exact"),
    ("6-01-009", "Komplement skupa",
     "finite_set_direct", {"concepts": ["complement"]}, "exact"),
    ("6-01-010", "Uređeni par",
     "finite_set_direct", {"concepts": ["ordered_pair"]}, "exact"),
    ("6-01-011", "Direktni/Dekartov proizvod skupova",
     "finite_set_direct", {"concepts": ["cartesian_product"]}, "exact"),
    ("6-02-001", "Skupovi N i N0",
     "number_set_membership", {"concepts": ["natural_sets"]}, "exact"),
    ("8-01-001", "Skupovi N, Z, Q, I i R i odnosi među njima",
     "number_set_membership", {"concepts": ["set_classification"]}, "exact"),
    ("8-01-002", "Iracionalni brojevi i beskonačni neperiodični decimalni zapis",
     "number_set_membership", {"concepts": ["irrational_recognition"]},
     "exact"),
    ("8-01-003", "Realni brojevi i brojevna osa",
     "number_set_membership", {"concepts": ["sqrt_between"]}, "exact"),
    # --- događaji, odluke, uzorak -----------------------------------------
    ("8-06-010", "Elementarni ishod i slučajni događaj",
     "event_probability_facts", {"concepts": ["elementary_outcomes"]},
     "exact"),
    ("8-06-011", "Siguran, nemoguć i slučajan događaj",
     "event_probability_facts", {"concepts": ["event_classification"]},
     "exact"),
    ("9-08-003", "Donošenje odluke na osnovu vjerovatnoće",
     "event_probability_facts", {"concepts": ["probability_decision"]},
     "exact"),
    ("9-08-009", "Populacija i uzorak",
     "event_probability_facts", {"concepts": ["population_sample"]}, "exact"),
    # --- finansijska aritmetika -------------------------------------------
    ("8-03-019", "Jednostavni kamatni račun",
     "financial_arithmetic_direct", {"concepts": ["simple_interest"]},
     "exact"),
    ("9-08-004", "Preračunavanje valuta",
     "financial_arithmetic_direct", {"concepts": ["currency_conversion"]},
     "exact"),
    ("9-08-005", "Kamata na štednju",
     "financial_arithmetic_direct", {"concepts": ["simple_interest"]},
     "exact"),
    ("9-08-006", "Kamata na kredit",
     "financial_arithmetic_direct", {"concepts": ["credit_repayment"]},
     "exact"),
    ("9-08-008", "Lični finansijski budžet",
     "financial_arithmetic_direct",
     {"concepts": ["budget_balance", "percent_change_price"]}, "exact"),
    # --- parametarska diskusija -------------------------------------------
    ("9-04-007", "Jednačina sa promjenljivim koeficijentom",
     "parametric_linear_discussion", {"concepts": ["parameter_value_solve"]},
     "exact"),
    ("9-04-022", "Diskusija linearne jednačine sa parametrom",
     "parametric_linear_discussion", {"concepts": ["parameter_case"]},
     "exact"),
    ("9-05-017", "Sistem sa parametrom",
     "parametric_linear_discussion",
     {"concepts": ["parametric_system_classification"]}, "exact"),
    # --- nejednačine ------------------------------------------------------
    ("9-04-013", "Ekvivalentne nejednačine",
     "linear_inequality_direct", {"concepts": ["equivalent_inequality"]},
     "exact"),
    ("9-04-015", "Nejednačina sa razlomcima",
     "linear_inequality_direct", {"concepts": ["fraction_inequality"]},
     "exact"),
    ("9-04-018", "Skup rješenja intervalom",
     "linear_inequality_direct", {"concepts": ["interval_solution"]},
     "exact"),
    # --- svojstva operacija -----------------------------------------------
    ("6-02-006", "Svojstva računskih operacija: komutativnost, asocijativnost, distributivnost",
     "operation_property_recognition",
     {"properties": ["commutativity_add", "commutativity_mul",
                     "associativity_add", "associativity_mul",
                     "distributivity"],
      "number_domain": "natural"}, "exact"),
    ("6-04-013", "Svojstva računskih operacija s razlomcima",
     "operation_property_recognition",
     {"properties": ["commutativity_add", "commutativity_mul",
                     "associativity_add", "associativity_mul",
                     "distributivity"],
      "number_domain": "fraction"}, "exact"),
    ("7-02-010", "Komutativnost i asocijativnost sabiranja",
     "operation_property_recognition",
     {"properties": ["commutativity_add", "associativity_add"],
      "number_domain": "integer"}, "exact"),
    ("7-02-014", "Distributivnost u skupu Z",
     "operation_property_recognition",
     {"properties": ["distributivity"], "number_domain": "integer"}, "exact"),
    ("7-03-013", "Svojstva računskih operacija u Q",
     "operation_property_recognition",
     {"properties": ["commutativity_add", "commutativity_mul",
                     "associativity_add", "associativity_mul",
                     "distributivity"],
      "number_domain": "fraction"}, "exact"),
    # --- pojmovi razlomaka ------------------------------------------------
    ("6-04-002", "Razlomak kao dio cjeline i kao količnik",
     "fraction_concept_direct", {"concepts": ["part_of_whole"]}, "exact"),
    ("6-04-003", "Pravi, nepravi i prividni razlomci; mješoviti brojevi",
     "fraction_concept_direct", {"concepts": ["fraction_types"]}, "exact"),
    ("6-04-007", "Svođenje razlomaka na zajednički nazivnik/imenilac",
     "fraction_concept_direct", {"concepts": ["common_denominator_numeric"]},
     "exact"),
    ("6-05-004", "Konačni i beskonačni/periodični decimalni brojevi",
     "fraction_concept_direct", {"concepts": ["decimal_type"]}, "exact"),
    # --- sličnost ---------------------------------------------------------
    ("8-03-008", "Proporcionalne duži",
     "similarity_direct", {"concepts": ["proportional_segments"]}, "exact"),
    ("8-03-013", "Pojam sličnosti i koeficijent sličnosti",
     "similarity_direct", {"concepts": ["similarity_coefficient"]}, "exact"),
    ("8-03-015", "Dužine i obimi sličnih trouglova",
     "similarity_direct",
     {"concepts": ["similarity_coefficient", "similar_perimeter"]}, "exact"),
    ("8-03-016", "Površine sličnih trouglova",
     "similarity_direct", {"concepts": ["similar_area"]}, "exact"),
    # --- uglovi mnogougla -------------------------------------------------
    ("8-08-003", "Zbir vanjskih uglova konveksnog mnogougla",
     "polygon_angle_direct", {"concepts": ["exterior_sum", "interior_sum"]},
     "exact"),
    # --- mjerne jedinice površine/zapremine (postojeća porodica) ----------
    ("7-05-018", "Mjerne jedinice za površinu",
     "unit_conversion_direct", {"dimensions": ["area"]}, "exact"),
    ("8-05-022", "Mjerne jedinice površine i zapremine",
     "unit_conversion_direct", {"dimensions": ["area", "volume"]}, "exact"),
    ("9-07-034", "Mjerne jedinice površine i zapremine u zadacima sa tijelima",
     "unit_conversion_direct", {"dimensions": ["area", "volume"]}, "exact"),
    # --- prava i linearna funkcija (tekstualno) ---------------------------
    ("8-02-004", "Udaljenost između dvije tačke",
     "coordinate_line_direct", {"concepts": ["point_distance"]}, "exact"),
    ("8-02-008", "Geometrijsko značenje koeficijenta k",
     "coordinate_line_direct", {"concepts": ["slope_meaning"]}, "exact"),
    ("8-02-009", "Geometrijsko značenje slobodnog člana n",
     "coordinate_line_direct", {"concepts": ["intercept_meaning"]}, "exact"),
    ("8-02-012", "Eksplicitni i implicitni oblik prave",
     "coordinate_line_direct", {"concepts": ["implicit_explicit"]}, "exact"),
    ("8-02-014", "Paralelne i presječne prave preko koeficijenta k",
     "coordinate_line_direct", {"concepts": ["parallel_lines"]}, "exact"),
    ("9-03-005", "Značenje koeficijenta k",
     "coordinate_line_direct", {"concepts": ["slope_meaning"]}, "exact"),
    ("9-03-006", "Značenje slobodnog člana n",
     "coordinate_line_direct", {"concepts": ["intercept_meaning"]}, "exact"),
    ("9-03-015", "Paralelne i podudarne prave prema koeficijentima",
     "coordinate_line_direct", {"concepts": ["parallel_lines"]}, "exact"),
    ("9-03-017", "Presjek dvije prave",
     "coordinate_line_direct", {"concepts": ["line_intersection"]}, "exact"),
    ("9-03-021", "Poređenje direktne, obrnute i linearne zavisnosti",
     "coordinate_line_direct", {"concepts": ["dependency_type"]}, "exact"),
]

ACTIVATIONS = ACTIVATIONS + _BATCH4_ACTIVATIONS

# ---------------------------------------------------------------------------
# KLASIFIKACIJA PREOSTALIH LEKCIJA (izvještaj, bez uticaja na ponašanje)
# ---------------------------------------------------------------------------

_VISUAL_KEYWORDS = ("na brojevnoj polupravoj",
                    "konstrukc", "crtanje", "crtež", "dijagram", "grafič",
                    "prikaz", "koordinat", "simetri", "preslikav", "izometri",
                    "vektor", "mreža", "čitanje i kritičko")
_WORD_PROBLEM_KEYWORDS = ("tekstualni", "primjena", "problemi")
_PROOF_KEYWORDS = ("dokaz", "teorema", "diskusija")
_NEEDS_EXTENSION = {
    # Kandidati čiji bi zadaci stali u POSTOJEĆE motore uz proširenje
    # parametara/oblika — namjerno NEaktivirani u ovoj fazi (Batch #2 je
    # aktivirao prethodnih 26; ovo je preostali rep).
    "9-04-013",   # ekvivalentne NEJEDNAČINE — treba nejednačinska varijanta
    "9-04-007",   # promjenljivi koeficijent — treba parametarska varijanta
    "6-07-007",   # prikaz rješenja na brojevnoj osi — treba vizuelni format
}
# Batch #3, Priority A: 6-07-007 stvarno traži prikaz na brojevnoj
# osi (vizuelno); 9-04-007 traži parametarsku diskusiju (nov
# kapacitet — pada u NEEDS_NEW preko ključne riječi).
_NEEDS_EXTENSION -= {"6-07-007", "9-04-007"}
# Batch #4: 9-04-013 je aktivirana porodicom linear_inequality_direct.
_NEEDS_EXTENSION -= {"9-04-013"}


def _classify(lesson_id, title, activated):
    lowered = title.lower()
    if lesson_id in activated:
        return "DETERMINISTIC_READY"
    if lesson_id in _NEEDS_EXTENSION:
        return "DETERMINISTIC_NEEDS_EXISTING_CAPABILITY_EXTENSION"
    if any(keyword in lowered for keyword in _PROOF_KEYWORDS):
        return "MODEL_ONLY_FOR_NOW"
    if any(keyword in lowered for keyword in _VISUAL_KEYWORDS):
        return "VISUAL_OR_CONSTRUCTION_REQUIRED"
    if any(keyword in lowered for keyword in _WORD_PROBLEM_KEYWORDS):
        return "MODEL_ONLY_FOR_NOW"
    if any(keyword in lowered for keyword in
           ("jednačin", "nejednačin", "razmjer", "proporc", "procent",
            "kamata", "stepen", "korijen", "polinom", "izraz", "funkcij",
            "frekvencij", "vjerovatno", "pretvaranje", "zaokruživanje")):
        return "DETERMINISTIC_NEEDS_NEW_CAPABILITY"
    return "MODEL_ONLY_FOR_NOW"


# ---------------------------------------------------------------------------
# DOKAZI IZ FAZE 2
# ---------------------------------------------------------------------------

def _load_exact_evidence():
    """lesson_id → sortirana lista item_id-jeva s relacijom `exact`."""
    if not FAZA2_XLSX.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        return {}
    workbook = openpyxl.load_workbook(FAZA2_XLSX, read_only=True, data_only=True)
    if "Mapiranje" not in workbook.sheetnames:
        return {}
    sheet = workbook["Mapiranje"]
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell or "") for cell in next(rows)]
    index = {name: position for position, name in enumerate(header)}
    lesson_column = next((name for name in header if "lesson" in name.lower()
                          or "lekcij" in name.lower()), None)
    item_column = next((name for name in header if name.lower() == "item_id"), None)
    relation_column = next((name for name in header
                            if "relation" in name.lower()), None)
    if not (lesson_column and item_column and relation_column):
        return {}
    evidence = {}
    for row in rows:
        relation = str(row[index[relation_column]] or "")
        if relation != "exact":
            continue
        lesson_id = str(row[index[lesson_column]] or "")
        item_id = str(row[index[item_column]] or "")
        if lesson_id and item_id:
            evidence.setdefault(lesson_id, set()).add(item_id)
    workbook.close()
    return {key: sorted(values) for key, values in evidence.items()}


# ---------------------------------------------------------------------------
# GRADNJA ARTEFAKATA
# ---------------------------------------------------------------------------

class OnboardingError(RuntimeError):
    pass


def _read(path):
    return json.loads(path.read_text(encoding="utf-8"))


def _write(path, payload):
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
                    encoding="utf-8", newline="\n")


def _canonical_lessons():
    payload = _read(TOPICS_PATH)
    return {lesson["id"]: (grade_key, lesson["oblast"], lesson["title"])
            for grade_key, grade in payload["grades"].items()
            for lesson in grade["lessons"]}


def build_assignment_rows(lessons, exact_evidence):
    rows = []
    for lesson_id, expected_title, family_id, parameters, provenance in ACTIVATIONS:
        if lesson_id not in lessons:
            raise OnboardingError(f"{lesson_id}: lekcija ne postoji u kurikulumu")
        _grade, _oblast, title = lessons[lesson_id]
        if title != expected_title:
            raise OnboardingError(
                f"{lesson_id}: naslov se ne poklapa s pregledanim "
                f"({title!r} != {expected_title!r})")
        if family_id not in NEW_FAMILIES:
            raise OnboardingError(f"{lesson_id}: nepoznata porodica {family_id}")
        evidence = list(exact_evidence.get(lesson_id, ()))
        if provenance == "pilot":
            evidence.append("F25-PILOT-READY")
        if not evidence:
            # Izričita oznaka aktivacije na osnovu jednoznačnog naslova —
            # nikad izmišljen NPP izvor.
            evidence.append(f"CANON-TITLE-{lesson_id}")
        rows.append({
            "lesson_id": lesson_id,
            "family_id": family_id,
            "enforcement_mode": "blocking",
            "activation_class": "READY",
            "evidence_ids": sorted(set(evidence)),
            "parameters": parameters,
            "level_bounds": _LEVEL_BOUNDS[family_id],
            "forbidden_neighbour_skills": [],
            "reviewer_note": _REVIEWER_NOTES[family_id],
        })
    return rows


def merge_sources():
    lessons = _canonical_lessons()
    exact_evidence = _load_exact_evidence()

    families_payload = _read(FAMILIES_PATH)
    families = families_payload["families"]
    generated_rows = build_assignment_rows(lessons, exact_evidence)

    for family_id, definition in NEW_FAMILIES.items():
        families[family_id] = definition
    families_payload["families"] = dict(sorted(families.items()))

    assignments_payload = _read(ASSIGNMENTS_PATH)
    activated_ids = {row["lesson_id"] for row in generated_rows}
    kept = [row for row in assignments_payload["assignments"]
            if row["lesson_id"] not in activated_ids
            and row["family_id"] not in NEW_FAMILIES]
    overlap = {row["lesson_id"] for row in kept} & activated_ids
    if overlap:
        raise OnboardingError(f"dodjela postoji dvaput: {sorted(overlap)}")
    assignments_payload["assignments"] = sorted(
        kept + generated_rows, key=lambda row: row["lesson_id"])
    assignments_payload["contract_version"] = CONTRACT_VERSION
    return families_payload, assignments_payload, lessons


def verify_generator_support():
    """Svaka blocking dodjela nove porodice MORA imati potpun generator."""
    from matbot import deterministic as registry
    from matbot.semantics import contracts as semantic_contracts

    semantic_contracts.reset_cache()
    unsupported = []
    for lesson_id, contract in sorted(semantic_contracts.all_contracts().items()):
        module = registry.GENERATORS.get(contract.family_id)
        if module is None:
            unsupported.append((lesson_id, contract.family_id, "nema generatora"))
        elif not module.supports(dict(contract.parameters)):
            unsupported.append((lesson_id, contract.family_id,
                                "parametri nisu podržani"))
    if unsupported:
        raise OnboardingError(f"generator ne pokriva: {unsupported}")


_ANALYSIS_RULES = (
    ("PROOF_REQUIRED", ("dokaz", "podudarnost", "teorema o", "talesova")),
    ("CONSTRUCTION_REQUIRED", ("konstrukc",)),
    ("VISUAL_REQUIRED", ("grafik", "dijagram", "crtanje", "prikaz",
                         "koordinatn", "brojevnoj", "simetri", "preslikav",
                         "čitanje")),
    ("STRUCTURED_WORD_PROBLEM_CANDIDATE", ("tekstualni zadaci",
                                           "tekstualni zadatak")),
    ("OPEN_ENDED_WORD_PROBLEM", ("problemsk", "praktični", "primjena u",
                                 "modeliranje")),
    ("SYMBOLIC_ALGEBRA_CANDIDATE", ("razlomljen", "algebarsk", "polinom",
                                    "izraz", "monom")),
    ("FORMULA_GEOMETRY_CANDIDATE", ("površina", "obim", "zapremina",
                                    "dijagonala", "visina", "presjek")),
    ("FINANCIAL_MATH_CANDIDATE", ("kamata", "valut", "budžet", "otplatn")),
    ("STATISTICS_CANDIDATE", ("frekvencij", "uzorak", "podataka", "sredina")),
    ("CLOSED_FORM_DETERMINISTIC_CANDIDATE",
     ("pretvaranje", "zaokruživanje", "vrijednost", "račun", "jednačin",
      "nejednačin", "procjena", "korijen", "stepen")),
    ("CONCEPTUAL_ONLY", ("pojam", "elementi", "svojstva", "vrste", "načini",
                         "definicija", "odnos", "skup", "označavanje")),
)


def _analysis_bucket(title):
    lowered = title.lower()
    for bucket, keywords in _ANALYSIS_RULES:
        if any(keyword in lowered for keyword in keywords):
            return bucket
    return "INSUFFICIENT_SEMANTIC_EVIDENCE"


def build_report(lessons):
    activated = {row[0]: row[2] for row in ACTIVATIONS}
    activated["6-04-009"] = "fraction_arithmetic_direct"
    activated["6-04-010"] = "fraction_arithmetic_direct"
    activated["6-04-011"] = "fraction_arithmetic_direct"
    activated["6-04-012"] = "fraction_arithmetic_direct"

    classes, by_grade, by_family, table = {}, {}, {}, []
    for lesson_id, (grade, oblast, title) in sorted(lessons.items()):
        classification = _classify(lesson_id, title, activated)
        classes[classification] = classes.get(classification, 0) + 1
        if classification == "DETERMINISTIC_READY":
            family = activated[lesson_id]
            by_family[family] = by_family.get(family, 0) + 1
            by_grade[grade] = by_grade.get(grade, 0) + 1
        entry = {"lesson_id": lesson_id, "grade": grade, "oblast": oblast,
                 "title": title, "class": classification,
                 "family": activated.get(lesson_id, "")}
        if classification == "MODEL_ONLY_FOR_NOW":
            entry["analysis"] = _analysis_bucket(title)
        table.append(entry)
    total = len(lessons)
    deterministic = classes.get("DETERMINISTIC_READY", 0)
    analysis_totals = {}
    for entry in table:
        bucket = entry.get("analysis")
        if bucket:
            analysis_totals[bucket] = analysis_totals.get(bucket, 0) + 1
    return {
        "model_only_analysis": dict(sorted(analysis_totals.items())),
        "contract_version": CONTRACT_VERSION,
        "total_lessons": total,
        "deterministic_lessons": deterministic,
        "deterministic_share": round(deterministic / total, 4),
        "class_totals": dict(sorted(classes.items())),
        "deterministic_by_grade": dict(sorted(by_grade.items())),
        "deterministic_by_family": dict(sorted(by_family.items())),
        "lessons": table,
    }


def main(argv=None):
    parser = argparse.ArgumentParser(description="Masovno determinističko uključivanje")
    parser.add_argument("--check", action="store_true")
    parser.add_argument("--report", action="store_true")
    args = parser.parse_args(argv)

    families_payload, assignments_payload, lessons = merge_sources()
    report = build_report(lessons)

    if args.check:
        current_families = _read(FAMILIES_PATH)
        current_assignments = _read(ASSIGNMENTS_PATH)
        current_report = _read(REPORT_PATH) if REPORT_PATH.exists() else None
        if (current_families != families_payload
                or current_assignments != assignments_payload
                or current_report != report):
            print("Artefakti NISU u koraku s tabelom aktivacija.", file=sys.stderr)
            return 1
        verify_generator_support()
        print("OK: artefakti su u koraku s tabelom aktivacija.")
        return 0

    _write(FAMILIES_PATH, families_payload)
    _write(ASSIGNMENTS_PATH, assignments_payload)
    _write(REPORT_PATH, report)

    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "build_lesson_semantics", ROOT / "scripts" / "build_lesson_semantics.py")
    build_lesson_semantics = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(build_lesson_semantics)
    build_lesson_semantics.write()
    verify_generator_support()

    if args.report:
        print(f"UKUPNO LEKCIJA:      {report['total_lessons']}")
        print(f"DETERMINISTIČKI:     {report['deterministic_lessons']} "
              f"({report['deterministic_share'] * 100:.1f} %)")
        for name, count in report["class_totals"].items():
            print(f"  {name:48s} {count}")
        print("Po razredu:", report["deterministic_by_grade"])
        for family, count in report["deterministic_by_family"].items():
            print(f"  {family:36s} {count}")
    print(f"OK: {len(assignments_payload['assignments'])} dodjela, "
          f"{len(families_payload['families'])} porodica")
    return 0


if __name__ == "__main__":
    sys.exit(main())
