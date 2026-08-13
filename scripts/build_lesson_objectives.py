"""Kompajlira PRIMARNU VJEŠTINU lekcije iz kanonske Faze-2 mapiranja.

    python scripts/build_lesson_objectives.py            # provjeri (ne piše)
    python scripts/build_lesson_objectives.py --write    # upiši artefakt

ZAŠTO POSTOJI (živi nalaz: semantički zanos susjedne vještine):
Lekcija „Skup racionalnih brojeva Q“ dobijala je zadatke čija je STVARNA
vještina sabiranje razlomaka. Zadatak jeste sadržavao racionalne brojeve, ali
nije ispitivao ono što lekcija predaje. Mjerena je bila samo pripadnost sesije
lekciji, a to nije isto što i semantička vjernost.

Uzrok NIJE bio nedostatak podataka. Kanonsko mapiranje NPP stavki na lekcije
(reference/curriculum/semantics/MATBOT_Faza2_Mapiranje.xlsx) već nosi, po
lekciji, ishode učenja i njihovu RELACIJU prema lekciji:

    exact        — ishod TE lekcije (primarna vještina)
    supporting   — pomoćni/preduslovni pojam (smije se pojaviti, ne smije biti cilj)
    prerequisite — isto, preduslov
    neighbour    — ishod SUSJEDNE lekcije (upravo ono u šta zadatak zanosi)

Mjereno nad kurikulumom: 152 od 184 model-podržanih lekcija dobijalo je u
promptu SAMO naslov. Za 89 njih ova evidencija postoji i nikad nije bila
korištena. Ovaj skript je izlaže — nijedan podatak se ne izmišlja.

FILTER PROTIV POGREŠNOG MAPIRANJA: sirova evidencija nije savršena (ima
`exact` redova očito prenesenih s druge lekcije). Zato `exact` ishod ulazi u
primarnu vještinu SAMO ako dijeli sadržajnu riječ s naslovom ili oblašću
lekcije. Filter je opšti i deterministički — nema izuzetka po lekciji.

MORFOLOGIJA (mjereni nalaz): prvo poređenje tražilo je DOSLOVNO istu riječ, pa
je u jeziku s bogatom deklinacijom odbacivalo tačnu evidenciju — „proširivanjem“
nije bilo isto što i naslovno „Proširivanje“. Tako je 141 ispravan `exact` red
nestao, a 58 lekcija ostalo bez ijednog mjerila. Poređenje sada priznaje da je
jedna riječ nastavak druge (zajednički prefiks), pa evidencija preživi promjenu
padeža. Rangiranje po broju poklapanja štiti od suprotne greške: kad se u istoj
lekciji nađe i tačan i prenesen red, prolazi onaj koji dijeli VIŠE sadržaja.

LEKCIJE BEZ IJEDNE NPP STAVKE: za njih se ne izmišlja ishod. Umjesto toga se iz
kanonskog spiska lekcija izvodi ono što se DA dokazati — koje SUSJEDNE lekcije
iste oblasti postoje i po čemu se razlikuju. To je granica opsega, ne ishod, pa
nosi `objective_source`/`objective_confidence` i vidljivo je u reviziji.
"""
import argparse
import io
import json
import re
import sys
import unicodedata
from collections import OrderedDict, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
XLSX = ROOT / "reference" / "curriculum" / "semantics" / "MATBOT_Faza2_Mapiranje.xlsx"
TOPICS = ROOT / "data" / "topics.json"
OUTPUT = ROOT / "data" / "lesson_objectives.compiled.json"

MAX_PRIMARY = 4
MAX_SUPPORTING = 3
MAX_EXCLUSIONS = 4
MAX_CHARS = 190

PRIMARY_RELATIONS = {"exact"}
SUPPORTING_RELATIONS = {"supporting", "prerequisite"}
NEIGHBOUR_RELATIONS = {"neighbour"}

# Riječi bez sadržaja — ne dokazuju vezu ishoda s lekcijom.
_STOPWORDS = frozenset("""
i ili te a u na za od do sa se je su bez kao što sto koji koja koje kojih
znati razumjeti moći umjeti uspješno pravilno datih date dati dato naći
primjenom pomoću prema pri kroz iz o s
""".split())
_WORD_RE = re.compile(r"[^\W\d_]{4,}", re.UNICODE)
# Rečenice u izvoru ponekad nose zalijepljen pojmovnik iza tačke:
# „… u skupu racionalnim postupkom. Pozitivni racionalni brojevi Negativni …“
_GLOSSARY_TAIL_RE = re.compile(r"\.\s+(?=[A-ZČĆŽŠĐ])")


def _fold(word):
    return "".join(
        c for c in unicodedata.normalize("NFKD", word.lower())
        if not unicodedata.combining(c))


def _content_words(text):
    return {_fold(w) for w in _WORD_RE.findall(text or "")
            if _fold(w) not in {_fold(s) for s in _STOPWORDS}}


def _clean(text):
    """Prva rečenica bez zalijepljenog pojmovnika, dužinski ograničena.

    Tekst ide U PROMPT, dakle u projektnu prozu — zato prolazi kroz istu
    determinističku normalizaciju terminologije kao svaki drugi vidljivi tekst
    (`matbot/terminology.py`). Kurikularni izvor ostaje netaknut."""
    from matbot.terminology import normalize_terminology

    body = " ".join(normalize_terminology(text or "").split())
    if not body:
        return ""
    head = _GLOSSARY_TAIL_RE.split(body)[0].strip(" ,;:.")
    if len(head) > MAX_CHARS:
        head = head[:MAX_CHARS].rsplit(" ", 1)[0] + "…"
    return head


def _related(left, right):
    """Iste riječi u različitom padežu su ISTA sadržajna riječ.

    Priznaje se samo nastavak (jedna riječ je prefiks druge, bar 4 znaka) ili
    dovoljno dug zajednički prefiks (6 znakova). Kraći zajednički prefiks bi
    spojio „prosti“ i „prostor“, što su različiti pojmovi."""
    if left == right:
        return True
    if len(left) >= 4 and len(right) >= 4 and (left.startswith(right)
                                               or right.startswith(left)):
        return True
    shared = 0
    for one, other in zip(left, right):
        if one != other:
            break
        shared += 1
    # Zajednički korijen uz KRATKE nastavke: „liniju“/„linija“ jesu ista riječ,
    # a „linija“/„linearna“ nisu — nastavak od tri znaka je granica padeža.
    return shared >= 5 and (len(left) - shared) <= 3 and (len(right) - shared) <= 3


def _content_overlap(statement, lesson_words):
    """Broj sadržajnih riječi ishoda koje pripadaju lekciji (0 = nepovezano)."""
    return sum(1 for word in _content_words(statement)
               if any(_related(word, known) for known in lesson_words))


def _shares_content(statement, lesson_words):
    """Deterministički filter pogrešnog mapiranja: bar jedna sadržajna riječ."""
    return _content_overlap(statement, lesson_words) > 0


def load_mapping():
    import openpyxl

    workbook = openpyxl.load_workbook(XLSX, read_only=True)
    sheet = workbook["Mapiranje"]
    header = [c for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: position for position, name in enumerate(header)}
    grouped = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        lesson_id = row[index["target_lesson_id"]]
        if not lesson_id:
            continue
        grouped[str(lesson_id).strip()].append({
            "text": _clean(row[index["source_text"]]),
            "relation": (row[index["relation"]] or "").strip().lower(),
            "confidence": (row[index["confidence"]] or "").strip().lower(),
            "item_id": row[index["item_id"]],
        })
    return grouped


def _distinguishing_words(lesson, siblings):
    """Riječi naslova koje NE dijeli većina susjeda iste oblasti.

    „uglovi“ stoji u skoro svakom naslovu oblasti o uglovima i zato ne dokazuje
    ništa; „konveksni“ dokazuje. Poklapanje samo na zajedničkoj riječi je slabiji
    dokaz, pa lekcija dobija `medium` — evidencija se ne briše (izvor zna i šire
    ishode koji su legitimni), ali je u reviziji odvojiva od jake."""
    own = _content_words(lesson["title"])
    family = [s for s in siblings if s.get("oblast") == lesson.get("oblast")]
    if len(family) <= 1:
        return own
    out = set()
    for word in own:
        shared = sum(1 for other in family
                     if any(_related(word, w) for w in _content_words(other["title"])))
        if shared / len(family) <= 0.34:
            out.add(word)
    return out


def _sibling_boundaries(lesson, siblings):
    """Granice opsega iz KANONSKOG spiska lekcija — dokaziv podatak, ne ishod.

    Susjed ulazi samo ako se lekcije UZAJAMNO razlikuju: svaka nosi bar jednu
    sadržajnu riječ koju druga nema. Time se nikad ne zabranjuje ono što je
    možda vlastita vještina lekcije (npr. uža i šira varijanta istog naslova).
    Zabrana govori šta ne smije biti CILJ zadatka; kao korak susjedna vještina
    ostaje dozvoljena, jer to prompt izričito razlikuje."""
    own = _content_words(lesson["title"])
    if not own:
        return []
    out = []
    for other in siblings:
        # Susjed je lekcija ISTE oblasti: samo tu je zamjena vještine stvarna.
        if other["id"] == lesson["id"] or other.get("oblast") != lesson.get("oblast"):
            continue
        theirs = _content_words(other["title"])
        if not theirs:
            continue
        own_only = [w for w in own if not any(_related(w, t) for t in theirs)]
        their_only = [w for w in theirs if not any(_related(w, o) for o in own)]
        if own_only and their_only:
            out.append(_clean(other["title"]))
    return out


def build():
    topics = json.loads(TOPICS.read_text(encoding="utf-8"))
    mapping = load_mapping()
    compiled = OrderedDict()
    stats = defaultdict(int)
    for grade, payload in sorted(topics["grades"].items()):
        for lesson in payload["lessons"]:
            lesson_id = lesson["id"]
            rows = mapping.get(lesson_id, [])
            lesson_words = _content_words(lesson["title"]) | _content_words(lesson["oblast"])
            title_words = _content_words(lesson["title"])
            scored, supporting, exclusions = [], [], []
            for row in rows:
                text = row["text"]
                if not text:
                    continue
                if row["relation"] in PRIMARY_RELATIONS:
                    score = _content_overlap(text, lesson_words)
                    if not score:
                        stats["dropped_unrelated_exact"] += 1
                        continue
                    # NASLOV NADJAČAVA OBLAST (mjereni nalaz): oblast „Skupovi
                    # tačaka, kružnica i krug“ poklopila se s ishodom o SKUPOVNIM
                    # operacijama, pa je lekcija o izlomljenoj liniji dobila tuđu
                    # vještinu. Oblast dijeli desetine lekcija i sama po sebi ne
                    # dokazuje ništa; red koji dodiruje naslov mora pobijediti red
                    # koji dodiruje samo oblast, bez obzira na broj poklapanja.
                    score = _content_overlap(text, title_words) * 100 + score
                    if not any(text == known for _, known, _ in scored):
                        scored.append((score, text, row["item_id"]))
                elif row["relation"] in SUPPORTING_RELATIONS:
                    if text not in supporting:
                        supporting.append(text)
                elif row["relation"] in NEIGHBOUR_RELATIONS:
                    if text not in exclusions:
                        exclusions.append(text)
            # Kad lekcija nosi i tačan i prenesen `exact` red, prolazi onaj koji
            # dijeli VIŠE sadržaja s lekcijom — rangiranje, ne prvi po redu.
            scored.sort(key=lambda item: -item[0])
            primary = [text for _, text, _ in scored[:MAX_PRIMARY]]
            evidence = [item for _, _, item in scored[:MAX_PRIMARY]]
            # ŠIROK ISHOD OBLASTI NIJE PRIMARNA VJEŠTINA LEKCIJE (živi nalaz, val 1).
            # Lekcija „Upoređivanje decimalnih brojeva“ dobila je kao primarnu
            # vještinu „izvoditi osnovne računske operacije s decimalnim
            # brojevima“ — istinit ishod TE OBLASTI, ali ne ono što lekcija
            # ispituje. Posljedica je bila dvostruka: mjerenje je proglašavalo
            # savršeno pogođene zadatke promašajem, a prompt je istu rečenicu
            # nudio kao OBAVEZAN cilj i time gurao generisanje s lekcije.
            # Dokaz koji ne dodiruje nijednu razlikovnu riječ naslova ostaje
            # dokaz VEZE, a ne dokaz VJEŠTINE — zato ide u pomoćne pojmove.
            marks = _distinguishing_words(lesson, payload["lessons"])
            if primary and not any(_content_overlap(text, marks) for text in primary):
                supporting = primary + supporting
                primary, evidence = [], []
                stats["primary_demoted_to_supporting"] += 1
            if primary:
                source, confidence = "npp_exact_mapping", "high"
            elif exclusions:
                source, confidence = "npp_neighbour_only", "medium"
            else:
                source, confidence = "canonical_lesson_scope", "low"
            if not exclusions and not primary:
                # NEMA dokazanog ishoda ne znači NEMA mjerila: kanonski spisak
                # lekcija i dalje dokazuje po čemu se ova lekcija razlikuje od
                # svojih susjeda u istoj oblasti.
                #
                # SAMO kad primarna vještina NE postoji. Lekcija s dokazanim
                # ishodom već nosi jače ograničenje („cilj mora biti baš ova
                # vještina“); dodavanje spiska zabrana preko toga mijenja prompt
                # za lekcije koje mjereno rade dobro, bez ijednog dokaza da im
                # treba. Svakoj lekciji NAJJAČE dostupno mjerilo, ne sva.
                exclusions = _sibling_boundaries(lesson, payload["lessons"])[:MAX_EXCLUSIONS]
                if exclusions and not primary:
                    stats["scope_from_siblings"] += 1
            if not primary and not exclusions:
                stats["no_usable_signal"] += 1
                continue
            compiled[lesson_id] = OrderedDict([
                ("grade", int(grade)),
                ("primary_skills", primary[:MAX_PRIMARY]),
                ("supporting_concepts", supporting[:MAX_SUPPORTING]),
                ("neighbour_exclusions", exclusions[:MAX_EXCLUSIONS]),
                ("evidence_ids", [e for e in evidence[:MAX_PRIMARY] if e]),
                ("objective_source", source),
                ("objective_confidence", confidence),
            ])
            stats["compiled"] += 1
            if primary:
                stats["with_primary"] += 1

            if exclusions:
                stats["with_exclusions"] += 1
    return compiled, stats


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true")
    args = parser.parse_args(argv)
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    compiled, stats = build()
    print(f"lekcija s kompajliranim ishodima: {stats['compiled']}")
    print(f"  s primarnom vještinom : {stats['with_primary']}")
    print(f"  širok ishod -> pomoćni : {stats['primary_demoted_to_supporting']}")
    print(f"  sa susjednim zabranama: {stats['with_exclusions']}")
    print(f"  odbačeni nepovezani `exact` redovi: {stats['dropped_unrelated_exact']}")
    print(f"  opseg iz susjednih lekcija: {stats['scope_from_siblings']}")
    print(f"  bez ijednog mjerila   : {stats['no_usable_signal']}")
    if args.write:
        document = OrderedDict([
            ("_readme", [
                "GENERISANO — ne uređivati ručno.",
                "Izvor: reference/curriculum/semantics/MATBOT_Faza2_Mapiranje.xlsx",
                "Generator: scripts/build_lesson_objectives.py",
                "primary_skills = ishodi TE lekcije (relation=exact, filtrirano)",
                "supporting_concepts = smiju se pojaviti, ne smiju biti CILJ",
                "neighbour_exclusions = ishodi SUSJEDNIH lekcija; nikad cilj zadatka",
            ]),
            ("schema_version", 1),
            ("lessons", compiled),
        ])
        OUTPUT.write_text(json.dumps(document, ensure_ascii=False, indent=1) + "\n",
                          encoding="utf-8")
        print(f"OK: {OUTPUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
