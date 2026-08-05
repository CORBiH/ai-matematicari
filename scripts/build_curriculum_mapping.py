"""Faza 2 kurikularne semantike: mapiranje KS/RS NPP stavki na 534 MAT-BOT lekcije.

    python scripts/build_curriculum_mapping.py            # izgradi Fazu 2 workbook
    python scripts/build_curriculum_mapping.py --report   # izgradi + sažetak na stdout
    python scripts/build_curriculum_mapping.py --dry-run  # sve provjere, bez pisanja

ULAZI (samo čitanje, nikad se ne mijenjaju):
  reference/curriculum/MATBOT_Sve_Lekcije_6_7_8_9.xlsx          (kanonskih 534)
  reference/curriculum/semantics/MATBOT_Faza1_KS_RS_NPP_Matematika.xlsx
  data/topics.json                                               (samo unakrsna provjera)

IZLAZ:
  reference/curriculum/semantics/MATBOT_Faza2_Mapiranje.xlsx

PRINCIPI (isti kao serverski validatori — vidi CLAUDE.md):
  • deterministički dokazi prije semantičke interpretacije; što se ne može
    dokazati ide u `needs_review`/`no_match`, NIKAD u tiho pogađanje;
  • generičke riječi (zadatak, primjena, pojam, računanje, rješavanje,
    svojstva…) same po sebi NIKAD nisu osnov mapiranja;
  • nijedna stavka Faze 1 se ne odbacuje: svaka dobija bar jedan red u listu
    Mapiranje (mapiran ili `no_match`);
  • sve odluke su podaci i višekratna pravila — bez grane po ID-ju lekcije;
  • nula mrežnih i nula model poziva; dvostruko pokretanje daje bajt-identičan
    fajl (fiksni metapodaci + normalizovani ZIP timestampovi).
"""
from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import io
import json
import re
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parent.parent
CANONICAL_XLSX = ROOT / "reference" / "curriculum" / "MATBOT_Sve_Lekcije_6_7_8_9.xlsx"
PHASE1_XLSX = ROOT / "reference" / "curriculum" / "semantics" / "MATBOT_Faza1_KS_RS_NPP_Matematika.xlsx"
TOPICS_JSON = ROOT / "data" / "topics.json"
OUTPUT_XLSX = ROOT / "reference" / "curriculum" / "semantics" / "MATBOT_Faza2_Mapiranje.xlsx"

EXPECTED_LESSON_COUNT = 534
EXPECTED_ITEM_COUNT = 573
# Kontrolni brojevi iz lista Kontrola Faze 1 — build pada ako se ne poklope.
EXPECTED_ITEMS_BY_SOURCE_GRADE = {
    ("KS_2018", 6): 77, ("KS_2018", 7): 99, ("KS_2018", 8): 137, ("KS_2018", 9): 110,
    ("RS_2014", 6): 31, ("RS_2014", 7): 26, ("RS_2014", 8): 48, ("RS_2014", 9): 45,
}

RELATIONS = ("exact", "supporting", "prerequisite", "neighbour", "advanced", "excluded")
CONFIDENCES = ("high", "medium", "low")
REVIEW_STATUSES = ("auto_high_confidence", "needs_review", "confirmed", "rejected",
                   "conflict", "no_match")
MAPPING_METHODS = ("exact_title", "normalized_title", "same_grade_area",
                   "curriculum_sequence", "terminology_alias", "semantic_overlap",
                   "manual_rule", "unresolved")

# Fiksni metapodaci dokumenta — uslov bajt-reproducibilnosti (bez "sada").
_FIXED_DOC_DATE = _dt.datetime(2026, 1, 1, 0, 0, 0)

# ---------------------------------------------------------------------------
# NORMALIZACIJA PISMA I OBLIKA
# ---------------------------------------------------------------------------

# Srpska ćirilica → latinica (digrafi prije pojedinačnih slova).
_CYR_DIGRAPHS = {"Љ": "Lj", "Њ": "Nj", "Џ": "Dž", "љ": "lj", "њ": "nj", "џ": "dž"}
_CYR_SINGLE = {
    "А": "A", "Б": "B", "В": "V", "Г": "G", "Д": "D", "Ђ": "Đ", "Е": "E", "Ж": "Ž",
    "З": "Z", "И": "I", "Ј": "J", "К": "K", "Л": "L", "М": "M", "Н": "N", "О": "O",
    "П": "P", "Р": "R", "С": "S", "Т": "T", "Ћ": "Ć", "У": "U", "Ф": "F", "Х": "H",
    "Ц": "C", "Ч": "Č", "Ш": "Š",
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "ђ": "đ", "е": "e", "ж": "ž",
    "з": "z", "и": "i", "ј": "j", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "ћ": "ć", "у": "u", "ф": "f", "х": "h",
    "ц": "c", "ч": "č", "ш": "š",
}


def transliterate(text: str) -> str:
    """Ćirilica → latinica; latinica prolazi nepromijenjena."""
    if not text:
        return ""
    out = []
    for ch in text:
        if ch in _CYR_DIGRAPHS:
            out.append(_CYR_DIGRAPHS[ch])
        else:
            out.append(_CYR_SINGLE.get(ch, ch))
    return "".join(out)


_WORD_RE = re.compile(r"[a-zšđčćž]+", re.IGNORECASE)

# Sufiksi se skidaju NAJDUŽI PRVI; osnova mora ostati >= 3 znaka. Namjerno grubo:
# služi isključivo poklapanju, nikad se ne prikazuje. Parovi koje mora spojiti
# zaključani su testom (sabiranje/sabirati, upoređivanje/uporediti, …).
_SUFFIXES = (
    "avanjima", "ivanjima", "ovanjima",
    "avanje", "ivanje", "ovanje", "avanja", "ivanja", "ovanja",
    "anjima", "enjima", "anjem", "enjem", "anja", "enja", "anje", "enje",
    "ijama", "ijema", "ijom", "ijama", "ijalna", "ijalni",
    "ostima", "osti", "ošću",
    "avati", "ivati", "ovati", "isati", "avaju", "ivaju", "ovaju",
    "ati", "iti", "eti", "uti", "aju", "uju", "aš", "eš", "iš",
    "ama", "ima", "ovima", "evima", "ovi", "evi", "ije", "ija", "iju", "ijo",
    "oga", "ega", "omu", "emu", "og", "eg", "im", "om", "ih", "oj",
    "a", "e", "i", "o", "u",
)


def stem(token: str) -> str:
    t = token.lower()
    for suffix in _SUFFIXES:
        if t.endswith(suffix) and len(t) - len(suffix) >= 3:
            t = t[: -len(suffix)]
            break
    return t


# Preslikavanje osnova na kanonsku osnovu (samo za poklapanje; original se čuva).
# Pokriva: eksplicitne KS aliase, ćirilične/ekavske varijante, nepravilne glagolske
# osnove koje grubi stemer ne spaja i RS→MAT-BOT terminologiju (kriterijum→pravilo).
ALIAS_STEMS = {
    "kut": "ugl", "kutov": "ugl", "ugao": "ugl", "uga": "ugl", "uglov": "ugl",
    "mnogokut": "mnogougl", "mnogougao": "mnogougl", "mnogougl": "mnogougl",
    "dirk": "tangent", "sekant": "sječic", "sjecic": "sječic",
    "brojnik": "brojioc", "brojil": "brojioc", "brojioc": "brojioc",
    "nazivnik": "imenioc", "imenil": "imenioc", "imenioc": "imenioc", "imenitelj": "imenioc",
    "postotak": "procent", "postotk": "procent", "postotn": "procent",
    "procen": "procent", "procenat": "procent", "procentn": "procent",
    "jednadžb": "jednačin", "jednacin": "jednačin", "jednačin": "jednačin",
    "nejednadžb": "nejednačin", "nejednacin": "nejednačin", "nejednačin": "nejednačin",
    "četvorougl": "četverougl", "četvorougao": "četverougl", "četverougao": "četverougl",
    "četverougl": "četverougl", "cetverougl": "četverougl", "cetvorougl": "četverougl",
    "kriterijum": "pravil", "kriterij": "pravil",
    "polumjer": "poluprečnik", "poluprecnik": "poluprečnik", "polupre čnik": "poluprečnik",
    "promjer": "prečnik", "precnik": "prečnik",
    "kugl": "lopt", "stošc": "kup", "stožc": "kup", "stozc": "kup", "cilindr": "valjk",
    "cilind": "valjk", "valjak": "valjk", "valjc": "valjk",
    "trokut": "trougl", "trougao": "trougl", "trougl": "trougl",
    "uporeb": "upoređ", "upored": "upoređ", "uspored": "upoređ", "upoređiv": "upoređ",
    "poređenj": "upoređ", "poretk": "upoređ", "poredak": "upoređ", "uređenj": "upoređ",
    "dijel": "dijelj", "dijeljenj": "dijelj", "podijel": "dijelj", "deljenj": "dijelj",
    "djeljenj": "dijelj",
    "množenj": "množ", "pomnož": "množ", "mnoz": "množ", "množ": "množ",
    "sabir": "sabir", "sabranj": "sabir", "zbrajanj": "sabir",
    "oduzm": "oduzim", "oduze": "oduzim", "oduzim": "oduzim",
    "prošir": "proširiv", "proširiv": "proširiv", "prosir": "proširiv", "prosiriv": "proširiv",
    "skrat": "skraćiv", "skrać": "skraćiv", "skraćen": "skraćiv", "skraćiv": "skraćiv",
    "vjerovatnoć": "vjerovatnoc", "vjerovatnoc": "vjerovatnoc", "vjerovatn": "vjerovatnoc",
    "kvadratn": "kvadrat", "korjen": "korijen", "koren": "korijen", "korij": "korijen",
    "korijen": "korijen",
    "stepen": "stepen", "stepenov": "stepen", "eksponent": "izložioc", "izložil": "izložioc",
    "izložioc": "izložioc", "izložic": "izložioc",
    "razmjer": "razmjer", "omjer": "razmjer",
    "djelitelj": "djelioc", "djelil": "djelioc", "djelioc": "djelioc", "delioc": "djelioc",
    "sadržil": "sadržioc", "sadržioc": "sadržioc", "sadrzil": "sadržioc",
    "sadrzioc": "sadržioc", "višekratnik": "sadržioc", "visekratnik": "sadržioc",
    "djeljivost": "djeljiv", "deljiv": "djeljiv", "djeljiv": "djeljiv",
    "sfer": "sfer", "grafik": "grafik", "graf": "grafik", "grafičk": "grafik",
    "tabelarn": "tabel", "tablic": "tabel", "tabel": "tabel",
    "dijagram": "dijagram", "histogram": "dijagram",
    "koordinat": "koordinat", "koordinatn": "koordinat",
    "apsolutn": "apsolutn", "modul": "apsolutn",
    "transformacij": "preslikav", "preslikavanj": "preslikav", "preslikav": "preslikav",
    "izometrijsk": "izometrij", "izometrij": "izometrij",
    "linearne": "linearn", "linearn": "linearn", "linearan": "linearn",
    "decimaln": "decimal", "decimal": "decimal",
    "razlomc": "razlomk", "razlomak": "razlomk", "razlomk": "razlomk", "razlomljен": "razlomljen",
    "prizm": "prizm", "piramid": "piramid",
    "zapremin": "zapremin", "volumen": "zapremin",
    "površin": "površin", "povrsin": "površin", "povšin": "površin",
    "simetral": "simetral", "simetrij": "simetrij", "simetričn": "simetrij",
    "osnosimetričn": "simetrij", "centralnosimetričn": "simetrij",
    "translacij": "translacij", "translir": "translacij",
    "rotacij": "rotacij", "rotir": "rotacij",
    "podudarn": "podudarn", "sličn": "sličn", "slicn": "sličn",
    "talesov": "tales", "tales": "tales", "pitagorin": "pitagor", "pitagor": "pitagor",
    "proporcij": "proporcij", "proporcionaln": "proporcionaln",
    "frekvencij": "frekvencij", "vennov": "venov", "venov": "venov",
}


# Generičke riječi — nikad same ne opravdavaju mapiranje (težina 0).
GENERIC_STEMS = {
    # eksplicitno traženo specifikacijom
    "zadatk", "zadac", "zadat", "zadatak", "primjen", "primijen", "primjenjiv",
    "pojm", "pojam",
    "računanj", "računsk", "račun", "izračunav", "izračun", "rješavanj", "rješenj",
    "riješ", "rješiv", "svojstv", "osobin",
    # kurikularni/pedagoški vokabular
    "učenik", "učenic", "učenj", "nastavnik", "nastavn", "gradiv", "znanj", "znat",
    "razumjet", "razumij", "shvatit", "shvat", "usvojit", "usvoj", "upoznat", "upozna",
    "poznavat", "poznaj", "moć", "umjet", "umij", "vještin", "sposobnost",
    "definisat", "definiš", "definicij", "objasnit", "objašnjav", "objašnjenj",
    "razlikovat", "razlikuj", "prepoznat", "prepoznav", "uočavat", "uočit", "uoč",
    "navest", "navod", "nabrojat", "nabroj", "koristit", "korišćenj", "korištenj",
    "upotrebljav", "upotrijeb", "izvodit", "izvođenj", "demonstrir", "formirat",
    "primjer", "primjerim", "život", "svakodnevn", "povezivanj", "povezat", "povež",
    "korelacij", "predmet", "sadržaj", "program", "tem", "čas", "ishod", "cilj",
    "obrad", "vježb", "sistematizacij", "pismen", "usmen", "mijenj", "mijen",
    "jednostavn", "jednostavnij", "prostij", "prostije", "složenij", "elementarn",
    "osnovn", "važn", "potrebn", "odgovarajuć", "različit", "raznovrsn", "dat", "zadan",
    "oblik", "obliku", "vrst", "vrijednost", "broj", "brojev", "brojn", "veličin",
    "način", "postupk", "postupak", "praviln", "uspješn", "moguć", "takođ",
    "odredit", "određiv", "određen", "utvrdit", "utvrđiv", "obrazlagat", "obrazlož",
    "iskaz", "tvrdnj", "teorem", "stav", "pravil",  # "pravil" generičko OSIM uz djeljivost (vidi dolje)
    "materijal", "sredstv", "pribor", "model", "slik", "insistir", "preporuč",
    "akcenat", "pažnj", "informativn", "praktičn", "problem", "problemsk",
    "matematik", "matematičk", "geometrij", "geometrijsk",
    # sveprisutni kvalifikatori bez diskriminacione moći
    "prirodn", "obrnut", "obratn", "pozitivn", "negativn", "nepoznat", "jednak",
    # puni glagolski/pedagoški oblici — kanonizuju se kroz stem+fold pri učitavanju
    "definisati", "objasniti", "objašnjavati", "razumjeti", "razumijevati",
    "shvatiti", "shvatati", "usvojiti", "usvajati", "upoznati", "upoznavati",
    "poznavati", "razlikovati", "prepoznati", "prepoznavati", "uočiti", "uočavati",
    "navesti", "navoditi", "nabrojati", "koristiti", "upotrebljavati", "izvoditi",
    "demonstrirati", "formirati", "povezivati", "zaključiti", "zaključivati",
    "moći", "umjeti", "znati", "etapa", "kontekst", "tačnost", "dobijen",
    "obrazlagati", "obrazložiti", "utvrditi", "utvrđivati", "odrediti",
    "određivati", "primijeniti", "primjenjivati", "izračunati", "izračunavati",
    "riješiti", "rješavati", "računati", "ispitivati", "operacija", "operacij",
    "interpretirati", "diskutovati", "rezultat", "procjenjivati",
}

# Funkcijske riječi (2-3 slova i veznici) — nikad tokeni sadržaja.
STOP_WORDS = {
    "i", "u", "s", "o", "a", "je", "su", "li", "ne", "da", "se", "sa", "na", "za",
    "po", "od", "do", "iz", "uz", "ka", "pri", "kroz", "te", "ili", "pa", "ni",
    "niti", "što", "sto", "kao", "kad", "ako", "gdje", "koji", "koja", "koje",
    "kojim", "kojih", "kojoj", "kojem", "ovaj", "ova", "ovo", "taj", "ta", "to",
    "npr", "itd", "dr", "tzv", "biti", "bit", "će", "ce", "bi", "im", "ih", "ga",
    "mu", "još", "jos", "već", "vec", "sve", "svi", "sva", "svaki", "svaka",
    "dva", "tri", "pet", "šest", "sest", "no",
    "zna", "znaju", "treba", "ima", "imaju", "bude", "može", "moze", "mogu",
}

# Iznimke od generičnosti: bigram konteksti u kojima generička osnova ipak nosi
# značenje (reusable pravilo, ne grana po lekciji).
_PRAVILA_DJELJIVOSTI = ("pravil", "djeljiv")

# Akcione osnove — operacija/radnja koju lekcija imenuje; težina 1 (objekti 2).
ACTION_STEMS = {
    "sabir", "oduzim", "množ", "dijelj", "upoređ", "proširiv", "skraćiv", "svođenj",
    "svod", "zaokruživ", "zaokruž", "pretvaranj", "pretvar", "preračunav", "prikaz",
    "prikazivanj", "čitanj", "čita", "zapisivanj", "zapis", "crtanj", "crta", "mjerenj",
    "mjer", "konstru", "konstrukcij", "konstruiš", "konstruis", "rastavljanj", "rastav",
    "izlučivanj", "izdvajanj", "grupisanj", "faktorizacij", "provjer", "uvrst",
    "modelir", "procjen", "procijen", "kvadrir",
    "stepenov", "korjenov", "poređaj", "ured",
}


def _fold(token_stem: str) -> str:
    return ALIAS_STEMS.get(token_stem, token_stem)


# Dodatni aliasi oblika koje grubi stemer razdvaja. Ključevi su POST-stem
# oblici (fold se primjenjuje NAKON stemera) — zato i nominativ na -lac i
# krnje osnove na -uga moraju biti navedeni eksplicitno.
ALIAS_STEMS.update({
    "skupovn": "skupov", "skupin": "skupov",
    # metoda zamjene ≡ metoda supstitucije (oba imena u upotrebi)
    "zamjen": "supstituc", "supstitucij": "supstituc",
    # nominativi na -lac (stemer ih ne dira, kosi padeži daju -oc osnovu)
    "djelilac": "djelioc", "sadržilac": "sadržioc", "sadrzilac": "sadržioc",
    "imenilac": "imenioc", "brojilac": "brojioc",
    # krnje osnove od -ugao (strip završnog 'o')
    "četvorouga": "četverougl", "četverouga": "četverougl", "cetverouga": "četverougl",
    "mnogouga": "mnogougl", "trouga": "trougl",
})

# Tabele se kanonizuju kroz ISTI stem+fold koji koristi tokenizacija — inače bi
# unos „definisati“ štitio samo tačno taj oblik, a stemer proizvodi „defin“.
GENERIC_STEMS = frozenset(
    {_fold(stem(entry)) for entry in GENERIC_STEMS} | set(GENERIC_STEMS)
)
ACTION_STEMS = frozenset(
    {_fold(stem(entry)) for entry in ACTION_STEMS} | set(ACTION_STEMS)
) - GENERIC_STEMS


def normalize_text(text: str) -> str:
    """Translit + NFC + mala slova + interpunkcija u razmake."""
    latin = transliterate(unicodedata.normalize("NFC", text or ""))
    return latin.lower()


@dataclass(frozen=True)
class Token:
    surface: str
    stem: str
    kind: str  # "object" | "action" | "generic"


def tokenize(text: str, alias_folds: set | None = None) -> list[Token]:
    tokens = []
    for match in _WORD_RE.finditer(normalize_text(text)):
        word = match.group(0)
        if len(word) < 2 or word in STOP_WORDS:
            continue
        raw_stem = stem(word)
        folded = _fold(raw_stem)
        if alias_folds is not None and folded != raw_stem:
            alias_folds.add(folded)
        if folded in GENERIC_STEMS:
            kind = "generic"
        elif folded in ACTION_STEMS:
            kind = "action"
        elif len(folded) < 3 and folded != "os":
            # Dvoznakovni ostaci (najčešće slova iz formula: 'bq', 'kx') nisu
            # pouzdani objekti; 'os' (brojevna osa) je jedini stvarni pojam.
            kind = "generic"
        else:
            kind = "object"
        tokens.append(Token(word, folded, kind))
    return tokens


def _content_stems(tokens: list[Token]) -> tuple[set, set]:
    objects = {t.stem for t in tokens if t.kind == "object"}
    actions = {t.stem for t in tokens if t.kind == "action"}
    # "pravila djeljivosti" — 'pravil' postaje objekat SAMO uz 'djeljiv'.
    stems_all = {t.stem for t in tokens}
    if _PRAVILA_DJELJIVOSTI[0] in {stem for t in tokens for stem in (t.stem,)} and \
            _PRAVILA_DJELJIVOSTI[1] in stems_all:
        objects = objects | {"pravil"}
    return objects, actions


# ---------------------------------------------------------------------------
# KANONSKE LEKCIJE
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class VariantProfile:
    """Jedna naslovna varijanta: stem-sekvenca (za fraze) + profil (za bodovanje)."""
    seq: tuple
    objects: frozenset
    actions: frozenset
    weight: int


@dataclass(frozen=True)
class Lesson:
    lesson_id: str
    grade: int
    oblast: str
    title: str
    profiles: tuple  # tuple[VariantProfile, ...]
    objects: frozenset  # unija svih varijanti (za brze presjeke)
    title_norm: str


def _title_variants(title: str) -> list[str]:
    """Naslov → varijante: cijeli, bez zagrada, alternativa razdvojenih kosom crtom,
    segmenti razdvojeni ';' / ' i ' se NE cijepaju (previše lažnih)."""
    base = re.sub(r"\([^)]*\)", " ", title)
    variants = {title, base}
    # a/b alternative unutar riječi ili fraza: uzmi obje strane.
    if "/" in base:
        variants.add(re.sub(r"(\S+)/(\S+)", r"\1", base))
        variants.add(re.sub(r"(\S+)/(\S+)", r"\2", base))
    # "X - svojstva" → i sam X.
    if " - " in base:
        variants.add(base.split(" - ")[0])
    return [v for v in variants if v.strip()]


def load_lessons() -> list[Lesson]:
    wb = openpyxl.load_workbook(CANONICAL_XLSX, data_only=True, read_only=True)
    ws = wb["Sve lekcije"]
    rows = list(ws.iter_rows(values_only=True))[1:]
    wb.close()
    lessons = []
    for razred, oblast, title, lesson_id in rows:
        if lesson_id is None:
            continue
        profiles = []
        for variant in _title_variants(str(title)):
            tokens = tokenize(variant)
            objects, actions = _content_stems(tokens)
            seq = tuple(t.stem for t in tokens if t.kind != "generic")
            weight = 2 * len(objects) + len(actions)
            if weight == 0:
                continue
            profiles.append(VariantProfile(
                seq=seq, objects=frozenset(objects), actions=frozenset(actions),
                weight=weight,
            ))
        # Deterministički redoslijed profila; bez profila = naslov bez sadržaja
        # (ne postoji u ovom kurikulumu, ali se brani).
        profiles.sort(key=lambda p: (-p.weight, p.seq))
        all_objects = frozenset().union(*(p.objects for p in profiles)) if profiles else frozenset()
        lessons.append(Lesson(
            lesson_id=str(lesson_id).strip(), grade=int(razred),
            oblast=str(oblast).strip(), title=str(title).strip(),
            profiles=tuple(profiles), objects=all_objects,
            title_norm=normalize_text(str(title)),
        ))
    return lessons


def crosscheck_topics_json(lessons: list[Lesson]) -> None:
    data = json.loads(TOPICS_JSON.read_text(encoding="utf-8"))
    json_ids = {
        (lesson["id"], str(grade), lesson["title"])
        for grade, grade_data in data["grades"].items()
        for lesson in grade_data["lessons"]
    }
    xlsx_ids = {(l.lesson_id, str(l.grade), l.title) for l in lessons}
    if json_ids != xlsx_ids:
        raise SystemExit("data/topics.json i kanonski workbook se ne poklapaju — stop.")


# ---------------------------------------------------------------------------
# STAVKE FAZE 1
# ---------------------------------------------------------------------------

# Sekcijska politika: content → puno mapiranje; methodology → najviše 'supporting';
# non_content → bez mape (no_match po dizajnu, uz obrazloženje).
SECTION_POLICY = {
    "sadrzaji_programa": "content",
    "tabelarni_pregled": "content",
    "posebni_ciljevi": "content",
    "metodicka_uputstva": "methodology",
    "didakticke_napomene": "methodology",
    "odgojni_zadaci": "non_content",
    "profil_nastavnika": "non_content",
    "opsti_ciljevi": "non_content",
}

_NON_CONTENT_REASON = {
    "opsti_ciljevi": "opšti vaspitno-obrazovni cilj, nije sadržaj lekcije",
    "odgojni_zadaci": "odgojni zadatak, nije matematički sadržaj lekcije",
    "profil_nastavnika": "uslovi/profil nastavnika, nije sadržaj lekcije",
}

# Kraj stavke često nosi naslov SLJEDEĆE teme ("… sadržioca. 5. Razlomci") ili
# rep "SADRŽAJI PROGRAMA …" — to se skida IZ TEKSTA ZA POKLAPANJE (original se
# uvijek čuva netaknut) i hrani sekvencijalni pratilac teme.
_TRAILING_THEME_RE = re.compile(
    r"[.;]\s*\(?(\d{1,2})[.)]\s+([A-ZŠĐČĆŽ][^.;]{2,80})\s*$"
)
_SADRZAJI_TAIL_RE = re.compile(r"SADRŽAJI\s+PROGRAMA.*$", re.IGNORECASE | re.DOTALL)
_CORRELATION_TAIL_RE = re.compile(
    r"(Povezivanje gradiva|Unutrašnja i međupredmetna).*$", re.IGNORECASE | re.DOTALL
)

_BROAD_ITEM_CHARS = 400
_HEAD_CLAUSE_CHARS = 220


@dataclass
class Item:
    item_id: str
    source_id: str
    grade: int
    page: int
    section: str
    theme: str
    item_type: str
    original_text: str
    latin_text: str
    parse_confidence: str
    source_review: str
    editor_note: str
    match_text: str = ""       # očišćen tekst za poklapanje
    head_text: str = ""        # vodeća klauzula (za bodovanje širokih stavki)
    inherited_theme: str = ""  # tema naslijeđena sekvencom kurikuluma
    is_broad: bool = False


def load_items() -> list[Item]:
    wb = openpyxl.load_workbook(PHASE1_XLSX, data_only=True, read_only=True)
    ws = wb["Stavke_NPP"]
    rows = list(ws.iter_rows(values_only=True))[1:]
    wb.close()
    items = []
    for r in rows:
        (item_id, source_id, grade, page, section, theme, item_type,
         original, latin, confidence, review, note) = r
        items.append(Item(
            item_id=str(item_id), source_id=str(source_id), grade=int(grade),
            page=int(page), section=str(section or ""), theme=str(theme or "").strip(),
            item_type=str(item_type or ""), original_text=str(original or ""),
            latin_text=str(latin or "") or transliterate(str(original or "")),
            parse_confidence=str(confidence or ""), source_review=str(review or ""),
            editor_note=str(note or ""),
        ))
    return items


def prepare_items(items: list[Item]) -> None:
    """Očisti tekst za poklapanje + sekvencijalno naslijedi temu (RS marker
    '… 5. Razlomci' na KRAJU stavke važi za SLJEDEĆE stavke, ne za tekuću).

    KS kolona 'Tema' je NEPOUZDANA (Faza 1 je čitav tabelarni niz razreda
    pripisala prvoj temi — npr. sve KS r8 stavke nose 'Realni brojevi'), pa se
    za KS NIKAD ne koristi kao kontekst; KS stavke same imenuju svoje objekte.
    RS 'Tema' uz posebne ciljeve jeste stvarna tema i koristi se."""
    current_theme: dict[tuple, str] = {}
    for item in sorted(items, key=lambda x: x.item_id):
        key = (item.source_id, item.grade)
        text = item.latin_text.strip()
        text = _SADRZAJI_TAIL_RE.sub(" ", text)
        next_theme = None
        match = _TRAILING_THEME_RE.search(text)
        if match:
            next_theme = match.group(2).strip()
            text = text[: match.start()] + "."
        if item.source_id == "RS_2014":
            item.inherited_theme = item.theme or current_theme.get(key, "")
        else:
            item.inherited_theme = ""
        if next_theme:
            current_theme[key] = next_theme
        item.is_broad = len(text) > _BROAD_ITEM_CHARS
        item.match_text = text
        head = _CORRELATION_TAIL_RE.sub(" ", text)
        item.head_text = head[:_HEAD_CLAUSE_CHARS]


# ---------------------------------------------------------------------------
# TABELA SUSJEDNIH (LAKO ZAMJENJIVIH) VJEŠTINA — višekratna pravila po NASLOVU,
# nikad po ID-ju lekcije. side_a/side_b su regexi nad normalizovanim naslovom.
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ConfusionPair:
    """Par lako zamjenjivih vještina. Svaka strana je UREĐENA torka obrazaca.

    KONVENCIJA REDOSLIJEDA (nosi je popravka defekta 1, Faza 2.5): prvi obrazac
    strane imenuje KANONSKU lekciju te strane, a kasniji su šire rezerve. Kad
    više lekcija pogodi istu stranu, bira se ona koja pogađa NAJRANIJI obrazac —
    ranije se biralo prvo po ID-ju lekcije, pa je za „dijeljenje s ostatkom“
    susjed ispao 6-03-002 (prva lekcija čiji naslov sadrži „djeljivost“) umjesto
    lekcije o PRAVILIMA djeljivosti. Redoslijed je podatak, ne grana po lekciji.
    """

    name: str
    side_a: tuple
    side_b: tuple
    note: str


def _side_rank(side, title_norm):
    """Indeks prvog obrasca strane koji pogađa naslov, ili None."""
    for index, pattern in enumerate(side):
        if re.search(pattern, title_norm):
            return index
    return None


def _canonical_neighbour(side, candidates):
    """Kandidat koji pogađa najraniji (najkanonskiji) obrazac te strane.

    Neriješeno se razrješava ID-jem lekcije — izbor ostaje determinističan."""
    best = None
    for lesson in candidates:
        rank = _side_rank(side, lesson.title_norm)
        if rank is None:
            continue
        key = (rank, lesson.lesson_id)
        if best is None or key < best[0]:
            best = (key, lesson)
    return best[1] if best is not None else None


CONFUSION_PAIRS = (
    ConfusionPair("pravila_djeljivosti_vs_djelioci", (r"pravila\s+djeljivosti",),
                  (r"djelilac/faktor", r"djelilac.*sadržilac", r"faktor i sadržilac"),
                  "primjena pravila djeljivosti nije traženje djelilaca/faktora"),
    ConfusionPair("djeljivost_vs_dijeljenje_s_ostatkom",
                  (r"pravila\s+djeljivosti", r"djeljivost"),
                  (r"dijeljenje s ostatkom",),
                  "djeljivost nije računanje količnika i ostatka"),
    ConfusionPair("prosti_vs_faktorizacija", (r"prosti i složeni",),
                  (r"rastavljanje.*proste faktore",),
                  "prepoznavanje prostih brojeva nije faktorizacija"),
    ConfusionPair("faktorizacija_vs_nzd_nzs", (r"rastavljanje.*proste faktore",),
                  (r"najveći zajednički djelilac", r"najmanji zajednički sadržilac",
                   r"nzd", r"nzs"),
                  "faktorizacija je alat, NZD/NZS je zasebna vještina"),
    ConfusionPair("pojam_razlomka_vs_kolicnik", (r"pojam razlomka",),
                  (r"razlomak kao dio cjeline i kao količnik",),
                  "pojam razlomka nije interpretacija količnikom"),
    ConfusionPair("vrste_razlomaka_vs_mjesoviti", (r"pravi, nepravi",),
                  (r"mješoviti broj",),
                  "klasifikacija razlomaka nije zapis mješovitim brojem"),
    ConfusionPair("prosirivanje_vs_skracivanje", (r"proširivanje razlomaka",),
                  (r"skraćivanje",),
                  "proširivanje i skraćivanje su suprotni smjerovi"),
    ConfusionPair("ekvivalencija_vs_uporedjivanje",
                  (r"proširivanje razlomaka", r"skraćivanje"),
                  (r"upoređivanje razlomaka",),
                  "ekvivalentni zapis nije poređenje veličina"),
    ConfusionPair("jednaki_vs_razliciti_imenioci",
                  (r"sabiranje i oduzimanje razlomaka jednakih",),
                  (r"sabiranje i oduzimanje razlomaka različitih",),
                  "jednaki i različiti imenioci su različite lekcije"),
    ConfusionPair("razlomci_vs_procenti",
                  (r"množenje razlomka", r"dijeljenje razlomka"),
                  (r"postotni zapis", r"postotak/procenat"),
                  "račun s razlomcima nije pretvaranje u procente"),
    ConfusionPair("racun_vs_jednacine_razlomci",
                  (r"sabiranje i oduzimanje razlomaka", r"množenje razlomka",
                   r"dijeljenje razlomka"),
                  (r"jednačine s razlomcima",
                   r"jednačine s množenjem i dijeljenjem razlomaka"),
                  "direktan račun nije rješavanje jednačine"),
    ConfusionPair("jednacina_vs_vrijednost_izraza", (r"jednačin",),
                  (r"brojevni izrazi", r"brojna vrijednost izraza",
                   r"izrazi s promjenljivim i brojna vrijednost"),
                  "rješavanje jednačine nije izračunavanje vrijednosti izraza"),
    ConfusionPair("obim_vs_povrsina_kruga", (r"broj π i obim kruga", r"obim kruga"),
                  (r"površina kruga",), "obim i površina kruga su različite formule"),
    ConfusionPair("pitagora_vs_korijen", (r"pitagorina teorema",),
                  (r"kvadratni korijen",), "Pitagorina teorema nije opšte korjenovanje"),
    ConfusionPair("prikupljanje_vs_citanje_podataka",
                  (r"prezentovanje podataka", r"tabela frekvencija",
                   r"prikazivanje.*podataka"),
                  (r"čitanje podataka", r"čitanje i kritičko"),
                  "prikupljanje/prikaz podataka nije čitanje i tumačenje dijagrama"),
    ConfusionPair("vjerovatnoca_vs_frekvencija", (r"vjerovatnoć",),
                  (r"frekvencij",), "vjerovatnoća nije statistička frekvencija"),
    ConfusionPair("konstrukcija_vs_racun", (r"konstrukcij",),
                  (r"površina", r"obim", r"zapremin"),
                  "konstruktivni zadatak nije računski zadatak"),
)


# ---------------------------------------------------------------------------
# BODOVANJE I ODLUKA
# ---------------------------------------------------------------------------

@dataclass
class Mapping:
    mapping_id: str
    item: Item
    lesson: Lesson | None
    relation: str
    confidence: str
    method: str
    evidence: str
    ambiguity_note: str
    review_status: str


_PREREQ_RE = re.compile(r"\bobnov\w*|\bponov\w*|ponavljanj\w*|prethodn\w+ znanj|"
                        r"naučen\w* u ranij\w* razred", re.IGNORECASE)
_PROOF_RE = re.compile(r"\bdokaz\w*|\bdokaž\w*", re.IGNORECASE)
_EXCLUDE_RE = re.compile(r"ne\s+treba\s+dokaziv\w*|bez\s+dokaza|ne\s+obrađuj\w*|"
                         r"obrnutu\s+tvrdnju\s+ne\s+treba", re.IGNORECASE)
_INFORMATIVE_RE = re.compile(r"\binformativn\w*", re.IGNORECASE)


def _seq_hit(item_stems: tuple, variant_seq: tuple) -> bool:
    """Varijanta kao gotovo-uzastopna podsekvenca sadržajnih osnova stavke.

    Dužina 2 traži STROGU susjednost (inače „upoređivanje … decimalni“ iz dva
    susjedna, RAZLIČITA sadržaja lažno spaja lekciju o poređenju decimalnih);
    dužina ≥3 dozvoljava ukupno jedan umetnut stem (PDF tabele znaju umetnuti
    riječ između dijelova istog naslova)."""
    n = len(variant_seq)
    if n < 2 or n > len(item_stems):
        return False
    max_skips = 0 if n == 2 else 1
    for start in range(len(item_stems)):
        if item_stems[start] != variant_seq[0]:
            continue
        position, skips = start + 1, 0
        matched = 1
        while matched < n and position < len(item_stems):
            if item_stems[position] == variant_seq[matched]:
                matched += 1
                position += 1
            else:
                skips += 1
                if skips > max_skips:
                    break
                position += 1
        if matched == n:
            return True
    return False


@dataclass(frozen=True)
class Candidate:
    lesson: Lesson
    score: float
    matched_objects: tuple
    matched_actions: tuple
    phrase: tuple | None      # varijanta naslova pogođena kao fraza
    used_theme: bool
    alias_used: bool


def _score_candidates(item: Item, lessons: list[Lesson]) -> list[Candidate]:
    """Deterministički kandidati za JEDNU stavku — bez nasumičnosti, bez modela."""
    alias_folds: set = set()
    tokens_full = tokenize(item.match_text, alias_folds)
    stems_full = tuple(t.stem for t in tokens_full if t.kind != "generic")
    head_source = item.head_text if item.is_broad else item.match_text
    objects, actions = _content_stems(tokenize(head_source))
    theme_objects, theme_actions = _content_stems(tokenize(item.inherited_theme))
    theme_only_objects = theme_objects - objects

    # Težina sadržaja same stavke — za preciznost (koliko stavke naslov objašnjava).
    # Za ŠIROKE stavke (tabele sadržaja) preciznost je besmislena, pa se ne koristi.
    item_weight = 2 * len(objects) + len(actions)

    candidates = []
    for lesson in lessons:
        phrase = None
        for profile in lesson.profiles:
            if _seq_hit(stems_full, profile.seq):
                phrase = profile.seq
                break

        best = None
        for profile in lesson.profiles:
            own_objects = profile.objects & objects
            own_actions = profile.actions & actions
            theme_hit = profile.objects & theme_only_objects
            hit = 2 * len(own_objects) + len(own_actions)
            used_theme = False
            if theme_hit and (own_objects or own_actions):
                # Tema iz sekvence samo DOPUNJUJE vlastiti pogodak stavke.
                hit += 2 * len(theme_hit)
                used_theme = True
            if hit == 0:
                continue
            recall = hit / profile.weight
            if item.is_broad or item_weight == 0:
                precision = 1.0
                score = recall
            else:
                # Geometrijska sredina recall/preciznost: stavka čiji SPECIFIČNI
                # pojmovi ostaju neobjašnjeni (npr. „najveći zajednički …“) ne
                # smije pobijediti kraći naslov samo zato što ga u cijelosti
                # pokriva (živi nalaz: NZD/NZS stavka → „Djelilac i sadržilac“).
                precision = min(hit / item_weight, 1.0)
                score = (recall * precision) ** 0.5
            # Jedan zajednički objekat bez akcije je slab dokaz — snizi
            # efektivni skor da trivijalni jednorječni naslovi ne dominiraju.
            # IZUZETAK: kad naslov objašnjava SVE sadržajne tokene stavke
            # (precision 1.0), taj jedan objekat je cijela stavka, ne šum.
            if (len(own_objects) + len(theme_hit) < 2 and not own_actions
                    and precision < 0.999):
                score *= 0.75
            # Naslov čija DEFINIŠUĆA radnja (npr. „Provjera …“) u stavci uopšte
            # ne postoji ne smije pobijediti čistim objektnim preklapanjem.
            if profile.actions and not own_actions:
                score *= 0.8
            entry = (score, own_objects | theme_hit, own_actions, used_theme)
            if best is None or entry[0] > best[0]:
                best = entry
        if phrase is None and best is None:
            continue
        if best is None:
            best = (0.0, frozenset(), frozenset(), False)
        candidates.append(Candidate(
            lesson=lesson, score=best[0],
            matched_objects=tuple(sorted(best[1])),
            matched_actions=tuple(sorted(best[2])),
            phrase=phrase, used_theme=best[3],
            alias_used=bool(alias_folds & (set(best[1]) | set(best[2])
                                           | set(phrase or ()))),
        ))
    candidates.sort(key=lambda c: (
        c.phrase is None, -c.score, c.lesson.grade != item.grade, c.lesson.lesson_id))
    return candidates


def _relation_for(item: Item, lesson: Lesson, default: str) -> tuple[str, str]:
    """(relacija, dodatna napomena) po višekratnim pravilima nad tekstom stavke.

    ŠIROKE stavke (tabele sadržaja) preskaču ova pravila: fraza „ne treba
    dokazivati“ negdje u dugom bloku ne smije preimenovati SVAKO poklapanje
    tog bloka — pravila važe samo kad se tekst stavke odnosi na jedan sadržaj."""
    if item.is_broad:
        return default, ""
    text = item.match_text
    if _EXCLUDE_RE.search(text):
        return "excluded", "izvor izričito isključuje (npr. 'ne treba dokazivati')"
    if _PREREQ_RE.search(text) and lesson.grade < item.grade:
        return "prerequisite", "izvor traži obnavljanje ranijeg gradiva nižeg razreda"
    if (_PROOF_RE.search(text) and "dokaz" not in lesson.title_norm
            and "podudarnost" not in lesson.title_norm):
        return "advanced", "izvor traži dokaz; lekcija ne pokriva dokaz"
    if _INFORMATIVE_RE.search(text):
        return "advanced", "izvor sadržaj uvodi samo informativno"
    return default, ""


def _grade_penalized_confidence(item: Item, lesson: Lesson, confidence: str) -> str:
    if lesson.grade == item.grade:
        return confidence
    return "medium" if confidence == "high" else "low"


def build_mappings(items: list[Item], lessons: list[Lesson]):
    """Glavna petlja: za svaku stavku bar jedan red (mapiran ili no_match)."""
    mappings: list[Mapping] = []
    issues: list[dict] = []
    by_grade: dict[int, list[Lesson]] = defaultdict(list)
    for lesson in lessons:
        by_grade[lesson.grade].append(lesson)
    lessons_by_id = {l.lesson_id: l for l in lessons}

    for item in sorted(items, key=lambda x: x.item_id):
        policy = SECTION_POLICY.get(item.section, "content")
        rows: list[tuple] = []  # (lesson|None, relation, confidence, method, evidence, ambiguity, status)

        if policy == "non_content":
            reason = _NON_CONTENT_REASON.get(item.section, "nije sadržaj lekcije")
            rows.append((None, "", "", "unresolved", reason, "", "no_match"))
            issues.append({
                "issue_type": "non_content_item", "item": item, "lesson": None,
                "description": f"{item.section}: {reason}",
                "severity": "information",
                "candidate_resolution": "ostaviti nemapirano; nije gradivo",
            })
        else:
            candidates = _score_candidates(item, lessons)
            seen_targets: set = set()

            # 1) FRAZNA POKLAPANJA NASLOVA — najjači deterministički dokaz.
            #    Dupli naslovi kroz razrede (npr. oblast identična u 6. i 7.):
            #    zadrži pogodak ISTOG razreda; tuđi razred samo kad svog nema.
            phrase_candidates = [c for c in candidates if c.phrase is not None]
            phrase_titles_same_grade = {
                c.lesson.title_norm for c in phrase_candidates
                if c.lesson.grade == item.grade
            }
            kept_phrases = []
            for candidate in phrase_candidates:
                if (candidate.lesson.grade != item.grade
                        and candidate.lesson.title_norm in phrase_titles_same_grade):
                    continue
                kept_phrases.append(candidate)
            # POTISKIVANJE PODVEDENIH FRAZA: „linearna jednačina“ pogođena
            # UNUTAR „sistem dvije linearne jednačine“ nije zaseban dokaz —
            # kraća fraza čiji je skup osnova podskup duže otpada.
            kept_phrases = [
                a for a in kept_phrases
                if not any(
                    a is not b and set(a.phrase) < set(b.phrase)
                    for b in kept_phrases
                )
            ]
            same_grade_phrase = any(
                c.lesson.grade == item.grade for c in kept_phrases)

            for candidate in kept_phrases:
                lesson = candidate.lesson
                if lesson.lesson_id in seen_targets:
                    continue
                seen_targets.add(lesson.lesson_id)
                relation, rel_note = _relation_for(item, lesson, "exact")
                if policy == "methodology" and relation == "exact":
                    relation, rel_note = "supporting", "metodičko uputstvo, ne središnja vještina"
                confidence = _grade_penalized_confidence(item, lesson, "high")
                method = "terminology_alias" if candidate.alias_used else "exact_title"
                evidence = ("naslov lekcije sadržan u tekstu stavke: „"
                            + " ".join(candidate.phrase) + "“")
                if rel_note:
                    evidence += f"; {rel_note}"
                ambiguity = ""
                if item.is_broad:
                    ambiguity = ("stavka je široka (sadrži tabelu sadržaja programa); "
                                 "poklapanje se odnosi na navedenu frazu")
                if lesson.grade != item.grade:
                    ambiguity = (ambiguity + "; " if ambiguity else "") + \
                        f"razred izvora {item.grade} ≠ razred lekcije {lesson.grade}"
                    issues.append({
                        "issue_type": "grade_mismatch", "item": item, "lesson": lesson,
                        "description": (f"stavka razreda {item.grade} mapirana na lekciju "
                                        f"{lesson.lesson_id} razreda {lesson.grade}"),
                        "severity": "warning",
                        "candidate_resolution": "stručno potvrditi razliku u rasporedu gradiva",
                    })
                status = ("auto_high_confidence" if confidence == "high"
                          and relation in ("exact", "supporting") else "needs_review")
                if relation in ("advanced", "excluded", "prerequisite"):
                    status = "needs_review"
                rows.append((lesson, relation, confidence, method, evidence, ambiguity, status))

            # 2) BODOVNA POKLAPANJA — samo gdje fraze nisu dale isti cilj.
            #    Kad frazni pogodak istog razreda postoji, bodovni kandidati iz
            #    DRUGIH oblasti moraju biti izuzetno jaki da bi opstali.
            phrase_areas = {c.lesson.oblast for c in kept_phrases
                            if c.lesson.grade == item.grade}
            scored = [c for c in candidates
                      if c.phrase is None and c.lesson.lesson_id not in seen_targets
                      and c.score > 0]
            if same_grade_phrase:
                scored = [c for c in scored
                          if c.lesson.oblast in phrase_areas
                          or (len(c.matched_objects) >= 2 and c.score >= 0.75)]
            filtered = []
            for candidate in scored:
                if candidate.lesson.grade != item.grade:
                    same_best = max((c.score for c in scored
                                     if c.lesson.grade == item.grade), default=0.0)
                    if same_grade_phrase or same_best >= 0.45 or candidate.score < 0.65:
                        continue
                filtered.append(candidate)

            def _strong(candidate):
                return (len(candidate.matched_objects) >= 2
                        or (len(candidate.matched_objects) >= 1
                            and len(candidate.matched_actions) >= 1))

            # POBJEDNICI (jasna, konačna pravila):
            #   • svi najbolji po skoru (unutar 0.001) su pobjednici;
            #   • pobjednik ISPOD najboljeg mora donijeti vlastiti dokaz radnje
            #     (poklopljenu akciju svog naslova) — „proširivanjem i
            #     skraćivanjem“ legitimno pokriva obje lekcije, ali „Grafička
            #     metoda“ ne smije pobijediti na stavci o supstituciji;
            #   • kad postoji frazni pogodak istog razreda, SVAKI bodovni
            #     pobjednik mora imati radnju ili skor ≥ 0.75 — fraza je stavku
            #     već objasnila, a band-rezerva se tada uopšte ne koristi;
            #   • pobjednici s IDENTIČNIM skupom pogođenih osnova su
            #     izjednačenje za stručni pregled, ne višestruka pokrivenost.
            eligible = [c for c in filtered if c.score >= 0.60 and _strong(c)]
            winners, tied_ids = [], set()
            if eligible:
                best_eligible = max(c.score for c in eligible)
                for candidate in eligible:
                    if candidate.score >= best_eligible - 0.001:
                        winners.append(candidate)
                    elif candidate.matched_actions:
                        winners.append(candidate)
                if same_grade_phrase:
                    winners = [c for c in winners
                               if c.matched_actions or c.score >= 0.75]
                winners = winners[:4]
                key_counts = Counter(
                    (c.matched_objects, c.matched_actions) for c in winners)
                tied_ids = {
                    c.lesson.lesson_id for c in winners
                    if key_counts[(c.matched_objects, c.matched_actions)] > 1
                }
            if winners:
                emit_list, is_tie = winners, False
            elif eligible and same_grade_phrase:
                # Fraza je odgovor; oslabljeni bodovni kandidati se ne emituju.
                emit_list, is_tie = [], False
            else:
                best_score = max((c.score for c in filtered), default=0.0)
                band = [c for c in filtered if best_score - c.score <= 0.10][:3]
                if same_grade_phrase:
                    band = [c for c in band
                            if c.matched_actions or c.score >= 0.75]
                emit_list = band
                is_tie = len(emit_list) > 1
            if emit_list:
                tie_band = emit_list
                for candidate in tie_band:
                    lesson, score = candidate.lesson, candidate.score
                    if score < 0.34:
                        continue
                    seen_targets.add(lesson.lesson_id)
                    candidate_tie = is_tie or lesson.lesson_id in tied_ids
                    if score >= 0.60 and not candidate_tie and _strong(candidate):
                        confidence = "high"
                    elif score >= 0.45:
                        confidence = "medium"
                    else:
                        confidence = "low"
                    if candidate.used_theme and confidence == "high":
                        confidence = "medium"  # tema iz sekvence nikad nije puni dokaz
                    confidence = _grade_penalized_confidence(item, lesson, confidence)
                    relation, rel_note = _relation_for(item, lesson, "exact")
                    if policy == "methodology" and relation == "exact":
                        relation, rel_note = "supporting", "metodičko uputstvo, ne središnja vještina"
                    if confidence == "low" and relation == "exact":
                        relation = "supporting"
                    method = ("curriculum_sequence" if candidate.used_theme
                              else "terminology_alias" if candidate.alias_used
                              else "normalized_title" if score >= 0.75
                              else "same_grade_area"
                              if lesson.grade == item.grade and score >= 0.45
                              else "semantic_overlap")
                    evidence = (f"pokrivenost naslova {score:.2f}; zajednički pojmovi: "
                                f"{', '.join(candidate.matched_objects[:5]) or '—'}")
                    if candidate.matched_actions:
                        evidence += f"; radnje: {', '.join(candidate.matched_actions[:3])}"
                    if candidate.used_theme:
                        evidence += f"; tema iz sekvence kurikuluma: „{item.inherited_theme}“"
                    if rel_note:
                        evidence += f"; {rel_note}"
                    ambiguity = ""
                    if candidate_tie:
                        others = [c.lesson.lesson_id for c in tie_band
                                  if c.lesson is not lesson
                                  and (is_tie or c.lesson.lesson_id in tied_ids)]
                        ambiguity = ("više uvjerljivih lekcija u istom pojasu bodova: "
                                     + ", ".join(others))
                    if lesson.grade != item.grade:
                        ambiguity = (ambiguity + "; " if ambiguity else "") + \
                            f"razred izvora {item.grade} ≠ razred lekcije {lesson.grade}"
                        issues.append({
                            "issue_type": "grade_mismatch", "item": item, "lesson": lesson,
                            "description": (f"stavka razreda {item.grade} mapirana na lekciju "
                                            f"{lesson.lesson_id} razreda {lesson.grade}"),
                            "severity": "warning",
                            "candidate_resolution": "stručno potvrditi razliku u rasporedu gradiva",
                        })
                    status = ("auto_high_confidence"
                              if confidence == "high" and not candidate_tie
                              and relation in ("exact", "supporting")
                              else "needs_review")
                    rows.append((lesson, relation, confidence, method, evidence,
                                 ambiguity, status))

            if not any(r[0] is not None for r in rows):
                rows.append((None, "", "", "unresolved",
                             "nijedna lekcija ne prelazi deterministički prag poklapanja",
                             "", "no_match"))
                own_objects, _own_actions = _content_stems(tokenize(item.head_text))
                if len(own_objects) >= 2 and not item.is_broad:
                    # Dovoljno specifična stavka bez ijedne lekcije — kandidat
                    # za lekciju koja u MAT-BOT katalogu možda nedostaje.
                    issues.append({
                        "issue_type": "possible_missing_lesson", "item": item,
                        "lesson": None,
                        "description": ("specifična sadržajna stavka bez lekcije: „"
                                        + item.match_text[:120] + "“"),
                        "severity": "information",
                        "candidate_resolution": "procijeniti da li kurikulum traži novu "
                                                "MAT-BOT lekciju",
                    })
                else:
                    issues.append({
                        "issue_type": "no_match_content_item", "item": item, "lesson": None,
                        "description": "sadržajna stavka bez odgovarajuće MAT-BOT lekcije",
                        "severity": "warning",
                        "candidate_resolution": "razmotriti novu lekciju ili proširenje postojeće",
                    })

            # 3) KONFLIKT: uska stavka s POUZDANIM exact pogodcima u >1 oblasti
            #    istog razreda — deterministe se protivrječe, čovjek odlučuje.
            #    RAČUNA SE PRIJE SUSJEDA (popravka defekta 2, Faza 2.5): sporno
            #    sidro ne smije izvesti nijedan susjed. Živi nalaz: KS stavka o
            #    DECIMALNOM zapisu razlomka imala je sporna sidra u dvije
            #    oblasti, pa je iz njih izveden susjed „množenje razlomaka“ —
            #    lažno visoko pouzdanje nad paketom koji ni sam nije razriješen.
            conflict_areas = {
                r[0].oblast for r in rows
                if r[0] and r[1] == "exact" and r[2] == "high"
                and r[0].grade == item.grade
            }
            disputed = not item.is_broad and len(conflict_areas) > 1
            if disputed:
                rows = [
                    (l, rel, conf, meth, ev, amb,
                     "conflict" if rel == "exact" and conf == "high" else st)
                    for (l, rel, conf, meth, ev, amb, st) in rows
                ]
                issues.append({
                    "issue_type": "contradictory_mapping", "item": item, "lesson": None,
                    "description": ("uska stavka ima pouzdane exact kandidate u više "
                                    "oblasti: " + ", ".join(sorted(conflict_areas))),
                    "severity": "warning",
                    "candidate_resolution": "stručni pregled; zadržati samo ispravnu oblast",
                })

            # 4) SUSJEDI IZ TABELE ZAMJENJIVIH VJEŠTINA — dokumentaciona zaštita:
            #    po jednoj (stavka, par) vezi, samo uz pouzdan i NESPORAN exact
            #    izvor, i uz kanonski izbor susjeda (vidi `_canonical_neighbour`).
            exact_lessons = [] if disputed else [
                r[0] for r in rows
                if r[0] is not None and r[1] == "exact" and r[2] == "high"
            ]
            for lesson in exact_lessons:
                for pair in CONFUSION_PAIRS:
                    for own_side, other_side in ((pair.side_a, pair.side_b),
                                                 (pair.side_b, pair.side_a)):
                        if _side_rank(own_side, lesson.title_norm) is None:
                            continue
                        neighbour = _canonical_neighbour(other_side, [
                            l for l in by_grade.get(lesson.grade, [])
                            if l.lesson_id != lesson.lesson_id
                            and l.lesson_id not in seen_targets
                        ])
                        if neighbour is not None:
                            seen_targets.add(neighbour.lesson_id)
                            rows.append((
                                neighbour, "neighbour", "high", "manual_rule",
                                f"susjedna vještina ({pair.name}): {pair.note}; "
                                f"stavka pripada lekciji {lesson.lesson_id}",
                                "", "auto_high_confidence",
                            ))
                        break  # jedna strana para je dovoljna
            if item.is_broad:
                broad_targets = [r[0].lesson_id for r in rows if r[0] and r[1] == "exact"]
                if len(broad_targets) > 6:
                    issues.append({
                        "issue_type": "broad_source_item", "item": item, "lesson": None,
                        "description": (f"široka stavka pokriva {len(broad_targets)} lekcija "
                                        "(sadrži tabelu sadržaja programa)"),
                        "severity": "information",
                        "candidate_resolution": "očekivano za KS tabele sadržaja; bez akcije",
                    })
            low_only = rows and all(
                (r[2] == "low") for r in rows if r[0] is not None and r[1] != "neighbour"
            ) and any(r[0] is not None and r[1] != "neighbour" for r in rows)
            if low_only:
                issues.append({
                    "issue_type": "ambiguous_low_confidence", "item": item, "lesson": None,
                    "description": "stavka ima samo kandidate niskog pouzdanja",
                    "severity": "warning",
                    "candidate_resolution": "stručni pregled prije upotrebe",
                })

        for sequence, (lesson, relation, confidence, method, evidence,
                       ambiguity, status) in enumerate(
                sorted(rows, key=lambda r: (
                    RELATIONS.index(r[1]) if r[1] in RELATIONS else 99,
                    r[0].lesson_id if r[0] else "zzz")), start=1):
            mappings.append(Mapping(
                mapping_id=f"M-{item.item_id}-{sequence:02d}",
                item=item, lesson=lesson, relation=relation,
                confidence=confidence, method=method, evidence=evidence,
                ambiguity_note=ambiguity, review_status=status,
            ))

    _lesson_side_issues(mappings, lessons, issues, lessons_by_id)
    return mappings, issues


def _lesson_side_issues(mappings, lessons, issues, lessons_by_id):
    """Praznine sa strane lekcija + KS/RS razlike u rasporedu + duplikati naslova."""
    per_lesson = defaultdict(list)
    for m in mappings:
        if m.lesson is not None:
            per_lesson[m.lesson.lesson_id].append(m)

    for lesson in lessons:
        rows = per_lesson.get(lesson.lesson_id, [])
        exact_sources = {m.item.source_id for m in rows if m.relation == "exact"}
        if not rows or not any(m.relation in ("exact", "supporting", "prerequisite",
                                              "advanced") for m in rows):
            issues.append({
                "issue_type": "lesson_without_source", "item": None, "lesson": lesson,
                "description": "nijedna KS/RS stavka ne pokriva ovu lekciju",
                "severity": ("blocking_for_contract_activation"
                             if lesson.oblast in PILOT_AREAS and lesson.grade == 6
                             else "warning"),
                "candidate_resolution": "provjeriti izvore ručno; lekcija možda proizlazi "
                                        "iz udžbenika, a ne iz NPP tabela",
            })
        elif len(exact_sources) == 1:
            missing = "RS_2014" if "KS_2018" in exact_sources else "KS_2018"
            issues.append({
                "issue_type": "single_source_lesson", "item": None, "lesson": lesson,
                "description": f"exact pokrivenost samo iz {next(iter(exact_sources))}; "
                               f"{missing} nema exact stavku",
                "severity": "information",
                "candidate_resolution": "prihvatljivo — unija kurikuluma se čuva",
            })

    # KS/RS razlika u rasporedu: ista lekcija exact-mapirana iz stavki različitih razreda.
    for lesson_id, rows in sorted(per_lesson.items()):
        grades_by_source = defaultdict(set)
        for m in rows:
            if m.relation == "exact":
                grades_by_source[m.item.source_id].add(m.item.grade)
        ks, rs = grades_by_source.get("KS_2018"), grades_by_source.get("RS_2014")
        if ks and rs and not (ks & rs):
            issues.append({
                "issue_type": "placement_difference_ks_rs", "item": None,
                "lesson": lessons_by_id[lesson_id],
                "description": f"KS mapira iz razreda {sorted(ks)}, RS iz {sorted(rs)}",
                "severity": "warning",
                "candidate_resolution": "razlika programa entiteta; zadržati uniju",
            })

    # Mogući duplikati lekcija: identičan normalizovan naslov u više razreda.
    by_title = defaultdict(list)
    for lesson in lessons:
        by_title[normalize_text(lesson.title)].append(lesson)
    for _title, group in sorted(by_title.items()):
        if len(group) > 1:
            issues.append({
                "issue_type": "possible_duplicate_lessons", "item": None, "lesson": group[0],
                "description": "identičan naslov lekcije u više razreda: "
                               + ", ".join(l.lesson_id for l in group),
                "severity": "information",
                "candidate_resolution": "namjerno ponavljanje gradiva po razredima; bez akcije",
            })

    # Terminološke dvosmislenosti: aliasi koji su PRAVILO NORMALIZACIJE (nisu
    # eksplicitno navedeni u izvoru) moraju biti vidljivi i kao nalaz.
    for canonical, alias, note in (
        ("pravila djeljivosti", "kriterijumi djeljivosti",
         "RS koristi „kriterijumi djeljivosti“; mapirano na pravila djeljivosti"),
        ("četverougao", "četvorougao",
         "RS ekavska varijanta naslova teme; mapirano na četverougao"),
        ("metoda supstitucije", "metoda zamjene",
         "oba imena metode u upotrebi; mapirano na metodu supstitucije"),
    ):
        issues.append({
            "issue_type": "terminology_ambiguity", "item": None, "lesson": None,
            "description": f"„{alias}“ ⇒ „{canonical}“ — {note}",
            "severity": "information",
            "candidate_resolution": "potvrditi kao kanonsko pravilo terminologije",
        })


# ---------------------------------------------------------------------------
# PILOT: kandidatska porodica iz NASLOVA (višekratna pravila, primijenjena na
# svih 534 — prikazuje se samo pilot 25).
# ---------------------------------------------------------------------------

PILOT_AREAS = ("Djeljivost brojeva", "Razlomci")

_FAMILY_RULES = (
    (r"pravila\s+djeljivosti", "apply_divisibility_rules", "primjena pravila djeljivosti"),
    (r"djeljivost\s+(zbira|razlike|proizvoda)", "divisibility_of_results",
     "djeljivost zbira/razlike/proizvoda"),
    (r"djeljivost\s+dekadskim", "decade_unit_divisibility", "djeljivost dekadskim jedinicama"),
    (r"djelilac/faktor|djelilac.*sadržilac", "divisors_and_multiples",
     "djelioci i sadržioci broja"),
    (r"relativno\s+prosti", "coprime_check", "provjera uzajamne prostosti"),
    (r"prosti\s+i\s+složeni", "prime_composite_classification",
     "klasifikacija prost/složen"),
    (r"rastavljanje.*proste\s+faktore", "prime_factorization", "faktorizacija na proste faktore"),
    (r"najveći zajednički|najmanji zajednički|nzd|nzs", "gcd_lcm_computation",
     "određivanje NZD/NZS"),
    (r"tekstualni\s+zadaci", "word_problems", "tekstualni (životni) zadaci"),
    (r"pojam\s+razlomka", "fraction_concept", "pojam razlomka i njegovih dijelova"),
    (r"dio\s+cjeline.*količnik", "fraction_meaning", "razlomak kao dio cjeline/količnik"),
    (r"pravi,\s*nepravi", "fraction_classification", "klasifikacija razlomaka"),
    (r"brojevnoj\s+polupravoj|brojevnoj\s+osi", "number_line_representation",
     "prikaz na brojevnoj polupravoj/osi"),
    (r"proširivanje", "fraction_equivalence_expand", "proširivanje razlomka"),
    (r"skraćivanje", "fraction_equivalence_reduce", "skraćivanje razlomka"),
    (r"zajednički\s+nazivnik|zajednički\s+imenilac", "common_denominator",
     "svođenje na zajednički imenilac"),
    (r"upoređivanje", "comparison_ordering", "upoređivanje/uređivanje vrijednosti"),
    (r"sabiranje i oduzimanje razlomaka jednakih", "fraction_add_sub_equal",
     "sabiranje/oduzimanje jednakih imenilaca"),
    (r"sabiranje i oduzimanje razlomaka različitih", "fraction_add_sub_unlike",
     "sabiranje/oduzimanje različitih imenilaca"),
    (r"množenje razlomka", "fraction_multiplication", "množenje razlomaka"),
    (r"dijeljenje razlomka", "fraction_division", "dijeljenje razlomaka"),
    (r"svojstva\s+računskih\s+operacija", "operation_properties",
     "svojstva računskih operacija"),
    (r"brojevni\s+izrazi", "numeric_expressions", "brojevni izrazi s više operacija"),
)


def candidate_family(title: str) -> tuple[str, str]:
    normalized = normalize_text(title)
    for pattern, family, skill in _FAMILY_RULES:
        if re.search(pattern, normalized):
            return family, skill
    return "", ""


# ---------------------------------------------------------------------------
# IZLAZNI WORKBOOK
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True)


def _sheet(wb, name, header, rows, widths=None, freeze="A2"):
    ws = wb.create_sheet(name)
    ws.append(list(header))
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="top")
    for row in rows:
        ws.append(["" if v is None else v for v in row])
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    if freeze:
        ws.freeze_panes = freeze
    return ws


def _readme_rows(input_hashes):
    text = [
        ("MAT-BOT — Faza 2: mapiranje KS/RS nastavnih planova na 534 MAT-BOT lekcije",),
        ("",),
        ("Svrha", "Povezati svaku stavku Faze 1 (KS 2018 + RS 2014) s postojećim "
                  "MAT-BOT lekcijama, uz očuvanje UNIJE oba kurikuluma. Stavka koja "
                  "postoji samo u jednom izvoru se NIKAD ne odbacuje."),
        ("Ulazi", "reference/curriculum/MATBOT_Sve_Lekcije_6_7_8_9.xlsx (kanonskih 534, "
                  "nepromijenjen) + reference/curriculum/semantics/"
                  "MATBOT_Faza1_KS_RS_NPP_Matematika.xlsx (573 stavke)."),
        ("SHA-256 ulaza", "; ".join(f"{name}={digest[:16]}…" for name, digest in input_hashes)),
        ("Status", "SVA mapiranja su PRIJEDLOZI za stručni pregled — nijedno nije "
                   "produkcijski ugovor i ništa se ne aktivira tiho. Nesigurno "
                   "mapiranje nosi needs_review/no_match, nikad tihi izbor."),
        ("",),
        ("Relacije", "exact = stavka direktno definiše šta lekcija ispituje; "
                     "supporting = terminologija/reprezentacija/kontekst, ne središnja "
                     "vještina; prerequisite = ranije predznanje (ne smije postati cilj "
                     "lekcije); neighbour = lako zamjenjiva susjedna vještina (ne smije "
                     "zamijeniti lekciju); advanced = iznad predviđenog opsega; "
                     "excluded = izvor izričito isključuje iz lekcije."),
        ("Pouzdanje", "high = frazno poklapanje naslova ili pokrivenost ≥0.60 uz isti "
                      "razred i bar jedan specifičan pojam; medium = jaka pokrivenost "
                      "ili naslijeđena tema sekvence; low = samo semantičko preklapanje "
                      "— uvijek needs_review."),
        ("Status pregleda", "auto_high_confidence (deterministički, bez izjednačenja), "
                            "needs_review, confirmed (samo čovjek), rejected (samo "
                            "čovjek), conflict (protivrječni kandidati), no_match."),
        ("KS+RS unija", "Lekcija pokrivena samo jednim izvorom ostaje pokrivena "
                        "(coverage KS_only/RS_only). Razlike u rasporedu po razredima "
                        "su zabilježene u Praznine_i_sukobi, nikad tiho razriješene."),
        ("Sukobi", "Uska stavka s exact kandidatima u više oblasti dobija status "
                   "conflict na svim exact redovima + zapis u Praznine_i_sukobi. "
                   "Široke KS stavke (tabele sadržaja programa) legitimno pokrivaju "
                   "više lekcija i označene su u ambiguity_note."),
        ("Generičke riječi", "zadatak, primjena, pojam, računanje, rješavanje, svojstva "
                             "i pedagoški vokabular imaju težinu 0 — nikad ne nose "
                             "mapiranje sami."),
        ("Metode", "exact_title → normalized_title → same_grade_area → "
                   "curriculum_sequence → terminology_alias → semantic_overlap → "
                   "manual_rule (sekcijska politika, tabela susjeda) → unresolved."),
        ("Reproducibilnost", "Ponovno pokretanje daje bajt-identičan fajl: fiksni "
                             "metapodaci dokumenta, normalizovani ZIP timestampovi, "
                             "stabilan redoslijed redova. Hash sadržaja je u Kontroli."),
        ("Alat", "scripts/build_curriculum_mapping.py (offline; nula mrežnih i nula "
                 "model poziva; --dry-run i --report podržani)."),
    ]
    return text


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _content_hash(sheets: dict) -> str:
    payload = json.dumps(
        {name: rows for name, (header, rows) in sorted(sheets.items())},
        ensure_ascii=False, sort_keys=True, default=str,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


_DCTERMS_RE = re.compile(
    rb'(<dcterms:(?:created|modified) xsi:type="dcterms:W3CDTF">)[^<]*(</dcterms:)'
)


def _normalize_zip(path: Path) -> None:
    """Prepiši ZIP s fiksnim timestampovima — uslov bajt-reproducibilnosti.

    openpyxl pri snimanju ponovo upiše `dcterms:modified` tekućim vremenom,
    bez obzira na postavljena svojstva — zato se docProps/core.xml normalizuje
    ovdje, na bajt nivou, na isti fiksni datum kao i ostatak dokumenta."""
    fixed = _FIXED_DOC_DATE.strftime("%Y-%m-%dT%H:%M:%SZ").encode("ascii")
    source = zipfile.ZipFile(path, "r")
    buffer = io.BytesIO()
    target = zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED)
    for info in sorted(source.infolist(), key=lambda i: i.filename):
        data = source.read(info.filename)
        if info.filename == "docProps/core.xml":
            data = _DCTERMS_RE.sub(rb"\g<1>" + fixed + rb"\g<2>", data)
        clean = zipfile.ZipInfo(info.filename, date_time=(1980, 1, 1, 0, 0, 0))
        clean.compress_type = zipfile.ZIP_DEFLATED
        clean.external_attr = info.external_attr
        target.writestr(clean, data)
    target.close()
    source.close()
    path.write_bytes(buffer.getvalue())


def build_workbook(lessons, items, mappings, issues, phase1_meta, out_path: Path,
                   dry_run=False):
    per_item = defaultdict(list)
    per_lesson = defaultdict(list)
    for m in mappings:
        per_item[m.item.item_id].append(m)
        if m.lesson is not None:
            per_lesson[m.lesson.lesson_id].append(m)

    input_hashes = [("kanonski", _sha256(CANONICAL_XLSX)), ("faza1", _sha256(PHASE1_XLSX))]
    sheets: dict[str, tuple] = {}

    # --- README ---
    sheets["README"] = (("MAT-BOT Faza 2",), _readme_rows(input_hashes))

    # --- Izvori ---
    mapped_by_source = Counter(
        m.item.source_id for m in mappings if m.lesson is not None)
    unresolved_by_source = Counter(
        m.item.source_id for m in mappings if m.review_status == "no_match")
    izvori_rows = []
    for row in phase1_meta["izvori"]:
        source_id = row[0]
        izvori_rows.append(list(row) + [
            phase1_meta["extracted_at"], row[2],
            mapped_by_source.get(source_id, 0), unresolved_by_source.get(source_id, 0),
        ])
    sheets["Izvori"] = (
        tuple(phase1_meta["izvori_header"]) + ("Datum ekstrakcije (Faza 1)",
                                               "Verzija izvora", "Mapiranih redova",
                                               "Neriješenih stavki"),
        izvori_rows,
    )

    # --- Lekcije_534 ---
    lesson_rows = []
    coverage_counter = Counter()
    for lesson in lessons:
        rows = per_lesson.get(lesson.lesson_id, [])
        ks_exact = sum(1 for m in rows if m.relation == "exact" and m.item.source_id == "KS_2018")
        rs_exact = sum(1 for m in rows if m.relation == "exact" and m.item.source_id == "RS_2014")
        supporting = sum(1 for m in rows if m.relation == "supporting")
        prerequisite = sum(1 for m in rows if m.relation == "prerequisite")
        neighbour = sum(1 for m in rows if m.relation == "neighbour")
        advanced = sum(1 for m in rows if m.relation == "advanced")
        excluded = sum(1 for m in rows if m.relation == "excluded")
        unresolved = sum(1 for m in rows if m.review_status == "needs_review")
        has_conflict = any(m.review_status == "conflict" for m in rows)
        if has_conflict:
            coverage = "conflict"
        elif ks_exact and rs_exact:
            coverage = "both_sources"
        elif ks_exact:
            coverage = "KS_only"
        elif rs_exact:
            coverage = "RS_only"
        elif supporting or prerequisite or advanced:
            coverage = "needs_review"
        else:
            coverage = "no_source_match"
        coverage_counter[coverage] += 1
        lesson_rows.append([
            lesson.lesson_id, lesson.grade, lesson.oblast, lesson.title,
            ks_exact, rs_exact, supporting, prerequisite, neighbour, advanced,
            excluded, unresolved, coverage,
            "auto" if coverage in ("both_sources", "KS_only", "RS_only") else "needs_review",
            "",
        ])
    sheets["Lekcije_534"] = (
        ("lesson_id", "grade", "area", "lesson_title", "KS_exact_count",
         "RS_exact_count", "supporting_count", "prerequisite_count",
         "neighbour_count", "advanced_count", "excluded_count", "unresolved_count",
         "coverage_status", "review_status", "reviewer_note"),
        lesson_rows,
    )

    # --- Stavke_NPP ---
    item_rows = []
    for item in sorted(items, key=lambda x: x.item_id):
        rows = per_item.get(item.item_id, [])
        mapped = [m for m in rows if m.lesson is not None]
        if any(m.relation == "exact" and m.review_status != "conflict" for m in mapped):
            status = "exact_mapped"
        elif any(m.review_status == "conflict" for m in mapped):
            status = "conflict"
        elif mapped:
            status = "mapped_non_exact"
        else:
            status = "no_match"
        note = ""
        if item.is_broad:
            note = "široka stavka (tabela sadržaja programa)"
        item_rows.append([
            item.item_id, item.source_id, item.grade, item.page, item.section,
            item.theme, item.item_type, item.original_text, item.latin_text,
            item.parse_confidence, item.source_review,
            len(mapped), status, note,
        ])
    sheets["Stavke_NPP"] = (
        ("item_id", "source_id", "grade", "pdf_page", "section", "source_topic",
         "item_type", "original_text", "normalized_latin_text",
         "extraction_confidence", "source_review_status",
         "mapping_count", "mapping_status", "mapping_note"),
        item_rows,
    )

    # --- Mapiranje ---
    mapping_rows = []
    for m in mappings:
        mapping_rows.append([
            m.mapping_id, m.item.item_id, m.item.source_id, m.item.grade,
            m.item.inherited_theme or m.item.theme,
            m.item.item_type, m.item.latin_text[:500], m.item.page,
            m.lesson.lesson_id if m.lesson else "",
            m.lesson.grade if m.lesson else "",
            m.lesson.oblast if m.lesson else "",
            m.lesson.title if m.lesson else "",
            m.relation, m.confidence, m.method, m.evidence[:300],
            m.ambiguity_note[:300], m.review_status, "",
        ])
    sheets["Mapiranje"] = (
        ("mapping_id", "item_id", "source_id", "source_grade", "source_topic",
         "source_item_type", "source_text", "source_page", "target_lesson_id",
         "target_grade", "target_area", "target_lesson_title", "relation",
         "confidence", "mapping_method", "evidence", "ambiguity_note",
         "review_status", "reviewer_note"),
        mapping_rows,
    )

    # --- Terminologija ---
    term_rows = list(phase1_meta["terminology_rows"])
    sheets["Terminologija"] = (
        ("canonical_term", "alias", "script", "source_id", "source_page",
         "context", "confidence", "review_status"),
        term_rows,
    )

    # --- Praznine_i_sukobi ---
    issue_rows = []
    for index, issue in enumerate(sorted(
            issues, key=lambda i: (i["issue_type"],
                                   i["item"].item_id if i["item"] else "",
                                   i["lesson"].lesson_id if i["lesson"] else "")), start=1):
        item, lesson = issue["item"], issue["lesson"]
        issue_rows.append([
            f"ISSUE-{index:04d}", issue["issue_type"],
            item.source_id if item else "", item.item_id if item else "",
            lesson.lesson_id if lesson else "",
            item.grade if item else (lesson.grade if lesson else ""),
            issue["description"], issue["candidate_resolution"],
            issue["severity"], "needs_review",
        ])
    sheets["Praznine_i_sukobi"] = (
        ("issue_id", "issue_type", "source_id", "item_id", "lesson_id", "grade",
         "description", "candidate_resolution", "severity", "review_status"),
        issue_rows,
    )

    # --- Pokrivenost ---
    def _pct(part, whole):
        return f"{(100.0 * part / whole):.1f}%" if whole else "0.0%"

    mapped_mappings = [m for m in mappings if m.lesson is not None]
    coverage_rows = [("— PO RAZREDU (lekcije) —", "", "", "")]
    for grade in (6, 7, 8, 9):
        grade_lessons = [r for r in lesson_rows if r[1] == grade]
        covered = [r for r in grade_lessons if r[12] in ("both_sources", "KS_only", "RS_only")]
        coverage_rows.append((f"razred {grade}", len(grade_lessons),
                              f"exact pokriveno {len(covered)}",
                              _pct(len(covered), len(grade_lessons))))
    coverage_rows.append(("— PO OBLASTI (lekcije bez exact pokrivenosti) —", "", "", ""))
    by_area = defaultdict(lambda: [0, 0])
    for r in lesson_rows:
        by_area[(r[1], r[2])][0] += 1
        if r[12] in ("no_source_match", "needs_review"):
            by_area[(r[1], r[2])][1] += 1
    for (grade, area), (total, uncovered) in sorted(by_area.items()):
        coverage_rows.append((f"r{grade} {area}", total, f"bez exact: {uncovered}",
                              _pct(total - uncovered, total)))
    coverage_rows.append(("— PO IZVORU (mapirani redovi) —", "", "", ""))
    for source in ("KS_2018", "RS_2014"):
        count = sum(1 for m in mapped_mappings if m.item.source_id == source)
        coverage_rows.append((source, count, "mapiranih redova", ""))
    coverage_rows.append(("— PO RELACIJI —", "", "", ""))
    relation_counts = Counter(m.relation for m in mapped_mappings)
    for relation in RELATIONS:
        coverage_rows.append((relation, relation_counts.get(relation, 0), "",
                              _pct(relation_counts.get(relation, 0), len(mapped_mappings))))
    coverage_rows.append(("— PO POUZDANJU —", "", "", ""))
    confidence_counts = Counter(m.confidence for m in mapped_mappings)
    for confidence in CONFIDENCES:
        coverage_rows.append((confidence, confidence_counts.get(confidence, 0), "",
                              _pct(confidence_counts.get(confidence, 0), len(mapped_mappings))))
    coverage_rows.append(("— PO STATUSU PREGLEDA (svi redovi) —", "", "", ""))
    status_counts = Counter(m.review_status for m in mappings)
    for status in REVIEW_STATUSES:
        coverage_rows.append((status, status_counts.get(status, 0), "",
                              _pct(status_counts.get(status, 0), len(mappings))))
    coverage_rows.append(("— POKRIVENOST LEKCIJA —", "", "", ""))
    for coverage_status in ("both_sources", "KS_only", "RS_only", "needs_review",
                            "no_source_match", "conflict"):
        coverage_rows.append((coverage_status, coverage_counter.get(coverage_status, 0),
                              "", _pct(coverage_counter.get(coverage_status, 0),
                                       len(lesson_rows))))
    unresolved_items = sum(1 for r in item_rows if r[12] == "no_match")
    coverage_rows.append(("— STAVKE —", "", "", ""))
    coverage_rows.append(("stavki ukupno", len(item_rows), "", ""))
    coverage_rows.append(("neriješene stavke (no_match)", unresolved_items, "",
                          _pct(unresolved_items, len(item_rows))))
    sheets["Pokrivenost"] = (("kategorija", "broj", "detalj", "procenat"), coverage_rows)

    # --- Pilot_25 ---
    pilot_rows = []
    for lesson in lessons:
        if lesson.grade != 6 or lesson.oblast not in PILOT_AREAS:
            continue
        rows = per_lesson.get(lesson.lesson_id, [])

        def _ids(relation, source=None):
            selected = [m.item.item_id for m in rows if m.relation == relation
                        and (source is None or m.item.source_id == source)]
            return "; ".join(selected[:8]) + ("…" if len(selected) > 8 else "")

        family, skill = candidate_family(lesson.title)
        unresolved = sum(1 for m in rows if m.review_status == "needs_review")
        has_issue = any(
            i["lesson"] is not None and i["lesson"].lesson_id == lesson.lesson_id
            and i["severity"] != "information" for i in issues)
        exact_total = sum(1 for m in rows if m.relation == "exact")
        priority = (1 if has_issue or exact_total == 0
                    else 2 if unresolved else 3)
        pilot_rows.append([
            lesson.lesson_id, lesson.oblast, lesson.title,
            _ids("exact", "KS_2018"), _ids("exact", "RS_2014"),
            _ids("supporting"), _ids("prerequisite"), _ids("neighbour"),
            _ids("excluded"), unresolved, family, skill, priority,
        ])
    pilot_rows.sort(key=lambda r: (r[12], r[0]))
    sheets["Pilot_25"] = (
        ("lesson_id", "area", "lesson_title", "KS_exact_items", "RS_exact_items",
         "supporting_items", "prerequisite_items", "neighbour_items",
         "excluded_items", "unresolved_count", "candidate_family",
         "candidate_core_skill", "review_priority"),
        pilot_rows,
    )

    # --- Kontrola ---
    checks = run_checks(lessons, items, mappings, item_rows, phase1_meta)
    content_digest = _content_hash(sheets)
    kontrola_rows = [(name, value, status) for name, value, status in checks]
    kontrola_rows.append(("hash sadržaja (svi listovi, prije upisa)",
                          content_digest, "INFO"))
    kontrola_rows.append(("reproducibilnost",
                          "dva uzastopna pokretanja daju bajt-identičan fajl "
                          "(fiksni metapodaci + normalizovan ZIP); dokaz: "
                          "sha256 fajla u izvještaju builda", "INFO"))
    sheets["Kontrola"] = (("provjera", "vrijednost", "status"), kontrola_rows)

    failed = [row for row in kontrola_rows if row[2] == "FAIL"]
    if failed:
        raise SystemExit(f"Kontrolne provjere pale: {failed}")

    if dry_run:
        return sheets, content_digest, None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "MAT-BOT Faza 2 builder"
    wb.properties.created = _FIXED_DOC_DATE
    wb.properties.modified = _FIXED_DOC_DATE
    wb.properties.lastModifiedBy = "build_curriculum_mapping.py"

    order = ("README", "Izvori", "Lekcije_534", "Stavke_NPP", "Mapiranje",
             "Terminologija", "Praznine_i_sukobi", "Pokrivenost", "Pilot_25",
             "Kontrola")
    widths_by_sheet = {
        "README": {"A": 22, "B": 120},
        "Mapiranje": {"A": 20, "G": 60, "L": 45, "P": 50, "Q": 35},
        "Stavke_NPP": {"H": 60, "I": 60},
        "Lekcije_534": {"D": 55},
        "Praznine_i_sukobi": {"G": 70, "H": 50},
        "Pilot_25": {"C": 50, "D": 30, "E": 30},
        "Pokrivenost": {"A": 45},
        "Kontrola": {"A": 55, "B": 60},
        "Izvori": {"F": 45, "G": 45},
        "Terminologija": {"F": 45},
    }
    for name in order:
        header, rows = sheets[name]
        _sheet(wb, name, header, rows, widths_by_sheet.get(name))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    _normalize_zip(out_path)
    return sheets, content_digest, _sha256(out_path)


# ---------------------------------------------------------------------------
# KONTROLNE PROVJERE (deterministe; FAIL obara build)
# ---------------------------------------------------------------------------

def run_checks(lessons, items, mappings, item_rows, phase1_meta):
    checks = []
    lesson_ids = {l.lesson_id for l in lessons}
    item_ids = {i.item_id for i in items}

    def check(name, ok, value):
        if not isinstance(value, (str, int, float, bool)) or isinstance(value, bool):
            value = str(value)
        checks.append((name, value, "PASS" if ok else "FAIL"))

    check("534 jedinstvene kanonske lekcije",
          len(lessons) == EXPECTED_LESSON_COUNT and len(lesson_ids) == EXPECTED_LESSON_COUNT,
          len(lessons))
    canonical_now = _sha256(CANONICAL_XLSX)
    check("kanonski workbook nepromijenjen tokom builda",
          canonical_now == phase1_meta["canonical_hash_at_start"], canonical_now[:16] + "…")
    bad_targets = sorted({m.lesson.lesson_id for m in mappings
                          if m.lesson and m.lesson.lesson_id not in lesson_ids})
    check("svaka ciljna lekcija postoji", not bad_targets, bad_targets or "sve postoje")
    bad_items = sorted({m.item.item_id for m in mappings if m.item.item_id not in item_ids})
    check("svaka mapirana stavka postoji", not bad_items, bad_items or "sve postoje")
    page_ok = all(i.page in phase1_meta["pages_by_source"][i.source_id] for i in items)
    check("svaka referenca stranice postoji u Fazi 1", page_ok,
          "sve stranice prisutne" if page_ok else "nedostaju stranice")
    bad_relation = sorted({m.relation for m in mappings
                           if m.lesson is not None and m.relation not in RELATIONS})
    check("relacija je iz dozvoljenog skupa", not bad_relation, bad_relation or "sve validne")
    bad_confidence = sorted({m.confidence for m in mappings
                             if m.lesson is not None and m.confidence not in CONFIDENCES})
    check("pouzdanje je iz dozvoljenog skupa", not bad_confidence,
          bad_confidence or "sva validna")
    bad_status = sorted({m.review_status for m in mappings
                         if m.review_status not in REVIEW_STATUSES})
    check("status pregleda je iz dozvoljenog skupa", not bad_status,
          bad_status or "svi validni")
    bad_method = sorted({m.method for m in mappings if m.method not in MAPPING_METHODS})
    check("metoda je iz dozvoljenog skupa", not bad_method, bad_method or "sve validne")
    mapping_ids = [m.mapping_id for m in mappings]
    check("nema duplih mapping_id", len(mapping_ids) == len(set(mapping_ids)),
          len(mapping_ids))
    items_with_rows = {m.item.item_id for m in mappings}
    check("nijedna stavka nije tiho odbačena",
          items_with_rows == item_ids and len(item_rows) == EXPECTED_ITEM_COUNT,
          f"{len(items_with_rows)}/{len(item_ids)} stavki ima red")
    actual_by_sg = Counter((i.source_id, i.grade) for i in items)
    check("zbirovi se poklapaju s Kontrolom Faze 1",
          dict(actual_by_sg) == EXPECTED_ITEMS_BY_SOURCE_GRADE, dict(actual_by_sg))
    sources_mapped = {m.item.source_id for m in mappings if m.lesson is not None}
    check("oba izvora (KS i RS) su zastupljena u mapiranjima",
          sources_mapped == {"KS_2018", "RS_2014"}, sorted(sources_mapped))
    no_match_ok = all(
        (m.lesson is None) == (m.review_status == "no_match")
        for m in mappings if m.review_status == "no_match" or m.lesson is None
    )
    check("no_match redovi nemaju cilj (i obrnuto)", no_match_ok, "dosljedno")
    return checks


# ---------------------------------------------------------------------------
# TERMINOLOGIJA — spajanje eksplicitnih KS aliasa + ćiriličnih RS oblika
# ---------------------------------------------------------------------------

# Kanonski (bosanski latinični) termin → osnova za pretragu RS ćiriličnog teksta.
_RS_TERM_TARGETS = (
    ("ugao", "ugl"), ("imenilac", "imenioc"), ("brojilac", "brojioc"),
    ("procenat", "procent"), ("kružnica", "kružnic"), ("djelilac", "djelioc"),
    ("sadržilac", "sadržioc"), ("jednačina", "jednačin"), ("nejednačina", "nejednačin"),
    ("razlomak", "razlomk"), ("trougao", "trougl"), ("četverougao", "četverougl"),
    ("površina", "površin"), ("zapremina", "zapremin"), ("upoređivanje", "upoređ"),
    ("djeljivost", "djeljiv"),
)


def build_terminology_rows(items, phase1_terms):
    rows = []
    # 1) Eksplicitni KS aliasi iz Faze 1 — čuvaju se svi, canonical = prvi termin.
    for source_id, page, term1, term2, context, _kind in phase1_terms:
        rows.append([term1, term2, "latinica", source_id, page, context,
                     "high", "auto_high_confidence"])
    # 2) Ćirilični oblici iz RS originala: prvi nalaz po (kanonski, alias).
    seen = set()
    for item in sorted((i for i in items if i.source_id == "RS_2014"),
                       key=lambda x: x.item_id):
        for token in re.findall(r"[Ѐ-ӿ]+", item.original_text):
            latin = transliterate(token)
            folded = _fold(stem(latin.lower()))
            for canonical, target_stem in _RS_TERM_TARGETS:
                if folded == target_stem:
                    key = (canonical, token.lower())
                    if key in seen:
                        continue
                    seen.add(key)
                    rows.append([canonical, token, "ćirilica", item.source_id, item.page,
                                 f"stavka {item.item_id}", "high", "auto_high_confidence"])
    # 3) Aliasi korišteni kao pravila normalizacije (izvedeni, traže pregled).
    rows.append(["pravila djeljivosti", "kriterijumi djeljivosti", "latinica",
                 "RS_2014", "", "pravilo normalizacije za poklapanje (RS→MAT-BOT)",
                 "medium", "needs_review"])
    rows.append(["četverougao", "četvorougao", "latinica", "RS_2014", "",
                 "pravilo normalizacije za poklapanje (ekavska varijanta)",
                 "medium", "needs_review"])
    rows.append(["metoda supstitucije", "metoda zamjene", "latinica", "RS_2014", "",
                 "pravilo normalizacije za poklapanje (oba imena metode)",
                 "medium", "needs_review"])
    rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[3])))
    return rows


# ---------------------------------------------------------------------------
# GLAVNI TOK
# ---------------------------------------------------------------------------

def load_phase1_meta(items):
    wb = openpyxl.load_workbook(PHASE1_XLSX, data_only=True, read_only=True)
    izvori = list(wb["Izvori"].iter_rows(values_only=True))
    terminologija = [r for r in list(wb["Terminologija"].iter_rows(values_only=True))[1:]]
    pages_rs = {r[2] for r in list(wb["Stranice_RS"].iter_rows(values_only=True))[1:]}
    pages_ks = {r[2] for r in list(wb["Stranice_KS"].iter_rows(values_only=True))[1:]}
    created = wb.properties.created
    wb.close()
    return {
        "izvori_header": izvori[0],
        "izvori": izvori[1:],
        "terminology_rows": build_terminology_rows(items, terminologija),
        "pages_by_source": {"RS_2014": pages_rs, "KS_2018": pages_ks},
        "extracted_at": created.date().isoformat() if created else "nepoznat",
        "canonical_hash_at_start": _sha256(CANONICAL_XLSX),
    }


def print_report(sheets, mappings, content_digest, file_digest):
    mapped = [m for m in mappings if m.lesson is not None]
    print("=== FAZA 2 — sažetak ===")
    print(f"stavki: {len({m.item.item_id for m in mappings})}  "
          f"redova mapiranja: {len(mappings)}  (s ciljem: {len(mapped)})")
    print("relacije:", dict(sorted(Counter(m.relation for m in mapped).items())))
    print("pouzdanje:", dict(sorted(Counter(m.confidence for m in mapped).items())))
    print("status:", dict(sorted(Counter(m.review_status for m in mappings).items())))
    _header, lesson_rows = sheets["Lekcije_534"]
    print("pokrivenost lekcija:", dict(sorted(Counter(r[12] for r in lesson_rows).items())))
    print("hash sadržaja:", content_digest[:32], "…")
    if file_digest:
        print("sha256 fajla:", file_digest)


def main(argv=None):
    parser = argparse.ArgumentParser(description="MAT-BOT Faza 2 kurikularno mapiranje")
    parser.add_argument("--dry-run", action="store_true",
                        help="sve provjere i izvještaj, bez pisanja fajla")
    parser.add_argument("--report", action="store_true", help="ispiši sažetak")
    parser.add_argument("--out", type=Path, default=OUTPUT_XLSX,
                        help="izlazna putanja (za testove reproducibilnosti)")
    args = parser.parse_args(argv)

    for path in (CANONICAL_XLSX, PHASE1_XLSX, TOPICS_JSON):
        if not path.exists():
            raise SystemExit(f"Ulaz ne postoji: {path}")

    lessons = load_lessons()
    crosscheck_topics_json(lessons)
    items = load_items()
    prepare_items(items)
    phase1_meta = load_phase1_meta(items)
    mappings, issues = build_mappings(items, lessons)
    sheets, content_digest, file_digest = build_workbook(
        lessons, items, mappings, issues, phase1_meta, args.out, dry_run=args.dry_run,
    )
    if args.report or args.dry_run:
        print_report(sheets, mappings, content_digest, file_digest)
    if not args.dry_run:
        print(f"OK: {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
