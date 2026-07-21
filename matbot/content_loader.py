"""Load NPP AI-tutor content from grade-specific Excel workbooks (grades 6–9).

Since 2026-07 all grades use ONE unified NPP workbook per grade
(``data/{g}_razred/AI_MATH_{g}_NPP_BASE_UNIFIED_SIMPLE.xlsx``). The primary
sheet ``NPP_TOPICS`` is structurally identical across grades; secondary sheets
(``VIDEO_LINKS``, ``AREA_SELECTOR`` …) diverge between the older grade-6 variant
and the newer 7/8/9 variant. To stay robust we read strictly BY HEADER NAME and
treat any variant-only column as optional, and we derive the oblast list from
``NPP_TOPICS`` itself (never from the divergent ``AREA_SELECTOR``).

Excel remains the single source of truth: this module only normalizes rows,
validates the few columns that exist in every variant, and caches loads per
grade. Nothing about topics/videos/oblasti is hardcoded.

``load_master_content`` returns::

    {
      "topics":        [row, ...],          # NPP topics, sheet order
      "topics_by_id":  {npp_topic_id: row},
      "topic_ids":     {npp_topic_id, ...},
      "videos_by_topic": {npp_topic_id: [video_row, ...]},
      "areas":         [{"oblast","area_order","topic_count","topics_with_video"}],
      "grade": int, "source_path": str, "columns": [...],
      # backwards-compatible keys expected by older callers/tests:
      "video_flow": None, "parent_report": None,
    }

Each topic ``row`` is the raw NPP row plus normalized aliases used downstream:
``topic`` = ``npp_topic_id``, ``oblast`` = ``oblast_ui``, ``display_name`` =
``tema_ui``.
"""
from __future__ import annotations

import re
import threading
from pathlib import Path
from typing import Any

import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent

SUPPORTED_GRADES = frozenset({6, 7, 8, 9})
DEFAULT_GRADE = 6

# --- NPP sheet / column names ---------------------------------------------------
NPP_TOPICS_SHEET = "NPP_TOPICS"
VIDEO_LINKS_SHEET = "VIDEO_LINKS"

# Columns present in NPP_TOPICS across ALL grade variants (safe to require).
REQUIRED_TOPIC_COLUMNS = frozenset(
    {"grade", "npp_topic_id", "oblast_ui", "tema_ui"}
)

# Only "video" lessons are offered as video recommendations (podsjetnik/other
# resource types are ignored for the reco list).
_VIDEO_LESSON_TYPE = "video"


class ContentLoadError(Exception):
    """Raised when content files, sheets, columns, or grade are invalid."""


def normalize_header(name: Any) -> str:
    if name is None:
        return ""
    s = str(name).strip().lower()
    return re.sub(r"[\s\-]+", "_", s)


def normalize_value(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return str(val)
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    return str(val).strip()


def normalize_grade(grade: Any = DEFAULT_GRADE) -> int:
    raw = normalize_value(grade)
    match = re.search(r"\d+", raw)
    g = int(match.group(0)) if match else DEFAULT_GRADE
    if g not in SUPPORTED_GRADES:
        supported = ", ".join(str(x) for x in sorted(SUPPORTED_GRADES))
        raise ContentLoadError(
            f"Nepodrzan razred: {raw or grade}. Podrzani razredi su: {supported}."
        )
    return g


def _grade_dir(grade: Any = DEFAULT_GRADE) -> Path:
    return _REPO_ROOT / "data" / f"{normalize_grade(grade)}_razred"


def master_path_for_grade(grade: Any = DEFAULT_GRADE) -> Path:
    g = normalize_grade(grade)
    return _grade_dir(g) / f"AI_MATH_{g}_NPP_BASE_UNIFIED_SIMPLE.xlsx"


# Backwards-compatible module constant (grade-6 default workbook path).
MASTER_PATH = master_path_for_grade(DEFAULT_GRADE)


def _read_sheet(wb, sheet_name: str) -> tuple[list[str], list[dict[str, str]]]:
    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return [], []

    headers = [normalize_header(h) for h in header_row]
    rows: list[dict[str, str]] = []
    for raw in row_iter:
        if raw is None:
            continue
        if all(c is None or str(c).strip() == "" for c in raw):
            continue
        row: dict[str, str] = {}
        for i, col in enumerate(headers):
            if not col:
                continue
            row[col] = normalize_value(raw[i]) if i < len(raw) else ""
        rows.append(row)
    return headers, rows


def _open_workbook(path: Path):
    if not path.exists():
        raise ContentLoadError(f"Excel fajl nije pronadjen: {path}")
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl raises varied types
        raise ContentLoadError(f"Ne mogu otvoriti Excel fajl {path}: {exc}") from exc


def _topic_row(raw: dict[str, str]) -> dict[str, str]:
    """Raw NPP_TOPICS row + downstream aliases (topic/oblast/display_name)."""
    row = dict(raw)
    row["topic"] = raw.get("npp_topic_id", "")
    row["oblast"] = raw.get("oblast_ui", "")
    row["display_name"] = raw.get("tema_ui", "")
    return row


def _video_sort_key(v: dict[str, str]) -> tuple[int, int]:
    """Order videos by recommendation_priority (g6) then lesson_order; missing
    values sort last. Both columns are read defensively (variant-dependent)."""
    def _int(val: str, default: int) -> int:
        try:
            return int(val)
        except (TypeError, ValueError):
            return default

    return (
        _int(v.get("recommendation_priority", ""), 99),
        _int(v.get("lesson_order", ""), 999),
    )


def _build_videos_by_topic(wb) -> dict[str, list[dict[str, str]]]:
    """Index VIDEO_LINKS by npp_topic_id (only lesson_type == 'video').

    Robust to the two schema variants: g6 has lesson_url/thinkific_lesson_id/
    recommendation_priority; g7/8/9 have recommendation_rule instead. We keep the
    raw row and only rely on always-present columns downstream."""
    if VIDEO_LINKS_SHEET not in wb.sheetnames:
        return {}
    _, rows = _read_sheet(wb, VIDEO_LINKS_SHEET)
    by_topic: dict[str, list[dict[str, str]]] = {}
    for r in rows:
        if normalize_value(r.get("lesson_type")).lower() != _VIDEO_LESSON_TYPE:
            continue
        tid = r.get("npp_topic_id", "")
        if not tid:
            continue
        by_topic.setdefault(tid, []).append(r)
    for tid in by_topic:
        by_topic[tid].sort(key=_video_sort_key)
    return by_topic


def _build_areas(topics: list[dict[str, str]]) -> list[dict[str, Any]]:
    """Oblast list derived from NPP_TOPICS (NOT AREA_SELECTOR, which diverges).

    Ordered by area_order (falling back to first-seen order). Carries topic_count
    and topics_with_video so the UI can render an oblast-first selector."""
    order: list[str] = []
    meta: dict[str, dict[str, Any]] = {}
    for t in topics:
        oblast = t.get("oblast", "")
        if not oblast:
            continue
        if oblast not in meta:
            order.append(oblast)
            ao = normalize_value(t.get("area_order"))
            meta[oblast] = {
                "oblast": oblast,
                "area_order": int(ao) if ao.isdigit() else 999,
                "topic_count": 0,
                "topics_with_video": 0,
            }
        meta[oblast]["topic_count"] += 1
        if normalize_value(t.get("has_thinkific_video")).upper() == "DA":
            meta[oblast]["topics_with_video"] += 1
    return sorted(
        (meta[o] for o in order),
        key=lambda m: (m["area_order"], order.index(m["oblast"])),
    )


THINKIFIC_RESOURCES_SHEET = "THINKIFIC_RESOURCES"


def _read_thinkific_resources(wb) -> list[dict[str, str]]:
    """Rows of THINKIFIC_RESOURCES, or [] when the sheet is absent.

    Read defensively: this sheet is optional and its columns differ between
    grade variants, so a missing sheet or column must never break loading.
    """
    if THINKIFIC_RESOURCES_SHEET not in wb.sheetnames:
        return []
    try:
        _cols, rows = _read_sheet(wb, THINKIFIC_RESOURCES_SHEET)
    except Exception:
        return []
    return [r for r in rows if any((v or "").strip() for v in r.values())]


def load_master_content(
    path: str | Path | None = None, grade: Any = DEFAULT_GRADE
) -> dict[str, Any]:
    g = normalize_grade(grade)
    path = Path(path) if path is not None else master_path_for_grade(g)
    wb = _open_workbook(path)
    try:
        if NPP_TOPICS_SHEET not in wb.sheetnames:
            raise ContentLoadError(
                f"NPP workbook nema obavezni sheet '{NPP_TOPICS_SHEET}'."
            )
        columns, all_rows = _read_sheet(wb, NPP_TOPICS_SHEET)
        missing = REQUIRED_TOPIC_COLUMNS - set(columns)
        if missing:
            raise ContentLoadError(
                f"Sheet '{NPP_TOPICS_SHEET}' nema obavezne kolone: {sorted(missing)}"
            )

        topics = [_topic_row(r) for r in all_rows if r.get("npp_topic_id")]
        topics_by_id: dict[str, dict[str, str]] = {}
        for row in topics:
            topics_by_id.setdefault(row["topic"], row)

        videos_by_topic = _build_videos_by_topic(wb)
        areas = _build_areas(topics)
        thinkific_resources = _read_thinkific_resources(wb)
    finally:
        wb.close()

    return {
        "topics": topics,
        "topics_by_id": topics_by_id,
        "topic_ids": set(topics_by_id),
        "videos_by_topic": videos_by_topic,
        # THINKIFIC_RESOURCES carries ``thinkific_lesson_id`` →
        # ``linked_npp_topic_ids``: the workbook's own runtime-id mapping. It was
        # never exposed, so a runtime lesson id could not be resolved even where
        # the sheet is populated.
        "thinkific_resources": thinkific_resources,
        "areas": areas,
        "columns": columns,
        "source_path": str(path),
        "grade": g,
        # backwards-compatible keys (old schema had these sheets; NPP does not)
        "video_flow": None,
        "parent_report": None,
    }


def load_thinkific_map(
    path: str | Path | None = None, grade: Any = DEFAULT_GRADE
) -> dict[str, Any]:
    """Compatibility shim for the old two-file schema.

    The NPP model keeps everything in one workbook, so there is no separate
    Thinkific map. Downstream validators only need ``topic_reference_ids`` to
    accept a detected topic; every NPP topic id is authoritative, so we expose
    the topic ids here and leave the lesson-lookup surfaces empty (Thinkific
    lesson entry is not part of the NPP MVP)."""
    g = normalize_grade(grade)
    master = load_master_content(path, grade=g) if path is not None else get_master(grade=g)
    return {
        "lessons": [],
        "mapped_topics": set(),
        "topic_reference": None,
        "topic_reference_ids": set(master["topic_ids"]),
        "columns": [],
        "source_path": master["source_path"],
        "grade": g,
    }


def validate_mapped_topics(
    master: dict[str, Any] | None = None,
    tmap: dict[str, Any] | None = None,
    grade: Any = DEFAULT_GRADE,
) -> list[str]:
    """Kept for API compatibility. NPP has no separate lesson map, so there are
    never orphaned mapped topics; always returns an empty list."""
    master = master if master is not None else get_master(grade=grade)
    tmap = tmap if tmap is not None else get_thinkific_map(grade=grade)
    topic_ids = master["topic_ids"]
    return sorted(t for t in tmap["mapped_topics"] if t not in topic_ids)


def videos_for_topic(topic_id: Any, master: dict[str, Any] | None = None) -> list[dict[str, str]]:
    """Ordered VIDEO_LINKS rows (lesson_type == 'video') for a topic, or []."""
    master = master if master is not None else get_master()
    tid = normalize_value(topic_id)
    if not tid:
        return []
    return list(master.get("videos_by_topic", {}).get(tid, []))


_lock = threading.Lock()
_master_cache: dict[int, dict[str, Any]] = {}
_tmap_cache: dict[int, dict[str, Any]] = {}


def get_master(
    path: str | Path | None = None,
    reload: bool = False,
    grade: Any = DEFAULT_GRADE,
) -> dict[str, Any]:
    if path is not None:
        return load_master_content(path, grade=grade)
    g = normalize_grade(grade)
    with _lock:
        if g not in _master_cache or reload:
            _master_cache[g] = load_master_content(grade=g)
        return _master_cache[g]


def get_thinkific_map(
    path: str | Path | None = None,
    reload: bool = False,
    grade: Any = DEFAULT_GRADE,
) -> dict[str, Any]:
    if path is not None:
        return load_thinkific_map(path, grade=grade)
    g = normalize_grade(grade)
    with _lock:
        if g in _tmap_cache and not reload:
            return _tmap_cache[g]
    # Build OUTSIDE the lock: load_thinkific_map → get_master takes _lock itself,
    # and threading.Lock is not reentrant (holding it here would deadlock).
    tm = load_thinkific_map(grade=g)
    with _lock:
        _tmap_cache[g] = tm
    return tm
